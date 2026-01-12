"""
TG-Matrix Backend Main Entry Point
Handles communication with Electron via stdin/stdout
"""
import sys
import json
import asyncio
import gc
import time
from typing import Dict, Any, Optional, List
from datetime import datetime, timedelta
from pathlib import Path
from database import db
from config import config
from telegram_client import TelegramClientManager
from ai_context_manager import ai_context
from ai_auto_chat import ai_auto_chat
from knowledge_base import search_engine, document_manager, media_manager
from message_queue import MessageQueue, MessagePriority
from error_handler import init_error_handler, handle_error, AppError, ErrorType
from message_ack import init_ack_manager, get_ack_manager
from backup_manager import BackupManager
from performance_monitor import init_performance_monitor
from validators import (
    validate_account, validate_keyword, validate_template, 
    validate_campaign, validate_group_url,
    AccountValidator, KeywordValidator, TemplateValidator,
    CampaignValidator, GroupValidator, ValidationError
)
from device_fingerprint import DeviceFingerprintGenerator
from proxy_manager import ProxyManager, ProxyConfig
from warmup_manager import WarmupManager
from proxy_rotation_manager import ProxyRotationManager, RotationReason, ProxyRotationConfig
from behavior_simulator import BehaviorSimulator, BehaviorConfig
from enhanced_health_monitor import EnhancedHealthMonitor, Anomaly
from queue_optimizer import QueueOptimizer
try:
    from error_recovery import ErrorRecoveryManager, RecoveryAction, ErrorCategory
except ImportError:
    # 如果 error_recovery 不存在，使用新的實現
    from error_recovery_manager import ErrorRecoveryManager
    RecoveryAction = None
    ErrorCategory = None
from auto_funnel_manager import auto_funnel, AutoFunnelManager
from vector_memory import vector_memory, VectorMemorySystem
from scheduler import scheduler, TaskScheduler
from cache_manager import init_cache_manager, get_cache_manager
from fulltext_search import init_search_engine, get_search_engine
from private_message_poller import private_message_poller
from group_join_service import group_join_service, GroupJoinService
from text_utils import safe_json_dumps, sanitize_text, sanitize_dict, format_chat_info, format_user_info
from connection_monitor import connection_monitor
from resource_discovery import resource_discovery, DiscoveredResource
from group_search_service import group_search_service
from discussion_watcher import discussion_watcher
from batch_operations import init_batch_operations, get_batch_ops, BatchOperationManager
from ad_template import init_ad_template_manager, get_ad_template_manager, SpintaxGenerator
from ad_manager import init_ad_manager, get_ad_manager
from ad_broadcaster import init_ad_broadcaster, get_ad_broadcaster
from ad_scheduler import init_ad_scheduler, get_ad_scheduler
from ad_analytics import init_ad_analytics, get_ad_analytics
from user_tracker import init_user_tracker, get_user_tracker
from user_analytics import init_user_analytics, get_user_analytics
from campaign_orchestrator import init_campaign_orchestrator, get_campaign_orchestrator
from multi_channel_stats import init_multi_channel_stats, get_multi_channel_stats
from multi_role_manager import init_multi_role_manager, get_multi_role_manager
from script_engine import init_script_engine, get_script_engine
from collaboration_coordinator import init_collaboration_coordinator, get_collaboration_coordinator


class BackendService:
    """Main backend service handling commands and events"""
    
    def __init__(self):
        self.is_monitoring = False
        self.running = True
        self.telegram_manager = TelegramClientManager(event_callback=self.send_event)
        self.background_tasks = []  # Track background tasks
        self.last_reset_date = None  # Track last daily reset date
        
        # Initialize message queue (will be connected to database in initialize())
        self.message_queue = None
        self.alert_manager = None
        
        # Proxy rotation manager (will be initialized in initialize())
        self.proxy_rotation_manager: Optional[ProxyRotationManager] = None
        
        # Enhanced health monitor (will be initialized in initialize())
        self.enhanced_health_monitor: Optional[EnhancedHealthMonitor] = None
        
        # Queue optimizer (will be initialized in initialize())
        self.queue_optimizer: Optional[QueueOptimizer] = None
        
        # Error recovery manager (will be initialized in initialize())
        self.error_recovery_manager: Optional[ErrorRecoveryManager] = None
        
        # Backup manager (will be initialized in initialize())
        self.backup_manager: Optional[Any] = None
        
        # Smart alert manager (will be initialized in initialize())
        self.smart_alert_manager: Optional[Any] = None
        
        # Cache for frequently accessed data (TTL: 30 seconds)
        self._cache: Dict[str, Any] = {}
        self._cache_timestamps: Dict[str, datetime] = {}
        self._cache_ttl = timedelta(seconds=30)
    
    def _invalidate_cache(self, cache_key: str):
        """Invalidate a specific cache entry"""
        self._cache.pop(cache_key, None)
        self._cache_timestamps.pop(cache_key, None)
    
    async def initialize(self):
        """Initialize the backend service"""
        # Initialize error handler
        def error_log_callback(error_type: str, message: str, details: Dict[str, Any]):
            """Callback for error logging"""
            log_type = "error"
            if error_type == ErrorType.NETWORK_ERROR.value:
                log_type = "warning"
            elif error_type == ErrorType.VALIDATION_ERROR.value:
                log_type = "warning"
            
            self.send_log(f"[{error_type}] {message}", log_type)
        
        init_error_handler(error_log_callback)
        
        # Initialize acknowledgment manager
        await init_ack_manager()
        
        # Initialize performance monitor
        def performance_event_callback(event_type: str, data: Any):
            """Callback for performance events"""
            if event_type == "performance-metric":
                # Send metric to frontend
                self.send_event("performance-metric", data)
            elif event_type == "performance-alert":
                # Send alert to frontend
                self.send_event("performance-alert", data)
                # Also log as warning
                alerts = data.get("alerts", [])
                if alerts:
                    self.send_log(f"Performance alert: {', '.join(alerts)}", "warning")
        
        performance_monitor = init_performance_monitor(performance_event_callback)
        await performance_monitor.start()
        
        # Initialize cache manager
        cache_manager = init_cache_manager(default_ttl=300)  # 5分鐘默認TTL
        await cache_manager.start_cleanup_task()
        self.send_log("Cache manager initialized", "info")
        
        # Initialize database
        await db.initialize()
        await db.connect()
        
        # Initialize full-text search engine
        try:
            from config import DATABASE_PATH
            search_engine = init_search_engine(str(DATABASE_PATH))
            # 異步重建索引（不阻塞啟動），如果資料庫損壞則跳過
            async def safe_rebuild_index():
                try:
                    await search_engine.rebuild_index()
                except Exception as e:
                    error_str = str(e).lower()
                    if "malformed" in error_str or "corrupt" in error_str or "database disk image" in error_str:
                        import sys
                        print(f"[Backend] Database corruption detected, skipping search index rebuild: {e}", file=sys.stderr)
                        self.send_log("資料庫損壞，跳過搜索索引重建", "warning")
                    else:
                        import sys
                        print(f"[Backend] Error rebuilding search index: {e}", file=sys.stderr)
            asyncio.create_task(safe_rebuild_index())
            self.send_log("全文搜索引擎已初始化", "success")
        except Exception as e:
            import sys
            print(f"[Backend] Failed to initialize search engine: {e}", file=sys.stderr)
            self.send_log(f"全文搜索引擎初始化失敗: {str(e)}", "warning")
        
        # Initialize migration manager (after database is ready)
        from migrations.migration_manager import init_migration_manager, get_migration_manager
        from pathlib import Path
        migrations_dir = Path(__file__).parent / "migrations"
        init_migration_manager(db, migrations_dir)
        migration_manager = get_migration_manager()
        if migration_manager:
            await migration_manager.initialize()
            # Apply pending migrations on startup
            pending = await migration_manager.get_pending_migrations()
            if pending:
                self.send_log(f"Found {len(pending)} pending migration(s), applying...", "info")
                success = await migration_manager.migrate()
                if success:
                    self.send_log("Migrations applied successfully", "success")
                else:
                    self.send_log("Some migrations may have failed", "warning")
        
        # Initialize alert manager (after database is ready)
        from alert_manager import init_alert_manager
        from smart_alert_manager import SmartAlertManager
        def alert_notification_callback(alert):
            """Callback for alert notifications"""
            # Send alert to frontend
            self.send_event("alert-triggered", {
                "id": alert.id,
                "type": alert.alert_type.value,
                "level": alert.level.value,
                "message": alert.message,
                "details": alert.details,
                "timestamp": alert.timestamp.isoformat()
            })
            # Also log
            log_type = "warning" if alert.level.value in ["warning", "info"] else "error"
            self.send_log(f"[告警] {alert.message}", log_type)
        
        self.alert_manager = init_alert_manager(db, alert_notification_callback)
        await self.alert_manager.start()
        
        # 備份管理器已在上面初始化
        # 如果需要啟動時備份，可以在這裡調用
        if self.backup_manager:
            try:
                await self.backup_manager.create_backup(backup_type='startup', compress=True)
                self.send_log("✓ 啟動時備份已創建", "success")
            except Exception as e:
                self.send_log(f"啟動時備份失敗: {str(e)}", "warning")
        
        # Cleanup old backups
        try:
            removed_count = backup_manager.cleanup_old_backups()
            if removed_count > 0:
                self.send_log(f"Cleaned up {removed_count} old backup(s)", "info")
        except Exception as e:
            self.send_log(f"Failed to cleanup old backups: {str(e)}", "warning")
        
        # Rotate and cleanup logs on startup
        try:
            log_rotator = get_log_rotator()
            rotated_files = log_rotator.rotate_all_logs()
            if rotated_files:
                self.send_log(f"Rotated {len(rotated_files)} log file(s)", "info")
            
            removed_logs = log_rotator.cleanup_old_logs()
            if removed_logs > 0:
                self.send_log(f"Cleaned up {removed_logs} old log file(s)", "info")
        except Exception as e:
            self.send_log(f"Failed to rotate/cleanup logs: {str(e)}", "warning")
        
        # Initialize queue optimizer (消息发送队列优化)
        self.queue_optimizer = QueueOptimizer(
            max_batch_size=10,
            batch_interval_seconds=5.0,
            min_send_interval=2.0,
            max_send_interval=10.0
        )
        
        # Initialize message queue with database and optimizer
        self.message_queue = MessageQueue(
            send_callback=self._queue_send_callback,
            database=db,
            optimizer=self.queue_optimizer
        )
        
        # Initialize proxy rotation manager (智能代理轮换)
        await self._initialize_proxy_rotation_manager()
        
        # Initialize enhanced health monitor (账户健康监控增强)
        await self._initialize_enhanced_health_monitor()
        
        # Initialize error recovery manager (错误恢复和自动重试)
        await self._initialize_error_recovery()
        
        # Initialize auto funnel manager (全自动销售漏斗)
        await self._initialize_auto_funnel()
        
        # Initialize AI auto chat (AI自动聊天)
        await self._initialize_ai_auto_chat()
        
        # Initialize vector memory system (向量化记忆系统)
        await self._initialize_vector_memory()
        
        # Initialize task scheduler (自动化任务调度器)
        await self._initialize_scheduler()
        
        # Initialize batch operations manager (批量操作系統)
        await self._initialize_batch_operations()
        
        # Initialize ad system (廣告發送系統)
        await self._initialize_ad_system()
        
        # Initialize user tracking system (用戶追蹤系統)
        await self._initialize_user_tracking()
        
        # Initialize campaign orchestrator (營銷活動協調器)
        await self._initialize_campaign_system()
        
        # Initialize multi-role collaboration system (多角色協作系統)
        await self._initialize_multi_role_system()
        
        # Register private message handlers for already logged-in Sender accounts
        await self._register_existing_sender_handlers()
        
        # Sync leads to user_profiles (one-time migration)
        await self._sync_leads_to_user_profiles()
    
    async def _sync_leads_to_user_profiles(self):
        """同步現有的 leads 到 user_profiles 表"""
        try:
            leads = await db.get_all_leads()
            synced_count = 0
            
            for lead in leads:
                user_id = str(lead.get('userId', lead.get('user_id', '')))
                if not user_id:
                    continue
                    
                # 檢查是否已存在
                profile = await db.get_user_profile(user_id)
                if not profile:
                    await db._connection.execute("""
                        INSERT INTO user_profiles 
                        (user_id, username, first_name, last_name, funnel_stage, interest_level, created_at)
                        VALUES (?, ?, ?, ?, 'new', 1, CURRENT_TIMESTAMP)
                    """, (
                        user_id,
                        lead.get('username', ''),
                        lead.get('firstName', lead.get('first_name', '')),
                        lead.get('lastName', lead.get('last_name', ''))
                    ))
                    synced_count += 1
            
            if synced_count > 0:
                await db._connection.commit()
                self.send_log(f"📊 已同步 {synced_count} 個 Lead 到漏斗系統", "info")
                
        except Exception as e:
            import sys
            print(f"[Backend] Error syncing leads to user_profiles: {e}", file=sys.stderr)
    
    async def _register_existing_sender_handlers(self):
        """為已登錄的發送帳號註冊私信處理器"""
        try:
            accounts = await db.get_all_accounts()
            for account in accounts:
                if account.get('status') == 'Online':
                    phone = account.get('phone')
                    account_role = account.get('role', 'Unassigned')
                    try:
                        await self.telegram_manager.register_private_message_handler(
                            phone=phone,
                            account_role=account_role
                        )
                        self.send_log(f"已為帳號 {phone} ({account_role}) 註冊私信處理器", "info")
                    except Exception as e:
                        self.send_log(f"註冊私信處理器失敗 ({phone}): {e}", "warning")
        except Exception as e:
            self.send_log(f"註冊現有發送帳號處理器錯誤: {e}", "warning")
    
    async def _initialize_auto_funnel(self):
        """Initialize auto funnel manager"""
        try:
            # Set callbacks
            auto_funnel.set_callbacks(
                send_callback=self._funnel_send_callback,
                log_callback=self.send_log,
                event_callback=self.send_event
            )
            
            # Start auto funnel
            await auto_funnel.start()
            self.send_log("[AutoFunnel] 全自动销售漏斗已启动", "success")
        except Exception as e:
            self.send_log(f"[AutoFunnel] 初始化失败: {e}", "error")
    
    async def _initialize_ai_auto_chat(self):
        """Initialize AI auto chat service"""
        try:
            # Initialize AI auto chat
            await ai_auto_chat.initialize()
            
            # Set callbacks
            async def ai_send_callback(account_phone: str, target_user_id: str, 
                                       message: str, source_group: str = None,
                                       username: str = None):
                """AI 自動回復發送回調"""
                try:
                    # 檢查用戶是否已互動（決定是否計入限額）
                    has_interacted = await self._user_has_interacted(target_user_id)
                    
                    # 檢查帳號限額（未互動用戶）
                    if not has_interacted:
                        account = await db.get_account_by_phone(account_phone)
                        if account:
                            if account.get('dailySendCount', 0) >= account.get('dailySendLimit', 50):
                                self.send_log(f"帳號 {account_phone} 已達每日發送限額，無法自動回復", "warning")
                                return False
                    
                    # 使用消息隊列發送
                    await self.message_queue.add_message(
                        phone=account_phone,
                        user_id=target_user_id,
                        text=message,
                        source_group=source_group,
                        priority=MessagePriority.NORMAL
                    )
                    
                    # 更新每日計數（僅未互動用戶）
                    if not has_interacted:
                        account = await db.get_account_by_phone(account_phone)
                        if account:
                            await db.update_account(account.get('id'), {
                                'dailySendCount': account.get('dailySendCount', 0) + 1
                            })
                    
                    # 保存 AI 回復到聊天歷史
                    await db.add_chat_message(
                        user_id=target_user_id,
                        role='assistant',
                        content=message,
                        account_phone=account_phone,
                        source_group=source_group
                    )
                    
                    return True
                except Exception as e:
                    self.send_log(f"AI 自動回復發送失敗: {e}", "error")
                    return False
            
            ai_auto_chat.set_callbacks(
                send_callback=ai_send_callback,
                log_callback=self.send_log,
                event_callback=self.send_event
            )
            
            self.send_log("[AIAutoChat] AI 自動聊天服務已初始化", "success")
        except Exception as e:
            self.send_log(f"[AIAutoChat] 初始化失败: {e}", "error")
    
    async def _funnel_send_callback(self, target_user_id: str, message: str, 
                                     is_follow_up: bool = False):
        """Callback for auto funnel to send messages"""
        try:
            # Find an available sender account
            accounts = await db.get_all_accounts()
            sender = None
            for acc in accounts:
                if acc['status'] == 'Online' and acc['role'] == 'Sender':
                    sender = acc
                    break
            
            if not sender:
                # Find any online account
                for acc in accounts:
                    if acc['status'] == 'Online':
                        sender = acc
                        break
            
            if sender:
                await self.message_queue.add_message(
                    phone=sender['phone'],
                    user_id=target_user_id,
                    text=message,
                    priority=MessagePriority.NORMAL
                )
                return True
            return False
        except Exception as e:
            self.send_log(f"[AutoFunnel] 发送失败: {e}", "error")
            return False
    
    async def _initialize_vector_memory(self):
        """Initialize vector memory system"""
        try:
            await vector_memory.initialize(use_neural=False)  # 默认使用简单嵌入
            self.send_log("[VectorMemory] 向量化记忆系统已启动", "success")
        except Exception as e:
            self.send_log(f"[VectorMemory] 初始化失败: {e}", "error")
    
    async def _initialize_scheduler(self):
        """Initialize task scheduler"""
        try:
            # Set callbacks
            scheduler.set_log_callback(self.send_log)
            scheduler.set_task_callback('follow_up', self._funnel_send_callback)
            
            # Start scheduler
            await scheduler.start()
            self.send_log("[Scheduler] 自动化任务调度器已启动", "success")
        except Exception as e:
            self.send_log(f"[Scheduler] 初始化失败: {e}", "error")
    
    async def _initialize_batch_operations(self):
        """Initialize batch operations manager"""
        try:
            self.batch_ops = await init_batch_operations(db, self.send_event)
            self.send_log("[BatchOps] 批量操作系統已啟動", "success")
        except Exception as e:
            self.send_log(f"[BatchOps] 初始化失敗: {e}", "error")
            self.batch_ops = None
    
    async def _initialize_ad_system(self):
        """Initialize ad system (廣告發送系統)"""
        try:
            # Initialize ad template manager
            await init_ad_template_manager(db)
            
            # Initialize ad manager
            await init_ad_manager(db, self.send_event)
            
            # Initialize ad broadcaster
            init_ad_broadcaster(
                telegram_manager=self.telegram_manager,
                db=db,
                event_callback=self.send_event,
                log_callback=self.send_log
            )
            
            # Initialize ad scheduler
            scheduler = init_ad_scheduler(
                event_callback=self.send_event,
                log_callback=self.send_log
            )
            await scheduler.start()
            
            # Initialize ad analytics
            init_ad_analytics(db)
            
            self.send_log("[AdSystem] 廣告發送系統已啟動", "success")
        except Exception as e:
            self.send_log(f"[AdSystem] 初始化失敗: {e}", "error")
    
    async def _initialize_user_tracking(self):
        """Initialize user tracking system (用戶追蹤系統)"""
        try:
            # Initialize user tracker
            await init_user_tracker(
                db=db,
                telegram_manager=self.telegram_manager,
                event_callback=self.send_event,
                log_callback=self.send_log
            )
            
            # Initialize user analytics
            init_user_analytics(db)
            
            self.send_log("[UserTracker] 用戶追蹤系統已啟動", "success")
        except Exception as e:
            self.send_log(f"[UserTracker] 初始化失敗: {e}", "error")
    
    async def _initialize_campaign_system(self):
        """Initialize campaign orchestrator and multi-channel stats"""
        try:
            # Initialize campaign orchestrator
            await init_campaign_orchestrator(
                db=db,
                event_callback=self.send_event,
                log_callback=self.send_log
            )
            
            # Initialize multi-channel stats
            init_multi_channel_stats(db)
            
            self.send_log("[Campaign] 營銷活動系統已啟動", "success")
        except Exception as e:
            self.send_log(f"[Campaign] 初始化失敗: {e}", "error")
    
    async def _initialize_multi_role_system(self):
        """Initialize multi-role collaboration system"""
        try:
            # Initialize multi-role manager
            await init_multi_role_manager(
                db=db,
                event_callback=self.send_event,
                log_callback=self.send_log
            )
            
            # Initialize script engine
            await init_script_engine(
                db=db,
                event_callback=self.send_event,
                log_callback=self.send_log
            )
            
            # Initialize collaboration coordinator
            await init_collaboration_coordinator(
                db=db,
                telegram_manager=telegram_manager,
                event_callback=self.send_event,
                log_callback=self.send_log
            )
            
            self.send_log("[MultiRole] 多角色協作系統已啟動", "success")
        except Exception as e:
            self.send_log(f"[MultiRole] 初始化失敗: {e}", "error")
    
    async def _initialize_enhanced_health_monitor(self):
        """Initialize enhanced health monitor"""
        try:
            # Create alert callback
            def alert_callback(anomaly: Anomaly):
                """告警回调函数"""
                # 发送告警事件到前端
                self.send_event("health-anomaly-detected", {
                    "account_id": anomaly.account_id,
                    "phone": anomaly.phone,
                    "anomaly_type": anomaly.anomaly_type,
                    "severity": anomaly.severity,
                    "message": anomaly.message,
                    "current_value": anomaly.current_value,
                    "threshold": anomaly.threshold,
                    "timestamp": anomaly.timestamp.isoformat(),
                    "details": anomaly.details
                })
                
                # 记录日志
                log_type = "error" if anomaly.severity == "critical" else "warning"
                self.send_log(f"[健康监控] 账户 {anomaly.phone}: {anomaly.message}", log_type)
                
                # 如果严重，也发送到告警管理器
                if anomaly.severity in ['high', 'critical']:
                    if self.alert_manager:
                        try:
                            from alert_manager import AlertType, AlertLevel
                            alert_type = AlertType.ACCOUNT_HEALTH
                            level = AlertLevel.CRITICAL if anomaly.severity == 'critical' else AlertLevel.WARNING
                            self.alert_manager.create_alert(
                                alert_type=alert_type,
                                level=level,
                                message=anomaly.message,
                                details={
                                    "account_id": anomaly.account_id,
                                    "phone": anomaly.phone,
                                    "anomaly_type": anomaly.anomaly_type,
                                    "current_value": anomaly.current_value,
                                    "threshold": anomaly.threshold,
                                    **anomaly.details
                                }
                            )
                        except Exception as e:
                            import sys
                            print(f"[EnhancedHealthMonitor] Error creating alert: {e}", file=sys.stderr)
            
            # Initialize enhanced health monitor
            self.enhanced_health_monitor = EnhancedHealthMonitor(
                alert_callback=alert_callback,
                check_interval_seconds=300  # 5 分钟检查一次
            )
            
            import sys
            print("[Backend] Enhanced health monitor initialized", file=sys.stderr)
        except Exception as e:
            import sys
            print(f"[Backend] Failed to initialize enhanced health monitor: {e}", file=sys.stderr)
            # Don't fail initialization if health monitor fails
            self.enhanced_health_monitor = None
    
    async def _initialize_proxy_rotation_manager(self):
        """Initialize proxy rotation manager"""
        try:
            # Callback to update account proxy in database
            async def update_proxy_callback(account_id: int, phone: str, new_proxy: str):
                """更新账户代理的回调函数"""
                await db.update_account(account_id, {"proxy": new_proxy})
                import sys
                print(f"[ProxyRotationManager] Updated proxy for account {phone}: {new_proxy[:30]}...", file=sys.stderr)
            
            # Initialize proxy rotation manager with empty pool (will be populated dynamically)
            self.proxy_rotation_manager = ProxyRotationManager(
                proxy_pool=[],  # Empty pool, will be populated from accounts
                config=None,  # Use default config
                health_check_callback=None
            )
            
            # Set update callback
            self.proxy_rotation_manager.update_proxy_callback = update_proxy_callback
            
            import sys
            print("[Backend] Proxy rotation manager initialized", file=sys.stderr)
        except Exception as e:
            import sys
            print(f"[Backend] Failed to initialize proxy rotation manager: {e}", file=sys.stderr)
            # Don't fail initialization if proxy rotation manager fails
            self.proxy_rotation_manager = None
    
    async def _initialize_error_recovery(self):
        """Initialize error recovery manager"""
        try:
            # Reconnect client callback
            async def reconnect_client(account_id: int, phone: str):
                """重新连接客户端的回调函数"""
                try:
                    # Disconnect and reconnect the client
                    if phone in self.telegram_manager.clients:
                        client = self.telegram_manager.clients[phone]
                        if client.is_connected:
                            await client.disconnect()
                        await client.connect()
                        return True
                    return False
                except Exception as e:
                    import sys
                    print(f"[ErrorRecovery] Failed to reconnect client for {phone}: {e}", file=sys.stderr)
                    return False
            
            # Rotate proxy callback
            async def rotate_proxy(account_id: int, phone: str):
                """切换代理的回调函数"""
                if self.proxy_rotation_manager:
                    try:
                        account = await db.get_account(account_id)
                        if account:
                            current_proxy = account.get('proxy')
                            new_proxy = await self.proxy_rotation_manager.rotate_proxy(
                                phone=phone,
                                current_proxy=current_proxy,
                                reason=RotationReason.ERROR
                            )
                            if new_proxy:
                                await db.update_account(account_id, {"proxy": new_proxy})
                                return new_proxy
                    except Exception as e:
                        import sys
                        print(f"[ErrorRecovery] Failed to rotate proxy for {phone}: {e}", file=sys.stderr)
                return None
            
            # Relogin callback
            async def relogin_account(account_id: int, phone: str):
                """重新登录账户的回调函数"""
                try:
                    account = await db.get_account(account_id)
                    if account:
                        result = await self.telegram_manager.login_account(
                            phone=phone,
                            api_id=account.get('apiId'),
                            api_hash=account.get('apiHash'),
                            proxy=account.get('proxy'),
                            two_factor_password=account.get('twoFactorPassword')
                        )
                        return result.get('success', False)
                except Exception as e:
                    import sys
                    print(f"[ErrorRecovery] Failed to relogin account {phone}: {e}", file=sys.stderr)
                return False
            
            # Initialize error recovery manager
            def log_callback(message: str, level: str = "info"):
                self.send_log(f"[錯誤恢復] {message}", level)
            
            # 兼容兩個版本的 ErrorRecoveryManager
            try:
                # 嘗試使用 error_recovery_manager (新版本，接受 log_callback)
                self.error_recovery_manager = ErrorRecoveryManager(log_callback=log_callback)
            except TypeError:
                # 如果失敗，可能是 error_recovery (舊版本，不接受 log_callback)
                # 使用舊版本的參數
                self.error_recovery_manager = ErrorRecoveryManager()
            
            import sys
            print("[Backend] Error recovery manager initialized", file=sys.stderr)
        except Exception as e:
            import sys
            print(f"[Backend] Failed to initialize error recovery manager: {e}", file=sys.stderr)
            # Don't fail initialization if error recovery manager fails
            self.error_recovery_manager = None
        
        # Initialize backup manager
        try:
            from backup_manager import BackupManager
            from config import DATABASE_PATH
            
            backup_dir = Path(DATABASE_PATH).parent / "backups"
            self.backup_manager = BackupManager(
                db_path=Path(DATABASE_PATH),
                backup_dir=backup_dir,
                log_callback=lambda msg, level="info": self.send_log(f"[備份] {msg}", level)
            )
            
            # 啟動定期備份（每24小時一次）
            await self.backup_manager.start_scheduled_backups(interval_hours=24)
            
            import sys
            print("[Backend] Backup manager initialized", file=sys.stderr)
            self.send_log("備份管理器已初始化（每24小時自動備份）", "success")
        except Exception as e:
            import sys
            print(f"[Backend] Failed to initialize backup manager: {e}", file=sys.stderr)
            self.backup_manager = None
        
        # Initialize smart alert manager
        try:
            from smart_alert_manager import SmartAlertManager
            self.smart_alert_manager = SmartAlertManager(db)
            import sys
            print("[Backend] Smart alert manager initialized", file=sys.stderr)
            self.send_log("智能告警管理器已初始化", "success")
        except Exception as e:
            import sys
            print(f"[Backend] Failed to initialize smart alert manager: {e}", file=sys.stderr)
            self.smart_alert_manager = None
    
    async def _user_has_interacted(self, user_id: str) -> bool:
        """
        檢查用戶是否已互動過（用於發送限額豁免）
        
        Args:
            user_id: 用戶ID
        
        Returns:
            True 如果用戶已互動過，False 否則
        """
        try:
            # 檢查聊天歷史中是否有用戶發送的消息
            cursor = await db._connection.execute("""
                SELECT COUNT(*) as count FROM chat_history 
                WHERE user_id = ? AND role = 'user'
            """, (user_id,))
            row = await cursor.fetchone()
            user_message_count = row['count'] if row else 0
            
            # 如果用戶發送過至少一條消息，視為已互動
            return user_message_count > 0
        except Exception as e:
            import sys
            print(f"[Backend] Error checking user interaction: {e}", file=sys.stderr)
            return False
    
    async def _queue_send_callback(self, phone: str, user_id: str, text: str, attachment: Optional[str] = None, source_group: Optional[str] = None, target_username: Optional[str] = None) -> Dict[str, Any]:
        """
        Callback function for MessageQueue to actually send messages via Telegram
        
        Args:
            phone: Account phone number
            user_id: Target user ID
            text: Message text
            attachment: Optional attachment path
            source_group: Optional source group ID/URL
            target_username: Optional target username (fallback)
            
        Returns:
            Dict with 'success' (bool) and optionally 'error' (str)
        """
        import sys
        print(f"[Backend] _queue_send_callback called: phone={phone}, user_id={user_id}, source_group={source_group}, target_username={target_username}, text={text[:50]}...", file=sys.stderr)
        self.send_log(f"正在發送消息到 {target_username or user_id}...", "info")
        
        try:
            # Check Warmup status before sending (防封)
            account = await db.get_account_by_phone(phone)
            if account:
                # Determine message type (simplified: assume "active" for now)
                message_type = "active"  # Could be "reply_only" if replying to a message
                
                # Check if sending is allowed
                warmup_check = WarmupManager.should_allow_send(account, message_type)
                
                if not warmup_check.get('allowed'):
                    reason = warmup_check.get('reason', 'Unknown reason')
                    stage_info = warmup_check.get('current_stage')
                    
                    import sys
                    print(f"[Backend] Warmup check failed for {phone}: {reason}", file=sys.stderr)
                    if stage_info:
                        print(f"[Backend] Current stage: {stage_info.get('stage_name')} (Stage {stage_info.get('stage')})", file=sys.stderr)
                        print(f"[Backend] Daily limit: {warmup_check.get('daily_limit')}", file=sys.stderr)
                    
                    return {
                        "success": False,
                        "error": f"Warmup限制: {reason}",
                        "warmup_info": warmup_check
                    }
            
            # Send message via Pyrogram
            import time
            send_start_time = time.time()
            
            result = await self.telegram_manager.send_message(
                phone=phone,
                user_id=user_id,
                text=text,
                attachment=attachment,
                source_group=source_group,
                target_username=target_username
            )
            
            send_latency = (time.time() - send_start_time) * 1000  # 转换为毫秒
            
            print(f"[Backend] telegram_manager.send_message result: {result}", file=sys.stderr)
            
            if result.get('success'):
                self.send_log(f"✓ 消息發送成功到 {user_id}", "success")
                # Record send performance
                from performance_monitor import get_performance_monitor
                try:
                    monitor = get_performance_monitor()
                    monitor.record_send_performance(phone, send_latency)
                except:
                    pass  # Performance monitor might not be initialized
                
                # Record health metrics (账户健康监控增强)
                if self.enhanced_health_monitor:
                    account = await db.get_account_by_phone(phone)
                    if account:
                        account_id = account.get('id')
                        self.enhanced_health_monitor.record_send_success(account_id, phone, send_latency)
                
                # Record proxy success (智能代理轮换)
                if self.proxy_rotation_manager:
                    account = await db.get_account_by_phone(phone)
                    if account:
                        current_proxy = account.get('proxy')
                        if current_proxy:
                            self.proxy_rotation_manager.record_proxy_success(current_proxy, send_latency)
                
                return result
            else:
                # Handle flood wait
                error = result.get('error', 'Unknown error')
                self.send_log(f"✗ 消息發送失敗: {error}", "error")
                print(f"[Backend] Message send failed: {error}", file=sys.stderr)
                
                # Record proxy error (智能代理轮换)
                if self.proxy_rotation_manager:
                    account = await db.get_account_by_phone(phone)
                    if account:
                        account_id = account.get('id')
                        current_proxy = account.get('proxy')
                        if current_proxy:
                            self.proxy_rotation_manager.record_proxy_error(current_proxy, error)
                            
                            # 如果是代理错误，尝试自动轮换
                            if 'Proxy' in error or 'proxy' in error or 'Connection' in error:
                                try:
                                    new_proxy = await self.proxy_rotation_manager.rotate_proxy(
                                        account_id=account_id,
                                        phone=phone,
                                        reason=RotationReason.ERROR,
                                        preferred_country=account.get('proxyCountry')
                                    )
                                    if new_proxy and new_proxy != current_proxy:
                                        # 更新数据库中的代理
                                        await db.update_account(account_id, {'proxy': new_proxy})
                                        self.send_log(f"账户 {phone} 代理已自动轮换: {current_proxy[:30]}... -> {new_proxy[:30]}...", "info")
                                except Exception as e:
                                    import sys
                                    print(f"[Backend] Failed to auto-rotate proxy: {e}", file=sys.stderr)
                
                # Handle error with recovery manager (错误恢复和自动重试机制)
                account = await db.get_account_by_phone(phone)
                account_id = account.get('id') if account else None
                
                if account_id and self.error_recovery_manager:
                    try:
                        # 处理错误并执行恢复动作
                        error_exception = Exception(error)
                        recovery_result = await self.error_recovery_manager.handle_error(
                            account_id=str(account_id),
                            phone=phone,
                            error=error_exception,
                            attempt=0,  # 这里应该从消息队列获取实际尝试次数
                            context={
                                "user_id": user_id,
                                "message_text": text[:100] if text else None
                            }
                        )
                        
                        # 记录恢复结果
                        if recovery_result.success:
                            self.error_recovery_manager.record_recovery_success(str(account_id), recovery_result.action_taken)
                            if recovery_result.action_taken != RecoveryAction.RETRY:
                                self.send_log(f"账户 {phone} 错误恢复成功: {recovery_result.message}", "info")
                        else:
                            self.error_recovery_manager.record_recovery_failure(str(account_id), recovery_result.action_taken)
                            self.send_log(f"账户 {phone} 错误恢复失败: {recovery_result.message}", "warning")
                        
                        # 如果需要等待，更新结果中的错误信息
                        if recovery_result.retry_after:
                            result['retry_after'] = recovery_result.retry_after
                            result['recovery_action'] = recovery_result.action_taken.value
                    except Exception as e:
                        import sys
                        print(f"[Backend] Error in error recovery: {e}", file=sys.stderr)
                
                # Record health metrics (账户健康监控增强)
                if self.enhanced_health_monitor and account:
                    account_id = account.get('id')
                    self.enhanced_health_monitor.record_send_failure(account_id, phone, error, send_latency)
                
                if 'Flood wait' in error:
                    # Extract wait time from error message
                    import re
                    wait_match = re.search(r'wait (\d+) seconds', error)
                    if wait_match:
                        wait_seconds = int(wait_match.group(1))
                        
                        # Record Flood Wait (账户健康监控增强)
                        if self.enhanced_health_monitor:
                            account = await db.get_account_by_phone(phone)
                            if account:
                                account_id = account.get('id')
                                self.enhanced_health_monitor.record_flood_wait(account_id, phone, wait_seconds)
                        
                        # Update rate limiter in message queue
                        if self.message_queue and phone in self.message_queue.rate_limiters:
                            await self.message_queue.rate_limiters[phone].set_flood_wait(wait_seconds)
                
                return result
                
        except Exception as e:
            error_msg = str(e)
            # Provide user-friendly error messages
            if "not connected" in error_msg.lower() or "client not" in error_msg.lower():
                friendly_msg = f"账户 {phone} 未连接。请先登录该账户。"
            elif "flood" in error_msg.lower():
                friendly_msg = f"账户 {phone} 触发限流保护。系统将自动等待后重试。"
            elif "banned" in error_msg.lower() or "deactivated" in error_msg.lower():
                friendly_msg = f"账户 {phone} 可能被封禁或已停用。请检查账户状态。"
            else:
                friendly_msg = f"发送消息失败 ({phone}): {error_msg}"
            
            self.send_log(friendly_msg, "error")
            return {
                "success": False,
                "error": friendly_msg
            }
        
        # Restore pending messages from database
        await self.message_queue.restore_from_database()
    
    async def _start_browsing_simulation(self, account_id: int, phone: str, group_urls: List[str]):
        """
        启动浏览行为模拟后台任务
        
        Args:
            account_id: 账户 ID
            phone: 电话号码
            group_urls: 群组 URL 列表
        """
        async def browsing_task():
            """浏览行为模拟任务"""
            try:
                # 获取行为模拟器
                behavior_simulator = self.telegram_manager.behavior_simulator
                
                # 获取客户端
                if phone not in self.telegram_manager.clients:
                    return
                client = self.telegram_manager.clients[phone]
                
                # 转换群组 URL 为 ID
                group_ids = []
                for group_url in group_urls:
                    try:
                        if isinstance(group_url, (int, str)) and str(group_url).lstrip('-').isdigit():
                            group_ids.append(int(group_url))
                        else:
                            chat = await client.get_chat(group_url)
                            group_ids.append(chat.id)
                    except Exception:
                        continue
                
                if not group_ids:
                    return
                
                # 持续运行浏览模拟
                while self.running:
                    try:
                        # 检查是否应该浏览
                        if behavior_simulator.should_browse_now(account_id):
                            # 模拟浏览
                            browse_result = await behavior_simulator.simulate_browsing(
                                client=client,
                                account_id=account_id,
                                group_ids=group_ids
                            )
                            
                            if browse_result.get('success'):
                                import sys
                                print(f"[BehaviorSimulator] Account {phone} browsed {browse_result.get('count', 0)} groups", file=sys.stderr)
                        
                        # 等待下次浏览（30-60 分钟）
                        delay = behavior_simulator.get_random_activity_delay()
                        await asyncio.sleep(delay)
                    
                    except asyncio.CancelledError:
                        break
                    except Exception as e:
                        import sys
                        print(f"[BehaviorSimulator] Error in browsing task for {phone}: {e}", file=sys.stderr)
                        # 等待一段时间后重试
                        await asyncio.sleep(300)  # 5 分钟后重试
            
            except Exception as e:
                import sys
                print(f"[BehaviorSimulator] Browsing task failed for {phone}: {e}", file=sys.stderr)
        
        # 启动后台任务
        task = asyncio.create_task(browsing_task())
        self.background_tasks.append(task)
        import sys
        print(f"[BehaviorSimulator] Started browsing simulation for account {phone}", file=sys.stderr)
        
        # Start background tasks
        self.background_tasks.append(asyncio.create_task(self.daily_reset_task()))
        self.background_tasks.append(asyncio.create_task(self.account_health_monitor_task()))
        self.background_tasks.append(asyncio.create_task(self.queue_cleanup_task()))
        self.background_tasks.append(asyncio.create_task(self.message_confirmation_timeout_task()))
        
        # Log startup
        await db.add_log("Backend service started", "info")
        print(safe_json_dumps({
            "event": "log-entry",
            "payload": {
                "id": int(datetime.now().timestamp() * 1000),
                "timestamp": datetime.now().isoformat() + "Z",
                "message": "Backend service started",
                "type": "info"
            }
        }), flush=True)
    
    async def safe_delete_session_file(self, session_path: Path, max_retries: int = 5, retry_delay: float = 0.5) -> bool:
        """
        Safely delete a session file with retry mechanism for Windows file locking issues
        
        Args:
            session_path: Path to the session file to delete
            max_retries: Maximum number of retry attempts
            retry_delay: Delay between retries in seconds
            
        Returns:
            True if file was deleted successfully, False otherwise
        """
        if not session_path.exists():
            return True  # File doesn't exist, consider it "deleted"
        
        import sys
        for attempt in range(max_retries):
            try:
                session_path.unlink()
                print(f"[Backend] Successfully deleted session file: {session_path} (attempt {attempt + 1})", file=sys.stderr)
                return True
            except PermissionError as e:
                if attempt < max_retries - 1:
                    print(f"[Backend] Session file locked, retrying in {retry_delay}s (attempt {attempt + 1}/{max_retries}): {session_path}", file=sys.stderr)
                    # Force garbage collection to release file handles
                    gc.collect()
                    await asyncio.sleep(retry_delay)
                    retry_delay *= 1.5  # Exponential backoff
                else:
                    print(f"[Backend] Failed to delete session file after {max_retries} attempts: {e}", file=sys.stderr)
                    return False
            except Exception as e:
                print(f"[Backend] Error deleting session file: {e}", file=sys.stderr)
                return False
        
        return False
    
    async def shutdown(self):
        """Shutdown the backend service"""
        self.running = False
        
        # Stop acknowledgment manager
        try:
            ack_manager = get_ack_manager()
            await ack_manager.stop()
        except Exception as e:
            print(f"[Backend] Error stopping ack manager: {e}", file=sys.stderr)
        
        # Disconnect all Telegram clients
        try:
            await self.telegram_manager.disconnect_all()
        except Exception as e:
            print(f"[Backend] Error disconnecting clients: {e}", file=sys.stderr)
        
        # Try to log shutdown (only if database is still connected)
        try:
            if db._connection is not None:
                await db.add_log("Backend service shutting down", "info")
                await db.close()
        except Exception as e:
            print(f"[Backend] Error during database shutdown: {e}", file=sys.stderr)
    
    def send_event(self, event_name: str, payload: Any, message_id: Optional[str] = None):
        """
        Send an event to Electron via stdout
        
        Args:
            event_name: Event name
            payload: Event payload
            message_id: Optional message ID for confirmation
        """
        message = {
            "event": event_name,
            "payload": payload
        }
        if message_id:
            message["message_id"] = message_id
        # 使用安全的 JSON 序列化，處理 emoji 和特殊字符
        try:
            print(safe_json_dumps(message), flush=True)
        except Exception as e:
            # 最後的備用方案：強制 ASCII 編碼
            print(json.dumps(sanitize_dict(message), ensure_ascii=True, default=str), flush=True)
    
    def send_log(self, message: str, log_type: str = "info"):
        """Send a log entry event"""
        log_entry = {
            "id": int(datetime.now().timestamp() * 1000),
            "timestamp": datetime.now().isoformat() + "Z",
            "message": message,
            "type": log_type
        }
        self.send_event("log-entry", log_entry)
    
    async def handle_command(self, command: str, payload: Any, request_id: Optional[str] = None):
        """Handle incoming commands"""
        try:
            # Register request for acknowledgment if request_id is provided
            ack_manager = get_ack_manager()
            if request_id:
                # Register the request
                await ack_manager.register_request(
                    command=command,
                    payload=payload,
                    callback=None,  # No callback needed for simple acknowledgment
                    timeout_seconds=30.0,
                    max_retries=0  # Don't retry, just acknowledge
                )
                
                # Send immediate acknowledgment
                self.send_event("command-ack", {
                    "request_id": request_id,
                    "command": command,
                    "status": "received"
                })
            
            # Route commands to handlers
            if command == "graceful-shutdown":
                await self.handle_graceful_shutdown()
                return  # Don't continue processing after shutdown
            
            elif command == "get-initial-state":
                await self.handle_get_initial_state()
            
            elif command == "add-account":
                await self.handle_add_account(payload)
            
            elif command == "login-account":
                await self.handle_login_account(payload)
            
            elif command == "check-account-status":
                await self.handle_check_account_status(payload)
            
            elif command == "update-account-data":
                await self.handle_update_account_data(payload)
            
            elif command == "bulk-assign-role":
                await self.handle_bulk_assign_role(payload)
            
            elif command == "bulk-assign-group":
                await self.handle_bulk_assign_group(payload)
            
            elif command == "bulk-delete-accounts":
                await self.handle_bulk_delete_accounts(payload)
            
            elif command == "remove-account":
                await self.handle_remove_account(payload)
            
            elif command == "start-monitoring":
                await self.handle_start_monitoring()
            
            elif command == "stop-monitoring":
                await self.handle_stop_monitoring()
            
            elif command == "one-click-start":
                await self.handle_one_click_start(payload)
            
            elif command == "one-click-stop":
                await self.handle_one_click_stop()
            
            elif command == "get-system-status":
                await self.handle_get_system_status()
            
            elif command == "learn-from-history":
                await self.handle_learn_from_history(payload)
            
            elif command == "get-knowledge-stats":
                await self.handle_get_knowledge_stats()
            
            elif command == "search-knowledge":
                await self.handle_search_knowledge(payload)
            
            elif command == "add-keyword-set":
                await self.handle_add_keyword_set(payload)
            
            elif command == "remove-keyword-set":
                await self.handle_remove_keyword_set(payload)
            
            elif command == "add-keyword":
                await self.handle_add_keyword(payload)
            
            elif command == "remove-keyword":
                await self.handle_remove_keyword(payload)
            
            elif command == "add-group":
                await self.handle_add_group(payload)
            
            elif command == "join-group":
                await self.handle_join_group(payload)
            
            elif command == "remove-group":
                await self.handle_remove_group(payload)
            
            elif command == "add-template":
                await self.handle_add_template(payload)
            
            elif command == "remove-template":
                await self.handle_remove_template(payload)
            
            elif command == "toggle-template-status":
                await self.handle_toggle_template_status(payload)
            
            elif command == "add-campaign":
                await self.handle_add_campaign(payload)
            
            elif command == "remove-campaign":
                await self.handle_remove_campaign(payload)
            
            elif command == "toggle-campaign-status":
                await self.handle_toggle_campaign_status(payload)
            
            elif command == "send-message":
                await self.handle_send_message(payload)
            
            elif command == "update-lead-status":
                await self.handle_update_lead_status(payload)
            
            elif command == "add-to-dnc":
                await self.handle_add_to_dnc(payload)
            
            elif command == "clear-logs":
                await self.handle_clear_logs()
            
            elif command == "load-accounts-from-excel":
                await self.handle_load_accounts_from_excel(payload)
            
            elif command == "export-leads-to-excel":
                await self.handle_export_leads_to_excel(payload)
            
            elif command == "reload-sessions-and-accounts":
                await self.handle_reload_sessions_and_accounts()
            elif command == "import-session":
                await self.handle_import_session(payload)
            elif command == "cleanup-session-files":
                await self.handle_cleanup_session_files()
            elif command == "cleanup-session-files":
                await self.handle_cleanup_session_files()
            elif command == "export-session":
                await self.handle_export_session(payload)
            elif command == "export-sessions-batch":
                await self.handle_export_sessions_batch(payload)
            
            elif command == "save-settings":
                await self.handle_save_settings(payload)
            
            elif command == "get-settings":
                await self.handle_get_settings()
            
            elif command == "get-queue-status":
                await self.handle_get_queue_status(payload)
            
            elif command == "clear-queue":
                await self.handle_clear_queue(payload)
            
            elif command == "pause-queue":
                await self.handle_pause_queue(payload)
            
            elif command == "resume-queue":
                await self.handle_resume_queue(payload)
            
            elif command == "delete-queue-message":
                await self.handle_delete_queue_message(payload)
            
            elif command == "update-queue-message-priority":
                await self.handle_update_queue_message_priority(payload)
            
            elif command == "get-queue-messages":
                await self.handle_get_queue_messages(payload)
            
            elif command == "get-logs":
                await self.handle_get_logs(payload)
            
            elif command == "export-logs":
                await self.handle_export_logs(payload)
            
            elif command == "create-backup":
                await self.handle_create_backup(payload)
            
            elif command == "restore-backup":
                await self.handle_restore_backup(payload)
            
            elif command == "list-backups":
                await self.handle_list_backups()
            
            elif command == "get-backup-info":
                await self.handle_get_backup_info()
            
            # ==================== Full-Text Search Commands ====================
            elif command == "search-chat-history":
                await self.handle_search_chat_history(payload)
            
            elif command == "search-leads":
                await self.handle_search_leads(payload)
            
            elif command == "rebuild-search-index":
                await self.handle_rebuild_search_index()
            
            elif command == "rotate-logs":
                await self.handle_rotate_logs()
            
            elif command == "get-log-stats":
                await self.handle_get_log_stats()
            
            elif command == "list-log-files":
                await self.handle_list_log_files()
            
            elif command == "get-performance-metrics":
                await self.handle_get_performance_metrics(payload)
            
            elif command == "get-performance-summary":
                await self.handle_get_performance_summary()
            
            elif command == "get-sending-stats":
                await self.handle_get_sending_stats(payload)
            
            elif command == "get-queue-length-history":
                await self.handle_get_queue_length_history(payload)
            
            elif command == "get-account-sending-comparison":
                await self.handle_get_account_sending_comparison(payload)
            
            elif command == "get-campaign-performance-stats":
                await self.handle_get_campaign_performance_stats(payload)
            
            elif command == "get-alerts":
                await self.handle_get_alerts(payload)
            
            elif command == "acknowledge-alert":
                await self.handle_acknowledge_alert(payload)
            
            elif command == "resolve-alert":
                await self.handle_resolve_alert(payload)
            
            elif command == "migration-status":
                await self.handle_migration_status(payload)
            
            elif command == "migrate":
                await self.handle_migrate(payload)
            
            elif command == "rollback-migration":
                await self.handle_rollback_migration(payload)
            
            # Local AI & Voice Services
            elif command == "test-local-ai":
                await self.handle_test_local_ai(payload)
            
            elif command == "test-tts-service":
                await self.handle_test_tts_service(payload)
            
            elif command == "test-stt-service":
                await self.handle_test_stt_service(payload)
            
            elif command == "save-ai-settings":
                await self.handle_save_ai_settings(payload)
            
            elif command == "generate-with-local-ai":
                await self.handle_generate_with_local_ai(payload)
            
            elif command == "text-to-speech":
                await self.handle_text_to_speech(payload)
            
            elif command == "speech-to-text":
                await self.handle_speech_to_text(payload)
            
            # Voice Clone Commands
            elif command == "upload-voice-sample":
                await self.handle_upload_voice_sample(payload)
            
            elif command == "delete-voice-sample":
                await self.handle_delete_voice_sample(payload)
            
            elif command == "preview-voice-sample":
                await self.handle_preview_voice_sample(payload)
            
            elif command == "generate-cloned-voice":
                await self.handle_generate_cloned_voice(payload)
            
            elif command == "list-voice-samples":
                await self.handle_list_voice_samples()
            
            # AI Auto Chat Commands
            elif command == "get-ai-chat-settings":
                await self.handle_get_ai_chat_settings()
            
            elif command == "update-ai-chat-settings":
                await self.handle_update_ai_chat_settings(payload)
            
            elif command == "get-chat-history":
                await self.handle_get_chat_history(payload)
            
            elif command == "get-user-context":
                await self.handle_get_user_context(payload)
            
            elif command == "generate-ai-response":
                await self.handle_generate_ai_response(payload)
            
            elif command == "add-ai-memory":
                await self.handle_add_ai_memory(payload)
            
            elif command == "get-ai-memories":
                await self.handle_get_ai_memories(payload)
            
            elif command == "analyze-conversation":
                await self.handle_analyze_conversation(payload)
            
            # Knowledge Base Commands
            elif command == "init-knowledge-base":
                await self.handle_init_knowledge_base()
            
            elif command == "get-knowledge-stats":
                await self.handle_get_knowledge_stats()
            
            elif command == "add-document":
                await self.handle_add_document(payload)
            
            elif command == "get-documents":
                await self.handle_get_documents(payload)
            
            elif command == "delete-document":
                await self.handle_delete_document(payload)
            
            elif command == "add-media":
                await self.handle_add_media(payload)
            
            elif command == "get-media":
                await self.handle_get_media(payload)
            
            elif command == "delete-media":
                await self.handle_delete_media(payload)
            
            elif command == "search-knowledge":
                await self.handle_search_knowledge(payload)
            
            elif command == "add-qa-pair":
                await self.handle_add_qa_pair(payload)
            
            elif command == "get-qa-pairs":
                await self.handle_get_qa_pairs(payload)
            
            elif command == "import-qa":
                await self.handle_import_qa(payload)
            
            elif command == "get-rag-context":
                await self.handle_get_rag_context(payload)
            
            # ==================== Auto Funnel Commands ====================
            elif command == "get-funnel-overview":
                await self.handle_get_funnel_overview()
            
            elif command == "analyze-user-message":
                await self.handle_analyze_user_message(payload)
            
            elif command == "transition-funnel-stage":
                await self.handle_transition_funnel_stage(payload)
            
            elif command == "get-user-journey":
                await self.handle_get_user_journey(payload)
            
            elif command == "batch-update-stages":
                await self.handle_batch_update_stages(payload)
            
            # ==================== Vector Memory Commands ====================
            elif command == "add-vector-memory":
                await self.handle_add_vector_memory(payload)
            
            elif command == "search-vector-memories":
                await self.handle_search_vector_memories(payload)
            
            elif command == "get-memory-context":
                await self.handle_get_memory_context(payload)
            
            elif command == "summarize-conversation":
                await self.handle_summarize_conversation(payload)
            
            elif command == "get-memory-stats":
                await self.handle_get_memory_stats(payload)
            
            # ==================== Telegram RAG Commands ====================
            elif command == "init-rag-system":
                await self.handle_init_rag_system()
            
            elif command == "get-rag-stats":
                await self.handle_get_rag_stats()
            
            elif command == "search-rag":
                await self.handle_search_rag(payload)
            
            elif command == "trigger-rag-learning":
                await self.handle_trigger_rag_learning(payload)
            
            elif command == "add-rag-knowledge":
                await self.handle_add_rag_knowledge(payload)
            
            elif command == "rag-feedback":
                await self.handle_rag_feedback(payload)
            
            elif command == "reindex-conversations":
                await self.handle_reindex_conversations(payload)
            
            elif command == "cleanup-rag-knowledge":
                await self.handle_cleanup_rag_knowledge(payload)
            
            # ==================== Resource Discovery Commands ====================
            elif command == "init-resource-discovery":
                await self.handle_init_resource_discovery()
            
            elif command == "search-resources":
                await self.handle_search_resources(payload)
            
            elif command == "get-resources":
                await self.handle_get_resources(payload)
            
            elif command == "get-resource-stats":
                await self.handle_get_resource_stats()
            
            elif command == "add-resource-manually":
                await self.handle_add_resource_manually(payload)
            
            elif command == "delete-resource":
                await self.handle_delete_resource(payload)
            
            elif command == "add-to-join-queue":
                await self.handle_add_to_join_queue(payload)
            
            elif command == "process-join-queue":
                await self.handle_process_join_queue(payload)
            
            elif command == "batch-join-resources":
                await self.handle_batch_join_resources(payload)
            
            elif command == "join-and-monitor-resource":
                await self.handle_join_and_monitor_resource(payload)
            
            elif command == "batch-join-and-monitor":
                await self.handle_batch_join_and_monitor(payload)
            
            elif command == "analyze-group-link":
                await self.handle_analyze_group_link(payload)
            
            elif command == "get-ollama-models":
                await self.handle_get_ollama_models(payload)
            
            elif command == "test-ollama-connection":
                await self.handle_test_ollama_connection(payload)
            
            elif command == "ollama-generate":
                await self.handle_ollama_generate(payload)
            
            elif command == "get-discovery-keywords":
                await self.handle_get_discovery_keywords()
            
            elif command == "add-discovery-keyword":
                await self.handle_add_discovery_keyword(payload)
            
            elif command == "get-discovery-logs":
                await self.handle_get_discovery_logs(payload)
            
            # ==================== Discussion Watcher Commands ====================
            elif command == "init-discussion-watcher":
                await self.handle_init_discussion_watcher()
            
            elif command == "discover-discussion":
                await self.handle_discover_discussion(payload)
            
            elif command == "discover-discussions-from-resources":
                await self.handle_discover_discussions_from_resources()
            
            elif command == "get-channel-discussions":
                await self.handle_get_channel_discussions(payload)
            
            elif command == "start-discussion-monitoring":
                await self.handle_start_discussion_monitoring(payload)
            
            elif command == "stop-discussion-monitoring":
                await self.handle_stop_discussion_monitoring(payload)
            
            elif command == "get-discussion-messages":
                await self.handle_get_discussion_messages(payload)
            
            elif command == "reply-to-discussion":
                await self.handle_reply_to_discussion(payload)
            
            elif command == "get-discussion-stats":
                await self.handle_get_discussion_stats()
            
            # ==================== Scheduler Commands ====================
            elif command == "schedule-follow-up":
                await self.handle_schedule_follow_up(payload)
            
            elif command == "get-pending-tasks":
                await self.handle_get_pending_tasks(payload)
            
            elif command == "cancel-scheduled-task":
                await self.handle_cancel_scheduled_task(payload)
            
            elif command == "get-scheduler-stats":
                await self.handle_get_scheduler_stats()
            
            # ==================== Monitoring Status Commands ====================
            elif command == "get-monitoring-status":
                await self.handle_get_monitoring_status()
            
            elif command == "check-monitoring-health":
                await self.handle_check_monitoring_health()
            
            # ==================== User CRM Commands ====================
            elif command == "get-user-profile":
                await self.handle_get_user_profile_full(payload)
            
            elif command == "update-user-crm":
                await self.handle_update_user_crm(payload)
            
            elif command == "add-user-tag":
                await self.handle_add_user_tag(payload)
            
            elif command == "remove-user-tag":
                await self.handle_remove_user_tag(payload)
            
            elif command == "get-user-tags":
                await self.handle_get_user_tags(payload)
            
            elif command == "rebuild-database":
                await self.handle_rebuild_database()
            
            # ==================== Chat History Commands ====================
            elif command == "get-chat-history-full":
                await self.handle_get_chat_history_full(payload)
            
            elif command == "get-chat-list":
                await self.handle_get_chat_list(payload)
            
            elif command == "send-ai-response":
                await self.handle_send_ai_response(payload)
            
            # ==================== User Management Commands ====================
            elif command == "get-users-with-profiles":
                await self.handle_get_users_with_profiles(payload)
            
            elif command == "get-funnel-stats":
                await self.handle_get_funnel_stats()
            
            elif command == "bulk-update-user-tags":
                await self.handle_bulk_update_user_tags(payload)
            
            elif command == "bulk-update-user-stage":
                await self.handle_bulk_update_user_stage(payload)
            
            elif command == "update-user-profile":
                await self.handle_update_user_profile(payload)
            
            # ==================== Batch Operations Commands ====================
            elif command == "batch-update-lead-status":
                await self.handle_batch_update_lead_status(payload)
            
            elif command == "batch-add-tag":
                await self.handle_batch_add_tag(payload)
            
            elif command == "batch-remove-tag":
                await self.handle_batch_remove_tag(payload)
            
            elif command == "batch-add-to-dnc":
                await self.handle_batch_add_to_dnc(payload)
            
            elif command == "batch-remove-from-dnc":
                await self.handle_batch_remove_from_dnc(payload)
            
            elif command == "batch-update-funnel-stage":
                await self.handle_batch_update_funnel_stage(payload)
            
            elif command == "batch-delete-leads":
                await self.handle_batch_delete_leads(payload)
            
            elif command == "undo-batch-operation":
                await self.handle_undo_batch_operation(payload)
            
            elif command == "get-batch-operation-history":
                await self.handle_get_batch_operation_history(payload)
            
            elif command == "get-all-tags":
                await self.handle_get_all_tags()
            
            elif command == "create-tag":
                await self.handle_create_tag(payload)
            
            elif command == "delete-tag":
                await self.handle_delete_tag(payload)
            
            elif command == "get-lead-tags":
                await self.handle_get_lead_tags(payload)
            
            # ==================== Ad System Commands (廣告發送系統) ====================
            # Ad Templates
            elif command == "create-ad-template":
                await self.handle_create_ad_template(payload)
            
            elif command == "update-ad-template":
                await self.handle_update_ad_template(payload)
            
            elif command == "delete-ad-template":
                await self.handle_delete_ad_template(payload)
            
            elif command == "get-ad-templates":
                await self.handle_get_ad_templates(payload)
            
            elif command == "toggle-ad-template-status":
                await self.handle_toggle_ad_template_status(payload)
            
            elif command == "preview-ad-template":
                await self.handle_preview_ad_template(payload)
            
            elif command == "validate-spintax":
                await self.handle_validate_spintax(payload)
            
            # Ad Schedules
            elif command == "create-ad-schedule":
                await self.handle_create_ad_schedule(payload)
            
            elif command == "update-ad-schedule":
                await self.handle_update_ad_schedule(payload)
            
            elif command == "delete-ad-schedule":
                await self.handle_delete_ad_schedule(payload)
            
            elif command == "get-ad-schedules":
                await self.handle_get_ad_schedules(payload)
            
            elif command == "toggle-ad-schedule-status":
                await self.handle_toggle_ad_schedule_status(payload)
            
            elif command == "run-ad-schedule-now":
                await self.handle_run_ad_schedule_now(payload)
            
            # Ad Sending
            elif command == "send-ad-now":
                await self.handle_send_ad_now(payload)
            
            elif command == "get-ad-send-logs":
                await self.handle_get_ad_send_logs(payload)
            
            # Ad Analytics
            elif command == "get-ad-overview-stats":
                await self.handle_get_ad_overview_stats(payload)
            
            elif command == "get-ad-template-stats":
                await self.handle_get_ad_template_stats(payload)
            
            elif command == "get-ad-schedule-stats":
                await self.handle_get_ad_schedule_stats(payload)
            
            elif command == "get-ad-account-stats":
                await self.handle_get_ad_account_stats(payload)
            
            elif command == "get-ad-group-stats":
                await self.handle_get_ad_group_stats(payload)
            
            elif command == "get-ad-daily-stats":
                await self.handle_get_ad_daily_stats(payload)
            
            # ==================== User Tracking Commands (用戶追蹤系統) ====================
            elif command == "add-user-to-track":
                await self.handle_add_user_to_track(payload)
            
            elif command == "add-user-from-lead":
                await self.handle_add_user_from_lead(payload)
            
            elif command == "remove-tracked-user":
                await self.handle_remove_tracked_user(payload)
            
            elif command == "get-tracked-users":
                await self.handle_get_tracked_users(payload)
            
            elif command == "update-user-value-level":
                await self.handle_update_user_value_level(payload)
            
            elif command == "track-user-groups":
                await self.handle_track_user_groups(payload)
            
            elif command == "batch-track-users":
                await self.handle_batch_track_users(payload)
            
            elif command == "get-user-groups":
                await self.handle_get_user_groups(payload)
            
            elif command == "get-high-value-groups":
                await self.handle_get_high_value_groups(payload)
            
            elif command == "get-tracking-stats":
                await self.handle_get_tracking_stats(payload)
            
            elif command == "get-tracking-logs":
                await self.handle_get_tracking_logs(payload)
            
            elif command == "get-user-value-distribution":
                await self.handle_get_user_value_distribution(payload)
            
            elif command == "get-group-overlap-analysis":
                await self.handle_get_group_overlap_analysis(payload)
            
            elif command == "get-tracking-effectiveness":
                await self.handle_get_tracking_effectiveness(payload)
            
            # ==================== Campaign & Stats Commands (整合優化) ====================
            # Campaigns
            elif command == "create-campaign":
                await self.handle_create_campaign(payload)
            
            elif command == "update-campaign":
                await self.handle_update_campaign(payload)
            
            elif command == "delete-campaign":
                await self.handle_delete_campaign(payload)
            
            elif command == "get-campaigns":
                await self.handle_get_campaigns(payload)
            
            elif command == "get-campaign":
                await self.handle_get_campaign(payload)
            
            elif command == "start-campaign":
                await self.handle_start_campaign(payload)
            
            elif command == "pause-campaign":
                await self.handle_pause_campaign(payload)
            
            elif command == "resume-campaign":
                await self.handle_resume_campaign(payload)
            
            elif command == "stop-campaign":
                await self.handle_stop_campaign(payload)
            
            elif command == "get-campaign-logs":
                await self.handle_get_campaign_logs(payload)
            
            # Multi-Channel Stats
            elif command == "get-unified-overview":
                await self.handle_get_unified_overview(payload)
            
            elif command == "get-daily-trends":
                await self.handle_get_daily_trends(payload)
            
            elif command == "get-channel-performance":
                await self.handle_get_channel_performance(payload)
            
            elif command == "get-funnel-analysis":
                await self.handle_get_funnel_analysis(payload)
            
            # ==================== Multi-Role Commands (多角色協作) ====================
            # Role Management
            elif command == "get-role-templates":
                await self.handle_get_role_templates(payload)
            
            elif command == "assign-role":
                await self.handle_assign_role(payload)
            
            elif command == "update-role":
                await self.handle_update_role(payload)
            
            elif command == "remove-role":
                await self.handle_remove_role(payload)
            
            elif command == "get-account-roles":
                await self.handle_get_account_roles(payload)
            
            elif command == "get-all-roles":
                await self.handle_get_all_roles(payload)
            
            elif command == "get-role-stats":
                await self.handle_get_role_stats(payload)
            
            # Script Management
            elif command == "get-script-templates":
                await self.handle_get_script_templates(payload)
            
            elif command == "create-script-template":
                await self.handle_create_script_template(payload)
            
            elif command == "delete-script-template":
                await self.handle_delete_script_template(payload)
            
            elif command == "start-script-execution":
                await self.handle_start_script_execution(payload)
            
            elif command == "run-script-execution":
                await self.handle_run_script_execution(payload)
            
            elif command == "stop-script-execution":
                await self.handle_stop_script_execution(payload)
            
            elif command == "get-active-executions":
                await self.handle_get_active_executions(payload)
            
            elif command == "get-execution-stats":
                await self.handle_get_execution_stats(payload)
            
            # Collaboration
            elif command == "create-collab-group":
                await self.handle_create_collab_group(payload)
            
            elif command == "add-collab-member":
                await self.handle_add_collab_member(payload)
            
            elif command == "get-collab-groups":
                await self.handle_get_collab_groups(payload)
            
            elif command == "update-collab-status":
                await self.handle_update_collab_status(payload)
            
            elif command == "get-collab-stats":
                await self.handle_get_collab_stats(payload)

            else:
                self.send_log(f"Unknown command: {command}", "warning")
        
        except Exception as e:
            # Use global error handler
            app_error = handle_error(e, {"command": command, "payload": payload})
            # Error is already logged by error handler
            
            # Send error acknowledgment if request_id provided
            if request_id:
                self.send_event("command-complete", {
                    "request_id": request_id,
                    "command": command,
                    "status": "error",
                    "error": str(app_error)
                })
            
            import traceback
            traceback.print_exc()
    
    # ==================== Command Handlers ====================
    
    async def handle_get_initial_state(self):
        """Handle get-initial-state command and restore monitoring state if needed"""
        try:
            # Fetch all data from database
            accounts = await db.get_all_accounts()
            keyword_sets = await db.get_all_keyword_sets()
            monitored_groups = await db.get_all_groups()
            campaigns = await db.get_all_campaigns()
            message_templates = await db.get_all_templates()
            leads = await db.get_all_leads()
            logs = await db.get_recent_logs(limit=100)
            
            # Format timestamps to ISO 8601
            for account in accounts:
                # Accounts don't have timestamps in the response
                pass
            
            for lead in leads:
                if isinstance(lead.get('timestamp'), str):
                    # Already a string, ensure ISO format
                    pass
                else:
                    lead['timestamp'] = datetime.fromisoformat(lead['timestamp']).isoformat() + "Z"
                
                # Format interaction timestamps
                for interaction in lead.get('interactionHistory', []):
                    if isinstance(interaction.get('timestamp'), str):
                        pass
                    else:
                        interaction['timestamp'] = datetime.fromisoformat(interaction['timestamp']).isoformat() + "Z"
            
            for log in logs:
                if isinstance(log.get('timestamp'), str):
                    pass
                else:
                    log['timestamp'] = datetime.fromisoformat(log['timestamp']).isoformat() + "Z"
            
            # Get settings
            settings = await db.get_all_settings()
            
            # Get monitoring state from database
            monitoring_config = await db.get_monitoring_config()
            is_monitoring = monitoring_config.get('isActive', False)
            
            # Restore monitoring state if it was active
            if is_monitoring and not self.is_monitoring:
                # Check if we can restore monitoring
                listener_accounts = [a for a in accounts if a.get('role') == 'Listener' and a.get('status') == 'Online']
                if listener_accounts and monitored_groups and keyword_sets:
                    # Auto-restore monitoring
                    try:
                        await self.handle_start_monitoring()
                        await db.add_log("Monitoring state restored from database", "info")
                    except Exception as e:
                        await db.add_log(f"Failed to restore monitoring: {str(e)}", "warning")
                        # Reset monitoring state if restore failed
                        await db.set_monitoring_active(False)
                        is_monitoring = False
                else:
                    # Cannot restore, reset state
                    await db.set_monitoring_active(False)
                    is_monitoring = False
                    await db.add_log("Cannot restore monitoring: missing requirements", "warning")
            
            self.is_monitoring = is_monitoring
            
            # Send initial state event
            self.send_event("initial-state", {
                "accounts": accounts,
                "keywordSets": keyword_sets,
                "monitoredGroups": monitored_groups,
                "campaigns": campaigns,
                "messageTemplates": message_templates,
                "leads": leads,
                "logs": logs,
                "settings": settings,
                "isMonitoring": is_monitoring
            })
            
            await db.add_log("Initial state sent to frontend", "success")
        
        except Exception as e:
            self.send_log(f"Error getting initial state: {str(e)}", "error")
    
    # ========== Partial Update Functions ==========
    # These functions send only the updated data instead of full state refresh
    
    async def send_keyword_sets_update(self):
        """Send only keyword sets update to frontend with deduplication and error handling"""
        try:
            keyword_sets = await db.get_all_keyword_sets()
            
            if not keyword_sets:
                # 如果沒有關鍵詞集，發送空數組
                self.send_event("keyword-sets-updated", {"keywordSets": []})
                return
            
            # 去重處理：確保沒有重複的關鍵詞集和關鍵詞
            seen_set_ids = set()  # 使用 ID 而不是名稱，因為名稱可能重複
            seen_set_names = {}  # 名稱 -> ID 映射，用於檢測重複名稱
            deduplicated_sets = []
            
            for keyword_set in keyword_sets:
                set_id = keyword_set.get('id')
                set_name = keyword_set.get('name', '')
                
                # 如果關鍵詞集 ID 已處理過，跳過（防止重複）
                if set_id in seen_set_ids:
                    continue
                seen_set_ids.add(set_id)
                
                # 如果關鍵詞集名稱已存在且 ID 不同，記錄警告但保留（因為可能確實有同名但不同的集）
                if set_name and set_name in seen_set_names:
                    if seen_set_names[set_name] != set_id:
                        import sys
                        print(f"[Backend] Warning: Duplicate keyword set name '{set_name}' with different IDs: {seen_set_names[set_name]} and {set_id}", file=sys.stderr)
                seen_set_names[set_name] = set_id
                
                # 對關鍵詞進行去重（基於 keyword + isRegex 組合）
                seen_keywords = set()
                unique_keywords = []
                for keyword in keyword_set.get('keywords', []):
                    keyword_text = keyword.get('keyword', '')
                    is_regex = keyword.get('isRegex', False)
                    keyword_id = keyword.get('id')
                    key = (keyword_text, is_regex)
                    
                    # 如果關鍵詞已存在，跳過（保留第一個）
                    if key in seen_keywords:
                        import sys
                        print(f"[Backend] Warning: Duplicate keyword '{keyword_text}' (isRegex={is_regex}) in set {set_id}, skipping", file=sys.stderr)
                        continue
                    
                    seen_keywords.add(key)
                    unique_keywords.append({
                        'id': keyword_id,
                        'keyword': keyword_text,
                        'isRegex': is_regex
                    })
                
                # 創建去重後的關鍵詞集
                deduplicated_set = {
                    'id': set_id,
                    'name': set_name,
                    'keywords': unique_keywords
                }
                deduplicated_sets.append(deduplicated_set)
            
            # 確保事件被發送
            import sys
            print(f"[Backend] Sending keyword-sets-updated event with {len(deduplicated_sets)} sets", file=sys.stderr)
            self.send_event("keyword-sets-updated", {"keywordSets": deduplicated_sets})
        except Exception as e:
            import sys
            print(f"[Backend] Error sending keyword sets update: {e}", file=sys.stderr)
            import traceback
            traceback.print_exc(file=sys.stderr)
            # 即使出錯，也嘗試發送一個空數組或最後已知的狀態，避免前端狀態卡住
            try:
                # 嘗試獲取一個簡化的狀態
                keyword_sets = await db.get_all_keyword_sets()
                self.send_event("keyword-sets-updated", {"keywordSets": keyword_sets if keyword_sets else []})
            except:
                # 如果連這個都失敗，至少發送空數組
                self.send_event("keyword-sets-updated", {"keywordSets": []})
    
    async def send_groups_update(self):
        """Send only monitored groups update to frontend"""
        try:
            groups = await db.get_all_groups()
            self.send_event("groups-updated", {"monitoredGroups": groups})
        except Exception as e:
            print(f"[Backend] Error sending groups update: {e}", file=sys.stderr)
    
    async def send_templates_update(self):
        """Send only message templates update to frontend"""
        try:
            templates = await db.get_all_templates()
            self.send_event("templates-updated", {"messageTemplates": templates})
        except Exception as e:
            print(f"[Backend] Error sending templates update: {e}", file=sys.stderr)
    
    async def send_campaigns_update(self):
        """Send only campaigns update to frontend"""
        try:
            campaigns = await db.get_all_campaigns()
            self.send_event("campaigns-updated", {"campaigns": campaigns})
        except Exception as e:
            print(f"[Backend] Error sending campaigns update: {e}", file=sys.stderr)
    
    async def send_leads_update(self):
        """Send only leads update to frontend"""
        try:
            leads = await db.get_all_leads()
            for lead in leads:
                if isinstance(lead.get('timestamp'), str):
                    pass
                else:
                    lead['timestamp'] = datetime.fromisoformat(lead['timestamp']).isoformat() + "Z"
                for interaction in lead.get('interactionHistory', []):
                    if isinstance(interaction.get('timestamp'), str):
                        pass
                    else:
                        interaction['timestamp'] = datetime.fromisoformat(interaction['timestamp']).isoformat() + "Z"
            self.send_event("leads-updated", {"leads": leads})
        except Exception as e:
            print(f"[Backend] Error sending leads update: {e}", file=sys.stderr)
    
    # ========== End Partial Update Functions ==========
    
    async def handle_add_account(self, payload: Dict[str, Any]):
        """Handle add-account command"""
        try:
            import sys
            print(f"[Backend] Handling add-account command for phone: {payload.get('phone', 'unknown')}", file=sys.stderr)
            
            # Validate account data
            is_valid, errors = validate_account(payload)
            if not is_valid:
                error_message = "验证失败: " + "; ".join(errors)
                print(f"[Backend] Validation failed: {errors}", file=sys.stderr)
                self.send_log(error_message, "error")
                self.send_event("account-validation-error", {
                    "errors": errors,
                    "account_data": payload
                })
                handle_error(
                    AppError(ErrorType.VALIDATION_ERROR, error_message, {"errors": errors}),
                    {"command": "add-account", "payload": payload}
                )
                return
            
            print(f"[Backend] Validation passed, checking for existing account/session...", file=sys.stderr)
            
            phone = payload.get('phone', '')
            
            # Check if account already exists in database
            existing_account = await db.get_account_by_phone(phone)
            if existing_account:
                existing_status = existing_account.get('status', 'Offline')
                existing_id = existing_account.get('id')
                
                # CRITICAL: If account is already Online, just return success - don't trigger any login
                if existing_status == "Online":
                    print(f"[Backend] Account {phone} already exists and is Online, skipping", file=sys.stderr)
                    # Send duplicate error to inform user
                    error_msg = f"账户已存在: 电话号码 {phone} 已经在系统中（状态: 在线）。无需重复添加。"
                    self.send_log(error_msg, "info")
                    self.send_event("account-validation-error", {
                        "errors": [error_msg],
                        "account_data": payload,
                        "error_type": "duplicate"
                    })
                    return
                
                # OPTIMIZATION: If account exists with "error" or "Offline" status, 
                # automatically trigger login instead of showing "account already exists" error
                if existing_status in ["error", "Offline"]:
                    print(f"[Backend] Account exists with status '{existing_status}', automatically triggering login...", file=sys.stderr)
                    
                    # Reset status to Offline and clear login state
                    await db.update_account(existing_id, {"status": "Offline"})
                    # Clear any login callbacks for this phone
                    self.telegram_manager.login_callbacks.pop(phone, None)
                    
                    # Update account data if provided (API ID, API Hash, etc.)
                    update_data = {}
                    if payload.get('apiId'):
                        update_data['apiId'] = payload['apiId']
                    if payload.get('apiHash'):
                        update_data['apiHash'] = payload['apiHash']
                    if payload.get('proxy'):
                        update_data['proxy'] = payload['proxy']
                    if payload.get('group'):
                        update_data['group'] = payload['group']
                    if payload.get('twoFactorPassword'):
                        update_data['twoFactorPassword'] = payload['twoFactorPassword']
                    if payload.get('role'):
                        update_data['role'] = payload['role']
                    
                    if update_data:
                        await db.update_account(existing_id, update_data)
                        print(f"[Backend] Updated account data for {phone}", file=sys.stderr)
                    
                    # Send updated accounts list
                    accounts = await db.get_all_accounts()
                    self._cache.pop("accounts", None)
                    self._cache_timestamps.pop("accounts", None)
                    self.send_event("accounts-updated", accounts)
                    
                    # Automatically trigger login
                    print(f"[Backend] Auto-triggering login for existing account {phone} (ID: {existing_id})", file=sys.stderr)
                    self.send_log(f"账户 {phone} 已存在，自动触发登录...", "info")
                    
                    # Trigger login asynchronously (don't block the response)
                    asyncio.create_task(self.handle_login_account({
                        "accountId": existing_id,
                        "phoneCode": None,
                        "phoneCodeHash": None
                    }))
                    
                    # Return success with a message indicating auto-login was triggered
                    return
                
                # Check if account has stuck status (Logging in... or Waiting Code)
                if existing_status in ['Logging in...', 'Waiting Code', 'Waiting 2FA']:
                    # CRITICAL: If account is in login process (especially Waiting Code), 
                    # DO NOT update account data or send events that might trigger re-login
                    # Just return success silently to prevent status loop
                    print(f"[Backend] Account {phone} is in login process (status: {existing_status}), skipping update to prevent status loop", file=sys.stderr)
                    
                    # Only update account data if there are actual changes AND status is not Waiting Code
                    # If status is Waiting Code, user is waiting for verification code - do NOT update
                    if existing_status != 'Waiting Code':
                        # For Logging in... or Waiting 2FA, allow minimal updates
                        update_data = {}
                        if payload.get('apiId'):
                            update_data['apiId'] = payload.get('apiId')
                        if payload.get('apiHash'):
                            update_data['apiHash'] = payload.get('apiHash')
                        
                        if update_data and existing_id:
                            await db.update_account(existing_id, update_data)
                            print(f"[Backend] Updated account data for {phone}", file=sys.stderr)
                            # Only send accounts-updated if we actually updated something
                            accounts = await db.get_all_accounts()
                            self._cache.pop("accounts", None)
                            self._cache_timestamps.pop("accounts", None)
                            self.send_event("accounts-updated", accounts)
                    
                    # Return success - account exists and is in login process
                    # DO NOT send accounts-updated event if status is Waiting Code to prevent loop
                    return
                
                # If account exists with other statuses (Online, Offline, error), show error
                error_msg = f"账户已存在: 电话号码 {phone} 已经在系统中（状态: {existing_status}）。如需更新账户信息，请使用更新功能。"
                print(f"[Backend] Account already exists in database: {phone}, status: {existing_status}", file=sys.stderr)
                self.send_log(error_msg, "error")
                self.send_event("account-validation-error", {
                    "errors": [error_msg],
                    "account_data": payload,
                    "error_type": "duplicate"
                })
                return
            
            # Check if session file exists (but account not in database)
            from config import SESSIONS_DIR
            safe_phone = phone.replace("+", "").replace("-", "").replace(" ", "")
            session_file = SESSIONS_DIR / f"{safe_phone}.session"
            session_journal = SESSIONS_DIR / f"{safe_phone}.session.journal"
            
            if session_file.exists() or session_journal.exists():
                # Session file exists but account not in database
                # This could be an orphaned session file
                print(f"[Backend] Warning: Session file exists for {phone} but account not in database", file=sys.stderr)
                
                # CRITICAL: Ensure any client using this session is disconnected first
                try:
                    await self.telegram_manager.remove_client(phone, wait_for_disconnect=True)
                    # Force garbage collection to release file handles
                    gc.collect()
                    await asyncio.sleep(0.3)  # Give OS time to release file handles
                except Exception as e:
                    print(f"[Backend] Warning: Could not remove client for {phone} (may not exist): {e}", file=sys.stderr)
                
                # Now try to delete the orphaned session files
                print(f"[Backend] Deleting orphaned session file: {session_file}", file=sys.stderr)
                session_deleted = await self.safe_delete_session_file(session_file)
                
                if session_journal.exists():
                    journal_deleted = await self.safe_delete_session_file(session_journal)
                else:
                    journal_deleted = True
                
                if not session_deleted or not journal_deleted:
                    # If we can't delete the session file, we should not add the account
                    # This prevents database inconsistency
                    error_msg = f"无法删除已存在的会话文件。请确保没有其他程序正在使用该文件，然后重试。"
                    print(f"[Backend] Cannot delete orphaned session file, aborting account addition", file=sys.stderr)
                    self.send_log(error_msg, "error")
                    self.send_event("account-validation-error", {
                        "errors": [error_msg],
                        "account_data": payload,
                        "error_type": "file_locked"
                    })
                    return
                
                print(f"[Backend] Orphaned session file deleted successfully", file=sys.stderr)
            
            print(f"[Backend] Adding account to database...", file=sys.stderr)
            
            # Generate device fingerprint for anti-ban (防封)
            device_config = DeviceFingerprintGenerator.generate_for_phone(phone)
            print(f"[Backend] Generated device fingerprint for {phone}: {device_config.get('device_model')} ({device_config.get('platform')})", file=sys.stderr)
            
            # Add device fingerprint to payload
            payload['deviceModel'] = device_config.get('device_model')
            payload['systemVersion'] = device_config.get('system_version')
            payload['appVersion'] = device_config.get('app_version')
            payload['langCode'] = device_config.get('lang_code')
            payload['platform'] = device_config.get('platform')
            payload['deviceId'] = device_config.get('device_id')
            
            # Get proxy country from phone number (if proxy is provided)
            if payload.get('proxy'):
                proxy_country = ProxyManager.get_country_from_phone(phone)
                if proxy_country:
                    payload['proxyCountry'] = proxy_country
                    payload['proxyType'] = 'residential'  # Default to residential proxy
            
            # Ensure status is set to Offline when adding account
            payload['status'] = 'Offline'
            
            account_id = await db.add_account(payload)
            print(f"[Backend] Account added successfully with ID: {account_id}", file=sys.stderr)
            
            # Double-check: ensure status is Offline (in case of any issues)
            await db.update_account(account_id, {"status": "Offline"})
            print(f"[Backend] Account status set to Offline for {phone}", file=sys.stderr)
            
            await db.add_log(f"Account added: {payload.get('phone')}", "success")
            self.send_log(f"账户添加成功: {payload.get('phone')}", "success")
            
            # Send updated accounts list
            accounts = await db.get_all_accounts()
            print(f"[Backend] Sending accounts-updated event with {len(accounts)} accounts", file=sys.stderr)
            
            # Invalidate cache
            self._cache.pop("accounts", None)
            self._cache_timestamps.pop("accounts", None)
            self.send_event("accounts-updated", accounts)
        
        except ValidationError as e:
            import sys
            print(f"[Backend] ValidationError: {e.message}", file=sys.stderr)
            self.send_log(f"验证错误: {e.message}", "error")
            self.send_event("account-validation-error", {
                "errors": [e.message],
                "field": e.field
            })
        except ValueError as e:
            # Handle specific errors like duplicate phone number
            import sys
            error_msg = str(e)
            print(f"[Backend] ValueError adding account: {error_msg}", file=sys.stderr)
            self.send_log(error_msg, "error")
            self.send_event("account-validation-error", {
                "errors": [error_msg],
                "account_data": payload,
                "error_type": "duplicate" if "已存在" in error_msg else "validation"
            })
        except Exception as e:
            import sys
            error_msg = str(e)
            print(f"[Backend] Exception adding account: {error_msg}", file=sys.stderr)
            import traceback
            traceback.print_exc(file=sys.stderr)
            
            # Provide user-friendly error message
            if "UNIQUE constraint failed: accounts.phone" in error_msg or "phone" in error_msg.lower():
                friendly_msg = f"账户已存在: 电话号码 {payload.get('phone', '')} 已经在系统中。如需更新账户信息，请使用更新功能。"
            else:
                friendly_msg = f"添加账户失败: {error_msg}"
            
            self.send_log(friendly_msg, "error")
            self.send_event("account-validation-error", {
                "errors": [friendly_msg],
                "account_data": payload,
                "error_type": "duplicate" if "已存在" in friendly_msg else "error"
            })
            handle_error(e, {"command": "add-account", "payload": payload})
    
    async def handle_login_account(self, payload: Any):
        """Handle login-account command with Pyrogram"""
        try:
            import sys
            print(f"[Backend] handle_login_account called with payload: {payload}", file=sys.stderr)
            
            # Payload can be account_id (int) or dict with account_id and login details
            if isinstance(payload, int):
                account_id = payload
                account = await db.get_account(account_id)
                phone_code = None
                phone_code_hash = None
                two_factor_password = None
            else:
                account_id = payload.get('accountId')
                account = await db.get_account(account_id)
                phone_code = payload.get('phoneCode')
                phone_code_hash = payload.get('phoneCodeHash')
                two_factor_password = payload.get('twoFactorPassword')
            
            if not account:
                error_msg = f"Account {account_id} not found"
                print(f"[Backend] {error_msg}", file=sys.stderr)
                self.send_log(error_msg, "error")
                return
            
            phone = account.get('phone')
            current_status = account.get('status', 'Offline')
            print(f"[Backend] Found account: {phone}, API ID: {account.get('apiId')}, API Hash: {'***' if account.get('apiHash') else 'None'}", file=sys.stderr)
            
            # CRITICAL: If account is already Online and no verification code is being submitted,
            # skip login to prevent database lock and unnecessary operations
            if current_status == 'Online' and not phone_code:
                print(f"[Backend] Account {phone} is already Online, skipping login", file=sys.stderr)
                # Just verify the account is still valid
                try:
                    status_result = await self.telegram_manager.check_account_status(phone)
                    if status_result.get('online'):
                        print(f"[Backend] Account {phone} verified online", file=sys.stderr)
                        return  # Already online, nothing to do
                except Exception as e:
                    print(f"[Backend] Error checking account status: {e}, will proceed with login", file=sys.stderr)
            
            # CRITICAL: If account is in 'Waiting Code' status and we're NOT submitting a code,
            # just return the existing hash (don't resend code)
            if current_status == 'Waiting Code' and not phone_code:
                if phone in self.telegram_manager.login_callbacks:
                    existing_hash = self.telegram_manager.login_callbacks[phone].get("phone_code_hash")
                    if existing_hash:
                        print(f"[Backend] Account {phone} already waiting for code, returning existing hash", file=sys.stderr)
                        self.send_event("login-requires-code", {
                            "accountId": account_id,
                            "phone": phone,
                            "phoneCodeHash": existing_hash,
                            "sendType": self.telegram_manager.login_callbacks[phone].get("send_type", "app"),
                            "message": "验证码已发送，请在 Telegram 应用中查看",
                            "canRetrySMS": False
                        })
                        return
                print(f"[Backend] Account {phone} status is 'Waiting Code' but no callback found, will resend code", file=sys.stderr)
            
            # Update status to "Logging in..."
            await db.update_account(account_id, {"status": "Logging in..."})
            accounts = await db.get_all_accounts()
            # Invalidate cache
            self._cache.pop("accounts", None)
            self._cache_timestamps.pop("accounts", None)
            self.send_event("accounts-updated", accounts)
            
            self.send_log(f"Login initiated for account {phone}", "info")
            print(f"[Backend] Status updated to 'Logging in...', calling telegram_manager.login_account", file=sys.stderr)
            
            # Login with Pyrogram (使用设备指纹防封)
            print(f"[Backend] Calling telegram_manager.login_account for {phone}", file=sys.stderr)
            
            # Get device fingerprint from account (if exists) or generate new one
            device_model = account.get('deviceModel')
            system_version = account.get('systemVersion')
            app_version = account.get('appVersion')
            lang_code = account.get('langCode')
            platform = account.get('platform')
            
            # If device fingerprint not in account, generate it (for existing accounts)
            if not device_model or not system_version or not app_version:
                device_config = DeviceFingerprintGenerator.generate_for_phone(phone, prefer_platform=platform)
                device_model = device_model or device_config.get('device_model')
                system_version = system_version or device_config.get('system_version')
                app_version = app_version or device_config.get('app_version')
                lang_code = lang_code or device_config.get('lang_code')
                platform = platform or device_config.get('platform')
                
                # Update account with device fingerprint
                await db.update_account(account_id, {
                    'deviceModel': device_model,
                    'systemVersion': system_version,
                    'appVersion': app_version,
                    'langCode': lang_code,
                    'platform': platform,
                    'deviceId': device_config.get('device_id')
                })
                print(f"[Backend] Generated and saved device fingerprint for {phone}: {device_model} ({platform})", file=sys.stderr)
            
            result = await self.telegram_manager.login_account(
                phone=phone,
                api_id=account.get('apiId'),
                api_hash=account.get('apiHash'),
                proxy=account.get('proxy'),
                two_factor_password=two_factor_password or account.get('twoFactorPassword'),
                phone_code=phone_code,
                phone_code_hash=phone_code_hash,
                # Device fingerprint parameters (防封)
                device_model=device_model,
                system_version=system_version,
                app_version=app_version,
                lang_code=lang_code,
                platform=platform
            )
            
            print(f"[Backend] login_account result: success={result.get('success')}, requires_code={result.get('requires_code')}, requires_2fa={result.get('requires_2fa')}", file=sys.stderr)
            
            if result.get('success'):
                if result.get('requires_code'):
                    # Need verification code
                    phone_code_hash = result.get('phone_code_hash')
                    send_type = result.get('send_type', 'unknown')
                    next_type = result.get('next_type')
                    message = result.get('message', f"验证码已发送到 {phone}")
                    
                    print(f"[Backend] Sending login-requires-code event for account {account_id} (phone: {phone}), phone_code_hash: {phone_code_hash[:8] if phone_code_hash else 'None'}...", file=sys.stderr)
                    print(f"[Backend] Code send type: {send_type}, next_type: {next_type}, message: {message}", file=sys.stderr)
                    
                    # Check if we can retry for SMS (if previous was app and enough time has passed)
                    can_retry_sms = result.get('canRetrySMS', False) or result.get('can_retry_sms', False)
                    wait_seconds = result.get('waitSeconds', None)
                    
                    self.send_event("login-requires-code", {
                        "accountId": account_id,
                        "phone": phone,
                        "phoneCodeHash": phone_code_hash,
                        "sendType": send_type,  # Include send type
                        "nextType": next_type,  # Include next type
                        "message": message,  # Include message
                        "canRetrySMS": can_retry_sms,  # Include canRetrySMS flag
                        "waitSeconds": wait_seconds  # Include wait seconds if available
                    })
                    print(f"[Backend] login-requires-code event sent successfully", file=sys.stderr)
                    # State: Requesting Code -> Waiting Code
                    await db.update_account(account_id, {"status": "Waiting Code"})
                    print(f"[Backend] Account status updated to 'Waiting Code'", file=sys.stderr)
                elif result.get('requires_2fa'):
                    # Need 2FA password
                    self.send_event("login-requires-2fa", {
                        "accountId": account_id,
                        "phone": phone
                    })
                    await db.update_account(account_id, {"status": "Waiting 2FA"})
                else:
                    # Successfully logged in
                    await db.update_account(account_id, {"status": result.get('status', 'Online')})
                    self.send_log(f"Account {phone} logged in successfully", "success")
                    
                    # 獲取帳號信息
                    account = await db.get_account(account_id)
                    account_role = account.get('role', 'Unassigned')
                    
                    # 為所有帳號註冊私信處理器（處理用戶回復）
                    try:
                        await self.telegram_manager.register_private_message_handler(
                            phone=phone,
                            account_role=account_role
                        )
                        self.send_log(f"已為帳號 {phone} ({account_role}) 註冊私信處理器", "success")
                    except Exception as e:
                            self.send_log(f"註冊私信處理器失敗: {e}", "warning")
                    
                    # Start Warmup if enabled and not already started (防封)
                    warmup_enabled = account.get('warmupEnabled', False)
                    warmup_start_date = account.get('warmupStartDate')
                    
                    if warmup_enabled and not warmup_start_date:
                        # Start Warmup
                        from datetime import datetime
                        warmup_info = WarmupManager.start_warmup(account_id, datetime.now())
                        
                        await db.update_account(account_id, {
                            'warmupStartDate': warmup_info['warmup_start_date'],
                            'warmupStage': warmup_info['warmup_stage'],
                            'warmupDaysCompleted': 0
                        })
                        
                        stage_info = warmup_info['current_stage_info']
                        self.send_log(
                            f"账户 {phone} Warmup 已启动: {stage_info.get('stage_name')} "
                            f"(每日限制: {stage_info.get('daily_limit')} 条)",
                            "info"
                        )
                        print(f"[Backend] Warmup started for {phone}: Stage {stage_info.get('stage')} - {stage_info.get('stage_name')}", file=sys.stderr)
                    elif warmup_enabled and warmup_start_date:
                        # Update Warmup progress
                        warmup_progress = WarmupManager.get_warmup_progress(account)
                        if warmup_progress.get('enabled') and warmup_progress.get('stage'):
                            stage_info = warmup_progress['stage']
                            await db.update_account(account_id, {
                                'warmupStage': stage_info.get('stage'),
                                'warmupDaysCompleted': stage_info.get('days_completed', 0)
                            })
                            print(f"[Backend] Warmup progress updated for {phone}: Stage {stage_info.get('stage')} - {stage_info.get('stage_name')}, Days: {stage_info.get('days_completed')}", file=sys.stderr)
            else:
                # Login failed
                import sys
                error_status = result.get('status', 'Error')
                error_message = result.get('message', '未知错误')
                flood_wait = result.get('flood_wait')
                
                print(f"[Backend] Login failed for {phone}: status={error_status}, message={error_message}", file=sys.stderr)
                
                # Handle FloodWait
                if flood_wait:
                    friendly_msg = f'账户 {phone} 登录失败：请求过于频繁，请等待 {flood_wait} 秒后重试。'
                    self.send_event("account-login-error", {
                        "accountId": account_id,
                        "phone": phone,
                        "status": error_status,
                        "message": error_message,
                        "friendlyMessage": friendly_msg,
                        "floodWait": flood_wait
                    })
                    await db.update_account(account_id, {"status": "Offline"})
                    return
                
                await db.update_account(account_id, {"status": error_status})
                
                # Provide user-friendly error message
                if "API ID and API Hash are required" in error_message:
                    friendly_msg = f"账户 {phone} 登录失败：API ID 或 API Hash 未填写。请在账户管理中填写正确的 API 凭证。"
                elif "Invalid verification code" in error_message or "PhoneCodeInvalid" in error_message or "PHONE_CODE_INVALID" in error_message.upper():
                    friendly_msg = f"账户 {phone} 登录失败：验证码错误。请检查验证码是否正确，然后重新输入。"
                    # Keep the login state so user can retry
                    # Don't close the dialog, let user try again
                elif "code expired" in error_message.lower() or "PHONE_CODE_EXPIRED" in error_message.upper() or result.get('code_expired'):
                    friendly_msg = f'账户 {phone} 登录失败：验证码已过期。请点击"重新发送"获取新的验证码。'
                    # Non-recoverable error: State: Logging in... -> Offline (need to restart)
                    await db.update_account(account_id, {"status": "Offline"})
                    # Clear login callbacks
                    self.telegram_manager.login_callbacks.pop(phone, None)
                    # Mark as code expired so frontend can handle it
                    self.send_event("account-login-error", {
                        "accountId": account_id,
                        "phone": phone,
                        "status": "Offline",
                        "message": error_message,
                        "friendlyMessage": friendly_msg,
                        "codeExpired": True
                    })
                    # Don't send duplicate error event below
                    accounts = await db.get_all_accounts()
                    self._cache.pop("accounts", None)
                    self._cache_timestamps.pop("accounts", None)
                    self.send_event("accounts-updated", accounts)
                    return
                elif "hash mismatch" in error_message.lower() or "hash" in error_message.lower() and "mismatch" in error_message.lower():
                    friendly_msg = f'账户 {phone} 登录失败：验证码哈希不匹配。这通常是因为客户端被重新创建。请点击"重新发送"获取新的验证码。'
                    # Non-recoverable error: State: Logging in... -> Offline (need to restart)
                    await db.update_account(account_id, {"status": "Offline"})
                    # Clear login callbacks
                    self.telegram_manager.login_callbacks.pop(phone, None)
                    # Mark as code expired so frontend can handle it
                    self.send_event("account-login-error", {
                        "accountId": account_id,
                        "phone": phone,
                        "status": "Offline",
                        "message": error_message,
                        "friendlyMessage": friendly_msg,
                        "codeExpired": True
                    })
                    accounts = await db.get_all_accounts()
                    self._cache.pop("accounts", None)
                    self._cache_timestamps.pop("accounts", None)
                    self.send_event("accounts-updated", accounts)
                    return
                elif "Invalid 2FA password" in error_message or "2FA" in error_message and "password" in error_message.lower():
                    friendly_msg = f"账户 {phone} 登录失败：2FA密码错误。请检查密码后重试。"
                elif "Flood wait" in error_message or "FLOOD_WAIT" in error_message.upper() or result.get('flood_wait'):
                    wait_time = result.get('flood_wait', 0)
                    friendly_msg = f"账户 {phone} 登录失败：请求过于频繁，请等待 {wait_time} 秒后重试。"
                elif "Invalid phone number" in error_message or "PHONE_NUMBER_INVALID" in error_message.upper():
                    friendly_msg = f"账户 {phone} 登录失败：电话号码格式不正确。请确保格式为 +国家代码+号码。"
                elif "超時" in error_message or "timeout" in error_message.lower():
                    friendly_msg = f"账户 {phone} 登录失败：连接超时。请检查网络连接或代理设置后重试。"
                    # Non-recoverable error: State: Logging in... -> Offline (need to restart)
                    await db.update_account(account_id, {"status": "Offline"})
                    # Clear login callbacks
                    self.telegram_manager.login_callbacks.pop(phone, None)
                elif "Proxy" in error_message or "proxy" in error_message.lower():
                    friendly_msg = f"账户 {phone} 登录失败：代理连接错误。请检查代理配置或暂时移除代理。"
                elif "Connection" in error_message or "connection" in error_message.lower():
                    friendly_msg = f"账户 {phone} 登录失败：网络连接错误。请检查网络连接或代理设置。"
                elif "Client connection lost" in error_message or "No valid client" in error_message:
                    friendly_msg = f'账户 {phone} 登录失败：客户端连接丢失。请点击"重新发送"获取新的验证码。'
                    # Non-recoverable error: State: Logging in... -> Offline (need to restart)
                    await db.update_account(account_id, {"status": "Offline"})
                    # Clear login callbacks
                    self.telegram_manager.login_callbacks.pop(phone, None)
                    # Mark as code expired so frontend can handle it
                    self.send_event("account-login-error", {
                        "accountId": account_id,
                        "phone": phone,
                        "status": "Offline",
                        "message": error_message,
                        "friendlyMessage": friendly_msg,
                        "codeExpired": True
                    })
                    accounts = await db.get_all_accounts()
                    self._cache.pop("accounts", None)
                    self._cache_timestamps.pop("accounts", None)
                    self.send_event("accounts-updated", accounts)
                    return
                else:
                    friendly_msg = f"账户 {phone} 登录失败：{error_message}"
                
                self.send_log(friendly_msg, "error")
                
                # Also send a specific error event for frontend
                self.send_event("account-login-error", {
                    "accountId": account_id,
                    "phone": phone,
                    "status": error_status,
                    "message": error_message,
                    "friendlyMessage": friendly_msg
                })
            
            # Update accounts list
            accounts = await db.get_all_accounts()
            # Invalidate cache
            self._cache.pop("accounts", None)
            self._cache_timestamps.pop("accounts", None)
            self.send_event("accounts-updated", accounts)
        
        except Exception as e:
            import sys
            import traceback
            error_msg = f"Error logging in account: {str(e)}"
            print(f"[Backend] {error_msg}", file=sys.stderr)
            traceback.print_exc(file=sys.stderr)
            self.send_log(error_msg, "error")
    
    async def handle_check_account_status(self, payload: int):
        """Handle check-account-status command with Pyrogram"""
        try:
            account_id = payload
            account = await db.get_account(account_id)
            
            if not account:
                self.send_log(f"账户 ID {account_id} 不存在。无法检查状态。", "error")
                return
            
            phone = account.get('phone')
            
            # Check status with Pyrogram
            status_result = await self.telegram_manager.check_account_status(phone)
            
            # Update database
            await db.update_account(account_id, {"status": status_result.get('status', 'Offline')})
            
            # Send status update event
            status_info = {
                "accountId": account_id,
                "status": status_result.get('status', 'Offline'),
                "online": status_result.get('online', False),
                "message": status_result.get('message', ''),
                "user": status_result.get('user')
            }
            
            self.send_event("account-status-updated", status_info)
            
            # Update accounts list
            accounts = await db.get_all_accounts()
            # Invalidate cache
            self._cache.pop("accounts", None)
            self._cache_timestamps.pop("accounts", None)
            self.send_event("accounts-updated", accounts)
        
        except Exception as e:
            self.send_log(f"Error checking account status: {str(e)}", "error")
    
    async def handle_update_account_data(self, payload: Dict[str, Any]):
        """Handle update-account-data command"""
        try:
            account_id = payload.get('id')
            updates = payload.get('updates', {})
            await db.update_account(account_id, updates)
            
            accounts = await db.get_all_accounts()
            # Invalidate cache
            self._cache.pop("accounts", None)
            self._cache_timestamps.pop("accounts", None)
            self.send_event("accounts-updated", accounts)
            await db.add_log(f"Account {account_id} updated", "success")
        
        except Exception as e:
            self.send_log(f"Error updating account: {str(e)}", "error")
    
    async def handle_bulk_assign_role(self, payload: Dict[str, Any]):
        """Handle bulk-assign-role command"""
        try:
            account_ids = payload.get('accountIds', [])
            role = payload.get('role')
            await db.bulk_update_accounts_role(account_ids, role)
            
            accounts = await db.get_all_accounts()
            # Invalidate cache
            self._cache.pop("accounts", None)
            self._cache_timestamps.pop("accounts", None)
            self.send_event("accounts-updated", accounts)
            await db.add_log(f"Bulk assigned role '{role}' to {len(account_ids)} accounts", "success")
        
        except Exception as e:
            self.send_log(f"Error bulk assigning role: {str(e)}", "error")
    
    async def handle_bulk_assign_group(self, payload: Dict[str, Any]):
        """Handle bulk-assign-group command"""
        try:
            account_ids = payload.get('accountIds', [])
            group = payload.get('group')
            await db.bulk_update_accounts_group(account_ids, group)
            
            accounts = await db.get_all_accounts()
            # Invalidate cache
            self._cache.pop("accounts", None)
            self._cache_timestamps.pop("accounts", None)
            self.send_event("accounts-updated", accounts)
            await db.add_log(f"Bulk assigned group '{group}' to {len(account_ids)} accounts", "success")
        
        except Exception as e:
            self.send_log(f"Error bulk assigning group: {str(e)}", "error")
    
    async def handle_bulk_delete_accounts(self, payload: Dict[str, Any]):
        """Handle bulk-delete-accounts command - completely remove accounts and all related resources"""
        try:
            import sys
            from pathlib import Path
            from config import SESSIONS_DIR
            
            account_ids = payload.get('accountIds', [])
            if not account_ids:
                self.send_log("未选择要删除的账户", "error")
                return
            
            print(f"[Backend] Bulk deleting {len(account_ids)} accounts", file=sys.stderr)
            
            # Delete from database (this also cleans up related data)
            deleted_phones = await db.bulk_delete_accounts(account_ids)
            print(f"[Backend] {len(deleted_phones)} accounts deleted from database", file=sys.stderr)
            
            # Remove clients from TelegramClientManager and delete session files
            for phone in deleted_phones:
                if phone:
                    # Ensure client is fully disconnected
                    try:
                        await self.ensure_client_disconnected(phone)
                        print(f"[Backend] Client disconnected and removed for {phone}", file=sys.stderr)
                    except Exception as e:
                        print(f"[Backend] Error removing client for {phone}: {e}", file=sys.stderr)
                    
                    # Delete session files with retry mechanism
                    try:
                        safe_phone = phone.replace("+", "").replace("-", "").replace(" ", "")
                        session_file = SESSIONS_DIR / f"{safe_phone}.session"
                        session_journal = SESSIONS_DIR / f"{safe_phone}.session.journal"
                        
                        if session_file.exists():
                            deleted = await self.safe_delete_session_file(session_file)
                            if deleted:
                                print(f"[Backend] Deleted session file: {session_file}", file=sys.stderr)
                            else:
                                print(f"[Backend] WARNING: Failed to delete session file after retries: {session_file}", file=sys.stderr)
                        
                        if session_journal.exists():
                            deleted = await self.safe_delete_session_file(session_journal)
                            if deleted:
                                print(f"[Backend] Deleted session journal: {session_journal}", file=sys.stderr)
                            else:
                                print(f"[Backend] WARNING: Failed to delete session journal after retries: {session_journal}", file=sys.stderr)
                    except Exception as e:
                        print(f"[Backend] Error deleting session files for {phone}: {e}", file=sys.stderr)
            
            # Update accounts list and send event
            accounts = await db.get_all_accounts()
            # Invalidate cache
            self._cache.pop("accounts", None)
            self._cache_timestamps.pop("accounts", None)
            self.send_event("accounts-updated", accounts)
            
            await db.add_log(f"批量删除了 {len(deleted_phones)} 个账户", "success")
            self.send_log(f"已删除 {len(deleted_phones)} 个账户", "success")
            print(f"[Backend] Bulk delete completed: {len(deleted_phones)} accounts removed", file=sys.stderr)
            
            accounts = await db.get_all_accounts()
            # Invalidate cache
            self._cache.pop("accounts", None)
            self._cache_timestamps.pop("accounts", None)
            self.send_event("accounts-updated", accounts)
            await db.add_log(f"Deleted {len(account_ids)} accounts", "success")
        
        except Exception as e:
            self.send_log(f"Error deleting accounts: {str(e)}", "error")
    
    async def handle_remove_account(self, payload: Dict[str, Any]):
        """Handle remove-account command - completely remove account and all related resources"""
        try:
            import sys
            from pathlib import Path
            from config import SESSIONS_DIR
            
            account_id = payload.get('id')
            if not account_id:
                self.send_log("账户ID不能为空", "error")
                return
            
            # Get account info before deleting (for cleanup)
            account = await db.get_account(account_id)
            if not account:
                self.send_log(f"账户 {account_id} 不存在", "error")
                return
            
            phone = account.get('phone')
            print(f"[Backend] Removing account {account_id} (phone: {phone})", file=sys.stderr)
            
            # 1. Delete from database (this also cleans up related data)
            deleted_phone = await db.delete_account(account_id)
            if not deleted_phone:
                self.send_log(f"删除账户 {account_id} 失败", "error")
                return
            
            print(f"[Backend] Account {account_id} deleted from database", file=sys.stderr)
            
            # 2. Remove client from TelegramClientManager (this will disconnect it)
            if phone:
                try:
                    await self.telegram_manager.remove_client(phone, wait_for_disconnect=True)
                    # Force garbage collection to release file handles
                    gc.collect()
                    await asyncio.sleep(0.3)  # Give OS time to release file handles
                    print(f"[Backend] Client disconnected and removed for {phone}", file=sys.stderr)
                except Exception as e:
                    print(f"[Backend] Error removing client from TelegramClientManager: {e}", file=sys.stderr)
                    # Continue with file deletion attempt anyway
            
            # 3. Delete session files (with retry mechanism)
            if phone:
                try:
                    # Sanitize phone number for filename
                    safe_phone = phone.replace("+", "").replace("-", "").replace(" ", "")
                    session_file = SESSIONS_DIR / f"{safe_phone}.session"
                    session_journal = SESSIONS_DIR / f"{safe_phone}.session.journal"
                    
                    # Delete session file with retry
                    if session_file.exists():
                        deleted = await self.safe_delete_session_file(session_file)
                        if deleted:
                            print(f"[Backend] Deleted session file: {session_file}", file=sys.stderr)
                        else:
                            print(f"[Backend] WARNING: Failed to delete session file after retries: {session_file}", file=sys.stderr)
                    
                    # Delete session journal file if exists
                    if session_journal.exists():
                        deleted = await self.safe_delete_session_file(session_journal)
                        if deleted:
                            print(f"[Backend] Deleted session journal file: {session_journal}", file=sys.stderr)
                        else:
                            print(f"[Backend] WARNING: Failed to delete session journal file after retries: {session_journal}", file=sys.stderr)
                    
                except Exception as e:
                    print(f"[Backend] Error deleting session files for {phone}: {e}", file=sys.stderr)
                    # Don't fail the whole operation if session file deletion fails
                    # The account is already deleted from database
            
            # 4. Update accounts list and send event
            accounts = await db.get_all_accounts()
            # Invalidate cache
            self._cache.pop("accounts", None)
            self._cache_timestamps.pop("accounts", None)
            self.send_event("accounts-updated", accounts)
            
            await db.add_log(f"账户 {phone} (ID: {account_id}) 已完全删除", "success")
            self.send_log(f"账户 {phone} 已完全删除", "success")
            print(f"[Backend] Account {account_id} ({phone}) completely removed", file=sys.stderr)
        
        except Exception as e:
            import sys
            import traceback
            error_msg = str(e)
            print(f"[Backend] Error removing account: {error_msg}", file=sys.stderr)
            traceback.print_exc(file=sys.stderr)
            self.send_log(f"删除账户失败: {error_msg}", "error")
    
    async def check_monitoring_configuration(self) -> Dict[str, Any]:
        """
        完整配置檢查 - 在啟動監控前檢測所有必要配置
        
        Returns:
            Dict containing all check results and recommendations
        """
        checks = {
            "passed": True,
            "critical_issues": [],
            "warnings": [],
            "info": [],
            "details": {}
        }
        
        # ========== 1. 檢查監控帳號 ==========
        accounts = await db.get_all_accounts()
        listener_accounts = [a for a in accounts if a.get('role') == 'Listener']
        online_listeners = [a for a in listener_accounts if a.get('status') == 'Online']
        
        checks["details"]["listener_accounts"] = {
            "total": len(listener_accounts),
            "online": len(online_listeners),
            "accounts": [{"phone": a.get('phone'), "status": a.get('status')} for a in listener_accounts]
        }
        
        if not listener_accounts:
            checks["passed"] = False
            checks["critical_issues"].append({
                "code": "NO_LISTENER",
                "message": "沒有監控帳號（Listener 角色）",
                "fix": "在「帳戶管理」中將帳號角色設為「Listener」"
            })
        elif not online_listeners:
            checks["passed"] = False
            checks["critical_issues"].append({
                "code": "LISTENER_OFFLINE",
                "message": "監控帳號全部離線",
                "fix": "點擊「登入」按鈕使監控帳號上線"
            })
        else:
            checks["info"].append(f"✓ {len(online_listeners)} 個監控帳號在線")
        
        # ========== 2. 檢查發送帳號 ==========
        sender_accounts = [a for a in accounts if a.get('role') == 'Sender']
        online_senders = [a for a in sender_accounts if a.get('status') == 'Online']
        
        checks["details"]["sender_accounts"] = {
            "total": len(sender_accounts),
            "online": len(online_senders),
            "accounts": [{"phone": a.get('phone'), "status": a.get('status'), 
                         "dailySendCount": a.get('dailySendCount', 0),
                         "dailySendLimit": a.get('dailySendLimit', 50)} for a in sender_accounts]
        }
        
        if not sender_accounts:
            checks["warnings"].append({
                "code": "NO_SENDER",
                "message": "沒有發送帳號（Sender 角色）",
                "fix": "在「帳戶管理」中將帳號角色設為「Sender」，否則無法發送消息"
            })
        elif not online_senders:
            checks["warnings"].append({
                "code": "SENDER_OFFLINE",
                "message": "發送帳號全部離線",
                "fix": "點擊「登入」按鈕使發送帳號上線，否則無法發送消息"
            })
        else:
            # Check if any sender has remaining quota
            available_senders = [s for s in online_senders 
                                if s.get('dailySendCount', 0) < s.get('dailySendLimit', 50)]
            if not available_senders:
                checks["warnings"].append({
                    "code": "SENDER_LIMIT_REACHED",
                    "message": "所有發送帳號已達每日發送限額",
                    "fix": "等待明天重置限額，或增加新的發送帳號"
                })
            else:
                checks["info"].append(f"✓ {len(available_senders)} 個發送帳號可用")
        
        # ========== 3. 檢查監控群組 ==========
        monitored_groups = await db.get_all_monitored_groups()
        
        checks["details"]["monitored_groups"] = {
            "total": len(monitored_groups),
            "groups": [{"url": g.get('url'), "keywordSetIds": g.get('keywordSetIds', [])} 
                      for g in monitored_groups]
        }
        
        if not monitored_groups:
            checks["passed"] = False
            checks["critical_issues"].append({
                "code": "NO_GROUPS",
                "message": "沒有監控群組",
                "fix": "在「自動化中心」添加要監控的群組 URL"
            })
        else:
            checks["info"].append(f"✓ {len(monitored_groups)} 個監控群組")
        
        # ========== 4. 檢查關鍵詞集 ==========
        keyword_sets = await db.get_all_keyword_sets()
        
        # 計算總關鍵詞數
        total_keywords = sum(len(ks.get('keywords', [])) for ks in keyword_sets)
        
        checks["details"]["keyword_sets"] = {
            "total": len(keyword_sets),
            "total_keywords": total_keywords,
            "sets": [{"id": ks.get('id'), "name": ks.get('name'), 
                     "keywords": ks.get('keywords', [])} for ks in keyword_sets]
        }
        
        if not keyword_sets:
            checks["passed"] = False
            checks["critical_issues"].append({
                "code": "NO_KEYWORDS",
                "message": "沒有關鍵詞集",
                "fix": "在「自動化中心」創建關鍵詞集並添加關鍵詞"
            })
        elif total_keywords == 0:
            checks["passed"] = False
            checks["critical_issues"].append({
                "code": "EMPTY_KEYWORDS",
                "message": "關鍵詞集沒有任何關鍵詞",
                "fix": "在關鍵詞集中添加要監控的關鍵詞"
            })
        else:
            checks["info"].append(f"✓ {len(keyword_sets)} 個關鍵詞集，共 {total_keywords} 個關鍵詞")
        
        # ========== 5. 檢查群組與關鍵詞綁定 ==========
        groups_without_keywords = [g for g in monitored_groups if not g.get('keywordSetIds')]
        
        if groups_without_keywords and monitored_groups:
            checks["warnings"].append({
                "code": "GROUP_NO_KEYWORD",
                "message": f"{len(groups_without_keywords)} 個群組未綁定關鍵詞集",
                "fix": "在「監控群組」中為群組勾選關鍵詞集"
            })
        
        # ========== 6. 檢查活動（Campaign）==========
        campaigns = await db.get_all_campaigns()
        active_campaigns = [c for c in campaigns if c.get('isActive')]
        
        checks["details"]["campaigns"] = {
            "total": len(campaigns),
            "active": len(active_campaigns),
            "campaigns": [{
                "id": c.get('id'), 
                "name": c.get('name'), 
                "isActive": c.get('isActive'),
                "sourceGroupIds": c.get('trigger', {}).get('sourceGroupIds', []),
                "keywordSetIds": c.get('trigger', {}).get('keywordSetIds', []),
                "templateId": c.get('actions', [{}])[0].get('templateId', 0) if c.get('actions') else 0
            } for c in campaigns]
        }
        
        if not campaigns:
            checks["warnings"].append({
                "code": "NO_CAMPAIGN",
                "message": "沒有活動（Campaign）",
                "fix": "在「自動化中心」創建活動並配置觸發器、模板"
            })
        elif not active_campaigns:
            checks["warnings"].append({
                "code": "NO_ACTIVE_CAMPAIGN",
                "message": "沒有啟用的活動",
                "fix": "在「活動列表」中開啟活動開關"
            })
        else:
            # 檢查活動配置是否完整
            for campaign in active_campaigns:
                issues = []
                # 正確訪問嵌套結構：trigger.sourceGroupIds, trigger.keywordSetIds
                trigger = campaign.get('trigger', {})
                actions = campaign.get('actions', [])
                
                source_group_ids = trigger.get('sourceGroupIds', [])
                keyword_set_ids = trigger.get('keywordSetIds', [])
                template_id = actions[0].get('templateId', 0) if actions else 0
                
                if not source_group_ids:
                    issues.append("未選擇來源群組")
                if not keyword_set_ids:
                    issues.append("未選擇關鍵詞集")
                if not template_id:
                    issues.append("未選擇消息模板")
                
                if issues:
                    checks["warnings"].append({
                        "code": "CAMPAIGN_INCOMPLETE",
                        "message": f"活動「{campaign.get('name')}」配置不完整: {', '.join(issues)}",
                        "fix": "在「自動化中心」完善活動配置"
                    })
            
            checks["info"].append(f"✓ {len(active_campaigns)} 個活動已啟用")
        
        # ========== 7. 檢查消息模板 ==========
        templates = await db.get_all_templates()
        active_templates = [t for t in templates if t.get('isActive', True)]
        
        checks["details"]["templates"] = {
            "total": len(templates),
            "active": len(active_templates)
        }
        
        if not templates:
            checks["warnings"].append({
                "code": "NO_TEMPLATE",
                "message": "沒有消息模板",
                "fix": "在「自動化中心」創建消息模板"
            })
        else:
            checks["info"].append(f"✓ {len(templates)} 個消息模板")
        
        # ========== 8. 檢查 AI 設置 ==========
        ai_settings = await db.get_ai_settings()
        ai_enabled = ai_settings.get('auto_chat_enabled', 0) == 1
        ai_greeting_enabled = ai_settings.get('auto_greeting', 0) == 1
        ai_mode = ai_settings.get('auto_chat_mode', 'semi')
        
        checks["details"]["ai_settings"] = {
            "auto_chat_enabled": ai_enabled,
            "auto_greeting": ai_greeting_enabled,
            "auto_chat_mode": ai_mode
        }
        
        if ai_enabled:
            mode_names = {'full': '全自動', 'semi': '半自動', 'assist': '輔助', 'keyword': '關鍵詞觸發'}
            checks["info"].append(f"✓ AI 自動聊天已開啟 (模式: {mode_names.get(ai_mode, ai_mode)})")
            if ai_greeting_enabled:
                checks["info"].append("✓ AI 自動問候已開啟")
        else:
            checks["info"].append("ℹ AI 自動聊天未開啟（可在設置中開啟）")
        
        # ========== 生成總結 ==========
        checks["summary"] = {
            "can_monitor": checks["passed"],
            "can_send_messages": len(online_senders) > 0 and len(active_campaigns) > 0,
            "critical_count": len(checks["critical_issues"]),
            "warning_count": len(checks["warnings"]),
            "info_count": len(checks["info"])
        }
        
        return checks
    
    async def handle_start_monitoring(self):
        """Handle start-monitoring command with Pyrogram"""
        try:
            if self.is_monitoring:
                self.send_log("Monitoring is already running", "warning")
                return
            
            # ========== 完整配置檢查 ==========
            self.send_log("正在檢查監控配置...", "info")
            try:
                config_check = await self.check_monitoring_configuration()
            except Exception as check_error:
                import traceback
                traceback.print_exc(file=sys.stderr)
                self.send_log(f"配置檢查時發生錯誤: {str(check_error)}", "error")
                self.send_event("monitoring-start-failed", {
                    "reason": "exception",
                    "message": f"配置檢查時發生錯誤: {str(check_error)}"
                })
                return
            
            # 發送配置檢查報告到前端
            self.send_event("monitoring-config-check", config_check)
            
            # 顯示檢查結果
            for info in config_check.get("info", []):
                self.send_log(info, "info")
            
            for warning in config_check.get("warnings", []):
                self.send_log(f"⚠ {warning['message']}", "warning")
            
            for issue in config_check.get("critical_issues", []):
                self.send_log(f"✗ {issue['message']}", "error")
            
            # 如果有嚴重問題，阻止啟動
            if not config_check.get("passed", False):
                self.send_log("配置檢查未通過，無法啟動監控。請修復上述問題。", "error")
                self.send_event("monitoring-start-failed", {
                    "reason": "config_check_failed",
                    "message": "配置檢查未通過，請修復問題後重試",
                    "issues": config_check.get("critical_issues", []),
                    "warnings": config_check.get("warnings", [])
                })
                return
            
            # 如果沒有發送能力，給出警告但繼續
            if not config_check.get("summary", {}).get("can_send_messages", False):
                self.send_log("⚠ 警告：監控可以運行，但沒有可用的發送配置。Lead 將被捕獲但不會自動發送消息。", "warning")
            
            # 從配置檢查中獲取數據
            accounts = await db.get_all_accounts()
            listener_accounts = [a for a in accounts if a.get('role') == 'Listener' and a.get('status') == 'Online']
            monitored_groups = await db.get_all_monitored_groups()
            keyword_sets = await db.get_all_keyword_sets()
            
            # ========== 新增：檢查監控號是否已加入群組 ==========
            group_urls = [g.get('url') for g in monitored_groups if g.get('url')]
            self.send_log(f"正在檢查 {len(listener_accounts)} 個監控賬號對 {len(group_urls)} 個群組的成員狀態...", "info")
            
            # Check membership for all listener accounts
            all_membership_reports = []
            groups_needing_join = []  # Groups that need to be joined
            groups_accessible = []    # Groups that are accessible
            groups_cannot_join = []   # Groups that cannot be joined with reasons
            
            for account in listener_accounts:
                phone = account.get('phone')
                try:
                    report = await self.telegram_manager.check_all_groups_membership(phone, group_urls)
                    all_membership_reports.append(report)
                    
                    # Collect groups that are accessible (already member)
                    for group_info in report.get("member_of", []):
                        if group_info["url"] not in [g["url"] for g in groups_accessible]:
                            groups_accessible.append(group_info)
                    
                    # Collect groups that need joining
                    for group_info in report.get("can_join", []):
                        if group_info["url"] not in [g["url"] for g in groups_needing_join]:
                            groups_needing_join.append(group_info)
                    
                    # Collect groups that cannot be joined with reasons
                    for group_info in report.get("cannot_join", []):
                        if group_info["url"] not in [g["url"] for g in groups_cannot_join]:
                            groups_cannot_join.append(group_info)
                    
                except Exception as e:
                    import sys
                    print(f"[Backend] Error checking membership for {phone}: {e}", file=sys.stderr)
            
            # Send status report to frontend
            status_report = {
                "total_groups": len(group_urls),
                "accessible_groups": len(groups_accessible),
                "groups_needing_join": len(groups_needing_join),
                "groups_cannot_join": len(groups_cannot_join),
                "accessible_list": groups_accessible,
                "needing_join_list": groups_needing_join,
                "cannot_join_list": groups_cannot_join,
                "accounts_checked": len(listener_accounts)
            }
            
            self.send_event("monitoring-status-report", status_report)
            
            # Log status
            if groups_accessible:
                accessible_names = [g.get("title", g.get("url", "Unknown")) for g in groups_accessible[:3]]
                self.send_log(f"✓ 可監控群組: {len(groups_accessible)} 個 ({', '.join(accessible_names)}{'...' if len(groups_accessible) > 3 else ''})", "success")
            
            if groups_needing_join:
                needing_names = [g.get("url", "Unknown") for g in groups_needing_join[:3]]
                self.send_log(f"⚠ 需要加入: {len(groups_needing_join)} 個群組 ({', '.join(needing_names)}{'...' if len(groups_needing_join) > 3 else ''})", "warning")
            
            # 顯示無法加入的群組及原因
            if groups_cannot_join:
                for group_info in groups_cannot_join:
                    reason = group_info.get("reason", "未知原因")
                    self.send_log(f"✗ 無法加入 {group_info.get('url')}: {reason}", "error")
            
            # If no groups are accessible, try to auto-join
            if not groups_accessible and groups_needing_join:
                self.send_log(f"監控號未加入任何群組，正在嘗試自動加入 {len(groups_needing_join)} 個群組...", "info")
                # The start_monitoring will try to join groups automatically
            
            # If no groups can be monitored even after potential joins
            if not groups_accessible and not groups_needing_join:
                # 生成詳細的錯誤信息
                error_details = []
                for g in groups_cannot_join:
                    error_details.append(f"{g.get('url')}: {g.get('reason', '未知原因')}")
                
                error_message = "無法啟動監控: 監控號無法訪問任何群組。"
                if error_details:
                    error_message += f"\n詳情: {'; '.join(error_details)}"
                
                self.send_log(error_message, "error")
                self.send_event("monitoring-start-failed", {
                    "reason": "no_accessible_groups",
                    "message": "無法啟動監控: 監控號無法訪問任何群組。請確保監控號已加入要監控的群組。",
                    "cannot_join_list": groups_cannot_join
                })
                return
            
            # ========== 結束新增 ==========
            
            # Start monitoring for each account
            async def on_lead_captured(lead_data):
                """Callback when a lead is captured - optimized with batch query"""
                try:
                    # Batch check: get lead and DNC status in one query
                    existing_lead, is_dnc = await db.check_lead_and_dnc(lead_data['user_id'])
                    
                    if is_dnc:
                        return
                    
                    should_greet = False  # 是否應該發送問候
                    
                    if existing_lead:
                        # Update existing lead
                        lead_id = existing_lead['id']
                        await db.add_interaction(
                            lead_id,
                            'Keyword Matched',
                            f"Matched keyword: {lead_data['triggered_keyword']}"
                        )
                        
                        # 確保 user_profile 存在（用於漏斗統計）
                        user_id_str = str(lead_data['user_id'])
                        existing_profile = await db.get_user_profile(user_id_str)
                        if not existing_profile:
                            await db._connection.execute("""
                                INSERT INTO user_profiles 
                                (user_id, username, first_name, last_name, funnel_stage, interest_level, created_at)
                                VALUES (?, ?, ?, ?, 'new', 1, CURRENT_TIMESTAMP)
                            """, (
                                user_id_str,
                                lead_data.get('username', ''),
                                lead_data.get('first_name', ''),
                                lead_data.get('last_name', '')
                            ))
                            await db._connection.commit()
                        
                        # 檢查是否已經問候過此用戶（檢查狀態）
                        lead_status = existing_lead.get('status', 'New')
                        if lead_status == 'New':
                            # 還沒有聯繫過，應該發送問候
                            should_greet = True
                            self.send_log(f"📌 現有 Lead @{lead_data.get('username')} 尚未問候，將發送問候", "info")
                    else:
                        # Create new lead - 優先使用 source_group_url（群組 URL）
                        source_group_value = lead_data.get('source_group_url') or lead_data.get('source_group')
                        lead_id = await db.add_lead({
                            'userId': lead_data['user_id'],
                            'username': lead_data.get('username'),
                            'firstName': lead_data.get('first_name'),
                            'lastName': lead_data.get('last_name'),
                            'sourceGroup': source_group_value,  # 存儲 URL 而不是 chat_id
                            'triggeredKeyword': lead_data['triggered_keyword'],
                            'onlineStatus': lead_data.get('online_status', 'Unknown')
                        })
                        
                        # Send event with properly formatted data for frontend
                        import datetime
                        self.send_event("lead-captured", {
                            "id": lead_id,
                            "userId": lead_data['user_id'],
                            "username": lead_data.get('username') or '',
                            "firstName": lead_data.get('first_name') or '',
                            "lastName": lead_data.get('last_name') or '',
                            "sourceGroup": source_group_value,  # 使用 URL
                            "triggeredKeyword": lead_data['triggered_keyword'],
                            "timestamp": lead_data.get('timestamp') or datetime.datetime.now().isoformat(),
                            "status": "New",
                            "onlineStatus": lead_data.get('online_status', 'Unknown'),
                            "interactionHistory": [],
                            "doNotContact": False
                        })
                        
                        self.send_log(f"✓ 新潛在客戶已捕獲: @{lead_data.get('username') or lead_data.get('first_name')}", "success")
                        await db.add_log(f"New lead captured: {lead_data.get('username') or lead_data.get('first_name')}", "success")
                        
                        # 同步到 user_profiles 表（用於漏斗統計）
                        user_id_str = str(lead_data['user_id'])
                        existing_profile = await db.get_user_profile(user_id_str)
                        if not existing_profile:
                            await db._connection.execute("""
                                INSERT INTO user_profiles 
                                (user_id, username, first_name, last_name, funnel_stage, interest_level, created_at)
                                VALUES (?, ?, ?, ?, 'new', 1, CURRENT_TIMESTAMP)
                            """, (
                                user_id_str,
                                lead_data.get('username', ''),
                                lead_data.get('first_name', ''),
                                lead_data.get('last_name', '')
                            ))
                            await db._connection.commit()
                            self.send_log(f"📊 已創建用戶資料: @{lead_data.get('username')}", "info")
                        
                        # 新 Lead 總是發送問候
                        should_greet = True
                    
                    # AI Auto Chat: Handle greeting (for new leads or existing leads not yet contacted)
                    if should_greet:
                        await self._handle_ai_auto_greeting(lead_data, lead_id)
                    
                    # Check for matching campaigns and execute them
                    await self.execute_matching_campaigns(lead_id, lead_data)
                
                except Exception as e:
                    import traceback
                    error_details = traceback.format_exc()
                    error_msg = f"Error processing captured lead: {str(e)}\n{error_details}"
                    print(f"[Backend] Error in lead_callback: {error_msg}", file=sys.stderr)
                    self.send_log(f"處理潛在客戶時出錯: {str(e)}", "error")
                    # 記錄詳細錯誤到數據庫
                    await db.add_log(f"Lead callback error: {str(e)}", "error")
            
            # Start monitoring for each listener account
            successful_starts = 0
            failed_accounts = []
            
            for account in listener_accounts:
                phone = account.get('phone')
                group_urls = [g.get('url') for g in monitored_groups]
                keyword_sets_list = [
                    {
                        "id": ks.get('id'),
                        "keywords": ks.get('keywords', [])
                    }
                    for ks in keyword_sets
                ]
                
                try:
                    import sys
                    print(f"[Backend] Attempting to start monitoring for account {phone}", file=sys.stderr)
                    print(f"[Backend] Group URLs: {group_urls}", file=sys.stderr)
                    print(f"[Backend] Keyword sets count: {len(keyword_sets_list)}", file=sys.stderr)
                    
                    # 嘗試啟動監控
                    result = await self.telegram_manager.start_monitoring(
                        phone=phone,
                        group_urls=group_urls,
                        keyword_sets=keyword_sets_list,
                        on_lead_captured=on_lead_captured
                    )
                    
                    print(f"[Backend] start_monitoring result for {phone}: {result} (type: {type(result)})", file=sys.stderr)
                    
                    # 檢查是否成功啟動（start_monitoring 返回 True 表示成功）
                    if result is True:
                        successful_starts += 1
                        print(f"[Backend] ✓ Successfully started monitoring for {phone}", file=sys.stderr)
                        
                        # 記錄監控的群組信息
                        if hasattr(self.telegram_manager, 'monitoring_info') and phone in self.telegram_manager.monitoring_info:
                            monitoring_info = self.telegram_manager.monitoring_info[phone]
                            monitored_urls = monitoring_info.get('group_urls', [])
                            if monitored_urls:
                                self.send_log(f"賬戶 {phone} 成功啟動監控，監控群組: {', '.join(monitored_urls)}", "success")
                        
                        # 啟動瀏覽行為模擬後台任務（行為模擬）
                        try:
                            await self._start_browsing_simulation(account.get('id'), phone, group_urls)
                        except Exception as sim_error:
                            self.send_log(f"啟動行為模擬失敗 ({phone}): {str(sim_error)}", "warning")
                    else:
                        failed_accounts.append(phone)
                        print(f"[Backend] ✗ Failed to start monitoring for {phone}, result: {result}", file=sys.stderr)
                        self.send_log(f"賬戶 {phone} 監控啟動失敗（返回值: {result}）", "warning")
                
                except Exception as account_error:
                    failed_accounts.append(phone)
                    import sys
                    import traceback
                    print(f"[Backend] ✗ Exception starting monitoring for {phone}: {account_error}", file=sys.stderr)
                    traceback.print_exc(file=sys.stderr)
                    self.send_log(f"賬戶 {phone} 監控啟動失敗: {str(account_error)}", "error")
            
            # 只有在至少一個賬戶成功啟動監控時，才設置全局監控狀態
            if successful_starts > 0:
                self.is_monitoring = True
                # Save monitoring state to database
                await db.set_monitoring_active(True)
                self.send_event("monitoring-status-changed", True)
                
                success_message = f"監控已啟動：{successful_starts} 個賬戶正在監控"
                if failed_accounts:
                    success_message += f"，{len(failed_accounts)} 個賬戶啟動失敗"
                
                await db.add_log(success_message, "success")
                self.send_log(success_message, "success")
                
                # 為所有在線帳號註冊私信處理器（確保 AI 可以回復用戶私信）
                all_accounts = await db.get_all_accounts()
                online_clients = {}
                for acc in all_accounts:
                    if acc.get('status') == 'Online':
                        acc_phone = acc.get('phone')
                        acc_role = acc.get('role', 'Unassigned')
                        try:
                            await self.telegram_manager.register_private_message_handler(
                                phone=acc_phone,
                                account_role=acc_role
                            )
                            print(f"[Backend] ✓ 已為帳號 {acc_phone} 註冊私信處理器", file=sys.stderr)
                            
                            # 收集在線客戶端用於輪詢
                            client = self.telegram_manager.get_client(acc_phone)
                            if client:
                                online_clients[acc_phone] = client
                        except Exception as e:
                            print(f"[Backend] ✗ 註冊私信處理器失敗 ({acc_phone}): {e}", file=sys.stderr)
                
                # 啟動私信輪詢服務（雙重保險機制）
                if online_clients:
                    try:
                        private_message_poller.event_callback = self.send_event
                        await private_message_poller.start_polling(online_clients)
                        self.send_log(f"🔄 私信輪詢服務已啟動，監控 {len(online_clients)} 個帳號", "success")
                    except Exception as poller_err:
                        print(f"[Backend] ✗ 啟動私信輪詢服務失敗: {poller_err}", file=sys.stderr)
                        self.send_log(f"私信輪詢服務啟動失敗: {poller_err}", "warning")
            else:
                # 所有賬戶都失敗，不設置監控狀態
                self.is_monitoring = False
                await db.set_monitoring_active(False)
                self.send_event("monitoring-status-changed", False)
                
                error_message = "監控啟動失敗：所有監聽賬戶都無法啟動監控"
                if failed_accounts:
                    error_message += f"。失敗的賬戶：{', '.join(failed_accounts)}"
                
                self.send_event("monitoring-start-failed", {
                    "reason": "all_accounts_failed",
                    "message": error_message,
                    "failed_accounts": failed_accounts
                })
                await db.add_log(error_message, "error")
                self.send_log(error_message, "error")
        
        except Exception as e:
            # 發生異常時，確保監控狀態為 False
            self.is_monitoring = False
            await db.set_monitoring_active(False)
            self.send_event("monitoring-status-changed", False)
            self.send_log(f"啟動監控時發生錯誤: {str(e)}", "error")
            self.send_event("monitoring-start-failed", {
                "reason": "exception",
                "message": f"啟動監控時發生錯誤: {str(e)}"
            })
    
    async def handle_stop_monitoring(self):
        """Handle stop-monitoring command"""
        try:
            self.is_monitoring = False
            # Save monitoring state to database
            await db.set_monitoring_active(False)
            
            # Stop monitoring for all listener accounts (keeps connection, just removes handlers)
            accounts = await db.get_all_accounts()
            listener_accounts = [a for a in accounts if a.get('role') == 'Listener']
            
            stopped_count = 0
            for account in listener_accounts:
                phone = account.get('phone')
                try:
                    # Use stop_monitoring to just remove handlers without disconnecting
                    await self.telegram_manager.stop_monitoring(phone)
                    stopped_count += 1
                except Exception as stop_error:
                    self.send_log(f"停止監控時發生錯誤 ({phone}): {str(stop_error)}", "warning")
            
            self.send_event("monitoring-status-changed", False)
            await db.add_log(f"Monitoring stopped for {stopped_count} account(s)", "info")
            self.send_log(f"監控已停止：{stopped_count} 個賬戶", "info")
        
        except Exception as e:
            self.send_log(f"Error stopping monitoring: {str(e)}", "error")
    
    async def handle_one_click_start(self, payload: Dict[str, Any] = None):
        """
        一鍵啟動 v2.0：強制驗證 → 重新連接 → 啟動監控 → 啟用 AI
        增強版：不信任資料庫狀態，驗證實際連接
        """
        import sys
        print(f"[Backend] === 一鍵啟動 v2.0 開始 ===", file=sys.stderr)
        
        results = {
            'accounts': {'success': 0, 'failed': 0, 'total': 0, 'details': []},
            'monitoring': {'success': False, 'message': '', 'groups': 0},
            'ai': {'success': False, 'message': ''},
            'overall_success': False
        }
        
        try:
            # === 步驟 0: 預檢查 ===
            self.send_event("one-click-start-progress", {
                "step": "precheck",
                "message": "🔍 正在進行預檢查...",
                "progress": 5
            })
            
            accounts = await db.get_all_accounts()
            results['accounts']['total'] = len(accounts)
            
            if not accounts:
                self.send_event("one-click-start-progress", {
                    "step": "error",
                    "message": "❌ 沒有配置任何帳號",
                    "progress": 100
                })
                results['monitoring']['message'] = "沒有配置帳號"
                self.send_event("one-click-start-result", results)
                return
            
            # === 步驟 1: 強制驗證並重新連接所有帳號 ===
            self.send_event("one-click-start-progress", {
                "step": "accounts",
                "message": "🔑 正在驗證並連接帳號...",
                "progress": 10
            })
            
            for idx, account in enumerate(accounts):
                phone = account.get('phone')
                account_id = account.get('id')
                api_id = account.get('apiId')
                api_hash = account.get('apiHash')
                db_status = account.get('status', 'Offline')
                
                progress = 10 + int((idx + 1) / len(accounts) * 25)
                self.send_event("one-click-start-progress", {
                    "step": "account_connecting",
                    "message": f"🔑 連接帳號 {phone} ({idx+1}/{len(accounts)})...",
                    "progress": progress
                })
                
                account_result = {
                    'phone': phone,
                    'success': False,
                    'message': ''
                }
                
                try:
                    if not api_id or not api_hash:
                        account_result['message'] = "未配置 API"
                        results['accounts']['failed'] += 1
                        results['accounts']['details'].append(account_result)
                        self.send_log(f"✗ {phone}: 未配置 API ID/Hash", "warning")
                        continue
                    
                    # 關鍵：檢查實際客戶端狀態，不信任資料庫
                    client = self.telegram_manager.get_client(phone)
                    is_actually_connected = False
                    
                    if client:
                        try:
                            is_actually_connected = client.is_connected
                            print(f"[Backend] {phone}: 客戶端存在, is_connected={is_actually_connected}", file=sys.stderr)
                        except:
                            is_actually_connected = False
                    else:
                        print(f"[Backend] {phone}: 客戶端不存在", file=sys.stderr)
                    
                    # 如果客戶端不存在或未連接，強制重新登入
                    if not client or not is_actually_connected:
                        self.send_log(f"🔄 {phone}: 重新連接中...", "info")
                        
                        # 強制登入（使用正確的方法名 login_account）
                        login_result = await self.telegram_manager.login_account(
                            phone=phone,
                            api_id=api_id,
                            api_hash=api_hash
                        )
                        
                        print(f"[Backend] {phone}: 登入結果 = {login_result}", file=sys.stderr)
                        
                        if login_result.get('success') or login_result.get('status') == 'Online':
                            # 驗證連接：嘗試 get_me()
                            try:
                                client = self.telegram_manager.get_client(phone)
                                if client and client.is_connected:
                                    me = await client.get_me()
                                    if me:
                                        account_result['success'] = True
                                        account_result['message'] = f"已連接 (@{me.username or me.first_name})"
                                        results['accounts']['success'] += 1
                                        await db.update_account(account_id, {"status": "Online"})
                                        self.send_log(f"✓ {phone}: 連接成功", "success")
                                    else:
                                        raise Exception("get_me() 返回空")
                                else:
                                    raise Exception("客戶端未正確連接")
                            except Exception as verify_err:
                                account_result['message'] = f"驗證失敗: {verify_err}"
                                results['accounts']['failed'] += 1
                                await db.update_account(account_id, {"status": "Offline"})
                                self.send_log(f"✗ {phone}: 驗證失敗 - {verify_err}", "error")
                        else:
                            error_msg = login_result.get('error', '登入失敗')
                            # 檢查是否需要驗證碼
                            if login_result.get('status') == 'Code Required':
                                account_result['message'] = "需要驗證碼，請手動登入"
                            else:
                                account_result['message'] = error_msg
                            results['accounts']['failed'] += 1
                            await db.update_account(account_id, {"status": "Offline"})
                            self.send_log(f"✗ {phone}: {account_result['message']}", "warning")
                    else:
                        # 客戶端已連接，驗證會話
                        try:
                            me = await client.get_me()
                            if me:
                                account_result['success'] = True
                                account_result['message'] = f"已在線 (@{me.username or me.first_name})"
                                results['accounts']['success'] += 1
                                self.send_log(f"✓ {phone}: 已在線", "success")
                            else:
                                raise Exception("會話無效")
                        except Exception as session_err:
                            # 會話無效，重新連接
                            self.send_log(f"🔄 {phone}: 會話過期，重新連接...", "warning")
                            try:
                                await client.disconnect()
                            except:
                                pass
                            
                            login_result = await self.telegram_manager.login_account(
                                phone=phone,
                                api_id=api_id,
                                api_hash=api_hash
                            )
                            
                            if login_result.get('success') or login_result.get('status') == 'Online':
                                account_result['success'] = True
                                account_result['message'] = "重新連接成功"
                                results['accounts']['success'] += 1
                                await db.update_account(account_id, {"status": "Online"})
                                self.send_log(f"✓ {phone}: 重新連接成功", "success")
                            else:
                                account_result['message'] = f"重連失敗"
                                results['accounts']['failed'] += 1
                                await db.update_account(account_id, {"status": "Offline"})
                                self.send_log(f"✗ {phone}: 重連失敗", "error")
                    
                except Exception as acc_err:
                    account_result['message'] = str(acc_err)
                    results['accounts']['failed'] += 1
                    try:
                        await db.update_account(account_id, {"status": "Offline"})
                    except:
                        pass
                    self.send_log(f"✗ {phone}: {acc_err}", "error")
                
                results['accounts']['details'].append(account_result)
            
            self.send_event("one-click-start-progress", {
                "step": "accounts_done",
                "message": f"✅ 帳號連接: {results['accounts']['success']}/{results['accounts']['total']}",
                "progress": 40
            })
            
            # 如果沒有成功連接的帳號，停止
            if results['accounts']['success'] == 0:
                self.send_event("one-click-start-progress", {
                    "step": "error",
                    "message": "❌ 沒有成功連接的帳號，無法啟動監控",
                    "progress": 100
                })
                results['monitoring']['message'] = "沒有可用帳號"
                self.send_event("one-click-start-result", results)
                return
            
            # === 步驟 1.5: 自動加入群組 ===
            self.send_event("one-click-start-progress", {
                "step": "groups",
                "message": "👥 正在檢查並加入群組...",
                "progress": 42
            })
            
            results['groups'] = {
                'success': [],
                'pending': [],
                'failed': []
            }
            
            try:
                groups = await db.get_all_groups()
                if groups:
                    # 找到一個可用的監控帳號
                    listener_account = None
                    for acc in results['accounts']['details']:
                        if acc.get('success'):
                            listener_account = acc.get('phone')
                            break
                    
                    if listener_account:
                        client = self.telegram_manager.get_client(listener_account)
                        if client and client.is_connected:
                            # 設置事件回調
                            group_join_service.event_callback = self.send_event
                            
                            # 批量加入群組
                            group_urls = [g.get('url') for g in groups if g.get('url')]
                            self.send_log(f"🔄 檢查 {len(group_urls)} 個群組...", "info")
                            
                            join_report = await group_join_service.join_multiple_groups(
                                client=client,
                                group_urls=group_urls,
                                delay_between=1.5,
                                auto_verify=True
                            )
                            
                            results['groups']['success'] = join_report['success']
                            results['groups']['pending'] = join_report['pending']
                            results['groups']['failed'] = join_report['failed']
                            
                            # 顯示詳細結果
                            success_count = len(join_report['success'])
                            pending_count = len(join_report['pending'])
                            failed_count = len(join_report['failed'])
                            
                            if success_count > 0:
                                self.send_log(f"✓ 成功加入/已在 {success_count} 個群組", "success")
                            if pending_count > 0:
                                self.send_log(f"⏳ {pending_count} 個群組等待審批", "info")
                            if failed_count > 0:
                                for fail in join_report['failed']:
                                    self.send_log(f"✗ {fail['url']}: {fail['error']}", "warning")
                        else:
                            self.send_log("⚠ 沒有可用客戶端檢查群組", "warning")
                    else:
                        self.send_log("⚠ 沒有可用帳號檢查群組", "warning")
                else:
                    self.send_log("ℹ 沒有配置監控群組", "info")
            except Exception as group_err:
                self.send_log(f"群組檢查錯誤: {group_err}", "warning")
            
            self.send_event("one-click-start-progress", {
                "step": "groups_done",
                "message": f"✅ 群組檢查完成: {len(results['groups']['success'])} 個可用",
                "progress": 48
            })
            
            # === 步驟 2: 啟用 AI 自動聊天（先於監控，讓配置檢查能看到 AI 狀態）===
            self.send_event("one-click-start-progress", {
                "step": "ai",
                "message": "🤖 正在啟用 AI...",
                "progress": 50
            })
            
            try:
                # 更新為啟用狀態和全自動模式
                await db.update_ai_settings({
                    'auto_chat_enabled': 1,
                    'auto_chat_mode': 'full',
                    'auto_greeting': 1
                })
                
                # 重新載入 AI 設置到內存
                await ai_auto_chat.initialize()
                
                # 檢查 AI 端點是否已配置
                ai_endpoint = ai_auto_chat.local_ai_endpoint
                if ai_endpoint:
                    self.send_log(f"✓ AI 端點: {ai_endpoint}", "success")
                else:
                    self.send_log("⚠ AI 端點未配置，將使用備用回覆", "warning")
                
                results['ai']['success'] = True
                results['ai']['message'] = f"AI 全自動模式已啟用" + (f" (端點: {ai_endpoint[:30]}...)" if ai_endpoint else " (備用回覆)")
                self.send_log("✓ AI 自動聊天已啟用 (全自動模式)", "success")
                
                # 發送 AI 設置更新事件
                self.send_event("ai-settings-updated", {
                    'auto_chat_enabled': True,
                    'auto_chat_mode': 'full',
                    'auto_greeting': True
                })
            except Exception as ai_err:
                results['ai']['message'] = str(ai_err)
                self.send_log(f"✗ AI 啟用錯誤: {ai_err}", "error")
            
            self.send_event("one-click-start-progress", {
                "step": "ai_done",
                "message": f"{'✅' if results['ai']['success'] else '❌'} {results['ai']['message']}",
                "progress": 60
            })
            
            # === 步驟 3: 啟動監控（在 AI 啟用之後，配置檢查能正確顯示 AI 狀態）===
            self.send_event("one-click-start-progress", {
                "step": "monitoring",
                "message": "📡 正在啟動監控...",
                "progress": 65
            })
            
            try:
                await self.handle_start_monitoring()
                if self.is_monitoring:
                    results['monitoring']['success'] = True
                    results['monitoring']['message'] = "監控已啟動"
                    
                    # 統計監控的群組數
                    groups = await db.get_all_groups()
                    results['monitoring']['groups'] = len(groups)
                else:
                    results['monitoring']['message'] = "監控啟動失敗，請檢查群組配置"
            except Exception as mon_err:
                results['monitoring']['message'] = str(mon_err)
                self.send_log(f"✗ 監控啟動錯誤: {mon_err}", "error")
            
            self.send_event("one-click-start-progress", {
                "step": "monitoring_done",
                "message": f"{'✅' if results['monitoring']['success'] else '⚠️'} {results['monitoring']['message']}",
                "progress": 80
            })
            
            # === 步驟 4: 啟動私信輪詢 ===
            self.send_event("one-click-start-progress", {
                "step": "poller",
                "message": "📬 正在啟動私信輪詢...",
                "progress": 90
            })
            
            try:
                # 收集所有在線客戶端
                online_clients = {}
                for acc in results['accounts']['details']:
                    if acc.get('success'):
                        phone = acc.get('phone')
                        client = self.telegram_manager.get_client(phone)
                        if client and client.is_connected:
                            online_clients[phone] = client
                
                if online_clients:
                    private_message_poller.event_callback = self.send_event
                    await private_message_poller.start_polling(online_clients)
                    self.send_log(f"✓ 私信輪詢已啟動，監控 {len(online_clients)} 個帳號", "success")
                else:
                    self.send_log("⚠ 沒有可用的客戶端用於私信輪詢", "warning")
            except Exception as poller_err:
                self.send_log(f"✗ 私信輪詢啟動錯誤: {poller_err}", "warning")
            
            # === 步驟 5: 啟動連接監控 ===
            self.send_event("one-click-start-progress", {
                "step": "connection_monitor",
                "message": "🔄 正在啟動連接監控...",
                "progress": 95
            })
            
            try:
                connection_monitor.set_telegram_manager(self.telegram_manager)
                connection_monitor.event_callback = self.send_event
                await connection_monitor.start(check_interval=60)
                self.send_log("✓ 連接監控已啟動（每60秒檢查）", "success")
            except Exception as monitor_err:
                self.send_log(f"⚠ 連接監控啟動錯誤: {monitor_err}", "warning")
            
            # === 步驟 6: 啟動漏斗自動流轉 ===
            self.send_event("one-click-start-progress", {
                "step": "funnel",
                "message": "🎯 正在啟動漏斗自動流轉...",
                "progress": 97
            })
            
            results['funnel'] = {
                'success': False,
                'message': ''
            }
            
            try:
                # 設置漏斗管理器回調
                async def funnel_send_callback(target_user_id: str, message: str, **kwargs):
                    """漏斗自動跟進發送回調"""
                    # 獲取任一在線帳號
                    for acc in results['accounts']['details']:
                        if acc.get('success'):
                            phone = acc.get('phone')
                            client = self.telegram_manager.get_client(phone)
                            if client and client.is_connected:
                                try:
                                    await client.send_message(int(target_user_id), message)
                                    self.send_log(f"[AutoFunnel] 已發送跟進消息給 {target_user_id}", "info")
                                    return True
                                except Exception as send_err:
                                    self.send_log(f"[AutoFunnel] 發送失敗: {send_err}", "warning")
                    return False
                
                auto_funnel.set_callbacks(
                    send_callback=funnel_send_callback,
                    log_callback=self.send_log,
                    event_callback=self.send_event
                )
                
                # 確保漏斗管理器已啟動
                if not auto_funnel.is_running:
                    await auto_funnel.start()
                
                results['funnel']['success'] = True
                results['funnel']['message'] = "漏斗自動流轉已啟用"
                self.send_log("✓ 漏斗自動流轉已啟動（每30分鐘檢查跟進）", "success")
                
            except Exception as funnel_err:
                results['funnel']['message'] = str(funnel_err)
                self.send_log(f"⚠ 漏斗管理器啟動錯誤: {funnel_err}", "warning")
            
            # === 完成 ===
            results['overall_success'] = (
                results['accounts']['success'] > 0 and
                results['monitoring']['success'] and
                results['ai']['success']
            )
            
            self.send_event("one-click-start-progress", {
                "step": "complete",
                "message": "🎉 一鍵啟動完成！" if results['overall_success'] else "⚠️ 部分功能啟動失敗",
                "progress": 100
            })
            
            self.send_event("one-click-start-result", results)
            
            summary = f"一鍵啟動完成 - 帳號: {results['accounts']['success']}/{results['accounts']['total']}, "
            summary += f"監控: {'✓' if results['monitoring']['success'] else '✗'}, "
            summary += f"AI: {'✓' if results['ai']['success'] else '✗'}"
            
            await db.add_log(summary, "success" if results['overall_success'] else "warning")
            self.send_log(summary, "success" if results['overall_success'] else "warning")
            
        except Exception as e:
            print(f"[Backend] 一鍵啟動錯誤: {e}", file=sys.stderr)
            import traceback
            traceback.print_exc(file=sys.stderr)
            
            self.send_event("one-click-start-result", {
                **results,
                'error': str(e),
                'overall_success': False
            })
            self.send_log(f"一鍵啟動失敗: {e}", "error")
    
    async def handle_one_click_stop(self):
        """
        一鍵停止：停止監控 → 關閉 AI
        """
        import sys
        print(f"[Backend] === 一鍵停止開始 ===", file=sys.stderr)
        
        try:
            # 停止連接監控
            try:
                await connection_monitor.stop()
            except:
                pass
            
            # 停止私信輪詢
            try:
                await private_message_poller.stop_polling()
            except:
                pass
            
            # 停止監控
            await self.handle_stop_monitoring()
            
            # 關閉 AI 自動聊天
            await db.update_ai_settings({
                'auto_chat_enabled': 0
            })
            
            self.send_event("ai-settings-updated", {
                'auto_chat_enabled': False
            })
            
            self.send_event("one-click-stop-result", {
                'success': True,
                'message': "所有服務已停止"
            })
            
            # 停止漏斗管理器
            try:
                await auto_funnel.stop()
            except:
                pass
            
            self.send_log("🛑 一鍵停止完成：監控已停止，AI 已關閉，連接監控已停止，漏斗管理已停止", "info")
            
        except Exception as e:
            self.send_log(f"一鍵停止錯誤: {e}", "error")
            self.send_event("one-click-stop-result", {
                'success': False,
                'error': str(e)
            })
    
    async def handle_get_system_status(self):
        """
        獲取系統狀態：帳號、監控、AI 等
        """
        try:
            # 獲取帳號狀態
            accounts = await db.get_all_accounts()
            online_count = sum(1 for a in accounts if a.get('status') == 'Online')
            
            # 獲取關鍵詞集
            keyword_sets = await db.get_all_keyword_sets()
            
            # 獲取群組
            groups = await db.get_all_groups()
            
            # 獲取活動
            campaigns = await db.get_all_campaigns()
            active_campaigns = sum(1 for c in campaigns if c.get('isActive'))
            
            # 獲取 AI 設置
            ai_settings = await db.get_ai_settings()
            ai_enabled = ai_settings.get('auto_chat_enabled', 0) == 1 if ai_settings else False
            ai_mode = ai_settings.get('auto_chat_mode', 'semi') if ai_settings else 'semi'
            
            # 獲取模板
            templates = await db.get_all_templates()
            active_templates = sum(1 for t in templates if t.get('isActive'))
            
            status = {
                'accounts': {
                    'total': len(accounts),
                    'online': online_count,
                    'offline': len(accounts) - online_count
                },
                'monitoring': {
                    'active': self.is_monitoring,
                    'groups': len(groups)
                },
                'ai': {
                    'enabled': ai_enabled,
                    'mode': ai_mode,
                    'endpoint': ai_settings.get('local_ai_endpoint', '') if ai_settings else ''
                },
                'keywords': {
                    'sets': len(keyword_sets),
                    'total': sum(len(ks.get('keywords', [])) for ks in keyword_sets)
                },
                'campaigns': {
                    'total': len(campaigns),
                    'active': active_campaigns
                },
                'templates': {
                    'total': len(templates),
                    'active': active_templates
                },
                'poller': {
                    'running': private_message_poller._running if hasattr(private_message_poller, '_running') else False
                }
            }
            
            self.send_event("system-status", status)
            
        except Exception as e:
            import sys
            print(f"[Backend] 獲取系統狀態錯誤: {e}", file=sys.stderr)
            self.send_event("system-status", {'error': str(e)})
    
    # ==================== 知識學習功能 ====================
    
    async def handle_learn_from_history(self, payload: Dict[str, Any]):
        """從歷史對話中學習知識"""
        try:
            from knowledge_learner import knowledge_learner
            
            # 初始化
            await knowledge_learner.initialize()
            
            user_id = payload.get('user_id')
            limit = payload.get('limit', 100)
            
            self.send_log("🎓 開始從歷史對話學習知識...", "info")
            
            total_learned = 0
            
            if user_id:
                # 學習特定用戶的對話
                cursor = await db._connection.execute("""
                    SELECT role, content, timestamp 
                    FROM chat_history 
                    WHERE user_id = ? 
                    ORDER BY timestamp ASC
                """, (user_id,))
                rows = await cursor.fetchall()
                
                if rows:
                    messages = [{'role': r['role'], 'content': r['content']} for r in rows]
                    profile = await db.get_user_profile(user_id)
                    outcome = profile.get('funnel_stage', 'new') if profile else 'new'
                    
                    result = await knowledge_learner.learn_from_conversation(
                        user_id=user_id,
                        messages=messages,
                        outcome=outcome
                    )
                    total_learned = result.get('total_knowledge', 0)
            else:
                # 學習所有成功對話
                cursor = await db._connection.execute("""
                    SELECT DISTINCT user_id, funnel_stage 
                    FROM user_profiles 
                    WHERE funnel_stage IN ('converted', 'interested', 'negotiating')
                    LIMIT ?
                """, (limit,))
                users = await cursor.fetchall()
                
                for user in users:
                    uid = user['user_id']
                    outcome = user['funnel_stage']
                    
                    # 獲取對話歷史
                    msg_cursor = await db._connection.execute("""
                        SELECT role, content, timestamp 
                        FROM chat_history 
                        WHERE user_id = ? 
                        ORDER BY timestamp ASC
                        LIMIT 50
                    """, (uid,))
                    rows = await msg_cursor.fetchall()
                    
                    if rows:
                        messages = [{'role': r['role'], 'content': r['content']} for r in rows]
                        result = await knowledge_learner.learn_from_conversation(
                            user_id=uid,
                            messages=messages,
                            outcome=outcome
                        )
                        total_learned += result.get('total_knowledge', 0)
            
            self.send_log(f"✓ 學習完成，共提取 {total_learned} 條知識", "success")
            self.send_event("learn-from-history-result", {
                'success': True,
                'total_learned': total_learned
            })
            
        except Exception as e:
            self.send_log(f"學習失敗: {e}", "error")
            self.send_event("learn-from-history-result", {
                'success': False,
                'error': str(e)
            })
    
    async def handle_get_knowledge_stats(self):
        """獲取知識庫統計"""
        try:
            from knowledge_learner import knowledge_learner
            await knowledge_learner.initialize()
            
            stats = await knowledge_learner.get_statistics()
            self.send_event("knowledge-stats", stats)
            
        except Exception as e:
            self.send_event("knowledge-stats", {'error': str(e)})
    
    async def handle_search_knowledge(self, payload: Dict[str, Any]):
        """搜索知識庫"""
        try:
            from knowledge_learner import knowledge_learner
            await knowledge_learner.initialize()
            
            query = payload.get('query', '')
            limit = payload.get('limit', 5)
            
            results = await knowledge_learner.search_knowledge(query, limit)
            
            self.send_event("knowledge-search-result", {
                'success': True,
                'query': query,
                'results': results
            })
            
        except Exception as e:
            self.send_event("knowledge-search-result", {
                'success': False,
                'error': str(e)
            })
    
    async def _handle_ai_auto_greeting(self, lead_data: Dict[str, Any], lead_id: int):
        """Handle AI auto greeting for new leads based on settings"""
        import sys
        print(f"[Backend] _handle_ai_auto_greeting called for lead_id={lead_id}, user={lead_data.get('username')}", file=sys.stderr)
        self.send_log(f"[AI] 開始處理自動問候 (Lead ID: {lead_id}, User: @{lead_data.get('username')})", "info")
        
        try:
            # Get AI settings
            settings = await db.get_ai_settings()
            print(f"[Backend] AI settings loaded: {settings}", file=sys.stderr)
            
            if not settings:
                self.send_log("[AI] AI 設置未配置，跳過自動問候", "warning")
                return
            
            # 使用正確的數據庫字段名稱 (整數 0/1)
            enabled = settings.get('auto_chat_enabled', 0) == 1
            auto_greeting = settings.get('auto_greeting', 0) == 1
            mode = settings.get('auto_chat_mode', 'semi')
            
            self.send_log(f"[AI] 設置檢查 - 啟用: {enabled}, 自動問候: {auto_greeting}, 模式: {mode}", "info")
            
            if not enabled:
                self.send_log("[AI] AI 自動聊天未啟用，跳過自動問候", "info")
                return
            
            if not auto_greeting:
                self.send_log("[AI] 自動問候未啟用，跳過自動問候", "info")
                return
            
            user_id = str(lead_data.get('user_id', ''))
            username = lead_data.get('username', '')
            first_name = lead_data.get('first_name', '')
            source_group = lead_data.get('source_group_url') or lead_data.get('source_group', '')
            
            # Get sender account (prefer sender role, fallback to listener)
            sender_phone = ''
            accounts = await db.get_all_accounts()
            
            # First try to find an online sender account (case-insensitive role check)
            for acc in accounts:
                role = str(acc.get('role', '')).lower()
                status = str(acc.get('status', '')).lower()
                if role == 'sender' and status == 'online':
                    sender_phone = acc.get('phone', '')
                    self.send_log(f"[AI] 找到發送帳號: {sender_phone}", "info")
                    break
            
            # If no sender, use any online account
            if not sender_phone:
                for acc in accounts:
                    status = str(acc.get('status', '')).lower()
                    if status == 'online':
                        sender_phone = acc.get('phone', '')
                        self.send_log(f"[AI] 使用在線帳號: {sender_phone}", "info")
                        break
            
            # Fallback to the monitoring account
            if not sender_phone:
                sender_phone = lead_data.get('account_phone', '')
            
            if not sender_phone:
                self.send_log("[AI] 沒有可用的發送帳號，跳過自動問候", "warning")
                return
            
            self.send_log(f"[AI] 準備發送問候給 @{username or first_name}，使用帳號: {sender_phone}", "info")
            
            # Generate greeting using AI (傳遞觸發關鍵詞用於個性化問候)
            triggered_keyword = lead_data.get('triggered_keyword', '')
            greeting = await ai_auto_chat.handle_auto_greeting(
                user_id=user_id,
                username=username,
                account_phone=sender_phone,
                source_group=source_group,
                first_name=first_name,
                triggered_keyword=triggered_keyword
            )
            
            if not greeting:
                self.send_log(f"[AI] 未能生成問候消息", "warning")
                return
            
            self.send_log(f"[AI] 生成問候: {greeting[:50]}...", "info")
            
            if mode == 'full':
                # Full auto mode: Send immediately using self.message_queue
                from message_queue import MessagePriority
                
                message_id = await self.message_queue.add_message(
                    phone=sender_phone,
                    user_id=user_id,
                    text=greeting,
                    priority=MessagePriority.HIGH  # High priority for greeting
                )
                
                self.send_log(f"[AI] ✓ 已自動發送問候給 @{username or first_name} (消息ID: {message_id})", "success")
                await db.add_interaction(lead_id, 'AI Auto Greeting', greeting)
                
                # Update lead status to "已聯繫"
                await db.update_lead(lead_id, {'status': 'Contacted'})
                self.send_event("leads-updated", await db.get_all_leads())
                
            elif mode == 'semi':
                # Semi-auto mode: Send to frontend for confirmation
                self.send_event("ai-greeting-suggestion", {
                    "leadId": lead_id,
                    "userId": user_id,
                    "username": username,
                    "firstName": first_name,
                    "sourceGroup": source_group,
                    "suggestedGreeting": greeting,
                    "accountPhone": sender_phone
                })
                self.send_log(f"[AI] 已生成問候建議給 @{username or first_name}，等待確認", "info")
            
        except Exception as e:
            import traceback
            import sys
            error_details = traceback.format_exc()
            error_msg = f"[AI] Error in auto greeting: {e}\n{error_details}"
            print(error_msg, file=sys.stderr)
            self.send_log(f"AI 自動問候出錯: {str(e)}", "error")
            await db.add_log(f"AI auto greeting error: {str(e)}", "error")
    
    async def execute_matching_campaigns(self, lead_id: int, lead_data: Dict[str, Any]):
        """Execute campaigns that match the captured lead"""
        try:
            # Get all active campaigns
            campaigns = await db.get_all_campaigns()
            # 支持兩種字段名：isActive (前端格式) 和 is_active (數據庫格式)
            active_campaigns = [c for c in campaigns if c.get('isActive') or c.get('is_active')]
            
            if not active_campaigns:
                self.send_log(f"[活動] 沒有啟用的活動，跳過執行。關鍵詞: {lead_data.get('triggered_keyword')}", "info")
                return
            
            self.send_log(f"[活動] 檢查 {len(active_campaigns)} 個啟用的活動，關鍵詞: {lead_data.get('triggered_keyword')}", "info")
            
            # Get lead details
            lead = await db.get_lead(lead_id)
            if not lead:
                return
            
            source_group_id = None
            # Find source group ID from URL (prefer source_group_url, fallback to source_group)
            monitored_groups = await db.get_all_monitored_groups()
            group_url_to_match = lead_data.get('source_group_url') or lead_data.get('source_group')
            
            for group in monitored_groups:
                # Try matching by URL first
                if group_url_to_match and str(group.get('url')) == str(group_url_to_match):
                    source_group_id = group.get('id')
                    self.send_log(f"找到匹配的群組: {group.get('url')} (ID: {source_group_id})", "info")
                    break
            
            if not source_group_id:
                self.send_log(f"警告: 無法找到匹配的群組，URL: {group_url_to_match}", "warning")
            
            # Get keyword set IDs that matched
            keyword_set_ids = []
            keyword_sets = await db.get_all_keyword_sets()
            triggered_keyword = lead_data.get('triggered_keyword', '')
            
            self.send_log(f"[活動] 查找匹配的關鍵詞集，觸發關鍵詞: '{triggered_keyword}'", "info")
            
            for ks in keyword_sets:
                for keyword in ks.get('keywords', []):
                    keyword_text = keyword.get('keyword', '')
                    is_regex = keyword.get('isRegex', False)
                    
                    # 檢查匹配（支持正則）
                    matched = False
                    if is_regex:
                        try:
                            import re
                            pattern = re.compile(keyword_text, re.IGNORECASE)
                            matched = bool(pattern.search(triggered_keyword))
                        except:
                            matched = keyword_text.lower() in triggered_keyword.lower()
                    else:
                        matched = keyword_text.lower() in triggered_keyword.lower()
                    
                    if matched:
                        keyword_set_ids.append(ks.get('id'))
                        self.send_log(f"[活動] 關鍵詞 '{keyword_text}' 匹配，關鍵詞集ID: {ks.get('id')}", "info")
                        break
            
            # Check each campaign
            for campaign in active_campaigns:
                trigger = campaign.get('trigger', {})
                source_group_ids = trigger.get('sourceGroupIds', [])
                campaign_keyword_set_ids = trigger.get('keywordSetIds', [])
                
                # 詳細日誌
                self.send_log(f"[活動檢查] 活動: {campaign.get('name')}, 來源群組IDs: {source_group_ids}, 關鍵詞集IDs: {campaign_keyword_set_ids}", "info")
                self.send_log(f"[活動檢查] Lead來源群組ID: {source_group_id}, Lead關鍵詞集IDs: {keyword_set_ids}", "info")
                
                # Check if campaign matches
                # If no source groups specified, match all groups
                matches_source = not source_group_ids or (source_group_id and source_group_id in source_group_ids)
                # If no keyword sets specified, match all keywords
                matches_keyword = not campaign_keyword_set_ids or any(ks_id in campaign_keyword_set_ids for ks_id in keyword_set_ids)
                
                self.send_log(f"[活動檢查] 匹配結果: 來源群組={matches_source}, 關鍵詞={matches_keyword}", "info")
                
                if matches_source and matches_keyword:
                    self.send_log(f"✓✓✓ 活動匹配成功: {campaign.get('name')}，開始執行", "success")
                    # Execute campaign
                    await self.execute_campaign(campaign, lead_id, lead_data)
                else:
                    self.send_log(f"✗ 活動不匹配: {campaign.get('name')} (來源群組: {matches_source}, 關鍵詞: {matches_keyword})", "info")
        
        except Exception as e:
            self.send_log(f"Error executing matching campaigns: {str(e)}", "error")
    
    async def execute_campaign(self, campaign: Dict[str, Any], lead_id: int, lead_data: Dict[str, Any]):
        """Execute a single campaign for a lead"""
        try:
            import random
            
            # Get action from campaign (actions is a list)
            actions = campaign.get('actions', [])
            if actions and len(actions) > 0:
                action = actions[0]
            else:
                # Fallback to direct campaign fields (for backward compatibility)
                action = {
                    'templateId': campaign.get('actionTemplateId'),
                    'minDelaySeconds': campaign.get('actionMinDelaySeconds', 30),
                    'maxDelaySeconds': campaign.get('actionMaxDelaySeconds', 120)
                }
            
            template_id = action.get('templateId')
            
            if not template_id:
                self.send_log(f"Campaign {campaign.get('name')} has no template", "warning")
                return
            
            # Get template
            templates = await db.get_all_templates()
            template = next((t for t in templates if t.get('id') == template_id), None)
            
            if not template or not template.get('isActive'):
                self.send_log(f"模板 ID {template_id} 不存在或未激活。请检查模板设置。", "warning")
                return
            
            # Generate message from template
            message = await self.generate_message_from_template(template, lead_data)
            
            if not message:
                self.send_log(f"Failed to generate message for campaign {campaign.get('name')}", "error")
                return
            
            # Calculate delay
            min_delay = action.get('minDelaySeconds', 30)
            max_delay = action.get('maxDelaySeconds', 120)
            delay = random.randint(min_delay, max_delay)
            
            # Schedule message sending using queue
            scheduled_time = datetime.now() + timedelta(seconds=delay)
            
            # Get sender accounts
            accounts = await db.get_all_accounts()
            sender_accounts = [a for a in accounts if a.get('role') == 'Sender' and a.get('status') == 'Online']
            
            if not sender_accounts:
                self.send_log(f"No online sender accounts available for campaign '{campaign.get('name')}'", "warning")
                await db.add_interaction(lead_id, 'Campaign Failed', "No online sender accounts available")
                return
            
            # Select account (round-robin or random)
            selected_account = random.choice(sender_accounts)
            
            # Add to message queue with scheduled time
            try:
                message_id = await self.message_queue.add_message(
                    phone=selected_account.get('phone'),
                    user_id=str(lead_data.get('user_id')),
                    text=message,
                    priority=MessagePriority.NORMAL,
                    scheduled_at=scheduled_time,
                    callback=self._on_message_sent_callback(lead_id)
                )
                
                # Update lead with campaign
                await db.update_lead(lead_id, {'campaignId': campaign.get('id')})
                await db.add_interaction(lead_id, 'Campaign Triggered', f"Campaign '{campaign.get('name')}' triggered, message queued (ID: {message_id})")
                
                self.send_log(f"Campaign '{campaign.get('name')}' triggered for lead {lead_id}, message queued (ID: {message_id})", "info")
            except Exception as e:
                self.send_log(f"Error queueing campaign message: {str(e)}", "error")
                await db.add_interaction(lead_id, 'Campaign Failed', f"Failed to queue message: {str(e)}")
        
        except Exception as e:
            self.send_log(f"Error executing campaign: {str(e)}", "error")
    
    async def generate_message_from_template(self, template: Dict[str, Any], lead_data: Dict[str, Any]) -> str:
        """Generate message from template using variable substitution"""
        try:
            import datetime
            import random
            
            prompt = template.get('prompt', '')
            
            # 用戶信息變量
            username = lead_data.get('username', '')
            first_name = lead_data.get('first_name', '')
            last_name = lead_data.get('last_name', '')
            name = first_name or username or 'User'  # 優先使用名字
            
            # 觸發信息變量
            keyword = lead_data.get('triggered_keyword', '')
            user_message = lead_data.get('user_message', lead_data.get('message', ''))
            source_group = lead_data.get('source_group', '')
            group_name = lead_data.get('group_name', source_group)
            
            # 時間變量
            now = datetime.datetime.now()
            current_time = now.strftime('%H:%M')
            current_date = now.strftime('%Y年%m月%d日')
            
            # 隨機表情
            random_emojis = ['😊', '🌟', '💫', '✨', '🎉', '👋', '💪', '🔥', '❤️', '🙌', '😄', '🤝']
            random_emoji = random.choice(random_emojis)
            
            # 變量替換
            message = prompt
            message = message.replace('{username}', username or 'User')
            message = message.replace('{firstName}', first_name)
            message = message.replace('{lastName}', last_name)
            message = message.replace('{name}', name)
            message = message.replace('{keyword}', keyword)
            message = message.replace('{message}', user_message)
            message = message.replace('{sourceGroup}', source_group)
            message = message.replace('{groupName}', group_name)
            message = message.replace('{triggeredKeyword}', keyword)  # 兼容舊變量
            message = message.replace('{time}', current_time)
            message = message.replace('{date}', current_date)
            message = message.replace('{random}', random_emoji)
            
            # 清理未替換的變量（用空字符串替換）
            import re
            message = re.sub(r'\{[^}]+\}', '', message)
            
            return message.strip()
        
        except Exception as e:
            self.send_log(f"Error generating message from template: {str(e)}", "error")
            return None
    
    async def send_campaign_message_after_delay(self, campaign: Dict[str, Any], lead_id: int, lead_data: Dict[str, Any], message: str, delay: int):
        """Send campaign message after delay"""
        try:
            import asyncio
            
            # Wait for delay
            await asyncio.sleep(delay)
            
            # Get sender accounts
            accounts = await db.get_all_accounts()
            sender_accounts = [a for a in accounts if a.get('role') == 'Sender' and a.get('status') == 'Online']
            
            if not sender_accounts:
                self.send_log(f"No online sender accounts available for campaign '{campaign.get('name')}'", "warning")
                await db.add_interaction(lead_id, 'Campaign Failed', "No online sender accounts available")
                return
            
            # Select account (round-robin or random)
            import random
            selected_account = random.choice(sender_accounts)
            
            # Check daily send limit (已互動用戶不受限額限制)
            user_id = lead_data.get('user_id')
            has_interacted = await self._user_has_interacted(user_id)
            
            if not has_interacted:
                # 未互動用戶需要檢查限額
                if selected_account.get('dailySendCount', 0) >= selected_account.get('dailySendLimit', 50):
                    self.send_log(f"Account {selected_account.get('phone')} reached daily send limit", "warning")
                    # Try another account
                    available_accounts = [a for a in sender_accounts if a.get('dailySendCount', 0) < a.get('dailySendLimit', 50)]
                    if available_accounts:
                        selected_account = random.choice(available_accounts)
                    else:
                        await db.add_interaction(lead_id, 'Campaign Failed', "All sender accounts reached daily limit")
                        return
            else:
                # 已互動用戶不受限額限制
                self.send_log(f"User {user_id} has interacted before, exempt from daily limit", "info")
            
            # Send message
            user_id = lead_data.get('user_id')
            result = await self.telegram_manager.send_message(
                phone=selected_account.get('phone'),
                user_id=user_id,
                text=message
            )
            
            if result.get('success'):
                # Update daily send count (已互動用戶不計入限額)
                if not has_interacted:
                    await db.update_account(selected_account.get('id'), {
                        'dailySendCount': selected_account.get('dailySendCount', 0) + 1
                    })
                
                # Update lead
                await db.update_lead_status(lead_id, 'Contacted')
                await db.update_lead(lead_id, {'assignedTemplateId': campaign.get('actions', [{}])[0].get('templateId')})
                await db.add_interaction(lead_id, 'Message Sent', f"Message sent via campaign '{campaign.get('name')}'")
                
                self.send_log(f"Campaign message sent to lead {lead_id} via account {selected_account.get('phone')}", "success")
                
                # Send event
                self.send_event("message-sent", {
                    "leadId": lead_id,
                    "campaignId": campaign.get('id'),
                    "success": True
                })
            else:
                error_msg = result.get('error', 'Unknown error')
                await db.add_interaction(lead_id, 'Campaign Failed', f"Failed to send message: {error_msg}")
                self.send_log(f"Failed to send campaign message: {error_msg}", "error")
        
        except Exception as e:
            self.send_log(f"Error sending campaign message: {str(e)}", "error")
    
    async def daily_reset_task(self):
        """Background task to reset daily send counts at midnight"""
        try:
            while self.running:
                now = datetime.now()
                # Calculate next midnight
                next_midnight = now.replace(hour=0, minute=0, second=0, microsecond=0)
                if next_midnight <= now:
                    next_midnight = next_midnight + timedelta(days=1)
                
                # Wait until midnight
                wait_seconds = (next_midnight - now).total_seconds()
                await asyncio.sleep(wait_seconds)
                
                # Reset daily send counts
                if self.running:
                    await self.reset_daily_send_counts()
                    self.last_reset_date = datetime.now().date()
        
        except asyncio.CancelledError:
            pass
        except Exception as e:
            self.send_log(f"Error in daily reset task: {str(e)}", "error")
    
    async def reset_daily_send_counts(self):
        """Reset daily send counts for all accounts"""
        try:
            accounts = await db.get_all_accounts()
            for account in accounts:
                await db.update_account(account.get('id'), {'dailySendCount': 0})
            
            self.send_log(f"Daily send counts reset for {len(accounts)} accounts", "info")
            await db.add_log(f"Daily send counts reset for {len(accounts)} accounts", "info")
            
            # Send updated accounts
            accounts = await db.get_all_accounts()
            # Invalidate cache
            self._cache.pop("accounts", None)
            self._cache_timestamps.pop("accounts", None)
            self.send_event("accounts-updated", accounts)
        
        except Exception as e:
            self.send_log(f"Error resetting daily send counts: {str(e)}", "error")
    
    async def account_health_monitor_task(self):
        """Background task to periodically check account health and status"""
        try:
            while self.running:
                # Wait 5 minutes between checks
                await asyncio.sleep(300)  # 5 minutes
                
                if not self.running:
                    break
                
                # Check all online accounts
                await self.check_all_accounts_health()
        
        except asyncio.CancelledError:
            pass
        except Exception as e:
            self.send_log(f"Error in account health monitor task: {str(e)}", "error")
    
    async def check_all_accounts_health(self):
        """Check health and status of all accounts"""
        try:
            accounts = await db.get_all_accounts()
            online_accounts = [a for a in accounts if a.get('status') == 'Online']
            
            if not online_accounts:
                return
            
            self.send_log(f"Checking health for {len(online_accounts)} online accounts", "info")
            
            for account in online_accounts:
                try:
                    # Check account status
                    phone = account.get('phone')
                    status_result = await self.telegram_manager.check_account_status(phone)
                    
                    # Calculate health score (原有方法)
                    health_score = await self.calculate_health_score(account, status_result)
                    
                    # 增强的健康分析（账户健康监控增强）
                    if self.enhanced_health_monitor:
                        health_analysis = await self.enhanced_health_monitor.analyze_account_health(
                            account.get('id'),
                            account
                        )
                        
                        # 使用增强的健康分析结果更新健康分数
                        if health_analysis.get('ban_risk_score') is not None:
                            # 根据封禁风险调整健康分数
                            ban_risk = health_analysis.get('ban_risk_score', 0.0)
                            adjusted_score = health_score * (1.0 - ban_risk * 0.5)  # 封禁风险越高，健康分数越低
                            health_score = max(0, min(100, int(adjusted_score)))
                        
                        # 发送健康分析事件
                        self.send_event("account-health-analysis", {
                            "account_id": account.get('id'),
                            "phone": phone,
                            **health_analysis
                        })
                    
                    # Update account
                    updates = {
                        'status': status_result.get('status', account.get('status')),
                        'healthScore': health_score
                    }
                    await db.update_account(account.get('id'), updates)
                
                except Exception as e:
                    self.send_log(f"Error checking health for account {account.get('phone')}: {str(e)}", "error")
            
            # Send updated accounts
            accounts = await db.get_all_accounts()
            # Invalidate cache
            self._cache.pop("accounts", None)
            self._cache_timestamps.pop("accounts", None)
            self.send_event("accounts-updated", accounts)
        
        except Exception as e:
            self.send_log(f"Error checking all accounts health: {str(e)}", "error")
    
    async def queue_cleanup_task(self):
        """Background task to clean up old queue messages"""
        while self.running:
            try:
                await asyncio.sleep(3600)  # Run every hour
                await db.cleanup_old_queue_messages(days=7)
            except Exception as e:
                self.send_log(f"Error cleaning up queue messages: {str(e)}", "error")
                await asyncio.sleep(60)
    
    async def calculate_health_score(self, account: Dict[str, Any], status_result: Dict[str, Any]) -> int:
        """Calculate health score for an account"""
        try:
            base_score = 100
            
            # Status penalty
            status = status_result.get('status', 'Offline')
            if status == 'Banned':
                return 0
            elif status == 'Proxy Error':
                base_score -= 30
            elif status == 'Offline':
                base_score -= 20
            elif status != 'Online':
                base_score -= 10
            
            # Daily send limit penalty
            daily_send_count = account.get('dailySendCount', 0)
            daily_send_limit = account.get('dailySendLimit', 50)
            if daily_send_limit > 0:
                send_ratio = daily_send_count / daily_send_limit
                if send_ratio >= 1.0:
                    base_score -= 20  # Reached limit
                elif send_ratio >= 0.8:
                    base_score -= 10  # Near limit
            
            # Ensure score is between 0 and 100
            health_score = max(0, min(100, base_score))
            
            return int(health_score)
        
        except Exception as e:
            self.send_log(f"Error calculating health score: {str(e)}", "error")
            return account.get('healthScore', 100)
    
    async def handle_save_settings(self, payload: Dict[str, Any]):
        """Handle save-settings command"""
        try:
            settings = payload.get('settings', {})
            for key, value in settings.items():
                await db.set_setting(key, value)
            
            await db.add_log("Settings saved", "success")
            self.send_log("Settings saved successfully", "success")
            
            # Send updated settings
            all_settings = await db.get_all_settings()
            self.send_event("settings-updated", all_settings)
        
        except Exception as e:
            self.send_log(f"Error saving settings: {str(e)}", "error")
    
    async def handle_get_settings(self):
        """Handle get-settings command"""
        try:
            settings = await db.get_all_settings()
            self.send_event("settings-loaded", settings)
        
        except Exception as e:
            self.send_log(f"Error loading settings: {str(e)}", "error")
    
    async def handle_get_queue_status(self, payload: Dict[str, Any]):
        """Handle get-queue-status command"""
        try:
            phone = payload.get('phone')  # Optional, if None returns all
            status = await self.message_queue.get_queue_status(phone)
            self.send_event("queue-status", status)
        
        except Exception as e:
            self.send_log(f"Error getting queue status: {str(e)}", "error")
    
    async def handle_clear_queue(self, payload: Dict[str, Any]):
        """Handle clear-queue command"""
        try:
            phone = payload.get('phone')
            status_str = payload.get('status')  # Optional: 'pending', 'failed', etc.
            
            if not phone:
                self.send_log("Phone number required to clear queue", "error")
                return
            
            # Convert status string to enum if provided
            status = None
            if status_str:
                from message_queue import MessageStatus
                status_map = {
                    'pending': MessageStatus.PENDING,
                    'processing': MessageStatus.PROCESSING,
                    'failed': MessageStatus.FAILED,
                    'retrying': MessageStatus.RETRYING
                }
                status = status_map.get(status_str.lower())
            
            await self.message_queue.clear_queue(phone, status)
            await db.add_log(f"Queue cleared for {phone}", "info")
            self.send_log(f"Queue cleared for {phone}", "success")
            
            # Send updated status
            status = await self.message_queue.get_queue_status(phone)
            self.send_event("queue-status", status)
        
        except Exception as e:
            self.send_log(f"Error clearing queue: {str(e)}", "error")
    
    async def handle_pause_queue(self, payload: Dict[str, Any]):
        """Handle pause-queue command"""
        try:
            phone = payload.get('phone')
            if not phone:
                self.send_log("Phone number required to pause queue", "error")
                return
            
            await self.message_queue.pause_queue(phone)
            self.send_log(f"Queue paused for {phone}", "info")
            
            # Send updated queue status
            queue_status = await self.message_queue.get_queue_status(phone)
            queue_status['paused'] = True
            self.send_event("queue-status", queue_status)
        
        except Exception as e:
            self.send_log(f"Error pausing queue: {str(e)}", "error")
    
    async def handle_resume_queue(self, payload: Dict[str, Any]):
        """Handle resume-queue command"""
        try:
            phone = payload.get('phone')
            if not phone:
                self.send_log("Phone number required to resume queue", "error")
                return
            
            await self.message_queue.resume_queue(phone)
            self.send_log(f"Queue resumed for {phone}", "success")
            
            # Send updated queue status
            queue_status = await self.message_queue.get_queue_status(phone)
            queue_status['paused'] = False
            self.send_event("queue-status", queue_status)
        
        except Exception as e:
            self.send_log(f"Error resuming queue: {str(e)}", "error")
    
    async def handle_delete_queue_message(self, payload: Dict[str, Any]):
        """Handle delete-queue-message command"""
        try:
            phone = payload.get('phone')
            message_id = payload.get('messageId')
            
            if not phone or not message_id:
                self.send_log("Phone number and message ID required", "error")
                return
            
            success = await self.message_queue.delete_message(phone, message_id)
            if success:
                self.send_log(f"Message {message_id} deleted from queue", "success")
                # Send updated queue status
                queue_status = await self.message_queue.get_queue_status(phone)
                self.send_event("queue-status", queue_status)
            else:
                self.send_log(f"Message {message_id} not found in queue", "warning")
        
        except Exception as e:
            self.send_log(f"Error deleting queue message: {str(e)}", "error")
    
    async def handle_update_queue_message_priority(self, payload: Dict[str, Any]):
        """Handle update-queue-message-priority command"""
        try:
            phone = payload.get('phone')
            message_id = payload.get('messageId')
            priority_str = payload.get('priority')  # 'HIGH', 'NORMAL', 'LOW'
            
            if not phone or not message_id or not priority_str:
                self.send_log("Phone number, message ID, and priority required", "error")
                return
            
            from message_queue import MessagePriority
            priority_map = {
                'HIGH': MessagePriority.HIGH,
                'NORMAL': MessagePriority.NORMAL,
                'LOW': MessagePriority.LOW
            }
            priority = priority_map.get(priority_str.upper())
            
            if not priority:
                self.send_log(f"Invalid priority: {priority_str}", "error")
                return
            
            success = await self.message_queue.update_message_priority(phone, message_id, priority)
            if success:
                self.send_log(f"Message {message_id} priority updated to {priority_str}", "success")
                # Send updated queue status
                queue_status = await self.message_queue.get_queue_status(phone)
                self.send_event("queue-status", queue_status)
            else:
                self.send_log(f"Message {message_id} not found in queue", "warning")
        
        except Exception as e:
            self.send_log(f"Error updating message priority: {str(e)}", "error")
    
    async def handle_get_queue_messages(self, payload: Dict[str, Any]):
        """Handle get-queue-messages command"""
        try:
            phone = payload.get('phone')  # Optional
            status_str = payload.get('status')  # Optional
            limit = payload.get('limit', 100)
            
            status = None
            if status_str:
                from message_queue import MessageStatus
                status_map = {
                    'pending': MessageStatus.PENDING,
                    'processing': MessageStatus.PROCESSING,
                    'failed': MessageStatus.FAILED,
                    'retrying': MessageStatus.RETRYING,
                    'completed': MessageStatus.COMPLETED
                }
                status = status_map.get(status_str.lower())
            
            messages = await self.message_queue.get_queue_messages(phone, status, limit)
            self.send_event("queue-messages", {
                "phone": phone,
                "messages": messages,
                "count": len(messages)
            })
        
        except Exception as e:
            self.send_log(f"Error getting queue messages: {str(e)}", "error")
    
    async def handle_get_logs(self, payload: Dict[str, Any]):
        """Handle get-logs command with filtering"""
        try:
            limit = payload.get('limit', 100)
            log_type = payload.get('type')  # Optional: 'info', 'success', 'warning', 'error'
            start_date = payload.get('startDate')  # Optional: ISO format
            end_date = payload.get('endDate')  # Optional: ISO format
            search_query = payload.get('search')  # Optional: search string
            
            logs = await db.get_recent_logs(
                limit=limit,
                log_type=log_type,
                start_date=start_date,
                end_date=end_date,
                search_query=search_query
            )
            
            # Format timestamps
            for log in logs:
                if isinstance(log.get('timestamp'), str):
                    pass  # Already a string
                else:
                    log['timestamp'] = datetime.fromisoformat(log['timestamp']).isoformat() + "Z"
            
            # Get total count
            total_count = await db.get_logs_count(
                log_type=log_type,
                start_date=start_date,
                end_date=end_date,
                search_query=search_query
            )
            
            self.send_event("logs-loaded", {
                "logs": logs,
                "total": total_count,
                "limit": limit
            })
        
        except Exception as e:
            self.send_log(f"Error getting logs: {str(e)}", "error")
    
    async def handle_export_logs(self, payload: Dict[str, Any]):
        """Handle export-logs command"""
        try:
            import openpyxl
            from openpyxl import Workbook
            from pathlib import Path
            
            log_type = payload.get('type')
            start_date = payload.get('startDate')
            end_date = payload.get('endDate')
            search_query = payload.get('search')
            
            # Get all matching logs (no limit for export)
            logs = await db.get_recent_logs(
                limit=10000,  # Large limit for export
                log_type=log_type,
                start_date=start_date,
                end_date=end_date,
                search_query=search_query
            )
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Logs"
            
            # Headers
            ws.append(["ID", "Timestamp", "Type", "Message"])
            
            # Data
            for log in logs:
                ws.append([
                    log.get('id'),
                    log.get('timestamp'),
                    log.get('type'),
                    log.get('message')
                ])
            
            # Save to file
            export_dir = config.TEMPLATES_DIR.parent / "exports"
            export_dir.mkdir(exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"logs_export_{timestamp}.xlsx"
            filepath = export_dir / filename
            
            wb.save(filepath)
            
            await db.add_log(f"Logs exported to {filename}", "success")
            self.send_log(f"Logs exported to {filename}", "success")
            
            # Send file path to frontend
            self.send_event("logs-exported", {
                "filepath": str(filepath),
                "filename": filename,
                "count": len(logs)
            })
        
        except Exception as e:
            self.send_log(f"Error exporting logs: {str(e)}", "error")
    
    async def handle_add_keyword_set(self, payload: Dict[str, Any]):
        """Handle add-keyword-set command"""
        try:
            name = payload.get('name')
            
            if not name:
                error_msg = "關鍵詞集名稱不能為空"
                self.send_log(error_msg, "error")
                self.send_event("keyword-set-error", {
                    "success": False,
                    "error": error_msg,
                    "name": name
                })
                return
            
            # Validate keyword set name
            is_valid, error = KeywordValidator.validate_keyword_set_name(name)
            if not is_valid:
                self.send_log(f"驗證錯誤: {error}", "error")
                self.send_event("keyword-set-validation-error", {
                    "errors": [error],
                    "name": name
                })
                self.send_event("keyword-set-error", {
                    "success": False,
                    "error": error,
                    "name": name
                })
                handle_error(
                    AppError(ErrorType.VALIDATION_ERROR, error, {"name": name}),
                    {"command": "add-keyword-set", "payload": payload}
                )
                return
            
            # Check if keyword set already exists
            try:
                existing_sets = await db.get_all_keyword_sets()
                for existing_set in existing_sets:
                    if existing_set.get('name') == name:
                        error_msg = f"關鍵詞集 '{name}' 已存在"
                        self.send_log(error_msg, "error")
                        self.send_event("keyword-set-error", {
                            "success": False,
                            "error": error_msg,
                            "name": name
                        })
                        return
            except Exception as check_err:
                import sys
                print(f"[Backend] Error checking existing keyword sets: {check_err}", file=sys.stderr)
                # Continue anyway, let the database handle the uniqueness constraint
            
            # Add keyword set
            try:
                keyword_set_id = await db.add_keyword_set(name)
                await db.add_log(f"關鍵詞集 '{name}' 已添加", "success")
                self.send_log(f"關鍵詞集 '{name}' 添加成功 (ID: {keyword_set_id})", "success")
                self._invalidate_cache("keyword_sets")
                await self.send_keyword_sets_update()
                # Send success event
                self.send_event("keyword-set-error", {
                    "success": True,
                    "message": f"關鍵詞集 '{name}' 添加成功",
                    "keywordSetId": keyword_set_id,
                    "name": name
                })
            except Exception as db_err:
                import sys
                import traceback
                error_str = str(db_err).lower()
                error_details = traceback.format_exc()
                print(f"[Backend] Database error adding keyword set: {error_details}", file=sys.stderr)
                
                # Check for specific database errors
                if "unique" in error_str or "already exists" in error_str:
                    error_msg = f"關鍵詞集 '{name}' 已存在"
                elif "database" in error_str and ("locked" in error_str or "corrupt" in error_str or "malformed" in error_str):
                    error_msg = f"數據庫錯誤：數據庫可能已損壞，請使用 rebuild_database.py 重建數據庫"
                elif "no such table" in error_str:
                    error_msg = f"數據庫錯誤：關鍵詞集表不存在，請重建數據庫"
                else:
                    error_msg = f"添加關鍵詞集失敗: {str(db_err)}"
                
                self.send_log(error_msg, "error")
                await db.add_log(f"添加關鍵詞集失敗: {error_msg}", "error")
                self.send_event("keyword-set-error", {
                    "success": False,
                    "error": error_msg,
                    "name": name,
                    "details": str(db_err)
                })
        
        except ValidationError as e:
            self.send_log(f"驗證錯誤: {e.message}", "error")
            self.send_event("keyword-set-validation-error", {
                "errors": [e.message],
                "field": e.field
            })
            self.send_event("keyword-set-error", {
                "success": False,
                "error": e.message,
                "field": e.field
            })
        except Exception as e:
            import sys
            import traceback
            error_details = traceback.format_exc()
            error_msg = f"添加關鍵詞集時發生未知錯誤: {str(e)}"
            print(f"[Backend] Unexpected error adding keyword set: {error_details}", file=sys.stderr)
            self.send_log(error_msg, "error")
            self.send_event("keyword-set-error", {
                "success": False,
                "error": error_msg,
                "details": str(e)
            })
            handle_error(e, {"command": "add-keyword-set", "payload": payload})
    
    async def handle_remove_keyword_set(self, payload: Dict[str, Any]):
        """Handle remove-keyword-set command - idempotent operation"""
        import sys
        try:
            set_id = payload.get('id')
            
            if set_id is None:
                error_msg = "關鍵詞集 ID 不能為空"
                print(f"[Backend] Remove keyword set error: {error_msg}", file=sys.stderr)
                self.send_log(error_msg, "error")
                self.send_event("keyword-set-error", {
                    "success": False,
                    "error": error_msg
                })
                return
            
            print(f"[Backend] Removing keyword set with ID: {set_id}", file=sys.stderr)
            
            # Delete from database (idempotent - won't error if already deleted)
            await db.delete_keyword_set(set_id)
            print(f"[Backend] Keyword set {set_id} delete operation completed", file=sys.stderr)
            
            # Invalidate cache and send update
            self._invalidate_cache("keyword_sets")
            print(f"[Backend] Cache invalidated, sending keyword sets update...", file=sys.stderr)
            
            await self.send_keyword_sets_update()
            print(f"[Backend] Keyword sets update sent", file=sys.stderr)
            
            # Send success event (always success since delete is idempotent)
            self.send_event("keyword-set-error", {
                "success": True,
                "message": f"關鍵詞集 {set_id} 刪除成功"
            })
        
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            error_msg = f"刪除關鍵詞集失敗: {str(e)}"
            print(f"[Backend] Error removing keyword set: {error_details}", file=sys.stderr)
            self.send_log(error_msg, "error")
            self.send_event("keyword-set-error", {
                "success": False,
                "error": error_msg,
                "details": str(e)
            })
    
    async def handle_add_keyword(self, payload: Dict[str, Any]):
        """Handle add-keyword command"""
        try:
            set_id = payload.get('setId')
            keyword = payload.get('keyword')
            is_regex = payload.get('isRegex', False)
            
            # Validate keyword
            is_valid, error = validate_keyword(keyword, is_regex)
            if not is_valid:
                self.send_log(f"Validation error: {error}", "error")
                self.send_event("keyword-validation-error", {
                    "errors": [error],
                    "keyword": keyword,
                    "is_regex": is_regex
                })
                handle_error(
                    AppError(ErrorType.VALIDATION_ERROR, error, {"keyword": keyword, "is_regex": is_regex}),
                    {"command": "add-keyword", "payload": payload}
                )
                return
            
            # 檢查關鍵詞是否已存在
            existing_keywords = await db.get_keywords_by_set(set_id)
            keyword_exists = any(
                k.get('keyword') == keyword and k.get('isRegex') == is_regex 
                for k in existing_keywords
            )
            
            if keyword_exists:
                self.send_log(f"Keyword '{keyword}' already exists in set {set_id}", "warning")
                # 仍然發送更新事件以確保前端狀態同步
                await self.send_keyword_sets_update()
                return
            
            keyword_id = await db.add_keyword(set_id, keyword, is_regex)
            await db.add_log(f"Keyword '{keyword}' added to set {set_id}", "success")
            self._invalidate_cache("keyword_sets")
            await self.send_keyword_sets_update()
        
        except ValidationError as e:
            self.send_log(f"Validation error: {e.message}", "error")
            self.send_event("keyword-validation-error", {
                "errors": [e.message],
                "field": e.field
            })
        except Exception as e:
            self.send_log(f"Error adding keyword: {str(e)}", "error")
            handle_error(e, {"command": "add-keyword", "payload": payload})
    
    async def handle_remove_keyword(self, payload: Dict[str, Any]):
        """Handle remove-keyword command"""
        try:
            keyword_id = payload.get('keywordId')
            set_id = payload.get('setId')
            
            if not keyword_id:
                self.send_log("Error: keywordId is required", "error")
                return
            
            await db.remove_keyword(keyword_id)
            await db.add_log(f"Keyword {keyword_id} removed", "success")
            
            # 確保發送更新事件
            self._invalidate_cache("keyword_sets")
            await self.send_keyword_sets_update()
        
        except Exception as e:
            self.send_log(f"Error removing keyword: {str(e)}", "error")
            # 即使發生錯誤，也發送更新事件以確保前端狀態同步
            try:
                await self.send_keyword_sets_update()
            except:
                pass
    
    async def handle_add_group(self, payload: Dict[str, Any]):
        """Handle add-group command"""
        try:
            url = payload.get('url')
            name = payload.get('name', url)  # Use URL as name if not provided
            keyword_set_ids = payload.get('keywordSetIds', [])
            
            # Validate group URL
            is_valid, error = validate_group_url(url)
            if not is_valid:
                self.send_log(f"Validation error: {error}", "error")
                self.send_event("group-validation-error", {
                    "errors": [error],
                    "url": url
                })
                handle_error(
                    AppError(ErrorType.VALIDATION_ERROR, error, {"url": url}),
                    {"command": "add-group", "payload": payload}
                )
                return
            
            # Validate group name (optional)
            if name:
                is_valid, error = GroupValidator.validate_group_name(name)
                if not is_valid:
                    self.send_log(f"Validation error: {error}", "error")
                    self.send_event("group-validation-error", {
                        "errors": [error],
                        "name": name
                    })
                    handle_error(
                        AppError(ErrorType.VALIDATION_ERROR, error, {"name": name}),
                        {"command": "add-group", "payload": payload}
                    )
                    return
            
            # ========== 新增：預檢查監控號入群狀態 ==========
            membership_status = None
            accounts = await db.get_all_accounts()
            listener_accounts = [a for a in accounts if a.get('role') == 'Listener' and a.get('status') == 'Online']
            
            if listener_accounts:
                # Check if any listener account is in this group
                for account in listener_accounts:
                    phone = account.get('phone')
                    try:
                        check_result = await self.telegram_manager.check_group_membership(phone, url)
                        if check_result.get("is_member"):
                            membership_status = {
                                "is_member": True,
                                "account": phone,
                                "chat_title": check_result.get("chat_title")
                            }
                            break
                        elif check_result.get("can_join"):
                            membership_status = {
                                "is_member": False,
                                "can_join": True,
                                "is_private": check_result.get("is_private", False),
                                "reason": check_result.get("reason")
                            }
                    except Exception as e:
                        import sys
                        print(f"[Backend] Error checking membership for {url}: {e}", file=sys.stderr)
            
            # Send membership status event
            if membership_status:
                self.send_event("group-membership-status", {
                    "url": url,
                    "status": membership_status
                })
                
                if membership_status.get("is_member"):
                    self.send_log(f"✓ 監控號已在群組中: {membership_status.get('chat_title', url)}", "success")
                elif membership_status.get("can_join"):
                    if membership_status.get("is_private"):
                        self.send_log(f"⚠ 監控號未加入此群組（私有群），需要手動加入: {url}", "warning")
                    else:
                        self.send_log(f"ℹ 監控號未加入此群組（公開群），啟動監控時將自動加入: {url}", "info")
            else:
                if listener_accounts:
                    self.send_log(f"⚠ 無法檢查群組成員狀態: {url}", "warning")
                else:
                    self.send_log(f"ℹ 沒有在線監控號，無法檢查群組狀態: {url}", "info")
            
            # ========== 結束新增 ==========
            
            # Check if group already exists
            existing = await db.get_group_by_url(url)
            if existing:
                # Update existing group
                group_id = await db.add_group(url, name, keyword_set_ids)
                await db.add_log(f"Group '{name}' updated (URL already exists)", "info")
            else:
                # Add new group
                group_id = await db.add_group(url, name, keyword_set_ids)
                await db.add_log(f"Group '{name}' added", "success")
            await self.send_groups_update()
        
        except ValidationError as e:
            self.send_log(f"Validation error: {e.message}", "error")
            self.send_event("group-validation-error", {
                "errors": [e.message],
                "field": e.field
            })
        except Exception as e:
            self.send_log(f"Error adding group: {str(e)}", "error")
            handle_error(e, {"command": "add-group", "payload": payload})
    
    async def handle_join_group(self, payload: Dict[str, Any]):
        """Handle join-group command - manually join a group with a specific account"""
        import sys
        print(f"[Backend] handle_join_group called with payload: {payload}", file=sys.stderr)
        
        try:
            phone = payload.get('phone')
            group_url = payload.get('groupUrl')
            
            print(f"[Backend] Extracted phone={phone}, group_url={group_url}", file=sys.stderr)
            
            if not phone or not group_url:
                self.send_log("缺少必要參數（電話或群組 URL）", "error")
                self.send_event("group-join-result", {
                    "success": False,
                    "phone": phone,
                    "groupUrl": group_url,
                    "error": "缺少必要參數"
                })
                return
            
            self.send_log(f"正在嘗試讓 {phone} 加入群組 {group_url}...", "info")
            
            # Use telegram manager to join the group
            result = await self.telegram_manager.join_group(phone, group_url)
            
            if result.get("success"):
                if result.get("already_member"):
                    self.send_log(f"✓ {phone} 已經在群組中: {result.get('chat_title', group_url)}", "success")
                else:
                    self.send_log(f"✓ {phone} 成功加入群組: {result.get('chat_title', group_url)}", "success")
                
                # Send success event
                self.send_event("group-join-result", {
                    "success": True,
                    "phone": phone,
                    "groupUrl": group_url,
                    "chatTitle": result.get("chat_title"),
                    "chatId": result.get("chat_id"),
                    "alreadyMember": result.get("already_member", False)
                })
            else:
                error_msg = result.get("error", "未知錯誤")
                self.send_log(f"✗ {phone} 加入群組失敗: {error_msg}", "error")
                
                # Send failure event
                self.send_event("group-join-result", {
                    "success": False,
                    "phone": phone,
                    "groupUrl": group_url,
                    "error": error_msg
                })
        
        except Exception as e:
            import traceback
            traceback.print_exc(file=sys.stderr)
            self.send_log(f"加入群組時發生錯誤: {str(e)}", "error")
            self.send_event("group-join-result", {
                "success": False,
                "phone": payload.get('phone'),
                "groupUrl": payload.get('groupUrl'),
                "error": str(e)
            })
    
    async def handle_remove_group(self, payload: Dict[str, Any]):
        """Handle remove-group command"""
        try:
            group_id = payload.get('id')
            await db.delete_group(group_id)
            await db.add_log(f"Group {group_id} removed", "success")
            await self.send_groups_update()
        
        except Exception as e:
            self.send_log(f"Error removing group: {str(e)}", "error")
    
    async def handle_add_template(self, payload: Dict[str, Any]):
        """Handle add-template command"""
        try:
            # Validate template data
            is_valid, errors = validate_template(payload)
            if not is_valid:
                error_message = "Validation failed: " + "; ".join(errors)
                self.send_log(error_message, "error")
                self.send_event("template-validation-error", {
                    "errors": errors,
                    "template_data": payload
                })
                handle_error(
                    AppError(ErrorType.VALIDATION_ERROR, error_message, {"errors": errors}),
                    {"command": "add-template", "payload": payload}
                )
                return
            
            name = payload.get('name')
            prompt = payload.get('prompt')
            
            # Check if template with same name already exists
            existing = await db.get_template_by_name(name)
            if existing:
                # Template already exists, don't create duplicate
                self.send_log(f"Template '{name}' already exists (ID: {existing['id']})", "warning")
                self.send_event("template-already-exists", {
                    "templateId": existing['id'],
                    "name": name,
                    "message": f"模板 '{name}' 已存在，未創建重複模板"
                })
            else:
                # Add new template
                template_id = await db.add_template(name, prompt)
                await db.add_log(f"Template '{name}' added", "success")
            
            await self.send_templates_update()
        
        except ValidationError as e:
            self.send_log(f"Validation error: {e.message}", "error")
            self.send_event("template-validation-error", {
                "errors": [e.message],
                "field": e.field
            })
        except Exception as e:
            self.send_log(f"Error adding template: {str(e)}", "error")
            handle_error(e, {"command": "add-template", "payload": payload})
    
    async def handle_remove_template(self, payload: Dict[str, Any]):
        """Handle remove-template command"""
        try:
            template_id = payload.get('id')
            await db.delete_template(template_id)
            await db.add_log(f"Template {template_id} removed", "success")
            await self.send_templates_update()
        
        except Exception as e:
            self.send_log(f"Error removing template: {str(e)}", "error")
    
    async def handle_toggle_template_status(self, payload: Dict[str, Any]):
        """Handle toggle-template-status command"""
        try:
            template_id = payload.get('id')
            await db.toggle_template_status(template_id)
            await db.add_log(f"Template {template_id} status toggled", "success")
            await self.send_templates_update()
        
        except Exception as e:
            self.send_log(f"Error toggling template status: {str(e)}", "error")
    
    async def handle_add_campaign(self, payload: Dict[str, Any]):
        """Handle add-campaign command"""
        try:
            # Validate campaign data
            is_valid, errors = validate_campaign(payload)
            if not is_valid:
                error_message = "Validation failed: " + "; ".join(errors)
                self.send_log(error_message, "error")
                self.send_event("campaign-validation-error", {
                    "errors": errors,
                    "campaign_data": payload
                })
                handle_error(
                    AppError(ErrorType.VALIDATION_ERROR, error_message, {"errors": errors}),
                    {"command": "add-campaign", "payload": payload}
                )
                return
            
            # 檢查活動是否已存在
            campaign_name = payload.get('name', '').strip()
            existing_campaigns = await db.get_all_campaigns()
            existing = next((c for c in existing_campaigns if c.get('name') == campaign_name), None)
            
            if existing:
                # 活動已存在，發送警告事件
                self.send_log(f"Campaign '{campaign_name}' already exists (ID: {existing.get('id')})", "warning")
                self.send_event("campaign-already-exists", {
                    "campaignId": existing.get('id'),
                    "name": campaign_name,
                    "message": f"活動 '{campaign_name}' 已存在，未創建重複活動"
                })
                # 仍然發送更新事件以確保前端狀態同步
                await self.send_campaigns_update()
                return
            
            campaign_id = await db.add_campaign(payload)
            await db.add_log(f"Campaign '{campaign_name}' added", "success")
            await self.send_campaigns_update()
        
        except ValidationError as e:
            self.send_log(f"Validation error: {e.message}", "error")
            self.send_event("campaign-validation-error", {
                "errors": [e.message],
                "field": e.field
            })
        except Exception as e:
            self.send_log(f"Error adding campaign: {str(e)}", "error")
            handle_error(e, {"command": "add-campaign", "payload": payload})
            # 即使出錯，也發送更新事件以確保前端狀態同步
            try:
                await self.send_campaigns_update()
            except:
                pass
    
    async def handle_remove_campaign(self, payload: Dict[str, Any]):
        """Handle remove-campaign command"""
        try:
            campaign_id = payload.get('id')
            await db.delete_campaign(campaign_id)
            await db.add_log(f"Campaign {campaign_id} removed", "success")
            await self.send_campaigns_update()
        
        except Exception as e:
            self.send_log(f"Error removing campaign: {str(e)}", "error")
    
    async def handle_toggle_campaign_status(self, payload: Dict[str, Any]):
        """Handle toggle-campaign-status command"""
        try:
            campaign_id = payload.get('id')
            await db.toggle_campaign_status(campaign_id)
            await db.add_log(f"Campaign {campaign_id} status toggled", "success")
            await self.send_campaigns_update()
        
        except Exception as e:
            self.send_log(f"Error toggling campaign status: {str(e)}", "error")
    
    async def handle_send_message(self, payload: Dict[str, Any]):
        """Handle send-message command with message queue"""
        try:
            lead_id = payload.get('leadId')
            account_phone = payload.get('accountPhone')
            user_id = payload.get('userId')
            source_group = payload.get('sourceGroup')  # 源群組（用於獲取用戶信息）
            message_text = payload.get('message')
            attachment = payload.get('attachment')
            priority = payload.get('priority', 'normal')  # high, normal, low
            scheduled_at = payload.get('scheduledAt')  # Optional ISO datetime string
            
            if not account_phone or not user_id or not message_text:
                self.send_log("Missing required parameters for sending message", "error")
                return
            
            # 嘗試獲取群組 URL（用於加入群組）
            source_group_url = source_group
            if source_group:
                try:
                    source_group_str = str(source_group)
                    
                    # 如果已經是 URL，直接使用
                    if 't.me/' in source_group_str or source_group_str.startswith('@'):
                        source_group_url = source_group_str
                        print(f"[Backend] source_group is already a URL: {source_group_url}", file=sys.stderr)
                    else:
                        # source_group 是 Telegram chat_id，需要從 monitoredGroups 查找對應的 URL
                        # monitoredGroups 存儲的是 URL，而不是 chat_id
                        # 我們需要從 lead 數據中獲取 source_group_url
                        if lead_id:
                            lead = await db.get_lead(lead_id)
                            if lead:
                                # 檢查 lead 中是否有存儲群組 URL
                                lead_source = lead.get('source_group') or lead.get('sourceGroup')
                                if lead_source and ('t.me/' in str(lead_source) or str(lead_source).startswith('@')):
                                    source_group_url = str(lead_source)
                                    print(f"[Backend] Found group URL from lead: {source_group_url}", file=sys.stderr)
                                else:
                                    # 如果 lead 中沒有 URL，嘗試從 monitoredGroups 查找
                                    # 由於 monitoredGroups 存儲的是 URL，而 source_group 是 chat_id
                                    # 我們無法直接匹配，只能使用第一個匹配的 URL
                                    monitored_groups = await db.get_all_monitored_groups()
                                    if monitored_groups:
                                        # 使用第一個可用的群組 URL 作為備選
                                        source_group_url = monitored_groups[0].get('url')
                                        print(f"[Backend] Using first monitored group URL as fallback: {source_group_url}", file=sys.stderr)
                        
                        if source_group_url == source_group:
                            print(f"[Backend] Could not find URL for chat_id {source_group}, using as-is", file=sys.stderr)
                except Exception as e:
                    print(f"[Backend] Error looking up group URL: {e}", file=sys.stderr)
            
            # 獲取用戶名作為備選（如果通過 userId 無法發送）
            target_username = None
            if lead_id:
                try:
                    lead = await db.get_lead(lead_id)
                    if lead:
                        target_username = lead.get('username')
                        if target_username:
                            print(f"[Backend] Got target username from lead: @{target_username}", file=sys.stderr)
                except Exception as e:
                    print(f"[Backend] Error getting username from lead: {e}", file=sys.stderr)
            
            # Convert priority
            if priority == 'high':
                msg_priority = MessagePriority.HIGH
            elif priority == 'low':
                msg_priority = MessagePriority.LOW
            else:
                msg_priority = MessagePriority.NORMAL
            
            # Parse scheduled time if provided
            scheduled_datetime = None
            if scheduled_at:
                try:
                    scheduled_datetime = datetime.fromisoformat(scheduled_at.replace('Z', '+00:00'))
                except:
                    pass
            
            # Add message to queue
            message_id = await self.message_queue.add_message(
                phone=account_phone,
                user_id=user_id,
                text=message_text,
                attachment=attachment,
                source_group=source_group_url,  # 使用 URL 而非 ID
                target_username=target_username,  # 用戶名作為備選
                priority=msg_priority,
                scheduled_at=scheduled_datetime,
                callback=self._on_message_sent_callback(lead_id)
            )
            
            await db.add_log(f"Message queued for lead {lead_id} (queue ID: {message_id})", "info")
            self.send_log(f"Message queued for lead {lead_id}", "success")
            
            # Send queued event
            self.send_event("message-queued", {
                "messageId": message_id,
                "leadId": lead_id,
                "accountPhone": account_phone,
                "userId": user_id
            })
        
        except Exception as e:
            self.send_log(f"Error queueing message: {str(e)}", "error")
    
    def _on_message_sent_callback(self, lead_id: int):
        """Create callback for when message is sent"""
        async def callback(message, result):
            if result.get('success'):
                await db.add_interaction(lead_id, 'Message Sent', message.text)
                await db.add_log(f"Message sent to lead {lead_id}", "success")
                
                # Update lead status if needed
                lead = await db.get_lead(lead_id)
                if lead and lead.get('status') == 'New':
                    await db.update_lead_status(lead_id, 'Contacted')
                
                # Send success event
                self.send_event("message-sent", {
                    "leadId": lead_id,
                    "accountPhone": message.phone,
                    "userId": message.user_id,
                    "success": True,
                    "messageId": message.id
                })
            else:
                error = result.get('error', 'Unknown error')
                await db.add_log(f"Failed to send message to lead {lead_id}: {error}", "error")
                
                # Send failure event
                self.send_event("message-sent", {
                    "leadId": lead_id,
                    "accountPhone": message.phone,
                    "userId": message.user_id,
                    "success": False,
                    "error": error,
                    "messageId": message.id
                })
        
        return callback
    
    async def handle_update_lead_status(self, payload: Dict[str, Any]):
        """Handle update-lead-status command"""
        try:
            lead_id = payload.get('leadId')
            new_status = payload.get('newStatus')
            await db.update_lead_status(lead_id, new_status)
            await db.add_interaction(lead_id, 'Status Change', f"Status changed to {new_status}")
            await db.add_log(f"Lead {lead_id} status updated to {new_status}", "success")
            await self.send_leads_update()
        
        except Exception as e:
            self.send_log(f"Error updating lead status: {str(e)}", "error")
    
    async def handle_add_to_dnc(self, payload: Dict[str, Any]):
        """Handle add-to-dnc command"""
        try:
            user_id = payload.get('userId')
            await db.add_to_dnc(user_id)
            await db.add_log(f"User {user_id} added to DNC list", "success")
            await self.send_leads_update()
        
        except Exception as e:
            self.send_log(f"Error adding to DNC: {str(e)}", "error")
    
    async def handle_clear_logs(self):
        """Handle clear-logs command"""
        try:
            await db.clear_logs()
            await db.add_log("Logs cleared", "info")
        
        except Exception as e:
            self.send_log(f"Error clearing logs: {str(e)}", "error")
    
    async def handle_load_accounts_from_excel(self, payload: Dict[str, Any]):
        """Handle load-accounts-from-excel command"""
        try:
            from openpyxl import load_workbook
            from pathlib import Path
            
            file_path = payload.get('filePath')
            if not file_path:
                self.send_log("No file path provided", "error")
                return
            
            file_path_obj = Path(file_path)
            if not file_path_obj.exists():
                self.send_log(f"File not found: {file_path}", "error")
                return
            
            # Load Excel file
            workbook = load_workbook(file_path_obj, data_only=True)
            sheet = workbook.active
            
            # Get headers (first row)
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value.lower().replace(' ', '_') if cell.value else '')
            
            # Map column names
            column_map = {
                'phone': ['phone', 'phone_number', '电话号码'],
                'api_id': ['api_id', 'apiid', 'api_id'],
                'api_hash': ['api_hash', 'apihash', 'api_hash'],
                'proxy': ['proxy', '代理'],
                'group': ['group', 'group_name', '分组'],
                'two_factor_password': ['two_factor_password', '2fa', '2fa_password', 'two_factor', '2fa密码'],
                'role': ['role', '角色']
            }
            
            # Find column indices
            column_indices = {}
            for key, possible_names in column_map.items():
                for idx, header in enumerate(headers):
                    if header in possible_names:
                        column_indices[key] = idx
                        break
            
            if 'phone' not in column_indices:
                self.send_log("Phone column not found in Excel file", "error")
                return
            
            # Read data rows
            imported_count = 0
            skipped_count = 0
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
                # Skip empty rows
                if not row[column_indices['phone']].value:
                    continue
                
                try:
                    account_data = {
                        'phone': str(row[column_indices['phone']].value).strip(),
                        'apiId': str(row[column_indices.get('api_id', 0)].value).strip() if column_indices.get('api_id') and row[column_indices['api_id']].value else '',
                        'apiHash': str(row[column_indices.get('api_hash', 0)].value).strip() if column_indices.get('api_hash') and row[column_indices['api_hash']].value else '',
                        'proxy': str(row[column_indices.get('proxy', 0)].value).strip() if column_indices.get('proxy') and row[column_indices['proxy']].value else '',
                        'group': str(row[column_indices.get('group', 0)].value).strip() if column_indices.get('group') and row[column_indices['group']].value else '',
                        'twoFactorPassword': str(row[column_indices.get('two_factor_password', 0)].value).strip() if column_indices.get('two_factor_password') and row[column_indices['two_factor_password']].value else '',
                        'role': str(row[column_indices.get('role', 0)].value).strip() if column_indices.get('role') and row[column_indices['role']].value else 'Unassigned'
                    }
                    
                    # Validate phone number
                    if not account_data['phone']:
                        skipped_count += 1
                        continue
                    
                    # Check if account already exists
                    existing = await db.get_account_by_phone(account_data['phone'])
                    if existing:
                        # Update existing account
                        await db.update_account(existing['id'], {
                            'apiId': account_data.get('apiId'),
                            'apiHash': account_data.get('apiHash'),
                            'proxy': account_data.get('proxy'),
                            'group': account_data.get('group'),
                            'role': account_data.get('role', 'Unassigned')
                        })
                        imported_count += 1
                    else:
                        # Add new account
                        await db.add_account(account_data)
                        imported_count += 1
                
                except Exception as e:
                    self.send_log(f"Error processing row {row_idx}: {str(e)}", "warning")
                    skipped_count += 1
                    continue
            
            # Send updated accounts
            accounts = await db.get_all_accounts()
            # Invalidate cache
            self._cache.pop("accounts", None)
            self._cache_timestamps.pop("accounts", None)
            self.send_event("accounts-updated", accounts)
            
            self.send_log(f"Imported {imported_count} accounts from Excel (skipped {skipped_count})", "success")
            await db.add_log(f"Imported {imported_count} accounts from Excel", "success")
        
        except Exception as e:
            self.send_log(f"Error loading accounts from Excel: {str(e)}", "error")
    
    async def handle_export_leads_to_excel(self, payload: Dict[str, Any]):
        """Handle export-leads-to-excel command"""
        try:
            from openpyxl import Workbook
            from pathlib import Path
            from datetime import datetime
            
            file_path = payload.get('filePath')
            leads = payload.get('leads', [])
            
            if not file_path:
                self.send_log("No file path provided", "error")
                return
            
            if not leads:
                self.send_log("No leads to export", "warning")
                return
            
            # Create workbook
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Leads"
            
            # Write headers
            headers = [
                'ID', 'User ID', 'Username', 'First Name', 'Last Name',
                'Source Group', 'Triggered Keyword', 'Status', 'Online Status',
                'Timestamp', 'Do Not Contact'
            ]
            sheet.append(headers)
            
            # Write data
            for lead in leads:
                row = [
                    lead.get('id', ''),
                    lead.get('userId', ''),
                    lead.get('username', ''),
                    lead.get('firstName', ''),
                    lead.get('lastName', ''),
                    lead.get('sourceGroup', ''),
                    lead.get('triggeredKeyword', ''),
                    lead.get('status', ''),
                    lead.get('onlineStatus', ''),
                    lead.get('timestamp', ''),
                    'Yes' if lead.get('doNotContact') else 'No'
                ]
                sheet.append(row)
            
            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save file
            file_path_obj = Path(file_path)
            workbook.save(file_path_obj)
            
            self.send_log(f"Exported {len(leads)} leads to Excel: {file_path}", "success")
            await db.add_log(f"Exported {len(leads)} leads to Excel", "success")
        
        except Exception as e:
            self.send_log(f"Error exporting leads to Excel: {str(e)}", "error")
    
    async def handle_reload_sessions_and_accounts(self):
        """Handle reload-sessions-and-accounts command - scan sessions directory and sync with database"""
        try:
            from pathlib import Path
            from config import SESSIONS_DIR
            import re
            
            self.send_log("Reloading sessions and accounts", "info")
            
            # Get all session files
            session_files = list(SESSIONS_DIR.glob("*.session"))
            
            if not session_files:
                self.send_log("No session files found", "info")
                accounts = await db.get_all_accounts()
                self.send_event("accounts-updated", accounts)
                return
            
            # Get all existing accounts
            existing_accounts = await db.get_all_accounts()
            existing_phones = {acc.get('phone') for acc in existing_accounts}
            
            imported_count = 0
            updated_count = 0
            
            # Process each session file
            for session_file in session_files:
                try:
                    # Extract phone number from filename (remove .session extension)
                    phone_number = session_file.stem
                    
                    # Try to normalize phone number (add + if missing)
                    if not phone_number.startswith('+'):
                        # Try to detect if it's a valid phone number
                        if phone_number.isdigit() and len(phone_number) >= 10:
                            # Assume it's a phone number without country code
                            # For now, keep as is - user can update manually
                            pass
                    
                    # Check if account exists
                    existing_account = await db.get_account_by_phone(phone_number)
                    
                    if existing_account:
                        # Update session file path if needed
                        updated_count += 1
                        self.send_log(f"Session found for existing account: {phone_number}", "info")
                    else:
                        # Create new account from session file
                        # Try to get account info from session file metadata if possible
                        # For now, create with minimal info
                        await db.add_account({
                            'phone': phone_number,
                            'api_id': '',  # Will need to be filled manually
                            'api_hash': '',  # Will need to be filled manually
                            'proxy': '',
                            'group': '',
                            'role': 'Unassigned',
                            'status': 'Offline'  # Will be updated when logged in
                        })
                        imported_count += 1
                        self.send_log(f"Created account from session file: {phone_number}", "info")
                
                except Exception as e:
                    self.send_log(f"Error processing session file {session_file.name}: {str(e)}", "warning")
                    continue
            
            # Check for accounts without session files
            accounts_without_sessions = []
            for account in existing_accounts:
                phone = account.get('phone')
                safe_phone = phone.replace("+", "").replace("-", "").replace(" ", "")
                session_file = SESSIONS_DIR / f"{safe_phone}.session"
                if not session_file.exists():
                    accounts_without_sessions.append(phone)
            
            # Send updated accounts
            accounts = await db.get_all_accounts()
            # Invalidate cache
            self._cache.pop("accounts", None)
            self._cache_timestamps.pop("accounts", None)
            self.send_event("accounts-updated", accounts)
            
            # Log summary
            summary = f"Reloaded sessions: {len(session_files)} session files found, {imported_count} new accounts created, {updated_count} existing accounts updated"
            if accounts_without_sessions:
                summary += f", {len(accounts_without_sessions)} accounts without session files"
            
            self.send_log(summary, "success")
            await db.add_log(summary, "info")
        
        except Exception as e:
            self.send_log(f"Error reloading sessions: {str(e)}", "error")
    
    async def handle_import_session(self, payload: Dict[str, Any]):
        """Handle import-session command - import a session file or package"""
        try:
            from pathlib import Path
            import shutil
            from session_package import SessionPackage, BatchSessionPackage
            
            file_path = payload.get('filePath')
            
            if not file_path:
                self.send_log("No file path provided", "error")
                return
            
            file_path = Path(file_path)
            if not file_path.exists():
                self.send_log(f"File not found: {file_path}", "error")
                return
            
            from config import SESSIONS_DIR
            
            # Check file type and handle accordingly
            if SessionPackage.is_package_file(file_path):
                # Import TG-Matrix session package (.tgpkg)
                success, message, account_data = SessionPackage.extract_package(file_path, SESSIONS_DIR)
                
                if not success:
                    self.send_log(message, "error")
                    self.send_event("session-import-result", {
                        "success": False,
                        "message": message
                    })
                    return
                
                # Check if account exists
                existing_account = await db.get_account_by_phone(account_data['phone'])
                if existing_account:
                    # Update existing account with new credentials
                    await db.update_account(existing_account['id'], {
                        'api_id': account_data['api_id'],
                        'api_hash': account_data['api_hash'],
                        'proxy': account_data.get('proxy', existing_account.get('proxy', '')),
                        'status': 'Offline'
                    })
                    self.send_log(f"✅ 賬戶已更新: {account_data['phone']}", "success")
                else:
                    # Create new account
                    await db.add_account(account_data)
                    self.send_log(f"✅ 賬戶已導入: {account_data['phone']}", "success")
                
                self.send_event("session-import-result", {
                    "success": True,
                    "message": f"Session package imported: {account_data['phone']}",
                    "phone": account_data['phone']
                })
                
            elif str(file_path).lower().endswith('.tgbatch'):
                # Import batch package
                success, message, accounts_list = BatchSessionPackage.extract_batch_package(file_path, SESSIONS_DIR)
                
                if not success:
                    self.send_log(message, "error")
                    self.send_event("session-import-result", {
                        "success": False,
                        "message": message
                    })
                    return
                
                imported_count = 0
                for account_data in accounts_list:
                    existing_account = await db.get_account_by_phone(account_data['phone'])
                    if existing_account:
                        await db.update_account(existing_account['id'], {
                            'api_id': account_data['api_id'],
                            'api_hash': account_data['api_hash'],
                            'proxy': account_data.get('proxy', ''),
                            'status': 'Offline'
                        })
                    else:
                        await db.add_account(account_data)
                    imported_count += 1
                
                self.send_log(f"✅ 批量導入完成: {imported_count} 個賬戶", "success")
                self.send_event("session-import-result", {
                    "success": True,
                    "message": f"Batch import complete: {imported_count} accounts",
                    "count": imported_count
                })
                
            elif SessionPackage.is_legacy_session(file_path):
                # Legacy .session file - requires API credentials
                phone_number = payload.get('phoneNumber', '')
                api_id = payload.get('apiId', '')
                api_hash = payload.get('apiHash', '')
                
                # Determine phone number from filename if not provided
                if not phone_number:
                    phone_number = file_path.stem.replace('+', '').replace('-', '').replace(' ', '')
                
                # Check if we have API credentials
                if not api_id or not api_hash:
                    # Try to get from existing account
                    existing_account = await db.get_account_by_phone(phone_number)
                    if existing_account and existing_account.get('api_id') and existing_account.get('api_hash'):
                        api_id = existing_account.get('api_id')
                        api_hash = existing_account.get('api_hash')
                    else:
                        # No credentials - notify frontend to ask user
                        self.send_event("session-import-needs-credentials", {
                            "filePath": str(file_path),
                            "phoneNumber": phone_number,
                            "message": "此 session 文件需要 API ID 和 API Hash 才能使用"
                        })
                        self.send_log("⚠️ 導入舊格式 session 需要提供 API ID 和 API Hash", "warning")
                        return
                
                # Copy session file
                target_path = SESSIONS_DIR / f"{phone_number}.session"
                shutil.copy2(file_path, target_path)
                
                # Create or update account
                existing_account = await db.get_account_by_phone(phone_number)
                if existing_account:
                    await db.update_account(existing_account['id'], {
                        'api_id': api_id,
                        'api_hash': api_hash,
                        'status': 'Offline'
                    })
                else:
                    await db.add_account({
                        'phone': phone_number,
                        'api_id': api_id,
                        'api_hash': api_hash,
                        'proxy': payload.get('proxy', ''),
                        'group': payload.get('group', ''),
                        'role': payload.get('role', 'Unassigned'),
                        'status': 'Offline'
                    })
                
                self.send_log(f"✅ Session 已導入: {phone_number}", "success")
                self.send_event("session-import-result", {
                    "success": True,
                    "message": f"Session imported: {phone_number}",
                    "phone": phone_number
                })
            else:
                self.send_log(f"不支持的文件格式: {file_path.suffix}", "error")
                self.send_event("session-import-result", {
                    "success": False,
                    "message": f"Unsupported file format: {file_path.suffix}"
                })
                return
            
            # Reload accounts
            accounts = await db.get_all_accounts()
            self._cache.pop("accounts", None)
            self._cache_timestamps.pop("accounts", None)
            self.send_event("accounts-updated", accounts)
        
        except Exception as e:
            self.send_log(f"Error importing session: {str(e)}", "error")
            self.send_event("session-import-result", {
                "success": False,
                "message": str(e)
            })
    
    async def handle_export_session(self, payload: Dict[str, Any]):
        """Handle export-session command - export a session package with credentials"""
        try:
            from pathlib import Path
            from session_package import SessionPackage
            
            phone_number = payload.get('phoneNumber')
            export_path = payload.get('filePath')
            export_format = payload.get('format', 'package')  # 'package' or 'legacy'
            
            if not phone_number:
                self.send_log("No phone number provided", "error")
                return
            
            if not export_path:
                self.send_log("No export path provided", "error")
                return
            
            # Get account data
            account = await db.get_account_by_phone(phone_number)
            if not account:
                self.send_log(f"Account not found: {phone_number}", "error")
                return
            
            # Find session file
            from config import SESSIONS_DIR
            safe_phone = phone_number.replace("+", "").replace("-", "").replace(" ", "")
            session_file = SESSIONS_DIR / f"{safe_phone}.session"
            
            if not session_file.exists():
                self.send_log(f"Session file not found for {phone_number}", "error")
                return
            
            export_path_obj = Path(export_path)
            
            if export_format == 'package' or export_path.endswith('.tgpkg'):
                # Export as TG-Matrix session package (recommended)
                success, message = SessionPackage.create_package(
                    session_file_path=session_file,
                    api_id=account.get('api_id', ''),
                    api_hash=account.get('api_hash', ''),
                    phone=phone_number,
                    output_path=export_path_obj,
                    proxy=account.get('proxy', ''),
                    role=account.get('role', 'Unassigned'),
                    group=account.get('group', ''),
                    daily_send_limit=account.get('daily_send_limit', 50),
                    notes=account.get('notes', '')
                )
                
                if success:
                    self.send_log(f"✅ Session 包已導出: {phone_number}", "success")
                    self.send_event("session-exported", {
                        "phoneNumber": phone_number,
                        "filePath": str(export_path_obj) if not str(export_path_obj).endswith('.tgpkg') else str(export_path_obj),
                        "format": "package"
                    })
                else:
                    self.send_log(f"導出失敗: {message}", "error")
            else:
                # Legacy export (just session file, not recommended)
                import shutil
                shutil.copy2(session_file, export_path_obj)
                self.send_log(f"⚠️ Session 文件已導出 (不包含 API 憑證): {phone_number}", "warning")
                self.send_event("session-exported", {
                    "phoneNumber": phone_number,
                    "filePath": export_path,
                    "format": "legacy",
                    "warning": "不包含 API 憑證，導入時需要重新輸入"
                })
        
        except Exception as e:
            self.send_log(f"Error exporting session: {str(e)}", "error")
    
    async def handle_export_sessions_batch(self, payload: Dict[str, Any]):
        """Handle batch export of multiple sessions"""
        try:
            from pathlib import Path
            from session_package import BatchSessionPackage
            from config import SESSIONS_DIR
            
            phone_numbers = payload.get('phoneNumbers', [])
            export_path = payload.get('filePath')
            
            if not phone_numbers:
                self.send_log("No accounts selected for export", "error")
                return
            
            if not export_path:
                self.send_log("No export path provided", "error")
                return
            
            # Get account data for all selected phones
            accounts_data = []
            for phone in phone_numbers:
                account = await db.get_account_by_phone(phone)
                if account:
                    accounts_data.append(account)
            
            if not accounts_data:
                self.send_log("No valid accounts found", "error")
                return
            
            export_path_obj = Path(export_path)
            success, message, count = BatchSessionPackage.create_batch_package(
                accounts_data=accounts_data,
                sessions_dir=SESSIONS_DIR,
                output_path=export_path_obj
            )
            
            if success:
                self.send_log(f"✅ 批量導出完成: {count} 個賬戶", "success")
                self.send_event("sessions-batch-exported", {
                    "count": count,
                    "filePath": str(export_path_obj)
                })
            else:
                self.send_log(f"批量導出失敗: {message}", "error")
        
        except Exception as e:
            self.send_log(f"Error batch exporting sessions: {str(e)}", "error")
    
    async def handle_create_backup(self, payload: Dict[str, Any]):
        """Handle create-backup command"""
        try:
            if not self.backup_manager:
                self.send_log("備份管理器未初始化", "error")
                return
            
            backup_type = payload.get('type', 'manual')
            compress = payload.get('compress', True)
            backup_path = await self.backup_manager.create_backup(
                backup_type=backup_type,
                compress=compress
            )
            
            self.send_log(f"✓ 備份創建成功: {backup_path.name}", "success")
            self.send_event("backup-created", {
                "backupPath": str(backup_path),
                "backupName": backup_path.name,
                "size": backup_path.stat().st_size,
                "type": backup_type
            })
        except Exception as e:
            app_error = handle_error(e, {"command": "create-backup", "payload": payload})
            self.send_log(f"創建備份失敗: {str(app_error)}", "error")
    
    async def handle_restore_backup(self, payload: Dict[str, Any]):
        """Handle restore-backup command"""
        try:
            from pathlib import Path
            backup_path_str = payload.get('backupPath')
            create_current_backup = payload.get('createCurrentBackup', True)
            
            if not backup_path_str:
                self.send_log("No backup path provided", "error")
                return
            
            backup_path = Path(backup_path_str)
            
            if not self.backup_manager:
                self.send_log("備份管理器未初始化", "error")
                return
            
            success = await self.backup_manager.restore_backup(backup_path, create_current_backup)
            
            if success:
                self.send_log(f"✓ 數據庫已從備份恢復: {backup_path.name}", "success")
                self.send_event("backup-restored", {
                    "backupPath": str(backup_path)
                })
                
                # Reload initial state after restore
                await self.handle_get_initial_state()
            else:
                self.send_log("恢復備份失敗", "error")
        except Exception as e:
            app_error = handle_error(e, {"command": "restore-backup", "payload": payload})
            self.send_log(f"Error restoring backup: {str(app_error)}", "error")
    
    async def handle_list_backups(self):
        """Handle list-backups command"""
        try:
            backup_manager = get_backup_manager()
            backups = backup_manager.list_backups()
            
            backup_list = []
            for backup in backups:
                backup_time = datetime.fromtimestamp(backup.stat().st_mtime)
                backup_list.append({
                    "name": backup.name,
                    "path": str(backup),
                    "size": backup.stat().st_size,
                    "sizeMB": round(backup.stat().st_size / (1024 * 1024), 2),
                    "timestamp": backup_time.isoformat()
                })
            
            self.send_event("backups-listed", {
                "backups": backup_list
            })
        except Exception as e:
            app_error = handle_error(e, {"command": "list-backups"})
            self.send_log(f"Error listing backups: {str(app_error)}", "error")
    
    async def handle_get_backup_info(self):
        """Handle get-backup-info command"""
        try:
            backup_manager = get_backup_manager()
            info = backup_manager.get_backup_info()
            self.send_event("backup-info", info)
        except Exception as e:
            app_error = handle_error(e, {"command": "get-backup-info"})
            self.send_log(f"Error getting backup info: {str(app_error)}", "error")
    
    async def handle_get_performance_summary(self):
        """Handle get-performance-summary command"""
        try:
            from performance_monitor import get_performance_monitor
            monitor = get_performance_monitor()
            if monitor:
                summary = monitor.get_performance_summary()
                self.send_event("performance-summary", summary)
            else:
                self.send_log("Performance monitor not initialized", "warning")
                self.send_event("performance-summary", {})
        except Exception as e:
            app_error = handle_error(e, {"command": "get-performance-summary"})
            self.send_log(f"Error getting performance summary: {str(app_error)}", "error")
            self.send_event("performance-summary", {})
    
    async def handle_get_performance_metrics(self, payload: Dict[str, Any]):
        """Handle get-performance-metrics command"""
        try:
            from performance_monitor import get_performance_monitor
            from datetime import datetime
            monitor = get_performance_monitor()
            if monitor:
                # Parse time filters from payload
                start_time = None
                end_time = None
                limit = payload.get('limit', 100)
                
                if payload.get('startTime'):
                    start_time = datetime.fromisoformat(payload['startTime'].replace('Z', '+00:00'))
                if payload.get('endTime'):
                    end_time = datetime.fromisoformat(payload['endTime'].replace('Z', '+00:00'))
                
                # Get metrics history with filters
                metrics = monitor.get_metrics_history(
                    start_time=start_time,
                    end_time=end_time,
                    limit=limit
                )
                
                # Convert datetime objects to ISO strings
                for metric in metrics:
                    if 'timestamp' in metric and isinstance(metric['timestamp'], datetime):
                        metric['timestamp'] = metric['timestamp'].isoformat()
                
                self.send_event("performance-metrics", {
                    "metrics": metrics
                })
            else:
                self.send_log("Performance monitor not initialized", "warning")
                self.send_event("performance-metrics", {"metrics": []})
        except Exception as e:
            app_error = handle_error(e, {"command": "get-performance-metrics", "payload": payload})
            self.send_log(f"Error getting performance metrics: {str(app_error)}", "error")
            self.send_event("performance-metrics", {"metrics": []})
    
    async def handle_get_sending_stats(self, payload: Dict[str, Any]):
        """Handle get-sending-stats command"""
        try:
            days = payload.get('days', 7)
            phone = payload.get('phone')  # Optional
            
            stats = await db.get_message_sending_stats(days, phone)
            self.send_event("sending-stats", {"stats": stats, "days": days, "phone": phone})
        except Exception as e:
            handle_error(e, {"command": "get-sending-stats", "payload": payload})
            self.send_log(f"Error getting sending stats: {str(e)}", "error")
    
    async def handle_get_queue_length_history(self, payload: Dict[str, Any]):
        """Handle get-queue-length-history command"""
        try:
            days = payload.get('days', 7)
            
            history = await db.get_queue_length_history(days)
            self.send_event("queue-length-history", {"history": history, "days": days})
        except Exception as e:
            handle_error(e, {"command": "get-queue-length-history", "payload": payload})
            self.send_log(f"Error getting queue length history: {str(e)}", "error")
    
    async def handle_get_account_sending_comparison(self, payload: Dict[str, Any]):
        """Handle get-account-sending-comparison command"""
        try:
            days = payload.get('days', 7)
            
            comparison = await db.get_account_sending_comparison(days)
            self.send_event("account-sending-comparison", {"comparison": comparison, "days": days})
        except Exception as e:
            handle_error(e, {"command": "get-account-sending-comparison", "payload": payload})
            self.send_log(f"Error getting account sending comparison: {str(e)}", "error")
    
    async def handle_get_campaign_performance_stats(self, payload: Dict[str, Any]):
        """Handle get-campaign-performance-stats command"""
        try:
            days = payload.get('days', 7)
            
            stats = await db.get_campaign_performance_stats(days)
            self.send_event("campaign-performance-stats", {"stats": stats, "days": days})
        except Exception as e:
            handle_error(e, {"command": "get-campaign-performance-stats", "payload": payload})
            self.send_log(f"Error getting campaign performance stats: {str(e)}", "error")
    
    async def handle_get_alerts(self, payload: Dict[str, Any]):
        """Handle get-alerts command"""
        try:
            limit = payload.get('limit', 50)
            level = payload.get('level')  # Optional: 'info', 'warning', 'error', 'critical'
            unresolved_only = payload.get('unresolvedOnly', False)
            
            if unresolved_only:
                alerts = await db.get_unresolved_alerts(limit)
            else:
                alerts = await db.get_recent_alerts(limit, level)
            
            self.send_event("alerts-loaded", {"alerts": alerts, "count": len(alerts)})
        except Exception as e:
            handle_error(e, {"command": "get-alerts", "payload": payload})
            self.send_log(f"Error getting alerts: {str(e)}", "error")
    
    async def handle_acknowledge_alert(self, payload: Dict[str, Any]):
        """Handle acknowledge-alert command"""
        try:
            alert_id = payload.get('alertId')
            if not alert_id:
                self.send_log("Alert ID required", "error")
                return
            
            await db.acknowledge_alert(alert_id)
            self.send_log(f"Alert {alert_id} acknowledged", "success")
            
            # Send updated alerts
            alerts = await db.get_recent_alerts(50)
            self.send_event("alerts-loaded", {"alerts": alerts, "count": len(alerts)})
        except Exception as e:
            handle_error(e, {"command": "acknowledge-alert", "payload": payload})
            self.send_log(f"Error acknowledging alert: {str(e)}", "error")
    
    async def handle_resolve_alert(self, payload: Dict[str, Any]):
        """Handle resolve-alert command"""
        try:
            alert_id = payload.get('alertId')
            if not alert_id:
                self.send_log("Alert ID required", "error")
                return
            
            await db.resolve_alert(alert_id)
            self.send_log(f"Alert {alert_id} resolved", "success")
            
            # Send updated alerts
            alerts = await db.get_recent_alerts(50)
            self.send_event("alerts-loaded", {"alerts": alerts, "count": len(alerts)})
        except Exception as e:
            handle_error(e, {"command": "resolve-alert", "payload": payload})
            self.send_log(f"Error resolving alert: {str(e)}", "error")
    
    async def handle_migration_status(self, payload: Dict[str, Any]):
        """Handle migration-status command"""
        try:
            from migrations.migration_manager import get_migration_manager
            migration_manager = get_migration_manager()
            
            if not migration_manager:
                self.send_log("Migration manager not initialized", "warning")
                self.send_event("migration-status", {
                    "error": "Migration manager not initialized"
                })
                return
            
            status = await migration_manager.status()
            self.send_event("migration-status", status)
        except Exception as e:
            handle_error(e, {"command": "migration-status", "payload": payload})
            self.send_log(f"Error getting migration status: {str(e)}", "error")
            self.send_event("migration-status", {"error": str(e)})
    
    async def handle_migrate(self, payload: Dict[str, Any]):
        """Handle migrate command"""
        try:
            from migrations.migration_manager import get_migration_manager
            migration_manager = get_migration_manager()
            
            if not migration_manager:
                self.send_log("Migration manager not initialized", "error")
                return
            
            target_version = payload.get('targetVersion')  # Optional
            
            success = await migration_manager.migrate(target_version)
            if success:
                status = await migration_manager.status()
                self.send_log("Migration completed successfully", "success")
                self.send_event("migration-completed", {
                    "message": "Migration completed successfully",
                    "status": status
                })
            else:
                self.send_log("Migration failed", "error")
                self.send_event("migration-completed", {
                    "error": "Migration failed"
                })
        except Exception as e:
            handle_error(e, {"command": "migrate", "payload": payload})
            self.send_log(f"Error running migration: {str(e)}", "error")
            self.send_event("migration-completed", {"error": str(e)})
    
    async def handle_rollback_migration(self, payload: Dict[str, Any]):
        """Handle rollback-migration command"""
        try:
            from migrations.migration_manager import get_migration_manager
            migration_manager = get_migration_manager()
            
            if not migration_manager:
                self.send_log("Migration manager not initialized", "error")
                return
            
            target_version = payload.get('targetVersion')
            if target_version is None:
                self.send_log("Missing targetVersion", "error")
                return
            
            success = await migration_manager.rollback(target_version)
            if success:
                status = await migration_manager.status()
                self.send_log(f"Rollback to version {target_version} completed successfully", "success")
                self.send_event("migration-rollback-completed", {
                    "message": f"Rollback to version {target_version} completed successfully",
                    "status": status
                })
            else:
                self.send_log("Rollback failed", "error")
                self.send_event("migration-rollback-completed", {
                    "error": "Rollback failed"
                })
        except Exception as e:
            handle_error(e, {"command": "rollback-migration", "payload": payload})
            self.send_log(f"Error rolling back migration: {str(e)}", "error")
            self.send_event("migration-rollback-completed", {"error": str(e)})
    
    async def run(self):
        """Main event loop - read commands from stdin"""
        await self.initialize()
        
        try:
            while self.running:
                # Read line from stdin (non-blocking)
                line = await asyncio.get_event_loop().run_in_executor(
                    None, sys.stdin.readline
                )
                
                if not line:
                    # EOF - stdin closed
                    break
                
                line = line.strip()
                if not line:
                    continue
                
                try:
                    # Parse JSON command
                    command_data = json.loads(line)
                    command = command_data.get('command')
                    payload = command_data.get('payload', {})
                    request_id = command_data.get('request_id')  # Optional request ID for acknowledgment
                    
                    # Handle command
                    await self.handle_command(command, payload, request_id)
                
                except json.JSONDecodeError as e:
                    self.send_log(f"Invalid JSON received: {str(e)}", "error")
                except Exception as e:
                    self.send_log(f"Unexpected error: {str(e)}", "error")
                    import traceback
                    traceback.print_exc()
        
        except KeyboardInterrupt:
            pass
        finally:
            await self.shutdown()


    # ==================== Local AI & Voice Services Handlers ====================
    
    async def handle_test_local_ai(self, payload: Dict[str, Any]):
        """Test connection to local AI service with detailed diagnostics"""
        import aiohttp
        import socket
        import time
        from urllib.parse import urlparse
        
        endpoint = payload.get('endpoint', 'http://localhost:3002')
        model = payload.get('model', '')
        
        diagnostics = {
            "endpoint": endpoint,
            "tcp_connection": False,
            "http_connection": False,
            "ai_response": False,
            "errors": []
        }
        
        try:
            # 解析端點
            parsed = urlparse(endpoint)
            host = parsed.hostname
            port = parsed.port or (443 if parsed.scheme == 'https' else 80)
            
            print(f"[AI Test] Testing connection to {host}:{port}...", file=sys.stderr)
            
            # 步驟 1: 測試 TCP 連接
            try:
                sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                sock.settimeout(5)
                result = sock.connect_ex((host, port))
                sock.close()
                
                if result == 0:
                    diagnostics["tcp_connection"] = True
                    print(f"[AI Test] ✓ TCP connection to {host}:{port} successful", file=sys.stderr)
                else:
                    diagnostics["errors"].append(f"TCP 連接失敗 (錯誤代碼: {result})")
                    print(f"[AI Test] ✗ TCP connection failed (error code: {result})", file=sys.stderr)
                    self.send_event("local-ai-test-result", {
                        "success": False,
                        "endpoint": endpoint,
                        "diagnostics": diagnostics,
                        "error": f"無法連接到 {host}:{port}。請檢查：\n1. AI 服務是否正在運行\n2. 防火牆是否允許連接\n3. 網絡是否正常"
                    })
                    return
            except socket.gaierror as e:
                diagnostics["errors"].append(f"DNS 解析失敗: {str(e)}")
                print(f"[AI Test] ✗ DNS resolution failed: {e}", file=sys.stderr)
                self.send_event("local-ai-test-result", {
                    "success": False,
                    "endpoint": endpoint,
                    "diagnostics": diagnostics,
                    "error": f"無法解析主機名 {host}。請檢查網絡設置"
                })
                return
            except socket.timeout:
                diagnostics["errors"].append("TCP 連接超時")
                print(f"[AI Test] ✗ TCP connection timeout", file=sys.stderr)
                self.send_event("local-ai-test-result", {
                    "success": False,
                    "endpoint": endpoint,
                    "diagnostics": diagnostics,
                    "error": f"連接 {host}:{port} 超時。請檢查防火牆設置"
                })
                return
            except Exception as e:
                diagnostics["errors"].append(f"TCP 連接錯誤: {str(e)}")
                print(f"[AI Test] ✗ TCP connection error: {e}", file=sys.stderr)
            
            # 步驟 2: 測試 HTTP 連接
            timeout = aiohttp.ClientTimeout(total=30, connect=5)
            async with aiohttp.ClientSession(timeout=timeout) as session:
                # 嘗試 GET 請求（Ollama 健康檢查）
                try:
                    health_url = endpoint.rstrip('/')
                    if not health_url.endswith('/api/tags') and not health_url.endswith('/v1/models'):
                        # 嘗試 Ollama 健康檢查端點
                        health_endpoints = [
                            f"{health_url}",
                            f"{health_url}/api/tags",
                            f"{health_url}/v1/models"
                        ]
                    else:
                        health_endpoints = [health_url]
                    
                    for health_endpoint in health_endpoints:
                        try:
                            async with session.get(health_endpoint, timeout=aiohttp.ClientTimeout(total=5)) as resp:
                                if resp.status in [200, 404, 405]:
                                    diagnostics["http_connection"] = True
                                    print(f"[AI Test] ✓ HTTP connection successful: {health_endpoint} (status: {resp.status})", file=sys.stderr)
                                    break
                        except:
                            continue
                except Exception as e:
                    diagnostics["errors"].append(f"HTTP 連接測試失敗: {str(e)}")
                    print(f"[AI Test] HTTP connection test failed: {e}", file=sys.stderr)
                
                # 步驟 3: 測試實際 AI 請求
                try:
                    chat_url = endpoint.rstrip('/')
                    if not chat_url.endswith('/v1/chat/completions'):
                        chat_url = chat_url.rstrip('/') + '/v1/chat/completions'
                    
                    test_payload = {
                        "messages": [{"role": "user", "content": "test"}],
                        "max_tokens": 10
                    }
                    if model:
                        test_payload["model"] = model
                    
                    print(f"[AI Test] Testing AI request to {chat_url}...", file=sys.stderr)
                    start_time = time.time()
                    
                    async with session.post(chat_url, json=test_payload, timeout=aiohttp.ClientTimeout(total=30)) as resp:
                        elapsed = time.time() - start_time
                        print(f"[AI Test] Response received in {elapsed:.2f}s, status: {resp.status}", file=sys.stderr)
                        
                        if resp.status == 200:
                            data = await resp.json()
                            if 'choices' in data or 'response' in data or 'content' in data:
                                diagnostics["ai_response"] = True
                                print(f"[AI Test] ✓ AI service responded successfully", file=sys.stderr)
                                self.send_event("local-ai-test-result", {
                                    "success": True,
                                    "endpoint": endpoint,
                                    "diagnostics": diagnostics,
                                    "response_time": elapsed
                                })
                                self.send_log(f"✓ 本地 AI 連接成功: {endpoint} (響應時間: {elapsed:.2f}秒)", "success")
                                return
                            else:
                                diagnostics["errors"].append(f"AI 響應格式異常: {list(data.keys())}")
                        else:
                            error_text = await resp.text()
                            diagnostics["errors"].append(f"HTTP {resp.status}: {error_text[:200]}")
                            print(f"[AI Test] ✗ AI service returned error: {resp.status}", file=sys.stderr)
                            
                except asyncio.TimeoutError:
                    elapsed = time.time() - start_time if 'start_time' in locals() else 30
                    diagnostics["errors"].append(f"AI 請求超時 ({elapsed:.1f}秒)")
                    print(f"[AI Test] ✗ AI request timeout after {elapsed:.2f}s", file=sys.stderr)
                except aiohttp.ClientConnectorError as e:
                    diagnostics["errors"].append(f"無法連接到端點: {str(e)}")
                    print(f"[AI Test] ✗ Connection error: {e}", file=sys.stderr)
                except Exception as e:
                    diagnostics["errors"].append(f"AI 請求錯誤: {str(e)}")
                    print(f"[AI Test] ✗ AI request error: {e}", file=sys.stderr)
            
            # 如果所有測試都失敗
            self.send_event("local-ai-test-result", {
                "success": False,
                "endpoint": endpoint,
                "diagnostics": diagnostics,
                "error": f"無法連接到本地 AI 服務。診斷信息：\n" + "\n".join(diagnostics["errors"])
            })
            self.send_log(f"✗ 本地 AI 連接失敗: {endpoint}", "error")
                
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"[AI Test] Unexpected error: {error_details}", file=sys.stderr)
            diagnostics["errors"].append(f"未預期的錯誤: {str(e)}")
            self.send_event("local-ai-test-result", {
                "success": False,
                "endpoint": endpoint,
                "diagnostics": diagnostics,
                "error": str(e)
            })
            self.send_log(f"✗ 本地 AI 測試錯誤: {str(e)}", "error")

    async def handle_test_tts_service(self, payload: Dict[str, Any]):
        """Test connection to TTS service (GPT-SoVITS)"""
        endpoint = payload.get('endpoint', 'http://localhost:9881')
        
        try:
            import aiohttp
            
            async with aiohttp.ClientSession(timeout=aiohttp.ClientTimeout(total=10)) as session:
                # Try common TTS endpoints
                test_urls = [
                    f"{endpoint}/",
                    f"{endpoint}/tts",
                    f"{endpoint}/api/tts",
                ]
                
                for test_url in test_urls:
                    try:
                        async with session.get(test_url) as response:
                            if response.status in [200, 404, 405]:
                                self.send_event("tts-test-result", {
                                    "success": True,
                                    "endpoint": endpoint,
                                    "status": response.status
                                })
                                self.send_log(f"✓ TTS 服務連接成功: {endpoint}", "success")
                                return
                    except:
                        continue
                
                self.send_event("tts-test-result", {
                    "success": False,
                    "endpoint": endpoint,
                    "error": "無法連接到 TTS 服務"
                })
                self.send_log(f"✗ TTS 服務連接失敗: {endpoint}", "error")
                
        except Exception as e:
            self.send_event("tts-test-result", {
                "success": False,
                "endpoint": endpoint,
                "error": str(e)
            })
            self.send_log(f"✗ TTS 測試錯誤: {str(e)}", "error")

    async def handle_test_stt_service(self, payload: Dict[str, Any]):
        """Test connection to STT service (Whisper)"""
        endpoint = payload.get('endpoint', 'http://localhost:9000')
        
        try:
            import aiohttp
            
            async with aiohttp.ClientSession(timeout=aiohttp.ClientTimeout(total=10)) as session:
                test_urls = [
                    f"{endpoint}/",
                    f"{endpoint}/transcribe",
                    f"{endpoint}/api/transcribe",
                ]
                
                for test_url in test_urls:
                    try:
                        async with session.get(test_url) as response:
                            if response.status in [200, 404, 405]:
                                self.send_event("stt-test-result", {
                                    "success": True,
                                    "endpoint": endpoint,
                                    "status": response.status
                                })
                                self.send_log(f"✓ STT 服務連接成功: {endpoint}", "success")
                                return
                    except:
                        continue
                
                self.send_event("stt-test-result", {
                    "success": False,
                    "endpoint": endpoint,
                    "error": "無法連接到 STT 服務"
                })
                self.send_log(f"✗ STT 服務連接失敗: {endpoint}", "error")
                
        except Exception as e:
            self.send_event("stt-test-result", {
                "success": False,
                "endpoint": endpoint,
                "error": str(e)
            })
            self.send_log(f"✗ STT 測試錯誤: {str(e)}", "error")

    async def handle_save_ai_settings(self, payload: Dict[str, Any]):
        """Save AI and voice service settings"""
        try:
            # Store settings in memory (for legacy compatibility)
            settings = {
                "apiType": payload.get('apiType', 'gemini'),
                "apiKey": payload.get('apiKey', ''),
                "endpoint": payload.get('endpoint', ''),
                "localAiEndpoint": payload.get('localAiEndpoint', 'http://localhost:3002'),
                "localAiModel": payload.get('localAiModel', ''),
                "ttsEndpoint": payload.get('ttsEndpoint', 'http://localhost:9881'),
                "ttsEnabled": payload.get('ttsEnabled', False),
                "ttsVoice": payload.get('ttsVoice', ''),
                "sttEndpoint": payload.get('sttEndpoint', 'http://localhost:9000'),
                "sttEnabled": payload.get('sttEnabled', False)
            }
            
            # Store in self for later use
            self.ai_settings = settings
            
            # 🔧 關鍵修復：同時保存到數據庫（使用 snake_case 鍵名）
            db_settings = {
                'local_ai_endpoint': payload.get('localAiEndpoint', ''),
                'local_ai_model': payload.get('localAiModel', '')
            }
            await db.update_ai_settings(db_settings)
            
            # 重新載入 AI 服務設置
            await ai_auto_chat.initialize()
            
            # 設置 AI 配置
            endpoint = payload.get('localAiEndpoint', '')
            model = payload.get('localAiModel', '')
            if endpoint:
                ai_auto_chat.set_ai_config(endpoint, model)
                self.send_log(f"✓ AI 端點已配置: {endpoint}", "success")
            
            self.send_event("ai-settings-saved", {"success": True})
            self.send_log("AI 和語音服務設置已保存到數據庫", "success")
            
        except Exception as e:
            import traceback
            traceback.print_exc(file=sys.stderr)
            self.send_event("ai-settings-saved", {"success": False, "error": str(e)})
            self.send_log(f"保存 AI 設置失敗: {str(e)}", "error")

    async def handle_generate_with_local_ai(self, payload: Dict[str, Any]):
        """Generate text using local AI service"""
        endpoint = payload.get('endpoint', getattr(self, 'ai_settings', {}).get('localAiEndpoint', 'http://localhost:3002'))
        prompt = payload.get('prompt', '')
        model = payload.get('model', '')
        
        try:
            import aiohttp
            
            # Try OpenAI-compatible API format
            request_data = {
                "model": model or "default",
                "messages": [{"role": "user", "content": prompt}],
                "max_tokens": 500,
                "temperature": 0.7
            }
            
            async with aiohttp.ClientSession(timeout=aiohttp.ClientTimeout(total=60)) as session:
                async with session.post(f"{endpoint}/v1/chat/completions", json=request_data) as response:
                    if response.status == 200:
                        result = await response.json()
                        generated_text = result.get('choices', [{}])[0].get('message', {}).get('content', '')
                        
                        self.send_event("local-ai-generation-result", {
                            "success": True,
                            "text": generated_text
                        })
                        return
                    else:
                        error_text = await response.text()
                        self.send_event("local-ai-generation-result", {
                            "success": False,
                            "error": f"API 返回 {response.status}: {error_text}"
                        })
                        
        except Exception as e:
            self.send_event("local-ai-generation-result", {
                "success": False,
                "error": str(e)
            })

    async def handle_text_to_speech(self, payload: Dict[str, Any]):
        """Convert text to speech using TTS service"""
        endpoint = payload.get('endpoint', getattr(self, 'ai_settings', {}).get('ttsEndpoint', 'http://localhost:9881'))
        text = payload.get('text', '')
        voice = payload.get('voice', '')
        
        try:
            import aiohttp
            
            # GPT-SoVITS API format
            request_data = {
                "text": text,
                "text_language": "zh",
                "ref_audio_path": voice if voice else None
            }
            
            async with aiohttp.ClientSession(timeout=aiohttp.ClientTimeout(total=120)) as session:
                async with session.post(f"{endpoint}/", json=request_data) as response:
                    if response.status == 200:
                        audio_data = await response.read()
                        import base64
                        audio_base64 = base64.b64encode(audio_data).decode('utf-8')
                        
                        self.send_event("tts-result", {
                            "success": True,
                            "audio": audio_base64,
                            "format": "wav"
                        })
                    else:
                        error_text = await response.text()
                        self.send_event("tts-result", {
                            "success": False,
                            "error": f"TTS 服務返回 {response.status}: {error_text}"
                        })
                        
        except Exception as e:
            self.send_event("tts-result", {
                "success": False,
                "error": str(e)
            })

    async def handle_speech_to_text(self, payload: Dict[str, Any]):
        """Convert speech to text using STT service"""
        endpoint = payload.get('endpoint', getattr(self, 'ai_settings', {}).get('sttEndpoint', 'http://localhost:9000'))
        audio_data = payload.get('audio', '')  # Base64 encoded audio
        
        try:
            import aiohttp
            import base64
            
            # Decode audio
            audio_bytes = base64.b64decode(audio_data)
            
            # Create form data for Whisper API
            data = aiohttp.FormData()
            data.add_field('file', audio_bytes, filename='audio.wav', content_type='audio/wav')
            
            async with aiohttp.ClientSession(timeout=aiohttp.ClientTimeout(total=120)) as session:
                async with session.post(f"{endpoint}/transcribe", data=data) as response:
                    if response.status == 200:
                        result = await response.json()
                        transcribed_text = result.get('text', '')
                        
                        self.send_event("stt-result", {
                            "success": True,
                            "text": transcribed_text
                        })
                    else:
                        error_text = await response.text()
                        self.send_event("stt-result", {
                            "success": False,
                            "error": f"STT 服務返回 {response.status}: {error_text}"
                        })
                        
        except Exception as e:
            self.send_event("stt-result", {
                "success": False,
                "error": str(e)
            })

    # ==================== Voice Clone Handlers ====================
    
    async def handle_upload_voice_sample(self, payload: Dict[str, Any]):
        """Upload and save a voice sample for cloning"""
        name = payload.get('name', 'unnamed')
        audio_data = payload.get('audioData', '')
        file_name = payload.get('fileName', 'audio.wav')
        file_type = payload.get('fileType', 'audio/wav')
        
        try:
            import base64
            import os
            
            # Create voice samples directory
            voice_dir = os.path.join(os.path.dirname(__file__), 'voice_samples')
            os.makedirs(voice_dir, exist_ok=True)
            
            # Generate unique filename
            import time
            unique_name = f"{int(time.time())}_{file_name}"
            file_path = os.path.join(voice_dir, unique_name)
            
            # Decode and save audio
            audio_bytes = base64.b64decode(audio_data)
            with open(file_path, 'wb') as f:
                f.write(audio_bytes)
            
            self.send_event("voice-sample-uploaded", {
                "success": True,
                "voiceId": str(int(time.time())),
                "name": name,
                "audioPath": unique_name,
                "filePath": file_path
            })
            self.send_log(f"✓ 聲音樣本已保存: {name}", "success")
            
        except Exception as e:
            self.send_event("voice-sample-uploaded", {
                "success": False,
                "error": str(e)
            })
            self.send_log(f"✗ 保存聲音樣本失敗: {str(e)}", "error")

    async def handle_delete_voice_sample(self, payload: Dict[str, Any]):
        """Delete a voice sample"""
        voice_id = payload.get('voiceId', '')
        
        try:
            import os
            
            # Find and delete the file
            voice_dir = os.path.join(os.path.dirname(__file__), 'voice_samples')
            
            # Try to find the file by ID (timestamp prefix)
            for filename in os.listdir(voice_dir):
                if filename.startswith(voice_id):
                    file_path = os.path.join(voice_dir, filename)
                    os.remove(file_path)
                    self.send_event("voice-sample-deleted", {"success": True, "voiceId": voice_id})
                    self.send_log(f"✓ 聲音樣本已刪除", "success")
                    return
            
            self.send_event("voice-sample-deleted", {"success": True, "voiceId": voice_id})
            
        except Exception as e:
            self.send_event("voice-sample-deleted", {
                "success": False,
                "error": str(e)
            })

    async def handle_preview_voice_sample(self, payload: Dict[str, Any]):
        """Preview a voice sample (send audio data back)"""
        voice_id = payload.get('voiceId', '')
        audio_path = payload.get('audioPath', '')
        
        try:
            import os
            import base64
            
            voice_dir = os.path.join(os.path.dirname(__file__), 'voice_samples')
            file_path = os.path.join(voice_dir, audio_path)
            
            if os.path.exists(file_path):
                with open(file_path, 'rb') as f:
                    audio_data = base64.b64encode(f.read()).decode('utf-8')
                
                self.send_event("voice-sample-preview", {
                    "success": True,
                    "voiceId": voice_id,
                    "audio": audio_data
                })
            else:
                self.send_event("voice-sample-preview", {
                    "success": False,
                    "error": "音頻文件不存在"
                })
                
        except Exception as e:
            self.send_event("voice-sample-preview", {
                "success": False,
                "error": str(e)
            })

    async def handle_generate_cloned_voice(self, payload: Dict[str, Any]):
        """Generate speech using a cloned voice via remote TTS service"""
        endpoint = payload.get('endpoint', '')
        text = payload.get('text', '')
        voice_id = payload.get('voiceId', '')
        audio_path = payload.get('audioPath', '')
        
        try:
            import aiohttp
            import base64
            import os
            
            # Read the reference audio file
            voice_dir = os.path.join(os.path.dirname(__file__), 'voice_samples')
            ref_audio_path = os.path.join(voice_dir, audio_path)
            
            if not os.path.exists(ref_audio_path):
                self.send_event("cloned-voice-generated", {
                    "success": False,
                    "error": "參考音頻文件不存在"
                })
                return
            
            # Read reference audio
            with open(ref_audio_path, 'rb') as f:
                ref_audio_data = base64.b64encode(f.read()).decode('utf-8')
            
            # GPT-SoVITS API format
            request_data = {
                "text": text,
                "text_language": "zh",
                "refer_wav_base64": ref_audio_data,
                "prompt_text": "",  # Can be set to reference text if available
                "prompt_language": "zh"
            }
            
            async with aiohttp.ClientSession(timeout=aiohttp.ClientTimeout(total=180)) as session:
                # Try different endpoint formats
                endpoints_to_try = [
                    f"{endpoint}/",
                    f"{endpoint}/tts",
                    f"{endpoint}/api/tts",
                    f"{endpoint}/v1/audio/speech"
                ]
                
                for api_url in endpoints_to_try:
                    try:
                        async with session.post(api_url, json=request_data) as response:
                            if response.status == 200:
                                content_type = response.headers.get('Content-Type', '')
                                
                                if 'audio' in content_type or 'octet-stream' in content_type:
                                    # Binary audio response
                                    audio_data = await response.read()
                                    audio_base64 = base64.b64encode(audio_data).decode('utf-8')
                                    
                                    self.send_event("cloned-voice-generated", {
                                        "success": True,
                                        "audio": audio_base64,
                                        "format": "wav"
                                    })
                                    self.send_log(f"✓ 使用克隆聲音生成語音成功", "success")
                                    return
                                else:
                                    # JSON response (might contain URL or base64)
                                    result = await response.json()
                                    if 'audio' in result:
                                        self.send_event("cloned-voice-generated", {
                                            "success": True,
                                            "audio": result['audio'],
                                            "format": result.get('format', 'wav')
                                        })
                                        self.send_log(f"✓ 使用克隆聲音生成語音成功", "success")
                                        return
                    except Exception as e:
                        continue
                
                self.send_event("cloned-voice-generated", {
                    "success": False,
                    "error": "無法連接到 TTS 服務或所有端點都失敗"
                })
                self.send_log(f"✗ TTS 生成失敗: 無法連接到服務", "error")
                        
        except Exception as e:
            self.send_event("cloned-voice-generated", {
                "success": False,
                "error": str(e)
            })
            self.send_log(f"✗ 克隆聲音生成失敗: {str(e)}", "error")

    async def handle_list_voice_samples(self):
        """List all saved voice samples"""
        try:
            import os
            
            voice_dir = os.path.join(os.path.dirname(__file__), 'voice_samples')
            
            if not os.path.exists(voice_dir):
                self.send_event("voice-samples-list", {"success": True, "samples": []})
                return
            
            samples = []
            for filename in os.listdir(voice_dir):
                file_path = os.path.join(voice_dir, filename)
                if os.path.isfile(file_path):
                    stat = os.stat(file_path)
                    samples.append({
                        "id": filename.split('_')[0],
                        "name": '_'.join(filename.split('_')[1:]).rsplit('.', 1)[0],
                        "audioPath": filename,
                        "size": stat.st_size,
                        "createdAt": stat.st_ctime
                    })
            
            self.send_event("voice-samples-list", {
                "success": True,
                "samples": samples
            })
            
        except Exception as e:
            self.send_event("voice-samples-list", {
                "success": False,
                "error": str(e)
            })

    # ==================== AI Auto Chat Handlers ====================
    
    async def handle_get_ai_chat_settings(self):
        """Get AI auto chat settings"""
        try:
            settings = await db.get_ai_settings()
            self.send_event("ai-chat-settings", {
                "success": True,
                "settings": settings
            })
        except Exception as e:
            self.send_event("ai-chat-settings", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_update_ai_chat_settings(self, payload: Dict[str, Any]):
        """Update AI auto chat settings"""
        try:
            settings = payload.get('settings', {})
            await db.update_ai_settings(settings)
            
            # Update AI auto chat service
            await ai_auto_chat.update_settings(settings)
            
            # Update AI endpoint if provided
            if 'localAiEndpoint' in payload:
                ai_auto_chat.set_ai_config(
                    payload.get('localAiEndpoint', ''),
                    payload.get('localAiModel', '')
                )
            
            self.send_event("ai-chat-settings-updated", {"success": True})
            self.send_log("AI 自動聊天設置已更新", "success")
        except Exception as e:
            self.send_event("ai-chat-settings-updated", {
                "success": False,
                "error": str(e)
            })
            self.send_log(f"更新 AI 設置失敗: {str(e)}", "error")
    
    async def handle_get_chat_history(self, payload: Dict[str, Any]):
        """Get chat history for a user"""
        try:
            user_id = payload.get('userId', '')
            limit = payload.get('limit', 50)
            
            if not user_id:
                self.send_event("chat-history", {"success": False, "error": "Missing userId"})
                return
            
            history = await db.get_chat_history(user_id, limit=limit)
            stats = await db.get_chat_stats(user_id)
            
            self.send_event("chat-history", {
                "success": True,
                "userId": user_id,
                "history": history,
                "stats": stats
            })
        except Exception as e:
            self.send_event("chat-history", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_get_user_context(self, payload: Dict[str, Any]):
        """Get full user context"""
        try:
            user_id = payload.get('userId', '')
            
            if not user_id:
                self.send_event("user-context", {"success": False, "error": "Missing userId"})
                return
            
            context = await ai_context.get_user_context(user_id)
            
            self.send_event("user-context", {
                "success": True,
                "userId": user_id,
                "context": context
            })
        except Exception as e:
            self.send_event("user-context", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_generate_ai_response(self, payload: Dict[str, Any]):
        """Generate AI response for a user message"""
        import time
        import sys
        start_time = time.time()
        
        print(f"[AI] handle_generate_ai_response called with payload keys: {payload.keys()}", file=sys.stderr)
        try:
            user_id = payload.get('userId', '')
            message = payload.get('message', '')
            system_prompt = payload.get('systemPrompt', '')
            endpoint = payload.get('localAiEndpoint', '')
            model = payload.get('localAiModel', '')
            
            print(f"[AI] Endpoint: {endpoint}, Model: {model}", file=sys.stderr)
            print(f"[AI] Message length: {len(message)}, System prompt length: {len(system_prompt)}", file=sys.stderr)
            
            if not message:
                self.send_event("ai-response", {"success": False, "error": "缺少消息內容"})
                return
            
            if not endpoint:
                self.send_event("ai-response", {"success": False, "error": "未配置 AI 服務端點"})
                return
            
            # 如果提供了端點，直接調用本地 AI
            if endpoint:
                print(f"[AI] Calling local AI service at {endpoint}...", file=sys.stderr)
                response = await self._call_local_ai(endpoint, model, system_prompt, message)
            else:
                # 使用 ai_auto_chat 服務
                print(f"[AI] Using ai_auto_chat service...", file=sys.stderr)
                ai_auto_chat.set_ai_config(endpoint, model)
                response = await ai_auto_chat.get_suggested_response(user_id, message)
            
            elapsed = time.time() - start_time
            print(f"[AI] AI generation completed in {elapsed:.2f}s", file=sys.stderr)
            
            if response:
                print(f"[AI] Successfully generated response (length: {len(response)})", file=sys.stderr)
                self.send_event("ai-response", {
                    "success": True,
                    "userId": user_id,
                    "response": response
                })
            else:
                print(f"[AI] AI returned empty response", file=sys.stderr)
                self.send_event("ai-response", {
                    "success": False,
                    "error": "AI 生成失敗，返回為空。請檢查服務配置和日誌"
                })
        except asyncio.TimeoutError:
            elapsed = time.time() - start_time
            error_msg = f"AI 生成超時（{elapsed:.1f}秒），請檢查服務連接"
            print(f"[AI] {error_msg}", file=sys.stderr)
            self.send_event("ai-response", {
                "success": False,
                "error": error_msg
            })
        except Exception as e:
            import traceback
            elapsed = time.time() - start_time
            error_details = traceback.format_exc()
            error_msg = f"AI 生成錯誤: {str(e)}"
            print(f"[AI] Error after {elapsed:.2f}s: {error_details}", file=sys.stderr)
            self.send_event("ai-response", {
                "success": False,
                "error": error_msg
            })
    
    async def _call_local_ai(self, endpoint: str, model: str, system_prompt: str, user_message: str) -> str:
        """直接調用本地/遠程 AI API"""
        import aiohttp
        import time
        import socket
        from urllib.parse import urlparse
        
        print(f"[AI] _call_local_ai called with endpoint: {endpoint}, model: {model}", file=sys.stderr)
        
        # 首先進行連接診斷
        try:
            parsed = urlparse(endpoint)
            host = parsed.hostname
            port = parsed.port or (443 if parsed.scheme == 'https' else 80)
            
            print(f"[AI] Diagnosing connection to {host}:{port}...", file=sys.stderr)
            
            # 測試 TCP 連接
            try:
                sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                sock.settimeout(5)
                result = sock.connect_ex((host, port))
                sock.close()
                
                if result == 0:
                    print(f"[AI] ✓ TCP connection to {host}:{port} successful", file=sys.stderr)
                else:
                    print(f"[AI] ✗ TCP connection to {host}:{port} failed (error code: {result})", file=sys.stderr)
                    raise Exception(f"無法連接到 AI 服務 {host}:{port}。請檢查：\n1. AI 服務是否正在運行\n2. 防火牆是否允許連接\n3. 網絡是否正常")
            except socket.gaierror as e:
                print(f"[AI] ✗ DNS resolution failed for {host}: {e}", file=sys.stderr)
                raise Exception(f"無法解析主機名 {host}。請檢查網絡設置或 DNS 配置")
            except socket.timeout:
                print(f"[AI] ✗ Connection timeout to {host}:{port}", file=sys.stderr)
                raise Exception(f"連接 {host}:{port} 超時。請檢查：\n1. AI 服務是否正在運行\n2. 防火牆是否阻塞了連接\n3. 網絡路由是否正確")
            except Exception as e:
                print(f"[AI] ✗ Connection test failed: {e}", file=sys.stderr)
                raise Exception(f"連接測試失敗: {str(e)}")
        except Exception as diag_error:
            # 診斷失敗，但繼續嘗試實際請求（可能診斷有誤）
            print(f"[AI] Connection diagnosis failed, but continuing: {diag_error}", file=sys.stderr)
        
        messages = []
        if system_prompt:
            messages.append({"role": "system", "content": system_prompt})
        messages.append({"role": "user", "content": user_message})
        
        # 嘗試 OpenAI 兼容格式
        request_body = {
            "messages": messages,
            "max_tokens": 500,
            "temperature": 0.7
        }
        if model:
            request_body["model"] = model
        
        # 增加超時時間到 90 秒（AI 生成可能需要更長時間）
        timeout = aiohttp.ClientTimeout(total=90, connect=10)
        
        try:
            start_time = time.time()
            async with aiohttp.ClientSession() as session:
                # 嘗試 /v1/chat/completions 端點
                chat_url = endpoint.rstrip('/')
                if not chat_url.endswith('/v1/chat/completions'):
                    chat_url = chat_url.rstrip('/') + '/v1/chat/completions'
                
                print(f"[AI] Attempting to call AI endpoint: {chat_url}", file=sys.stderr)
                print(f"[AI] Request body: model={model}, messages={len(messages)}, max_tokens=500", file=sys.stderr)
                
                try:
                    request_start = time.time()
                    async with session.post(chat_url, json=request_body, timeout=timeout) as resp:
                        connect_time = time.time() - request_start
                        print(f"[AI] Connection established in {connect_time:.2f}s, status: {resp.status}", file=sys.stderr)
                        
                        if resp.status == 200:
                            data_start = time.time()
                            data = await resp.json()
                            data_time = time.time() - data_start
                            total_time = time.time() - start_time
                            
                            print(f"[AI] Response received in {data_time:.2f}s, total time: {total_time:.2f}s", file=sys.stderr)
                            
                            if 'choices' in data and len(data['choices']) > 0:
                                content = data['choices'][0].get('message', {}).get('content', '')
                                print(f"[AI] ✓ Successfully generated response (length: {len(content)})", file=sys.stderr)
                                return content
                            else:
                                print(f"[AI] ✗ Response missing 'choices' field. Full response: {data}", file=sys.stderr)
                                raise Exception(f"AI 服務返回了無效的響應格式: {list(data.keys())}")
                        else:
                            error_text = await resp.text()
                            print(f"[AI] ✗ Error response (status {resp.status}): {error_text[:500]}", file=sys.stderr)
                            raise Exception(f"AI 服務返回錯誤 (HTTP {resp.status}): {error_text[:200]}")
                            
                except asyncio.TimeoutError:
                    elapsed = time.time() - start_time
                    print(f"[AI] ✗ Request timeout after {elapsed:.2f}s for endpoint: {chat_url}", file=sys.stderr)
                    raise Exception(f"AI 服務響應超時（{elapsed:.1f}秒）。可能原因：\n1. AI 服務響應過慢\n2. 網絡延遲過高\n3. 模型加載中\n請檢查 AI 服務狀態")
                except aiohttp.ClientConnectorError as e:
                    elapsed = time.time() - start_time
                    print(f"[AI] ✗ Connection error after {elapsed:.2f}s: {e}", file=sys.stderr)
                    raise Exception(f"無法連接到 AI 服務 ({host}:{port})。請檢查：\n1. AI 服務是否正在運行\n2. 防火牆是否允許連接\n3. 端點地址是否正確")
                except aiohttp.ClientError as e:
                    elapsed = time.time() - start_time
                    print(f"[AI] ✗ Client error after {elapsed:.2f}s: {e}", file=sys.stderr)
                    # 如果 /v1/chat/completions 失敗，嘗試直接端點
                    if chat_url != endpoint:
                        print(f"[AI] Trying direct endpoint: {endpoint}", file=sys.stderr)
                        try:
                            async with session.post(endpoint, json=request_body, timeout=timeout) as resp2:
                                if resp2.status == 200:
                                    data = await resp2.json()
                                    # 處理各種響應格式
                                    if 'choices' in data:
                                        return data['choices'][0].get('message', {}).get('content', '')
                                    elif 'response' in data:
                                        return data['response']
                                    elif 'content' in data:
                                        return data['content']
                                    elif 'text' in data:
                                        return data['text']
                                else:
                                    error_text = await resp2.text()
                                    print(f"[AI] Direct endpoint error (status {resp2.status}): {error_text[:200]}", file=sys.stderr)
                        except Exception as e2:
                            print(f"[AI] Direct endpoint also failed: {e2}", file=sys.stderr)
                    raise Exception(f"網絡錯誤: {str(e)}")
                    
        except asyncio.TimeoutError:
            raise Exception("AI 服務響應超時，請檢查服務連接或增加超時時間")
        except aiohttp.ClientError as e:
            error_msg = str(e)
            print(f"[AI] Network error: {error_msg}", file=sys.stderr)
            raise Exception(f"無法連接到 AI 服務 ({endpoint}): {error_msg}")
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"[AI] Unexpected error: {error_details}", file=sys.stderr)
            raise
            raise
    
    async def handle_add_ai_memory(self, payload: Dict[str, Any]):
        """Add an AI memory for a user"""
        try:
            user_id = payload.get('userId', '')
            memory_type = payload.get('memoryType', 'fact')
            content = payload.get('content', '')
            importance = payload.get('importance', 0.5)
            
            if not user_id or not content:
                self.send_event("ai-memory-added", {"success": False, "error": "Missing userId or content"})
                return
            
            memory_id = await db.add_ai_memory(user_id, memory_type, content, importance)
            
            self.send_event("ai-memory-added", {
                "success": True,
                "memoryId": memory_id
            })
            self.send_log(f"為用戶 {user_id} 添加了記憶", "success")
        except Exception as e:
            self.send_event("ai-memory-added", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_get_ai_memories(self, payload: Dict[str, Any]):
        """Get AI memories for a user"""
        try:
            user_id = payload.get('userId', '')
            memory_type = payload.get('memoryType')
            limit = payload.get('limit', 10)
            
            if not user_id:
                self.send_event("ai-memories", {"success": False, "error": "Missing userId"})
                return
            
            memories = await db.get_ai_memories(user_id, memory_type, limit)
            
            self.send_event("ai-memories", {
                "success": True,
                "userId": user_id,
                "memories": memories
            })
        except Exception as e:
            self.send_event("ai-memories", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_analyze_conversation(self, payload: Dict[str, Any]):
        """Analyze conversation and provide insights"""
        try:
            user_id = payload.get('userId', '')
            lead_id = payload.get('leadId')
            
            if not user_id:
                self.send_event("conversation-analysis", {"success": False, "error": "Missing userId"})
                return
            
            # Get conversation history
            history = await db.get_chat_history(user_id, limit=50)
            
            # Analyze conversation
            analysis = await ai_context.analyze_conversation_stage(user_id, history)
            
            self.send_event("conversation-analysis", {
                "success": True,
                "userId": user_id,
                "leadId": lead_id,
                "analysis": analysis
            })
        except Exception as e:
            self.send_event("conversation-analysis", {
                "success": False,
                "error": str(e)
            })

    # ==================== Knowledge Base Handlers ====================
    
    async def handle_init_knowledge_base(self):
        """Initialize knowledge base"""
        try:
            await search_engine.initialize()
            stats = await search_engine.get_stats()
            
            self.send_event("knowledge-base-initialized", {
                "success": True,
                "stats": stats
            })
            self.send_log("知識庫初始化成功", "success")
        except Exception as e:
            self.send_event("knowledge-base-initialized", {
                "success": False,
                "error": str(e)
            })
            self.send_log(f"知識庫初始化失敗: {str(e)}", "error")
    
    async def handle_get_knowledge_stats(self):
        """Get knowledge base statistics"""
        try:
            stats = await search_engine.get_stats()
            self.send_event("knowledge-stats", {
                "success": True,
                "stats": stats
            })
        except Exception as e:
            self.send_event("knowledge-stats", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_add_document(self, payload: Dict[str, Any]):
        """Add a document to knowledge base"""
        try:
            file_path = payload.get('filePath')
            title = payload.get('title')
            category = payload.get('category', 'general')
            tags = payload.get('tags', [])
            content = payload.get('content')  # For direct text input
            
            if content and title:
                result = await document_manager.add_document_from_text(
                    title=title,
                    content=content,
                    category=category,
                    tags=tags
                )
            elif file_path:
                result = await document_manager.add_document(
                    file_path=file_path,
                    title=title,
                    category=category,
                    tags=tags
                )
            else:
                result = {"success": False, "error": "No file or content provided"}
            
            self.send_event("document-added", result)
            if result.get('success'):
                self.send_log(f"文檔已添加: {result.get('title')}", "success")
        except Exception as e:
            self.send_event("document-added", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_get_documents(self, payload: Dict[str, Any]):
        """Get all documents"""
        try:
            category = payload.get('category')
            documents = await document_manager.get_all_documents(category)
            
            self.send_event("documents-list", {
                "success": True,
                "documents": documents
            })
        except Exception as e:
            self.send_event("documents-list", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_delete_document(self, payload: Dict[str, Any]):
        """Delete a document"""
        try:
            doc_id = payload.get('id')
            await document_manager.delete_document(doc_id)
            
            self.send_event("document-deleted", {"success": True, "id": doc_id})
            self.send_log("文檔已刪除", "success")
        except Exception as e:
            self.send_event("document-deleted", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_add_media(self, payload: Dict[str, Any]):
        """Add image or video to media library"""
        try:
            media_type = payload.get('mediaType', 'image')
            file_path = payload.get('filePath')
            base64_data = payload.get('base64Data')
            name = payload.get('name')
            category = payload.get('category', 'general')
            tags = payload.get('tags', [])
            description = payload.get('description')
            
            if media_type == 'image':
                result = await media_manager.add_image(
                    file_path=file_path,
                    base64_data=base64_data,
                    name=name,
                    category=category,
                    tags=tags,
                    description=description
                )
            else:
                result = await media_manager.add_video(
                    file_path=file_path,
                    base64_data=base64_data,
                    name=name,
                    category=category,
                    tags=tags,
                    description=description
                )
            
            self.send_event("media-added", result)
            if result.get('success'):
                self.send_log(f"媒體已添加: {result.get('name')}", "success")
        except Exception as e:
            self.send_event("media-added", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_get_media(self, payload: Dict[str, Any]):
        """Get media resources"""
        try:
            media_type = payload.get('mediaType')
            category = payload.get('category')
            
            media = await media_manager.get_all_media(media_type, category)
            
            self.send_event("media-list", {
                "success": True,
                "media": media
            })
        except Exception as e:
            self.send_event("media-list", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_delete_media(self, payload: Dict[str, Any]):
        """Delete a media resource"""
        try:
            media_id = payload.get('id')
            await media_manager.delete_media(media_id)
            
            self.send_event("media-deleted", {"success": True, "id": media_id})
            self.send_log("媒體已刪除", "success")
        except Exception as e:
            self.send_event("media-deleted", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_search_knowledge(self, payload: Dict[str, Any]):
        """Search knowledge base"""
        try:
            query = payload.get('query', '')
            include_docs = payload.get('includeDocs', True)
            include_images = payload.get('includeImages', True)
            include_videos = payload.get('includeVideos', True)
            limit = payload.get('limit', 10)
            
            results = await search_engine.search(
                query=query,
                include_docs=include_docs,
                include_images=include_images,
                include_videos=include_videos,
                limit=limit
            )
            
            self.send_event("knowledge-search-results", {
                "success": True,
                "query": query,
                "results": results
            })
        except Exception as e:
            self.send_event("knowledge-search-results", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_add_qa_pair(self, payload: Dict[str, Any]):
        """Add a QA pair"""
        try:
            question = payload.get('question')
            answer = payload.get('answer')
            category = payload.get('category', 'general')
            keywords = payload.get('keywords', [])
            media_ids = payload.get('mediaIds', [])
            
            if not question or not answer:
                self.send_event("qa-pair-added", {"success": False, "error": "Question and answer required"})
                return
            
            qa_id = await search_engine.add_qa_pair(
                question=question,
                answer=answer,
                category=category,
                keywords=keywords,
                media_ids=media_ids
            )
            
            self.send_event("qa-pair-added", {
                "success": True,
                "id": qa_id
            })
            self.send_log("問答對已添加", "success")
        except Exception as e:
            self.send_event("qa-pair-added", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_get_qa_pairs(self, payload: Dict[str, Any]):
        """Get all QA pairs"""
        try:
            category = payload.get('category')
            qa_pairs = await search_engine.get_all_qa_pairs(category)
            
            self.send_event("qa-pairs-list", {
                "success": True,
                "qaPairs": qa_pairs
            })
        except Exception as e:
            self.send_event("qa-pairs-list", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_import_qa(self, payload: Dict[str, Any]):
        """Import QA pairs from file"""
        try:
            file_path = payload.get('filePath')
            file_type = payload.get('fileType', 'csv')
            
            if file_type == 'csv':
                result = await search_engine.import_qa_from_csv(file_path)
            else:
                result = await search_engine.import_qa_from_json(file_path)
            
            self.send_event("qa-imported", result)
            if result.get('success'):
                self.send_log(f"導入了 {result.get('imported', 0)} 條問答對", "success")
        except Exception as e:
            self.send_event("qa-imported", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_get_rag_context(self, payload: Dict[str, Any]):
        """Get RAG context for AI prompt"""
        try:
            query = payload.get('query', '')
            max_chunks = payload.get('maxChunks', 3)
            max_tokens = payload.get('maxTokens', 2000)
            
            context = await search_engine.build_rag_context(
                query=query,
                max_chunks=max_chunks,
                max_tokens=max_tokens
            )
            
            # Also get relevant media
            media = await search_engine.find_relevant_media(query, limit=3)
            
            self.send_event("rag-context", {
                "success": True,
                "context": context,
                "relevantMedia": media
            })
        except Exception as e:
            self.send_event("rag-context", {
                "success": False,
                "error": str(e)
            })
    
    # ==================== Auto Funnel Handlers ====================
    
    async def handle_get_funnel_overview(self):
        """獲取漏斗總覽"""
        try:
            overview = await auto_funnel.get_funnel_overview()
            self.send_event("funnel-overview", {
                "success": True,
                **overview
            })
        except Exception as e:
            self.send_event("funnel-overview", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_analyze_user_message(self, payload: Dict[str, Any]):
        """分析用戶消息並確定漏斗階段"""
        try:
            user_id = payload.get('userId', '')
            message = payload.get('message', '')
            is_from_user = payload.get('isFromUser', True)
            
            result = await auto_funnel.analyze_message(user_id, message, is_from_user)
            
            self.send_event("message-analyzed", {
                "success": True,
                **result
            })
        except Exception as e:
            self.send_event("message-analyzed", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_transition_funnel_stage(self, payload: Dict[str, Any]):
        """手動轉換漏斗階段"""
        try:
            user_id = payload.get('userId', '')
            new_stage = payload.get('stage', '')
            reason = payload.get('reason', '手動設置')
            
            result = await auto_funnel.transition_stage(user_id, new_stage, reason)
            
            self.send_event("stage-transitioned", result)
        except Exception as e:
            self.send_event("stage-transitioned", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_get_user_journey(self, payload: Dict[str, Any]):
        """獲取用戶漏斗旅程"""
        try:
            user_id = payload.get('userId', '')
            journey = await auto_funnel.get_user_journey(user_id)
            
            self.send_event("user-journey", {
                "success": True,
                "userId": user_id,
                "journey": journey
            })
        except Exception as e:
            self.send_event("user-journey", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_batch_update_stages(self, payload: Dict[str, Any]):
        """批量更新用戶階段"""
        try:
            user_ids = payload.get('userIds', [])
            new_stage = payload.get('stage', '')
            reason = payload.get('reason', '批量更新')
            
            result = await auto_funnel.batch_update_stages(user_ids, new_stage, reason)
            
            self.send_event("batch-stages-updated", {
                "success": True,
                **result
            })
        except Exception as e:
            self.send_event("batch-stages-updated", {
                "success": False,
                "error": str(e)
            })
    
    # ==================== Vector Memory Handlers ====================
    
    async def handle_add_vector_memory(self, payload: Dict[str, Any]):
        """添加向量記憶"""
        try:
            user_id = payload.get('userId', '')
            content = payload.get('content', '')
            memory_type = payload.get('type', 'conversation')
            importance = payload.get('importance', 0.5)
            
            memory_id = await vector_memory.add_memory(
                user_id=user_id,
                content=content,
                memory_type=memory_type,
                importance=importance
            )
            
            self.send_event("memory-added", {
                "success": True,
                "memoryId": memory_id
            })
        except Exception as e:
            self.send_event("memory-added", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_search_vector_memories(self, payload: Dict[str, Any]):
        """搜索向量記憶"""
        try:
            user_id = payload.get('userId', '')
            query = payload.get('query', '')
            limit = payload.get('limit', 5)
            memory_type = payload.get('type')
            
            memories = await vector_memory.search_memories(
                user_id=user_id,
                query=query,
                limit=limit,
                memory_type=memory_type
            )
            
            self.send_event("memories-searched", {
                "success": True,
                "memories": memories
            })
        except Exception as e:
            self.send_event("memories-searched", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_get_memory_context(self, payload: Dict[str, Any]):
        """獲取記憶上下文"""
        try:
            user_id = payload.get('userId', '')
            message = payload.get('message', '')
            max_tokens = payload.get('maxTokens', 1500)
            
            context = await vector_memory.build_context_from_memory(
                user_id=user_id,
                current_message=message,
                max_tokens=max_tokens
            )
            
            self.send_event("memory-context", {
                "success": True,
                "context": context
            })
        except Exception as e:
            self.send_event("memory-context", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_summarize_conversation(self, payload: Dict[str, Any]):
        """生成對話摘要"""
        try:
            user_id = payload.get('userId', '')
            max_messages = payload.get('maxMessages', 50)
            
            result = await vector_memory.summarize_conversation(
                user_id=user_id,
                max_messages=max_messages
            )
            
            self.send_event("conversation-summarized", result)
        except Exception as e:
            self.send_event("conversation-summarized", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_get_memory_stats(self, payload: Dict[str, Any]):
        """獲取記憶統計"""
        try:
            user_id = payload.get('userId', '')
            stats = await vector_memory.get_user_memory_stats(user_id)
            
            self.send_event("memory-stats", {
                "success": True,
                "userId": user_id,
                **stats
            })
        except Exception as e:
            self.send_event("memory-stats", {
                "success": False,
                "error": str(e)
            })
    
    # ==================== Telegram RAG Handlers ====================
    
    async def handle_init_rag_system(self):
        """初始化 Telegram RAG 系統"""
        try:
            from telegram_rag_system import telegram_rag
            from chat_history_indexer import chat_indexer
            
            # 初始化 RAG 系統
            await telegram_rag.initialize()
            
            # 初始化索引服務
            await chat_indexer.initialize()
            
            # 啟動後台索引
            await chat_indexer.start_background_indexing()
            
            self.send_log("✓ Telegram RAG 系統初始化完成", "success")
            self.send_event("rag-initialized", {"success": True})
            
        except Exception as e:
            self.send_log(f"RAG 系統初始化失敗: {e}", "error")
            self.send_event("rag-initialized", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_get_rag_stats(self):
        """獲取 RAG 系統統計"""
        try:
            from telegram_rag_system import telegram_rag
            from chat_history_indexer import chat_indexer
            
            rag_stats = await telegram_rag.get_statistics()
            indexer_stats = await chat_indexer.get_indexing_statistics()
            
            self.send_event("rag-stats", {
                "success": True,
                "rag": rag_stats,
                "indexer": indexer_stats
            })
            
        except Exception as e:
            self.send_event("rag-stats", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_search_rag(self, payload: Dict[str, Any]):
        """搜索 RAG 知識庫"""
        try:
            from telegram_rag_system import telegram_rag, KnowledgeType
            
            query = payload.get('query', '')
            limit = payload.get('limit', 5)
            knowledge_type = payload.get('knowledgeType')
            
            # 轉換知識類型
            kt = None
            if knowledge_type:
                try:
                    kt = KnowledgeType(knowledge_type)
                except:
                    pass
            
            results = await telegram_rag.search(query, limit=limit, knowledge_type=kt)
            
            # 格式化結果
            formatted_results = []
            for r in results:
                formatted_results.append({
                    'id': r.item.id,
                    'type': r.item.knowledge_type.value,
                    'question': r.item.question,
                    'answer': r.item.answer,
                    'successScore': r.item.success_score,
                    'useCount': r.item.use_count,
                    'similarity': r.similarity,
                    'source': r.source
                })
            
            self.send_event("rag-search-result", {
                "success": True,
                "query": query,
                "results": formatted_results
            })
            
        except Exception as e:
            self.send_event("rag-search-result", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_trigger_rag_learning(self, payload: Dict[str, Any]):
        """觸發 RAG 學習"""
        try:
            from telegram_rag_system import telegram_rag, ConversationOutcome
            from chat_history_indexer import chat_indexer
            
            user_id = payload.get('userId')
            account_phone = payload.get('accountPhone', '')
            outcome = payload.get('outcome', 'unknown')
            
            if user_id:
                # 學習特定用戶的對話
                await chat_indexer.on_conversation_ended(
                    user_id=user_id,
                    account_phone=account_phone,
                    outcome=outcome
                )
                
                self.send_event("rag-learning-triggered", {
                    "success": True,
                    "userId": user_id
                })
            else:
                # 批量處理待索引的對話
                result = await chat_indexer.index_pending_conversations()
                
                self.send_event("rag-learning-triggered", {
                    "success": True,
                    "conversationsProcessed": result.get('conversations_processed', 0),
                    "knowledgeExtracted": result.get('knowledge_extracted', 0)
                })
            
        except Exception as e:
            self.send_event("rag-learning-triggered", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_add_rag_knowledge(self, payload: Dict[str, Any]):
        """手動添加 RAG 知識"""
        try:
            from telegram_rag_system import telegram_rag, KnowledgeType
            
            knowledge_type = payload.get('type', 'qa')
            question = payload.get('question', '')
            answer = payload.get('answer', '')
            context = payload.get('context', '')
            
            # 轉換類型
            try:
                kt = KnowledgeType(knowledge_type)
            except:
                kt = KnowledgeType.QA
            
            knowledge_id = await telegram_rag.add_manual_knowledge(
                knowledge_type=kt,
                question=question,
                answer=answer,
                context=context
            )
            
            self.send_event("rag-knowledge-added", {
                "success": True,
                "knowledgeId": knowledge_id
            })
            
        except Exception as e:
            self.send_event("rag-knowledge-added", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_rag_feedback(self, payload: Dict[str, Any]):
        """記錄 RAG 知識反饋"""
        try:
            from telegram_rag_system import telegram_rag
            
            knowledge_id = payload.get('knowledgeId')
            is_positive = payload.get('isPositive', True)
            feedback_text = payload.get('feedbackText', '')
            
            await telegram_rag.record_feedback(
                knowledge_id=knowledge_id,
                is_positive=is_positive,
                feedback_text=feedback_text
            )
            
            self.send_event("rag-feedback-recorded", {
                "success": True,
                "knowledgeId": knowledge_id
            })
            
        except Exception as e:
            self.send_event("rag-feedback-recorded", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_reindex_conversations(self, payload: Dict[str, Any]):
        """重新索引對話"""
        try:
            from chat_history_indexer import chat_indexer
            
            high_value_only = payload.get('highValueOnly', False)
            days = payload.get('days', 30)
            
            if high_value_only:
                result = await chat_indexer.reindex_high_value_conversations(days=days)
            else:
                result = await chat_indexer.index_pending_conversations()
            
            self.send_log(f"重新索引完成: 處理 {result.get('conversations_processed', 0)} 個對話", "success")
            
            self.send_event("rag-reindex-complete", {
                "success": True,
                **result
            })
            
        except Exception as e:
            self.send_event("rag-reindex-complete", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_cleanup_rag_knowledge(self, payload: Dict[str, Any]):
        """清理 RAG 知識庫"""
        try:
            from telegram_rag_system import telegram_rag
            
            min_score = payload.get('minScore', 0.2)
            days_old = payload.get('daysOld', 30)
            merge_similar = payload.get('mergeSimilar', True)
            
            result = {
                'deleted': 0,
                'merged': 0
            }
            
            # 清理低質量知識
            result['deleted'] = await telegram_rag.cleanup_low_quality_knowledge(
                min_score=min_score,
                days_old=days_old
            )
            
            # 合併相似知識
            if merge_similar:
                result['merged'] = await telegram_rag.merge_similar_knowledge()
            
            self.send_log(f"RAG 知識庫清理完成: 刪除 {result['deleted']} 條, 合併 {result['merged']} 條", "success")
            
            self.send_event("rag-cleanup-complete", {
                "success": True,
                **result
            })
            
        except Exception as e:
            self.send_event("rag-cleanup-complete", {
                "success": False,
                "error": str(e)
            })
    
    # ==================== Scheduler Handlers ====================
    
    async def handle_schedule_follow_up(self, payload: Dict[str, Any]):
        """排程跟進任務"""
        try:
            user_id = payload.get('userId', '')
            scheduled_at_str = payload.get('scheduledAt', '')
            message_template = payload.get('messageTemplate')
            task_type = payload.get('taskType', 'reminder')
            
            # Parse datetime
            scheduled_at = datetime.fromisoformat(scheduled_at_str.replace('Z', '+00:00'))
            
            task_id = await scheduler.schedule_follow_up(
                user_id=user_id,
                scheduled_at=scheduled_at,
                message_template=message_template,
                task_type=task_type
            )
            
            self.send_event("follow-up-scheduled", {
                "success": True,
                "taskId": task_id
            })
        except Exception as e:
            self.send_event("follow-up-scheduled", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_get_pending_tasks(self, payload: Dict[str, Any]):
        """獲取待執行任務"""
        try:
            limit = payload.get('limit', 50)
            tasks = await scheduler.get_pending_tasks(limit=limit)
            
            self.send_event("pending-tasks", {
                "success": True,
                "tasks": tasks
            })
        except Exception as e:
            self.send_event("pending-tasks", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_cancel_scheduled_task(self, payload: Dict[str, Any]):
        """取消排程任務"""
        try:
            task_id = payload.get('taskId')
            await scheduler.cancel_task(task_id)
            
            self.send_event("task-cancelled", {
                "success": True,
                "taskId": task_id
            })
        except Exception as e:
            self.send_event("task-cancelled", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_get_scheduler_stats(self):
        """獲取調度器統計"""
        try:
            stats = await scheduler.get_scheduler_stats()
            
            self.send_event("scheduler-stats", {
                "success": True,
                **stats
            })
        except Exception as e:
            self.send_event("scheduler-stats", {
                "success": False,
                "error": str(e)
            })
    
    # ==================== User CRM Handlers ====================
    
    async def handle_get_user_profile_full(self, payload: Dict[str, Any]):
        """獲取完整用戶資料"""
        try:
            user_id = payload.get('userId', '')
            
            # Get user profile
            profile = await db.get_user_profile(user_id)
            
            # Get user CRM data
            cursor = await db._connection.execute(
                "SELECT * FROM user_crm WHERE user_id = ?", (user_id,)
            )
            crm_row = await cursor.fetchone()
            crm_data = dict(crm_row) if crm_row else {}
            
            # Get user tags
            cursor = await db._connection.execute(
                "SELECT tag, tag_type, confidence FROM user_tags WHERE user_id = ?",
                (user_id,)
            )
            tags = await cursor.fetchall()
            tags_list = [dict(t) for t in tags]
            
            # Get funnel history
            history = await db.get_funnel_history(user_id, limit=10)
            
            # Get memory stats
            stats = await vector_memory.get_user_memory_stats(user_id)
            
            self.send_event("user-profile-full", {
                "success": True,
                "userId": user_id,
                "profile": profile,
                "crm": crm_data,
                "tags": tags_list,
                "funnelHistory": history,
                "memoryStats": stats
            })
        except Exception as e:
            self.send_event("user-profile-full", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_update_user_crm(self, payload: Dict[str, Any]):
        """更新用戶 CRM 資料"""
        try:
            user_id = payload.get('userId', '')
            data = payload.get('data', {})
            
            # Build update query
            fields = ['company', 'industry', 'job_title', 'phone', 'email', 
                      'website', 'location', 'budget_range', 'pain_points', 'goals']
            
            updates = []
            values = []
            
            for field in fields:
                if field in data:
                    updates.append(f"{field} = ?")
                    values.append(data[field])
            
            if updates:
                values.append(user_id)
                
                # Check if record exists
                cursor = await db._connection.execute(
                    "SELECT 1 FROM user_crm WHERE user_id = ?", (user_id,)
                )
                exists = await cursor.fetchone()
                
                if exists:
                    await db._connection.execute(f"""
                        UPDATE user_crm SET {', '.join(updates)}, updated_at = CURRENT_TIMESTAMP
                        WHERE user_id = ?
                    """, values)
                else:
                    # Insert new record
                    insert_fields = [f for f in fields if f in data]
                    insert_values = [data[f] for f in insert_fields]
                    placeholders = ', '.join(['?' for _ in insert_fields])
                    
                    await db._connection.execute(f"""
                        INSERT INTO user_crm (user_id, {', '.join(insert_fields)})
                        VALUES (?, {placeholders})
                    """, [user_id] + insert_values)
                
                await db._connection.commit()
            
            self.send_event("crm-updated", {
                "success": True,
                "userId": user_id
            })
        except Exception as e:
            self.send_event("crm-updated", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_add_user_tag(self, payload: Dict[str, Any]):
        """添加用戶標籤"""
        try:
            user_id = payload.get('userId', '')
            tag = payload.get('tag', '')
            tag_type = payload.get('tagType', 'custom')
            
            await db._connection.execute("""
                INSERT OR IGNORE INTO user_tags (user_id, tag, tag_type)
                VALUES (?, ?, ?)
            """, (user_id, tag, tag_type))
            await db._connection.commit()
            
            self.send_event("tag-added", {
                "success": True,
                "userId": user_id,
                "tag": tag
            })
        except Exception as e:
            self.send_event("tag-added", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_remove_user_tag(self, payload: Dict[str, Any]):
        """移除用戶標籤"""
        try:
            user_id = payload.get('userId', '')
            tag = payload.get('tag', '')
            
            await db._connection.execute("""
                DELETE FROM user_tags WHERE user_id = ? AND tag = ?
            """, (user_id, tag))
            await db._connection.commit()
            
            self.send_event("tag-removed", {
                "success": True,
                "userId": user_id,
                "tag": tag
            })
        except Exception as e:
            self.send_event("tag-removed", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_get_user_tags(self, payload: Dict[str, Any]):
        """獲取用戶標籤"""
        try:
            user_id = payload.get('userId', '')
            
            cursor = await db._connection.execute("""
                SELECT ut.tag, ut.tag_type, ut.confidence, ut.auto_assigned, ut.created_at,
                       td.color, td.description
                FROM user_tags ut
                LEFT JOIN tag_definitions td ON ut.tag = td.tag
                WHERE ut.user_id = ?
            """, (user_id,))
            
            rows = await cursor.fetchall()
            tags = [dict(row) for row in rows]
            
            self.send_event("user-tags", {
                "success": True,
                "userId": user_id,
                "tags": tags
            })
        except Exception as e:
            self.send_event("user-tags", {
                "success": False,
                "error": str(e)
            })
    
    # ==================== Full-Text Search Handlers ====================
    
    async def handle_search_chat_history(self, payload: Dict[str, Any]):
        """全文搜索聊天記錄"""
        try:
            from fulltext_search import get_search_engine
            from datetime import datetime as dt
            
            query = payload.get('query', '')
            if not query:
                self.send_event("search-chat-history", {
                    "success": False,
                    "error": "搜索查詢不能為空"
                })
                return
            
            user_id = payload.get('userId')
            account_phone = payload.get('accountPhone')
            role = payload.get('role')
            limit = payload.get('limit', 50)
            offset = payload.get('offset', 0)
            
            # 解析日期範圍
            date_from = None
            date_to = None
            if payload.get('dateFrom'):
                try:
                    date_from = dt.fromisoformat(payload['dateFrom'])
                except:
                    pass
            if payload.get('dateTo'):
                try:
                    date_to = dt.fromisoformat(payload['dateTo'])
                except:
                    pass
            
            search_engine = get_search_engine()
            results = await search_engine.search_chat_history(
                query=query,
                user_id=user_id,
                account_phone=account_phone,
                role=role,
                date_from=date_from,
                date_to=date_to,
                limit=limit,
                offset=offset
            )
            
            self.send_event("search-chat-history", {
                "success": True,
                "query": query,
                "results": results,
                "count": len(results),
                "hasMore": len(results) == limit
            })
        except Exception as e:
            self.send_event("search-chat-history", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_search_leads(self, payload: Dict[str, Any]):
        """全文搜索 Lead"""
        try:
            from fulltext_search import get_search_engine
            from datetime import datetime as dt
            
            query = payload.get('query', '')
            if not query:
                self.send_event("search-leads", {
                    "success": False,
                    "error": "搜索查詢不能為空"
                })
                return
            
            status = payload.get('status')
            limit = payload.get('limit', 50)
            offset = payload.get('offset', 0)
            
            # 解析日期範圍
            date_from = None
            date_to = None
            if payload.get('dateFrom'):
                try:
                    date_from = dt.fromisoformat(payload['dateFrom'])
                except:
                    pass
            if payload.get('dateTo'):
                try:
                    date_to = dt.fromisoformat(payload['dateTo'])
                except:
                    pass
            
            search_engine = get_search_engine()
            results = await search_engine.search_leads(
                query=query,
                status=status,
                date_from=date_from,
                date_to=date_to,
                limit=limit,
                offset=offset
            )
            
            self.send_event("search-leads", {
                "success": True,
                "query": query,
                "results": results,
                "count": len(results),
                "hasMore": len(results) == limit
            })
        except Exception as e:
            self.send_event("search-leads", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_rebuild_search_index(self):
        """重建搜索索引"""
        try:
            from fulltext_search import get_search_engine
            
            search_engine = get_search_engine()
            await search_engine.rebuild_index()
            
            self.send_log("搜索索引重建完成", "success")
            self.send_event("search-index-rebuilt", {
                "success": True,
                "message": "搜索索引重建完成"
            })
        except Exception as e:
            self.send_log(f"重建搜索索引失敗: {str(e)}", "error")
            self.send_event("search-index-rebuilt", {
                "success": False,
                "error": str(e)
            })
    
    # ==================== Analytics Handlers ====================
    
    async def handle_analyze_funnel(self, payload: Dict[str, Any]):
        """分析轉化漏斗"""
        try:
            from analytics_engine import AnalyticsEngine
            
            days = payload.get('days', 30)
            start_date = payload.get('startDate')
            end_date = payload.get('endDate')
            
            # 轉換日期
            start_dt = None
            end_dt = None
            if start_date:
                try:
                    start_dt = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
                except:
                    pass
            if end_date:
                try:
                    end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
                except:
                    pass
            
            engine = AnalyticsEngine(db)
            result = await engine.analyze_funnel(
                days=days,
                start_date=start_dt,
                end_date=end_dt
            )
            
            self.send_event("funnel-analysis", {
                "success": True,
                **result
            })
        except Exception as e:
            self.send_event("funnel-analysis", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_analyze_user_journey(self, payload: Dict[str, Any]):
        """分析用戶旅程"""
        try:
            from analytics_engine import AnalyticsEngine
            
            user_id = payload.get('userId', '')
            if not user_id:
                self.send_event("user-journey-analysis", {
                    "success": False,
                    "error": "用戶ID不能為空"
                })
                return
            
            engine = AnalyticsEngine(db)
            result = await engine.analyze_user_journey(user_id)
            
            self.send_event("user-journey-analysis", {
                "success": True,
                **result
            })
        except Exception as e:
            self.send_event("user-journey-analysis", {
                "success": False,
                "error": str(e)
            })
    
    # ==================== Monitoring Status Handlers ====================
    
    async def handle_get_monitoring_status(self):
        """獲取監控狀態詳情"""
        try:
            accounts = await db.get_all_accounts()
            listener_accounts = [a for a in accounts if a.get('role') == 'Listener']
            sender_accounts = [a for a in accounts if a.get('role') == 'Sender']
            
            # 檢查監控帳號狀態
            monitoring_status = []
            for account in listener_accounts:
                phone = account.get('phone')
                status = account.get('status', 'Offline')
                
                # 檢查是否有註冊處理器
                has_handler = phone in self.telegram_manager.message_handlers if hasattr(self.telegram_manager, 'message_handlers') else False
                
                # 檢查監控信息
                monitoring_info = None
                if hasattr(self.telegram_manager, 'monitoring_info') and phone in self.telegram_manager.monitoring_info:
                    info = self.telegram_manager.monitoring_info[phone]
                    monitoring_info = {
                        'chatIds': list(info.get('chat_ids', [])),
                        'groupUrls': info.get('group_urls', []),
                        'keywordSetCount': len(info.get('keyword_sets', []))
                    }
                
                monitoring_status.append({
                    'phone': phone,
                    'status': status,
                    'hasHandler': has_handler,
                    'monitoringInfo': monitoring_info
                })
            
            # 檢查發送帳號狀態
            sender_status = []
            for account in sender_accounts:
                phone = account.get('phone')
                status = account.get('status', 'Offline')
                
                # 檢查是否有私信處理器
                has_private_handler = False
                if hasattr(private_message_handler, 'private_handlers'):
                    has_private_handler = phone in private_message_handler.private_handlers
                
                sender_status.append({
                    'phone': phone,
                    'status': status,
                    'hasPrivateHandler': has_private_handler
                })
            
            # 獲取監控配置
            monitored_groups = await db.get_all_monitored_groups()
            keyword_sets = await db.get_all_keyword_sets()
            campaigns = await db.get_all_campaigns()
            active_campaigns = [c for c in campaigns if c.get('isActive') or c.get('is_active')]
            
            self.send_event("monitoring-status", {
                "success": True,
                "isMonitoring": self.is_monitoring,
                "listenerAccounts": monitoring_status,
                "senderAccounts": sender_status,
                "monitoredGroups": len(monitored_groups),
                "keywordSets": len(keyword_sets),
                "activeCampaigns": len(active_campaigns),
                "totalCampaigns": len(campaigns)
            })
        except Exception as e:
            self.send_event("monitoring-status", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_check_monitoring_health(self):
        """檢查監控健康狀態"""
        try:
            issues = []
            warnings = []
            
            # 檢查監控帳號
            accounts = await db.get_all_accounts()
            listener_accounts = [a for a in accounts if a.get('role') == 'Listener']
            online_listeners = [a for a in listener_accounts if a.get('status') == 'Online']
            
            if not listener_accounts:
                issues.append("沒有配置監控帳號（Listener 角色）")
            elif not online_listeners:
                issues.append(f"有 {len(listener_accounts)} 個監控帳號，但沒有在線的")
            
            # 檢查發送帳號
            sender_accounts = [a for a in accounts if a.get('role') == 'Sender']
            online_senders = [a for a in sender_accounts if a.get('status') == 'Online']
            
            if not sender_accounts:
                warnings.append("沒有配置發送帳號（Sender 角色）")
            elif not online_senders:
                warnings.append(f"有 {len(sender_accounts)} 個發送帳號，但沒有在線的")
            
            # 檢查監控配置
            monitored_groups = await db.get_all_monitored_groups()
            if not monitored_groups:
                issues.append("沒有配置監控群組")
            
            keyword_sets = await db.get_all_keyword_sets()
            if not keyword_sets:
                issues.append("沒有配置關鍵詞集")
            else:
                # 檢查關鍵詞集是否為空
                empty_sets = [ks for ks in keyword_sets if not ks.get('keywords')]
                if empty_sets:
                    warnings.append(f"有 {len(empty_sets)} 個關鍵詞集為空")
            
            # 檢查活動配置
            campaigns = await db.get_all_campaigns()
            active_campaigns = [c for c in campaigns if c.get('isActive') or c.get('is_active')]
            if not active_campaigns:
                warnings.append("沒有啟用的活動（即使捕獲到 Lead 也不會自動發送）")
            
            # 檢查處理器註冊
            handler_issues = []
            for account in online_listeners:
                phone = account.get('phone')
                if hasattr(self.telegram_manager, 'message_handlers'):
                    if phone not in self.telegram_manager.message_handlers:
                        handler_issues.append(f"監控帳號 {phone} 未註冊群組消息處理器")
            
            for account in online_senders:
                phone = account.get('phone')
                if hasattr(private_message_handler, 'private_handlers'):
                    if phone not in private_message_handler.private_handlers:
                        handler_issues.append(f"發送帳號 {phone} 未註冊私信處理器")
            
            if handler_issues:
                issues.extend(handler_issues)
            
            self.send_event("monitoring-health", {
                "success": True,
                "isHealthy": len(issues) == 0,
                "issues": issues,
                "warnings": warnings,
                "summary": {
                    "listenerAccounts": len(listener_accounts),
                    "onlineListeners": len(online_listeners),
                    "senderAccounts": len(sender_accounts),
                    "onlineSenders": len(online_senders),
                    "monitoredGroups": len(monitored_groups),
                    "keywordSets": len(keyword_sets),
                    "activeCampaigns": len(active_campaigns)
                }
            })
        except Exception as e:
            self.send_event("monitoring-health", {
                "success": False,
                "error": str(e)
            })
    
    # ==================== Database Management ====================
    
    async def handle_rebuild_database(self):
        """重建數據庫（會刪除所有數據）"""
        import shutil
        from pathlib import Path
        
        try:
            db_path = Path(config.DATABASE_URL)
            db_dir = db_path.parent
            
            self.send_log("開始重建數據庫...", "info")
            
            # 步驟 1: 備份現有數據庫
            if db_path.exists():
                backup_path = db_dir / f"tgmatrix_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
                try:
                    shutil.copy2(db_path, backup_path)
                    self.send_log(f"數據庫已備份到: {backup_path.name}", "success")
                except Exception as e:
                    self.send_log(f"備份失敗: {str(e)}", "warning")
            
            # 步驟 2: 關閉當前數據庫連接
            try:
                await db.close()
            except:
                pass
            
            # 步驟 3: 刪除舊數據庫文件
            try:
                if db_path.exists():
                    db_path.unlink()
                # 刪除 WAL 和 SHM 文件
                wal_path = Path(str(db_path) + "-wal")
                shm_path = Path(str(db_path) + "-shm")
                if wal_path.exists():
                    wal_path.unlink()
                if shm_path.exists():
                    shm_path.unlink()
                self.send_log("舊數據庫文件已刪除", "info")
            except Exception as e:
                self.send_log(f"刪除舊數據庫失敗: {str(e)}", "error")
                self.send_event("database-rebuild-result", {
                    "success": False,
                    "error": f"刪除舊數據庫失敗: {str(e)}"
                })
                return
            
            # 步驟 4: 重新初始化數據庫
            try:
                await db.initialize()
                await db.connect()
                
                # 驗證數據庫完整性
                cursor = await db._connection.execute("PRAGMA integrity_check")
                result = await cursor.fetchone()
                if result and result[0] == 'ok':
                    self.send_log("數據庫重建成功，完整性檢查通過", "success")
                else:
                    self.send_log(f"數據庫重建完成，但完整性檢查警告: {result[0] if result else 'Unknown'}", "warning")
                
                # 重新初始化全文搜索索引
                try:
                    from fulltext_search import init_search_engine
                    from config import DATABASE_PATH
                    search_engine = await init_search_engine(str(DATABASE_PATH))
                    self.send_log("全文搜索索引已重建", "success")
                except Exception as e:
                    self.send_log(f"全文搜索索引重建失敗（可選）: {str(e)}", "warning")
                
                self.send_event("database-rebuild-result", {
                    "success": True,
                    "message": "數據庫重建成功"
                })
                
                # 發送初始狀態事件，讓前端刷新
                await self.handle_get_initial_state()
                
            except Exception as e:
                import traceback
                error_details = traceback.format_exc()
                print(f"[Backend] Database rebuild error: {error_details}", file=sys.stderr)
                self.send_log(f"數據庫重建失敗: {str(e)}", "error")
                self.send_event("database-rebuild-result", {
                    "success": False,
                    "error": str(e)
                })
                
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"[Backend] Database rebuild error: {error_details}", file=sys.stderr)
            self.send_event("database-rebuild-result", {
                "success": False,
                "error": str(e)
            })
    
    # ==================== Chat History Handlers ====================
    
    async def handle_get_chat_history_full(self, payload: Dict[str, Any]):
        """獲取完整聊天記錄（支持分頁和緩存）"""
        try:
            import sys
            user_id = payload.get('userId', '')
            limit = payload.get('limit', 50)
            offset = payload.get('offset', 0)
            
            print(f"[Backend] Getting chat history for user {user_id}, limit={limit}, offset={offset}", file=sys.stderr)
            
            # 優化：只查詢一次，獲取 limit+1 條來判斷是否有更多
            history = await db.get_chat_history_paginated(
                user_id=user_id,
                limit=limit + 1,  # 多取一條來判斷是否有更多
                offset=offset
            )
            
            # 判斷是否有更多消息
            has_more = len(history) > limit
            # 只返回 limit 條
            if has_more:
                history = history[:limit]
            
            # 獲取總數（使用更快的查詢，只在需要時執行）
            # 如果 offset=0 且沒有更多消息，總數就是當前數量
            if offset == 0 and not has_more:
                total_count = len(history)
            else:
                # 需要查詢總數
                try:
                    cursor = await db._connection.execute("""
                        SELECT COUNT(*) as total FROM chat_history WHERE user_id = ?
                    """, (user_id,))
                    total_row = await cursor.fetchone()
                    total_count = total_row['total'] if total_row else len(history)
                except Exception as count_err:
                    print(f"[Backend] Error getting total count: {count_err}", file=sys.stderr)
                    total_count = len(history)  # 降級：使用當前數量
            
            # 獲取用戶資料
            profile = await db.get_user_profile(user_id)
            
            # 獲取用戶標籤
            cursor = await db._connection.execute("""
                SELECT tag, tag_type, confidence FROM user_tags WHERE user_id = ?
            """, (user_id,))
            tags = [dict(row) for row in await cursor.fetchall()]
            
            # 格式化消息（反轉順序，從舊到新）
            formatted_messages = []
            for msg in reversed(history):  # 反轉為從舊到新
                formatted_messages.append({
                    "id": msg.get('id'),
                    "role": msg.get('role'),
                    "content": msg.get('content'),
                    "timestamp": msg.get('timestamp'),
                    "accountPhone": msg.get('account_phone'),
                    "sourceGroup": msg.get('source_group'),
                })
            
            print(f"[Backend] Sending chat history: {len(formatted_messages)} messages, hasMore={has_more}, total={total_count}", file=sys.stderr)
            
            self.send_event("chat-history-full", {
                "success": True,
                "userId": user_id,
                "messages": formatted_messages,
                "profile": profile,
                "tags": tags,
                "total": total_count,
                "hasMore": has_more  # 添加 hasMore 字段
            })
        except Exception as e:
            import traceback
            import sys
            error_details = traceback.format_exc()
            print(f"[Backend] Error getting chat history: {error_details}", file=sys.stderr)
            self.send_event("chat-history-full", {
                "success": False,
                "error": str(e),
                "userId": payload.get('userId', '')
            })
    
    async def handle_get_chat_list(self, payload: Dict[str, Any]):
        """獲取聊天列表（所有有對話的用戶）"""
        try:
            limit = payload.get('limit', 50)
            offset = payload.get('offset', 0)
            search_query = payload.get('search', '')
            funnel_stage = payload.get('funnelStage')
            
            # 構建查詢
            query = """
                SELECT DISTINCT 
                    ch.user_id,
                    up.username,
                    up.first_name,
                    up.last_name,
                    up.funnel_stage,
                    up.interest_level,
                    up.last_interaction,
                    (SELECT content FROM chat_history ch2 
                     WHERE ch2.user_id = ch.user_id 
                     ORDER BY ch2.timestamp DESC LIMIT 1) as last_message,
                    (SELECT timestamp FROM chat_history ch2 
                     WHERE ch2.user_id = ch.user_id 
                     ORDER BY ch2.timestamp DESC LIMIT 1) as last_message_time,
                    (SELECT COUNT(*) FROM chat_history ch3 
                     WHERE ch3.user_id = ch.user_id AND ch3.role = 'user') as unread_count
                FROM chat_history ch
                LEFT JOIN user_profiles up ON ch.user_id = up.user_id
                WHERE 1=1
            """
            params = []
            
            if search_query:
                query += " AND (up.username LIKE ? OR up.first_name LIKE ? OR ch.content LIKE ?)"
                search_pattern = f"%{search_query}%"
                params.extend([search_pattern, search_pattern, search_pattern])
            
            if funnel_stage:
                query += " AND up.funnel_stage = ?"
                params.append(funnel_stage)
            
            query += " ORDER BY last_message_time DESC LIMIT ? OFFSET ?"
            params.extend([limit, offset])
            
            cursor = await db._connection.execute(query, params)
            rows = await cursor.fetchall()
            
            chats = []
            for row in rows:
                chats.append({
                    "userId": row['user_id'],
                    "username": row['username'],
                    "firstName": row['first_name'],
                    "lastName": row['last_name'],
                    "funnelStage": row['funnel_stage'] or 'new',
                    "interestLevel": row['interest_level'] or 1,
                    "lastInteraction": row['last_interaction'],
                    "lastMessage": row['last_message'],
                    "lastMessageTime": row['last_message_time'],
                    "unreadCount": row['unread_count'] or 0,
                })
            
            # 獲取總數
            count_query = """
                SELECT COUNT(DISTINCT ch.user_id) as total
                FROM chat_history ch
                LEFT JOIN user_profiles up ON ch.user_id = up.user_id
                WHERE 1=1
            """
            count_params = []
            
            if search_query:
                count_query += " AND (up.username LIKE ? OR up.first_name LIKE ? OR ch.content LIKE ?)"
                search_pattern = f"%{search_query}%"
                count_params.extend([search_pattern, search_pattern, search_pattern])
            
            if funnel_stage:
                count_query += " AND up.funnel_stage = ?"
                count_params.append(funnel_stage)
            
            count_cursor = await db._connection.execute(count_query, count_params)
            count_row = await count_cursor.fetchone()
            total = count_row['total'] if count_row else 0
            
            self.send_event("chat-list", {
                "success": True,
                "chats": chats,
                "total": total,
                "limit": limit,
                "offset": offset
            })
        except Exception as e:
            self.send_event("chat-list", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_send_ai_response(self, payload: Dict[str, Any]):
        """發送 AI 生成的回復（用於半自動模式）"""
        try:
            user_id = payload.get('userId', '')
            message = payload.get('message', '')
            account_phone = payload.get('accountPhone')
            source_group = payload.get('sourceGroup')
            username = payload.get('username', '')
            
            if not account_phone:
                # 自動選擇發送帳號
                accounts = await db.get_all_accounts()
                sender_accounts = [a for a in accounts if a.get('role') == 'Sender' and a.get('status') == 'Online']
                if sender_accounts:
                    import random
                    account_phone = random.choice(sender_accounts).get('phone')
                else:
                    raise ValueError("沒有可用的發送帳號")
            
            # 檢查用戶是否已互動（決定是否計入限額）
            has_interacted = await self._user_has_interacted(user_id)
            
            # 檢查帳號限額（未互動用戶）
            if not has_interacted:
                account = await db.get_account_by_phone(account_phone)
                if account:
                    if account.get('dailySendCount', 0) >= account.get('dailySendLimit', 50):
                        raise ValueError(f"帳號 {account_phone} 已達每日發送限額")
            
            # 發送消息
            result = await self.telegram_manager.send_message(
                phone=account_phone,
                user_id=user_id,
                text=message,
                source_group=source_group
            )
            
            if result.get('success'):
                # 保存 AI 回復到聊天歷史
                await db.add_chat_message(
                    user_id=user_id,
                    role='assistant',
                    content=message,
                    account_phone=account_phone,
                    source_group=source_group
                )
                
                # 更新每日計數（僅未互動用戶）
                if not has_interacted:
                    account = await db.get_account_by_phone(account_phone)
                    if account:
                        await db.update_account(account.get('id'), {
                            'dailySendCount': account.get('dailySendCount', 0) + 1
                        })
                
                # 記錄互動
                await db._connection.execute("""
                    INSERT INTO user_interactions 
                    (user_id, interaction_type, direction, content, account_phone, platform)
                    VALUES (?, 'message', 'outbound', ?, ?, 'telegram')
                """, (user_id, message, account_phone))
                await db._connection.commit()
                
                self.send_event("ai-response-sent", {
                    "success": True,
                    "userId": user_id,
                    "message": message,
                    "accountPhone": account_phone
                })
            else:
                raise Exception(result.get('error', '發送失敗'))
                
        except Exception as e:
            self.send_event("ai-response-sent", {
                "success": False,
                "error": str(e)
            })

    # ==================== User Management Handlers ====================
    
    async def handle_get_users_with_profiles(self, payload: Dict[str, Any]):
        """獲取用戶列表（含畫像），支持篩選"""
        try:
            result = await db.get_users_with_profiles(
                stage=payload.get('stage'),
                tags=payload.get('tags'),
                interest_min=payload.get('interestMin'),
                interest_max=payload.get('interestMax'),
                search=payload.get('search'),
                limit=payload.get('limit', 50),
                offset=payload.get('offset', 0)
            )
            
            self.send_event("users-with-profiles", result)
            
        except Exception as e:
            import traceback
            traceback.print_exc(file=sys.stderr)
            self.send_event("users-with-profiles", {
                "users": [],
                "total": 0,
                "error": str(e)
            })
    
    async def handle_get_funnel_stats(self):
        """獲取漏斗統計"""
        try:
            stats = await db.get_detailed_funnel_stats()
            self.send_event("funnel-stats", stats)
            
        except Exception as e:
            import traceback
            traceback.print_exc(file=sys.stderr)
            self.send_event("funnel-stats", {
                "stages": {},
                "tags": [],
                "interest_distribution": {},
                "error": str(e)
            })
    
    async def handle_bulk_update_user_tags(self, payload: Dict[str, Any]):
        """批量更新用戶標籤"""
        try:
            user_ids = payload.get('userIds', [])
            tags = payload.get('tags', '')
            action = payload.get('action', 'add')  # add, remove, set
            
            if not user_ids:
                raise ValueError("請選擇要更新的用戶")
            
            await db.bulk_update_user_tags(user_ids, tags, action)
            
            self.send_log(f"已更新 {len(user_ids)} 個用戶的標籤", "success")
            self.send_event("bulk-update-complete", {
                "success": True,
                "type": "tags",
                "count": len(user_ids)
            })
            
        except Exception as e:
            self.send_log(f"批量更新標籤失敗: {str(e)}", "error")
            self.send_event("bulk-update-complete", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_bulk_update_user_stage(self, payload: Dict[str, Any]):
        """批量更新用戶階段"""
        try:
            user_ids = payload.get('userIds', [])
            stage = payload.get('stage', '')
            
            if not user_ids:
                raise ValueError("請選擇要更新的用戶")
            if not stage:
                raise ValueError("請選擇目標階段")
            
            await db.bulk_update_user_stage(user_ids, stage)
            
            self.send_log(f"已將 {len(user_ids)} 個用戶更新為 {stage} 階段", "success")
            self.send_event("bulk-update-complete", {
                "success": True,
                "type": "stage",
                "count": len(user_ids)
            })
            
        except Exception as e:
            self.send_log(f"批量更新階段失敗: {str(e)}", "error")
            self.send_event("bulk-update-complete", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_update_user_profile(self, payload: Dict[str, Any]):
        """更新單個用戶畫像"""
        try:
            user_id = payload.get('userId')
            data = payload.get('data', {})
            
            if not user_id:
                raise ValueError("用戶 ID 不能為空")
            
            await db.update_user_profile(user_id, data)
            
            # 如果有階段更新，使用專門的方法
            if 'funnel_stage' in data:
                await db.set_user_funnel_stage(user_id, data['funnel_stage'])
            
            self.send_log(f"已更新用戶 {user_id} 的畫像", "success")
            self.send_event("user-profile-updated", {
                "success": True,
                "userId": user_id
            })
            
        except Exception as e:
            self.send_log(f"更新用戶畫像失敗: {str(e)}", "error")
            self.send_event("user-profile-updated", {
                "success": False,
                "error": str(e)
            })
    
    # ==================== Batch Operations Handlers ====================
    
    async def handle_batch_update_lead_status(self, payload: Dict[str, Any]):
        """批量更新 Lead 狀態"""
        try:
            lead_ids = payload.get('leadIds', [])
            new_status = payload.get('newStatus')
            
            if not lead_ids:
                self.send_event("batch-operation-result", {
                    "success": False,
                    "error": "未選擇任何 Lead"
                })
                return
            
            if not new_status:
                self.send_event("batch-operation-result", {
                    "success": False,
                    "error": "未指定新狀態"
                })
                return
            
            batch_ops = get_batch_ops()
            if not batch_ops:
                self.send_event("batch-operation-result", {
                    "success": False,
                    "error": "批量操作系統未初始化"
                })
                return
            
            result = await batch_ops.batch_update_lead_status(lead_ids, new_status)
            
            if result.get('success'):
                self.send_log(f"批量更新狀態完成: {result.get('successCount')}/{len(lead_ids)} 成功", "success")
            else:
                self.send_log(f"批量更新狀態失敗: {result.get('error')}", "error")
            
            self.send_event("batch-operation-result", result)
            
        except Exception as e:
            self.send_log(f"批量更新狀態失敗: {str(e)}", "error")
            self.send_event("batch-operation-result", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_batch_add_tag(self, payload: Dict[str, Any]):
        """批量添加標籤"""
        try:
            lead_ids = payload.get('leadIds', [])
            tag = payload.get('tag')
            
            if not lead_ids:
                self.send_event("batch-operation-result", {
                    "success": False,
                    "error": "未選擇任何 Lead"
                })
                return
            
            if not tag:
                self.send_event("batch-operation-result", {
                    "success": False,
                    "error": "未指定標籤"
                })
                return
            
            batch_ops = get_batch_ops()
            if not batch_ops:
                self.send_event("batch-operation-result", {
                    "success": False,
                    "error": "批量操作系統未初始化"
                })
                return
            
            result = await batch_ops.batch_add_tag(lead_ids, tag)
            
            if result.get('success'):
                self.send_log(f"批量添加標籤完成: {result.get('successCount')}/{len(lead_ids)} 成功", "success")
            else:
                self.send_log(f"批量添加標籤失敗: {result.get('error')}", "error")
            
            self.send_event("batch-operation-result", result)
            
        except Exception as e:
            self.send_log(f"批量添加標籤失敗: {str(e)}", "error")
            self.send_event("batch-operation-result", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_batch_remove_tag(self, payload: Dict[str, Any]):
        """批量移除標籤"""
        try:
            lead_ids = payload.get('leadIds', [])
            tag = payload.get('tag')
            
            if not lead_ids:
                self.send_event("batch-operation-result", {
                    "success": False,
                    "error": "未選擇任何 Lead"
                })
                return
            
            if not tag:
                self.send_event("batch-operation-result", {
                    "success": False,
                    "error": "未指定標籤"
                })
                return
            
            batch_ops = get_batch_ops()
            if not batch_ops:
                self.send_event("batch-operation-result", {
                    "success": False,
                    "error": "批量操作系統未初始化"
                })
                return
            
            result = await batch_ops.batch_remove_tag(lead_ids, tag)
            
            if result.get('success'):
                self.send_log(f"批量移除標籤完成: {result.get('successCount')}/{len(lead_ids)} 成功", "success")
            else:
                self.send_log(f"批量移除標籤失敗: {result.get('error')}", "error")
            
            self.send_event("batch-operation-result", result)
            
        except Exception as e:
            self.send_log(f"批量移除標籤失敗: {str(e)}", "error")
            self.send_event("batch-operation-result", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_batch_add_to_dnc(self, payload: Dict[str, Any]):
        """批量添加到 DNC 列表"""
        try:
            lead_ids = payload.get('leadIds', [])
            
            if not lead_ids:
                self.send_event("batch-operation-result", {
                    "success": False,
                    "error": "未選擇任何 Lead"
                })
                return
            
            batch_ops = get_batch_ops()
            if not batch_ops:
                self.send_event("batch-operation-result", {
                    "success": False,
                    "error": "批量操作系統未初始化"
                })
                return
            
            result = await batch_ops.batch_add_to_dnc(lead_ids)
            
            if result.get('success'):
                self.send_log(f"批量添加到 DNC 完成: {result.get('successCount')}/{len(lead_ids)} 成功", "success")
            else:
                self.send_log(f"批量添加到 DNC 失敗: {result.get('error')}", "error")
            
            self.send_event("batch-operation-result", result)
            
        except Exception as e:
            self.send_log(f"批量添加到 DNC 失敗: {str(e)}", "error")
            self.send_event("batch-operation-result", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_batch_remove_from_dnc(self, payload: Dict[str, Any]):
        """批量從 DNC 列表移除"""
        try:
            lead_ids = payload.get('leadIds', [])
            
            if not lead_ids:
                self.send_event("batch-operation-result", {
                    "success": False,
                    "error": "未選擇任何 Lead"
                })
                return
            
            batch_ops = get_batch_ops()
            if not batch_ops:
                self.send_event("batch-operation-result", {
                    "success": False,
                    "error": "批量操作系統未初始化"
                })
                return
            
            result = await batch_ops.batch_remove_from_dnc(lead_ids)
            
            if result.get('success'):
                self.send_log(f"批量從 DNC 移除完成: {result.get('successCount')}/{len(lead_ids)} 成功", "success")
            else:
                self.send_log(f"批量從 DNC 移除失敗: {result.get('error')}", "error")
            
            self.send_event("batch-operation-result", result)
            
        except Exception as e:
            self.send_log(f"批量從 DNC 移除失敗: {str(e)}", "error")
            self.send_event("batch-operation-result", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_batch_update_funnel_stage(self, payload: Dict[str, Any]):
        """批量更新漏斗階段"""
        try:
            lead_ids = payload.get('leadIds', [])
            new_stage = payload.get('newStage')
            
            if not lead_ids:
                self.send_event("batch-operation-result", {
                    "success": False,
                    "error": "未選擇任何 Lead"
                })
                return
            
            if not new_stage:
                self.send_event("batch-operation-result", {
                    "success": False,
                    "error": "未指定新階段"
                })
                return
            
            batch_ops = get_batch_ops()
            if not batch_ops:
                self.send_event("batch-operation-result", {
                    "success": False,
                    "error": "批量操作系統未初始化"
                })
                return
            
            result = await batch_ops.batch_update_funnel_stage(lead_ids, new_stage)
            
            if result.get('success'):
                self.send_log(f"批量更新漏斗階段完成: {result.get('successCount')}/{len(lead_ids)} 成功", "success")
            else:
                self.send_log(f"批量更新漏斗階段失敗: {result.get('error')}", "error")
            
            self.send_event("batch-operation-result", result)
            
        except Exception as e:
            self.send_log(f"批量更新漏斗階段失敗: {str(e)}", "error")
            self.send_event("batch-operation-result", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_batch_delete_leads(self, payload: Dict[str, Any]):
        """批量刪除 Lead"""
        try:
            lead_ids = payload.get('leadIds', [])
            
            if not lead_ids:
                self.send_event("batch-operation-result", {
                    "success": False,
                    "error": "未選擇任何 Lead"
                })
                return
            
            batch_ops = get_batch_ops()
            if not batch_ops:
                self.send_event("batch-operation-result", {
                    "success": False,
                    "error": "批量操作系統未初始化"
                })
                return
            
            result = await batch_ops.batch_delete_leads(lead_ids)
            
            if result.get('success'):
                self.send_log(f"批量刪除完成: {result.get('successCount')}/{len(lead_ids)} 成功", "success")
            else:
                self.send_log(f"批量刪除失敗: {result.get('error')}", "error")
            
            self.send_event("batch-operation-result", result)
            
        except Exception as e:
            self.send_log(f"批量刪除失敗: {str(e)}", "error")
            self.send_event("batch-operation-result", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_undo_batch_operation(self, payload: Dict[str, Any]):
        """撤銷批量操作"""
        try:
            operation_id = payload.get('operationId')
            
            if not operation_id:
                self.send_event("batch-undo-result", {
                    "success": False,
                    "error": "未指定操作 ID"
                })
                return
            
            batch_ops = get_batch_ops()
            if not batch_ops:
                self.send_event("batch-undo-result", {
                    "success": False,
                    "error": "批量操作系統未初始化"
                })
                return
            
            result = await batch_ops.undo_operation(operation_id)
            
            if result.get('success'):
                self.send_log(f"撤銷操作成功: {operation_id}", "success")
            else:
                self.send_log(f"撤銷操作失敗: {result.get('error')}", "error")
            
            self.send_event("batch-undo-result", result)
            
        except Exception as e:
            self.send_log(f"撤銷操作失敗: {str(e)}", "error")
            self.send_event("batch-undo-result", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_get_batch_operation_history(self, payload: Dict[str, Any]):
        """獲取批量操作歷史"""
        try:
            limit = payload.get('limit', 50)
            offset = payload.get('offset', 0)
            operation_type = payload.get('operationType')
            
            batch_ops = get_batch_ops()
            if not batch_ops:
                self.send_event("batch-operation-history", {
                    "success": False,
                    "error": "批量操作系統未初始化"
                })
                return
            
            result = await batch_ops.get_operation_history(limit, offset, operation_type)
            
            self.send_event("batch-operation-history", result)
            
        except Exception as e:
            self.send_log(f"獲取操作歷史失敗: {str(e)}", "error")
            self.send_event("batch-operation-history", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_get_all_tags(self):
        """獲取所有標籤"""
        try:
            batch_ops = get_batch_ops()
            if not batch_ops:
                self.send_event("all-tags", {
                    "success": False,
                    "error": "批量操作系統未初始化"
                })
                return
            
            result = await batch_ops.get_all_tags()
            
            self.send_event("all-tags", result)
            
        except Exception as e:
            self.send_log(f"獲取標籤列表失敗: {str(e)}", "error")
            self.send_event("all-tags", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_create_tag(self, payload: Dict[str, Any]):
        """創建新標籤"""
        try:
            name = payload.get('name')
            color = payload.get('color', '#3B82F6')
            
            if not name:
                self.send_event("tag-created", {
                    "success": False,
                    "error": "標籤名稱不能為空"
                })
                return
            
            batch_ops = get_batch_ops()
            if not batch_ops:
                self.send_event("tag-created", {
                    "success": False,
                    "error": "批量操作系統未初始化"
                })
                return
            
            result = await batch_ops.create_tag(name, color)
            
            if result.get('success'):
                self.send_log(f"創建標籤成功: {name}", "success")
            else:
                self.send_log(f"創建標籤失敗: {result.get('error')}", "error")
            
            self.send_event("tag-created", result)
            
        except Exception as e:
            self.send_log(f"創建標籤失敗: {str(e)}", "error")
            self.send_event("tag-created", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_delete_tag(self, payload: Dict[str, Any]):
        """刪除標籤"""
        try:
            name = payload.get('name')
            
            if not name:
                self.send_event("tag-deleted", {
                    "success": False,
                    "error": "標籤名稱不能為空"
                })
                return
            
            batch_ops = get_batch_ops()
            if not batch_ops:
                self.send_event("tag-deleted", {
                    "success": False,
                    "error": "批量操作系統未初始化"
                })
                return
            
            result = await batch_ops.delete_tag(name)
            
            if result.get('success'):
                self.send_log(f"刪除標籤成功: {name}", "success")
            else:
                self.send_log(f"刪除標籤失敗: {result.get('error')}", "error")
            
            self.send_event("tag-deleted", result)
            
        except Exception as e:
            self.send_log(f"刪除標籤失敗: {str(e)}", "error")
            self.send_event("tag-deleted", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_get_lead_tags(self, payload: Dict[str, Any]):
        """獲取 Lead 的標籤"""
        try:
            lead_id = payload.get('leadId')
            
            if not lead_id:
                self.send_event("lead-tags", {
                    "success": False,
                    "error": "Lead ID 不能為空"
                })
                return
            
            batch_ops = get_batch_ops()
            if not batch_ops:
                self.send_event("lead-tags", {
                    "success": False,
                    "error": "批量操作系統未初始化"
                })
                return
            
            result = await batch_ops.get_lead_tags(lead_id)
            result['leadId'] = lead_id
            
            self.send_event("lead-tags", result)
            
        except Exception as e:
            self.send_log(f"獲取 Lead 標籤失敗: {str(e)}", "error")
            self.send_event("lead-tags", {
                "success": False,
                "error": str(e)
            })
    
    # ==================== Ad System Handlers (廣告發送系統) ====================
    
    async def handle_create_ad_template(self, payload: Dict[str, Any]):
        """創建廣告模板"""
        try:
            template_manager = get_ad_template_manager()
            if not template_manager:
                self.send_event("ad-template-created", {"success": False, "error": "廣告系統未初始化"})
                return
            
            result = await template_manager.create_template(
                name=payload.get('name', ''),
                content=payload.get('content', ''),
                media_type=payload.get('mediaType', 'text'),
                media_file_id=payload.get('mediaFileId'),
                media_path=payload.get('mediaPath')
            )
            
            if result.get('success'):
                self.send_log(f"廣告模板已創建: {result.get('name')}", "success")
            
            self.send_event("ad-template-created", result)
            
        except Exception as e:
            self.send_event("ad-template-created", {"success": False, "error": str(e)})
    
    async def handle_update_ad_template(self, payload: Dict[str, Any]):
        """更新廣告模板"""
        try:
            template_manager = get_ad_template_manager()
            if not template_manager:
                self.send_event("ad-template-updated", {"success": False, "error": "廣告系統未初始化"})
                return
            
            template_id = payload.get('templateId')
            updates = payload.get('updates', {})
            
            result = await template_manager.update_template(template_id, updates)
            
            self.send_event("ad-template-updated", result)
            
        except Exception as e:
            self.send_event("ad-template-updated", {"success": False, "error": str(e)})
    
    async def handle_delete_ad_template(self, payload: Dict[str, Any]):
        """刪除廣告模板"""
        try:
            template_manager = get_ad_template_manager()
            if not template_manager:
                self.send_event("ad-template-deleted", {"success": False, "error": "廣告系統未初始化"})
                return
            
            template_id = payload.get('templateId')
            result = await template_manager.delete_template(template_id)
            
            if result.get('success'):
                self.send_log(f"廣告模板已刪除: ID {template_id}", "info")
            
            self.send_event("ad-template-deleted", result)
            
        except Exception as e:
            self.send_event("ad-template-deleted", {"success": False, "error": str(e)})
    
    async def handle_get_ad_templates(self, payload: Dict[str, Any]):
        """獲取廣告模板列表"""
        try:
            template_manager = get_ad_template_manager()
            if not template_manager:
                self.send_event("ad-templates", {"success": False, "error": "廣告系統未初始化"})
                return
            
            active_only = payload.get('activeOnly', False)
            templates = await template_manager.get_all_templates(active_only)
            
            self.send_event("ad-templates", {
                "success": True,
                "templates": [t.to_dict() for t in templates]
            })
            
        except Exception as e:
            self.send_event("ad-templates", {"success": False, "error": str(e)})
    
    async def handle_toggle_ad_template_status(self, payload: Dict[str, Any]):
        """切換廣告模板狀態"""
        try:
            template_manager = get_ad_template_manager()
            if not template_manager:
                self.send_event("ad-template-toggled", {"success": False, "error": "廣告系統未初始化"})
                return
            
            template_id = payload.get('templateId')
            result = await template_manager.toggle_template_status(template_id)
            
            self.send_event("ad-template-toggled", result)
            
        except Exception as e:
            self.send_event("ad-template-toggled", {"success": False, "error": str(e)})
    
    async def handle_preview_ad_template(self, payload: Dict[str, Any]):
        """預覽廣告模板變體"""
        try:
            template_manager = get_ad_template_manager()
            if not template_manager:
                self.send_event("ad-template-preview", {"success": False, "error": "廣告系統未初始化"})
                return
            
            template_id = payload.get('templateId')
            count = payload.get('count', 5)
            
            result = await template_manager.preview_template(template_id, count)
            
            self.send_event("ad-template-preview", result)
            
        except Exception as e:
            self.send_event("ad-template-preview", {"success": False, "error": str(e)})
    
    async def handle_validate_spintax(self, payload: Dict[str, Any]):
        """驗證 Spintax 語法"""
        try:
            content = payload.get('content', '')
            result = SpintaxGenerator.preview_variants(content, 5)
            result['success'] = result.get('valid', False)
            
            self.send_event("spintax-validated", result)
            
        except Exception as e:
            self.send_event("spintax-validated", {"success": False, "error": str(e)})
    
    async def handle_create_ad_schedule(self, payload: Dict[str, Any]):
        """創建廣告計劃"""
        try:
            ad_manager = get_ad_manager()
            if not ad_manager:
                self.send_event("ad-schedule-created", {"success": False, "error": "廣告系統未初始化"})
                return
            
            result = await ad_manager.create_schedule(
                template_id=payload.get('templateId'),
                name=payload.get('name', ''),
                target_groups=payload.get('targetGroups', []),
                send_mode=payload.get('sendMode', 'scheduled'),
                schedule_type=payload.get('scheduleType', 'once'),
                assigned_accounts=payload.get('assignedAccounts', []),
                schedule_time=payload.get('scheduleTime'),
                interval_minutes=payload.get('intervalMinutes', 60),
                trigger_keywords=payload.get('triggerKeywords', []),
                account_strategy=payload.get('accountStrategy', 'single')
            )
            
            if result.get('success'):
                self.send_log(f"廣告計劃已創建: {result.get('name')}", "success")
                # Reload triggers if needed
                scheduler = get_ad_scheduler()
                if scheduler:
                    await scheduler.reload_triggers()
            
            self.send_event("ad-schedule-created", result)
            
        except Exception as e:
            self.send_event("ad-schedule-created", {"success": False, "error": str(e)})
    
    async def handle_update_ad_schedule(self, payload: Dict[str, Any]):
        """更新廣告計劃"""
        try:
            ad_manager = get_ad_manager()
            if not ad_manager:
                self.send_event("ad-schedule-updated", {"success": False, "error": "廣告系統未初始化"})
                return
            
            schedule_id = payload.get('scheduleId')
            updates = payload.get('updates', {})
            
            result = await ad_manager.update_schedule(schedule_id, updates)
            
            if result.get('success'):
                scheduler = get_ad_scheduler()
                if scheduler:
                    await scheduler.reload_triggers()
            
            self.send_event("ad-schedule-updated", result)
            
        except Exception as e:
            self.send_event("ad-schedule-updated", {"success": False, "error": str(e)})
    
    async def handle_delete_ad_schedule(self, payload: Dict[str, Any]):
        """刪除廣告計劃"""
        try:
            ad_manager = get_ad_manager()
            if not ad_manager:
                self.send_event("ad-schedule-deleted", {"success": False, "error": "廣告系統未初始化"})
                return
            
            schedule_id = payload.get('scheduleId')
            result = await ad_manager.delete_schedule(schedule_id)
            
            if result.get('success'):
                self.send_log(f"廣告計劃已刪除: ID {schedule_id}", "info")
                scheduler = get_ad_scheduler()
                if scheduler:
                    await scheduler.reload_triggers()
            
            self.send_event("ad-schedule-deleted", result)
            
        except Exception as e:
            self.send_event("ad-schedule-deleted", {"success": False, "error": str(e)})
    
    async def handle_get_ad_schedules(self, payload: Dict[str, Any]):
        """獲取廣告計劃列表"""
        try:
            ad_manager = get_ad_manager()
            if not ad_manager:
                self.send_event("ad-schedules", {"success": False, "error": "廣告系統未初始化"})
                return
            
            active_only = payload.get('activeOnly', False)
            schedules = await ad_manager.get_all_schedules(active_only)
            
            self.send_event("ad-schedules", {
                "success": True,
                "schedules": [s.to_dict() for s in schedules]
            })
            
        except Exception as e:
            self.send_event("ad-schedules", {"success": False, "error": str(e)})
    
    async def handle_toggle_ad_schedule_status(self, payload: Dict[str, Any]):
        """切換廣告計劃狀態"""
        try:
            ad_manager = get_ad_manager()
            if not ad_manager:
                self.send_event("ad-schedule-toggled", {"success": False, "error": "廣告系統未初始化"})
                return
            
            schedule_id = payload.get('scheduleId')
            result = await ad_manager.toggle_schedule_status(schedule_id)
            
            if result.get('success'):
                scheduler = get_ad_scheduler()
                if scheduler:
                    await scheduler.reload_triggers()
            
            self.send_event("ad-schedule-toggled", result)
            
        except Exception as e:
            self.send_event("ad-schedule-toggled", {"success": False, "error": str(e)})
    
    async def handle_run_ad_schedule_now(self, payload: Dict[str, Any]):
        """立即執行廣告計劃"""
        try:
            scheduler = get_ad_scheduler()
            if not scheduler:
                self.send_event("ad-schedule-run-result", {"success": False, "error": "廣告排程器未初始化"})
                return
            
            schedule_id = payload.get('scheduleId')
            result = await scheduler.run_schedule_now(schedule_id)
            
            self.send_event("ad-schedule-run-result", result)
            
        except Exception as e:
            self.send_event("ad-schedule-run-result", {"success": False, "error": str(e)})
    
    async def handle_send_ad_now(self, payload: Dict[str, Any]):
        """立即發送廣告"""
        try:
            broadcaster = get_ad_broadcaster()
            if not broadcaster:
                self.send_event("ad-send-result", {"success": False, "error": "廣告發送器未初始化"})
                return
            
            result = await broadcaster.send_now(
                template_id=payload.get('templateId'),
                target_groups=payload.get('targetGroups', []),
                account_phones=payload.get('accountPhones', []),
                account_strategy=payload.get('accountStrategy', 'rotate')
            )
            
            self.send_event("ad-send-result", result)
            
        except Exception as e:
            self.send_event("ad-send-result", {"success": False, "error": str(e)})
    
    async def handle_get_ad_send_logs(self, payload: Dict[str, Any]):
        """獲取廣告發送記錄"""
        try:
            ad_manager = get_ad_manager()
            if not ad_manager:
                self.send_event("ad-send-logs", {"success": False, "error": "廣告系統未初始化"})
                return
            
            result = await ad_manager.get_send_logs(
                limit=payload.get('limit', 100),
                offset=payload.get('offset', 0),
                template_id=payload.get('templateId'),
                schedule_id=payload.get('scheduleId'),
                status=payload.get('status')
            )
            
            self.send_event("ad-send-logs", result)
            
        except Exception as e:
            self.send_event("ad-send-logs", {"success": False, "error": str(e)})
    
    async def handle_get_ad_overview_stats(self, payload: Dict[str, Any]):
        """獲取廣告總覽統計"""
        try:
            analytics = get_ad_analytics()
            if not analytics:
                self.send_event("ad-overview-stats", {"success": False, "error": "廣告分析未初始化"})
                return
            
            days = payload.get('days', 7)
            result = await analytics.get_overview_stats(days)
            
            self.send_event("ad-overview-stats", result)
            
        except Exception as e:
            self.send_event("ad-overview-stats", {"success": False, "error": str(e)})
    
    async def handle_get_ad_template_stats(self, payload: Dict[str, Any]):
        """獲取模板統計"""
        try:
            analytics = get_ad_analytics()
            if not analytics:
                self.send_event("ad-template-stats", {"success": False, "error": "廣告分析未初始化"})
                return
            
            template_id = payload.get('templateId')
            result = await analytics.get_template_stats(template_id)
            
            self.send_event("ad-template-stats", result)
            
        except Exception as e:
            self.send_event("ad-template-stats", {"success": False, "error": str(e)})
    
    async def handle_get_ad_schedule_stats(self, payload: Dict[str, Any]):
        """獲取計劃統計"""
        try:
            analytics = get_ad_analytics()
            if not analytics:
                self.send_event("ad-schedule-stats", {"success": False, "error": "廣告分析未初始化"})
                return
            
            schedule_id = payload.get('scheduleId')
            result = await analytics.get_schedule_stats(schedule_id)
            
            self.send_event("ad-schedule-stats", result)
            
        except Exception as e:
            self.send_event("ad-schedule-stats", {"success": False, "error": str(e)})
    
    async def handle_get_ad_account_stats(self, payload: Dict[str, Any]):
        """獲取帳號統計"""
        try:
            analytics = get_ad_analytics()
            if not analytics:
                self.send_event("ad-account-stats", {"success": False, "error": "廣告分析未初始化"})
                return
            
            days = payload.get('days', 7)
            result = await analytics.get_account_stats(days)
            
            self.send_event("ad-account-stats", result)
            
        except Exception as e:
            self.send_event("ad-account-stats", {"success": False, "error": str(e)})
    
    async def handle_get_ad_group_stats(self, payload: Dict[str, Any]):
        """獲取群組統計"""
        try:
            analytics = get_ad_analytics()
            if not analytics:
                self.send_event("ad-group-stats", {"success": False, "error": "廣告分析未初始化"})
                return
            
            days = payload.get('days', 7)
            result = await analytics.get_group_stats(days)
            
            self.send_event("ad-group-stats", result)
            
        except Exception as e:
            self.send_event("ad-group-stats", {"success": False, "error": str(e)})
    
    async def handle_get_ad_daily_stats(self, payload: Dict[str, Any]):
        """獲取每日統計"""
        try:
            analytics = get_ad_analytics()
            if not analytics:
                self.send_event("ad-daily-stats", {"success": False, "error": "廣告分析未初始化"})
                return
            
            days = payload.get('days', 30)
            result = await analytics.get_daily_stats(days)
            
            self.send_event("ad-daily-stats", result)
            
        except Exception as e:
            self.send_event("ad-daily-stats", {"success": False, "error": str(e)})
    
    # ==================== User Tracking Handlers (用戶追蹤系統) ====================
    
    async def handle_add_user_to_track(self, payload: Dict[str, Any]):
        """添加用戶到追蹤列表"""
        try:
            tracker = get_user_tracker()
            if not tracker:
                self.send_event("user-added-to-track", {"success": False, "error": "用戶追蹤系統未初始化"})
                return
            
            result = await tracker.add_user_to_track(
                user_id=payload.get('userId'),
                username=payload.get('username'),
                first_name=payload.get('firstName'),
                last_name=payload.get('lastName'),
                source=payload.get('source', 'manual'),
                source_group_id=payload.get('sourceGroupId'),
                notes=payload.get('notes')
            )
            
            if result.get('success'):
                self.send_log(f"用戶已添加到追蹤列表", "success")
            
            self.send_event("user-added-to-track", result)
            
        except Exception as e:
            self.send_event("user-added-to-track", {"success": False, "error": str(e)})
    
    async def handle_add_user_from_lead(self, payload: Dict[str, Any]):
        """從 Lead 添加用戶到追蹤"""
        try:
            tracker = get_user_tracker()
            if not tracker:
                self.send_event("user-added-from-lead", {"success": False, "error": "用戶追蹤系統未初始化"})
                return
            
            lead_id = payload.get('leadId')
            result = await tracker.add_user_from_lead(lead_id)
            
            self.send_event("user-added-from-lead", result)
            
        except Exception as e:
            self.send_event("user-added-from-lead", {"success": False, "error": str(e)})
    
    async def handle_remove_tracked_user(self, payload: Dict[str, Any]):
        """移除追蹤用戶"""
        try:
            tracker = get_user_tracker()
            if not tracker:
                self.send_event("user-removed", {"success": False, "error": "用戶追蹤系統未初始化"})
                return
            
            user_id = payload.get('userId')
            result = await tracker.remove_user(user_id)
            
            if result.get('success'):
                self.send_log(f"用戶已從追蹤列表移除", "info")
            
            self.send_event("user-removed", result)
            
        except Exception as e:
            self.send_event("user-removed", {"success": False, "error": str(e)})
    
    async def handle_get_tracked_users(self, payload: Dict[str, Any]):
        """獲取追蹤用戶列表"""
        try:
            tracker = get_user_tracker()
            if not tracker:
                self.send_event("tracked-users", {"success": False, "error": "用戶追蹤系統未初始化"})
                return
            
            result = await tracker.get_all_tracked_users(
                value_level=payload.get('valueLevel'),
                status=payload.get('status'),
                limit=payload.get('limit', 100),
                offset=payload.get('offset', 0)
            )
            
            self.send_event("tracked-users", result)
            
        except Exception as e:
            self.send_event("tracked-users", {"success": False, "error": str(e)})
    
    async def handle_update_user_value_level(self, payload: Dict[str, Any]):
        """更新用戶價值等級"""
        try:
            tracker = get_user_tracker()
            if not tracker:
                self.send_event("user-value-updated", {"success": False, "error": "用戶追蹤系統未初始化"})
                return
            
            result = await tracker.update_user_value_level(
                user_id=payload.get('userId'),
                value_level=payload.get('valueLevel')
            )
            
            self.send_event("user-value-updated", result)
            
        except Exception as e:
            self.send_event("user-value-updated", {"success": False, "error": str(e)})
    
    async def handle_track_user_groups(self, payload: Dict[str, Any]):
        """追蹤用戶群組"""
        try:
            tracker = get_user_tracker()
            if not tracker:
                self.send_event("user-tracking-completed", {"success": False, "error": "用戶追蹤系統未初始化"})
                return
            
            result = await tracker.track_user_groups(
                user_id=payload.get('userId'),
                account_phone=payload.get('accountPhone')
            )
            
            # Event is sent by tracker
            
        except Exception as e:
            self.send_event("user-tracking-failed", {"success": False, "error": str(e)})
    
    async def handle_batch_track_users(self, payload: Dict[str, Any]):
        """批量追蹤用戶"""
        try:
            tracker = get_user_tracker()
            if not tracker:
                self.send_event("batch-tracking-completed", {"success": False, "error": "用戶追蹤系統未初始化"})
                return
            
            result = await tracker.batch_track_users(
                user_ids=payload.get('userIds', []),
                account_phone=payload.get('accountPhone')
            )
            
            # Event is sent by tracker
            
        except Exception as e:
            self.send_event("batch-tracking-completed", {"success": False, "error": str(e)})
    
    async def handle_get_user_groups(self, payload: Dict[str, Any]):
        """獲取用戶群組"""
        try:
            tracker = get_user_tracker()
            if not tracker:
                self.send_event("user-groups", {"success": False, "error": "用戶追蹤系統未初始化"})
                return
            
            result = await tracker.get_user_groups(payload.get('userId'))
            
            self.send_event("user-groups", result)
            
        except Exception as e:
            self.send_event("user-groups", {"success": False, "error": str(e)})
    
    async def handle_get_high_value_groups(self, payload: Dict[str, Any]):
        """獲取高價值群組"""
        try:
            tracker = get_user_tracker()
            if not tracker:
                self.send_event("high-value-groups", {"success": False, "error": "用戶追蹤系統未初始化"})
                return
            
            result = await tracker.get_high_value_groups(
                limit=payload.get('limit', 50)
            )
            
            self.send_event("high-value-groups", result)
            
        except Exception as e:
            self.send_event("high-value-groups", {"success": False, "error": str(e)})
    
    async def handle_get_tracking_stats(self, payload: Dict[str, Any]):
        """獲取追蹤統計"""
        try:
            tracker = get_user_tracker()
            if not tracker:
                self.send_event("tracking-stats", {"success": False, "error": "用戶追蹤系統未初始化"})
                return
            
            result = await tracker.get_tracking_stats()
            
            self.send_event("tracking-stats", result)
            
        except Exception as e:
            self.send_event("tracking-stats", {"success": False, "error": str(e)})
    
    async def handle_get_tracking_logs(self, payload: Dict[str, Any]):
        """獲取追蹤日誌"""
        try:
            tracker = get_user_tracker()
            if not tracker:
                self.send_event("tracking-logs", {"success": False, "error": "用戶追蹤系統未初始化"})
                return
            
            result = await tracker.get_tracking_logs(
                user_id=payload.get('userId'),
                limit=payload.get('limit', 50),
                offset=payload.get('offset', 0)
            )
            
            self.send_event("tracking-logs", result)
            
        except Exception as e:
            self.send_event("tracking-logs", {"success": False, "error": str(e)})
    
    async def handle_get_user_value_distribution(self, payload: Dict[str, Any]):
        """獲取用戶價值分佈"""
        try:
            analytics = get_user_analytics()
            if not analytics:
                self.send_event("user-value-distribution", {"success": False, "error": "用戶分析未初始化"})
                return
            
            result = await analytics.get_user_value_distribution()
            
            self.send_event("user-value-distribution", result)
            
        except Exception as e:
            self.send_event("user-value-distribution", {"success": False, "error": str(e)})
    
    async def handle_get_group_overlap_analysis(self, payload: Dict[str, Any]):
        """獲取群組重疊分析"""
        try:
            analytics = get_user_analytics()
            if not analytics:
                self.send_event("group-overlap-analysis", {"success": False, "error": "用戶分析未初始化"})
                return
            
            result = await analytics.get_group_overlap_analysis(
                min_overlap=payload.get('minOverlap', 2)
            )
            
            self.send_event("group-overlap-analysis", result)
            
        except Exception as e:
            self.send_event("group-overlap-analysis", {"success": False, "error": str(e)})
    
    async def handle_get_tracking_effectiveness(self, payload: Dict[str, Any]):
        """獲取追蹤效率"""
        try:
            analytics = get_user_analytics()
            if not analytics:
                self.send_event("tracking-effectiveness", {"success": False, "error": "用戶分析未初始化"})
                return
            
            result = await analytics.get_tracking_effectiveness()
            
            self.send_event("tracking-effectiveness", result)
            
        except Exception as e:
            self.send_event("tracking-effectiveness", {"success": False, "error": str(e)})
    
    # ==================== Campaign Handlers (營銷活動協調器) ====================
    
    async def handle_create_campaign(self, payload: Dict[str, Any]):
        """創建營銷活動"""
        try:
            orchestrator = get_campaign_orchestrator()
            if not orchestrator:
                self.send_event("campaign-created", {"success": False, "error": "營銷活動系統未初始化"})
                return
            
            result = await orchestrator.create_campaign(
                name=payload.get('name', ''),
                description=payload.get('description', ''),
                phases=payload.get('phases'),
                target_groups=payload.get('targetGroups', []),
                assigned_accounts=payload.get('assignedAccounts', []),
                keywords=payload.get('keywords', []),
                ad_template_id=payload.get('adTemplateId'),
                settings=payload.get('settings', {})
            )
            
            if result.get('success'):
                self.send_log(f"營銷活動已創建: {result.get('name')}", "success")
            
            self.send_event("campaign-created", result)
            
        except Exception as e:
            self.send_event("campaign-created", {"success": False, "error": str(e)})
    
    async def handle_update_campaign(self, payload: Dict[str, Any]):
        """更新營銷活動"""
        try:
            orchestrator = get_campaign_orchestrator()
            if not orchestrator:
                self.send_event("campaign-updated", {"success": False, "error": "營銷活動系統未初始化"})
                return
            
            result = await orchestrator.update_campaign(
                campaign_id=payload.get('campaignId'),
                updates=payload.get('updates', {})
            )
            
            self.send_event("campaign-updated", result)
            
        except Exception as e:
            self.send_event("campaign-updated", {"success": False, "error": str(e)})
    
    async def handle_delete_campaign(self, payload: Dict[str, Any]):
        """刪除營銷活動"""
        try:
            orchestrator = get_campaign_orchestrator()
            if not orchestrator:
                self.send_event("campaign-deleted", {"success": False, "error": "營銷活動系統未初始化"})
                return
            
            result = await orchestrator.delete_campaign(payload.get('campaignId'))
            
            if result.get('success'):
                self.send_log("營銷活動已刪除", "info")
            
            self.send_event("campaign-deleted", result)
            
        except Exception as e:
            self.send_event("campaign-deleted", {"success": False, "error": str(e)})
    
    async def handle_get_campaigns(self, payload: Dict[str, Any]):
        """獲取營銷活動列表"""
        try:
            orchestrator = get_campaign_orchestrator()
            if not orchestrator:
                self.send_event("campaigns", {"success": False, "error": "營銷活動系統未初始化"})
                return
            
            result = await orchestrator.get_all_campaigns(
                status=payload.get('status'),
                limit=payload.get('limit', 50),
                offset=payload.get('offset', 0)
            )
            
            self.send_event("campaigns", result)
            
        except Exception as e:
            self.send_event("campaigns", {"success": False, "error": str(e)})
    
    async def handle_get_campaign(self, payload: Dict[str, Any]):
        """獲取單個營銷活動"""
        try:
            orchestrator = get_campaign_orchestrator()
            if not orchestrator:
                self.send_event("campaign", {"success": False, "error": "營銷活動系統未初始化"})
                return
            
            campaign = await orchestrator.get_campaign(payload.get('campaignId'))
            
            if campaign:
                self.send_event("campaign", {"success": True, "campaign": campaign.to_dict()})
            else:
                self.send_event("campaign", {"success": False, "error": "活動不存在"})
            
        except Exception as e:
            self.send_event("campaign", {"success": False, "error": str(e)})
    
    async def handle_start_campaign(self, payload: Dict[str, Any]):
        """啟動營銷活動"""
        try:
            orchestrator = get_campaign_orchestrator()
            if not orchestrator:
                self.send_event("campaign-started", {"success": False, "error": "營銷活動系統未初始化"})
                return
            
            result = await orchestrator.start_campaign(payload.get('campaignId'))
            
            # Event is sent by orchestrator
            
        except Exception as e:
            self.send_event("campaign-started", {"success": False, "error": str(e)})
    
    async def handle_pause_campaign(self, payload: Dict[str, Any]):
        """暫停營銷活動"""
        try:
            orchestrator = get_campaign_orchestrator()
            if not orchestrator:
                self.send_event("campaign-paused", {"success": False, "error": "營銷活動系統未初始化"})
                return
            
            result = await orchestrator.pause_campaign(payload.get('campaignId'))
            
            self.send_event("campaign-paused", result)
            
        except Exception as e:
            self.send_event("campaign-paused", {"success": False, "error": str(e)})
    
    async def handle_resume_campaign(self, payload: Dict[str, Any]):
        """恢復營銷活動"""
        try:
            orchestrator = get_campaign_orchestrator()
            if not orchestrator:
                self.send_event("campaign-resumed", {"success": False, "error": "營銷活動系統未初始化"})
                return
            
            result = await orchestrator.resume_campaign(payload.get('campaignId'))
            
            # Event is sent by start_campaign
            
        except Exception as e:
            self.send_event("campaign-resumed", {"success": False, "error": str(e)})
    
    async def handle_stop_campaign(self, payload: Dict[str, Any]):
        """停止營銷活動"""
        try:
            orchestrator = get_campaign_orchestrator()
            if not orchestrator:
                self.send_event("campaign-stopped", {"success": False, "error": "營銷活動系統未初始化"})
                return
            
            result = await orchestrator.stop_campaign(payload.get('campaignId'))
            
            # Event is sent by orchestrator
            
        except Exception as e:
            self.send_event("campaign-stopped", {"success": False, "error": str(e)})
    
    async def handle_get_campaign_logs(self, payload: Dict[str, Any]):
        """獲取營銷活動日誌"""
        try:
            orchestrator = get_campaign_orchestrator()
            if not orchestrator:
                self.send_event("campaign-logs", {"success": False, "error": "營銷活動系統未初始化"})
                return
            
            result = await orchestrator.get_campaign_logs(
                campaign_id=payload.get('campaignId'),
                limit=payload.get('limit', 50)
            )
            
            self.send_event("campaign-logs", result)
            
        except Exception as e:
            self.send_event("campaign-logs", {"success": False, "error": str(e)})
    
    async def handle_get_unified_overview(self, payload: Dict[str, Any]):
        """獲取統一概覽"""
        try:
            stats = get_multi_channel_stats()
            if not stats:
                self.send_event("unified-overview", {"success": False, "error": "統計系統未初始化"})
                return
            
            result = await stats.get_unified_overview(
                days=payload.get('days', 7)
            )
            
            self.send_event("unified-overview", result)
            
        except Exception as e:
            self.send_event("unified-overview", {"success": False, "error": str(e)})
    
    async def handle_get_daily_trends(self, payload: Dict[str, Any]):
        """獲取每日趨勢"""
        try:
            stats = get_multi_channel_stats()
            if not stats:
                self.send_event("daily-trends", {"success": False, "error": "統計系統未初始化"})
                return
            
            result = await stats.get_daily_trends(
                days=payload.get('days', 30)
            )
            
            self.send_event("daily-trends", result)
            
        except Exception as e:
            self.send_event("daily-trends", {"success": False, "error": str(e)})
    
    async def handle_get_channel_performance(self, payload: Dict[str, Any]):
        """獲取渠道效能"""
        try:
            stats = get_multi_channel_stats()
            if not stats:
                self.send_event("channel-performance", {"success": False, "error": "統計系統未初始化"})
                return
            
            result = await stats.get_channel_performance()
            
            self.send_event("channel-performance", result)
            
        except Exception as e:
            self.send_event("channel-performance", {"success": False, "error": str(e)})
    
    async def handle_get_funnel_analysis(self, payload: Dict[str, Any]):
        """獲取漏斗分析"""
        try:
            stats = get_multi_channel_stats()
            if not stats:
                self.send_event("funnel-analysis", {"success": False, "error": "統計系統未初始化"})
                return
            
            result = await stats.get_funnel_analysis()
            
            self.send_event("funnel-analysis", result)
            
        except Exception as e:
            self.send_event("funnel-analysis", {"success": False, "error": str(e)})
    
    # ==================== Multi-Role Handlers (多角色協作) ====================
    
    async def handle_get_role_templates(self, payload: Dict[str, Any]):
        """獲取角色模板"""
        try:
            role_manager = get_multi_role_manager()
            if not role_manager:
                self.send_event("role-templates", {"success": False, "error": "角色管理器未初始化"})
                return
            
            templates = role_manager.get_role_templates()
            self.send_event("role-templates", {"success": True, "templates": templates})
            
        except Exception as e:
            self.send_event("role-templates", {"success": False, "error": str(e)})
    
    async def handle_assign_role(self, payload: Dict[str, Any]):
        """分配角色"""
        try:
            role_manager = get_multi_role_manager()
            if not role_manager:
                self.send_event("role-assigned", {"success": False, "error": "角色管理器未初始化"})
                return
            
            result = await role_manager.assign_role(
                account_phone=payload.get('accountPhone'),
                role_type=payload.get('roleType'),
                role_name=payload.get('roleName', ''),
                personality=payload.get('personality'),
                speaking_style=payload.get('speakingStyle'),
                emoji_frequency=payload.get('emojiFrequency'),
                response_speed=payload.get('responseSpeed'),
                custom_prompt=payload.get('customPrompt'),
                bio=payload.get('bio')
            )
            
            if result.get('success'):
                self.send_log(f"已分配角色: {payload.get('roleType')}", "success")
            
            self.send_event("role-assigned", result)
            
        except Exception as e:
            self.send_event("role-assigned", {"success": False, "error": str(e)})
    
    async def handle_update_role(self, payload: Dict[str, Any]):
        """更新角色"""
        try:
            role_manager = get_multi_role_manager()
            if not role_manager:
                self.send_event("role-updated", {"success": False, "error": "角色管理器未初始化"})
                return
            
            result = await role_manager.update_role(
                role_id=payload.get('roleId'),
                updates=payload.get('updates', {})
            )
            
            self.send_event("role-updated", result)
            
        except Exception as e:
            self.send_event("role-updated", {"success": False, "error": str(e)})
    
    async def handle_remove_role(self, payload: Dict[str, Any]):
        """移除角色"""
        try:
            role_manager = get_multi_role_manager()
            if not role_manager:
                self.send_event("role-removed", {"success": False, "error": "角色管理器未初始化"})
                return
            
            result = await role_manager.remove_role(payload.get('roleId'))
            
            self.send_event("role-removed", result)
            
        except Exception as e:
            self.send_event("role-removed", {"success": False, "error": str(e)})
    
    async def handle_get_account_roles(self, payload: Dict[str, Any]):
        """獲取帳號角色"""
        try:
            role_manager = get_multi_role_manager()
            if not role_manager:
                self.send_event("account-roles", {"success": False, "error": "角色管理器未初始化"})
                return
            
            roles = await role_manager.get_account_roles(payload.get('accountPhone'))
            self.send_event("account-roles", {
                "success": True,
                "roles": [r.to_dict() for r in roles]
            })
            
        except Exception as e:
            self.send_event("account-roles", {"success": False, "error": str(e)})
    
    async def handle_get_all_roles(self, payload: Dict[str, Any]):
        """獲取所有角色"""
        try:
            role_manager = get_multi_role_manager()
            if not role_manager:
                self.send_event("all-roles", {"success": False, "error": "角色管理器未初始化"})
                return
            
            result = await role_manager.get_all_roles(
                role_type=payload.get('roleType'),
                active_only=payload.get('activeOnly', True)
            )
            
            self.send_event("all-roles", result)
            
        except Exception as e:
            self.send_event("all-roles", {"success": False, "error": str(e)})
    
    async def handle_get_role_stats(self, payload: Dict[str, Any]):
        """獲取角色統計"""
        try:
            role_manager = get_multi_role_manager()
            if not role_manager:
                self.send_event("role-stats", {"success": False, "error": "角色管理器未初始化"})
                return
            
            result = await role_manager.get_role_stats()
            self.send_event("role-stats", result)
            
        except Exception as e:
            self.send_event("role-stats", {"success": False, "error": str(e)})
    
    async def handle_get_script_templates(self, payload: Dict[str, Any]):
        """獲取劇本模板"""
        try:
            script_engine = get_script_engine()
            if not script_engine:
                self.send_event("script-templates", {"success": False, "error": "劇本引擎未初始化"})
                return
            
            result = await script_engine.get_all_templates(
                scenario=payload.get('scenario'),
                active_only=payload.get('activeOnly', True)
            )
            
            self.send_event("script-templates", result)
            
        except Exception as e:
            self.send_event("script-templates", {"success": False, "error": str(e)})
    
    async def handle_create_script_template(self, payload: Dict[str, Any]):
        """創建劇本模板"""
        try:
            script_engine = get_script_engine()
            if not script_engine:
                self.send_event("script-template-created", {"success": False, "error": "劇本引擎未初始化"})
                return
            
            result = await script_engine.create_template(
                name=payload.get('name'),
                description=payload.get('description', ''),
                scenario=payload.get('scenario', 'custom'),
                stages=payload.get('stages', []),
                required_roles=payload.get('requiredRoles', []),
                min_roles=payload.get('minRoles', 2),
                duration_minutes=payload.get('durationMinutes', 10)
            )
            
            self.send_event("script-template-created", result)
            
        except Exception as e:
            self.send_event("script-template-created", {"success": False, "error": str(e)})
    
    async def handle_delete_script_template(self, payload: Dict[str, Any]):
        """刪除劇本模板"""
        try:
            script_engine = get_script_engine()
            if not script_engine:
                self.send_event("script-template-deleted", {"success": False, "error": "劇本引擎未初始化"})
                return
            
            result = await script_engine.delete_template(payload.get('templateId'))
            self.send_event("script-template-deleted", result)
            
        except Exception as e:
            self.send_event("script-template-deleted", {"success": False, "error": str(e)})
    
    async def handle_start_script_execution(self, payload: Dict[str, Any]):
        """啟動劇本執行"""
        try:
            script_engine = get_script_engine()
            if not script_engine:
                self.send_event("script-execution-created", {"success": False, "error": "劇本引擎未初始化"})
                return
            
            result = await script_engine.start_execution(
                template_id=payload.get('templateId'),
                group_id=payload.get('groupId'),
                target_user_id=payload.get('targetUserId'),
                target_username=payload.get('targetUsername'),
                assigned_roles=payload.get('assignedRoles', {})
            )
            
            self.send_event("script-execution-created", result)
            
        except Exception as e:
            self.send_event("script-execution-created", {"success": False, "error": str(e)})
    
    async def handle_run_script_execution(self, payload: Dict[str, Any]):
        """運行劇本執行"""
        try:
            script_engine = get_script_engine()
            if not script_engine:
                self.send_event("script-execution-started", {"success": False, "error": "劇本引擎未初始化"})
                return
            
            result = await script_engine.run_execution(payload.get('executionId'))
            # Event sent by engine
            
        except Exception as e:
            self.send_event("script-execution-started", {"success": False, "error": str(e)})
    
    async def handle_stop_script_execution(self, payload: Dict[str, Any]):
        """停止劇本執行"""
        try:
            script_engine = get_script_engine()
            if not script_engine:
                self.send_event("script-execution-stopped", {"success": False, "error": "劇本引擎未初始化"})
                return
            
            result = await script_engine.stop_execution(
                execution_id=payload.get('executionId'),
                outcome=payload.get('outcome', 'stopped')
            )
            
        except Exception as e:
            self.send_event("script-execution-stopped", {"success": False, "error": str(e)})
    
    async def handle_get_active_executions(self, payload: Dict[str, Any]):
        """獲取活躍執行"""
        try:
            script_engine = get_script_engine()
            if not script_engine:
                self.send_event("active-executions", {"success": False, "error": "劇本引擎未初始化"})
                return
            
            result = await script_engine.get_active_executions()
            self.send_event("active-executions", result)
            
        except Exception as e:
            self.send_event("active-executions", {"success": False, "error": str(e)})
    
    async def handle_get_execution_stats(self, payload: Dict[str, Any]):
        """獲取執行統計"""
        try:
            script_engine = get_script_engine()
            if not script_engine:
                self.send_event("execution-stats", {"success": False, "error": "劇本引擎未初始化"})
                return
            
            result = await script_engine.get_execution_stats()
            self.send_event("execution-stats", result)
            
        except Exception as e:
            self.send_event("execution-stats", {"success": False, "error": str(e)})
    
    async def handle_create_collab_group(self, payload: Dict[str, Any]):
        """創建協作群組"""
        try:
            coordinator = get_collaboration_coordinator()
            if not coordinator:
                self.send_event("collab-group-created", {"success": False, "error": "協作協調器未初始化"})
                return
            
            result = await coordinator.create_collab_group(
                group_title=payload.get('groupTitle'),
                creator_phone=payload.get('creatorPhone'),
                purpose=payload.get('purpose', 'conversion'),
                target_user_id=payload.get('targetUserId'),
                target_username=payload.get('targetUsername')
            )
            
        except Exception as e:
            self.send_event("collab-group-created", {"success": False, "error": str(e)})
    
    async def handle_add_collab_member(self, payload: Dict[str, Any]):
        """添加協作成員"""
        try:
            coordinator = get_collaboration_coordinator()
            if not coordinator:
                self.send_event("collab-member-added", {"success": False, "error": "協作協調器未初始化"})
                return
            
            result = await coordinator.add_member(
                collab_id=payload.get('collabId'),
                account_phone=payload.get('accountPhone'),
                role_type=payload.get('roleType')
            )
            
        except Exception as e:
            self.send_event("collab-member-added", {"success": False, "error": str(e)})
    
    async def handle_get_collab_groups(self, payload: Dict[str, Any]):
        """獲取協作群組"""
        try:
            coordinator = get_collaboration_coordinator()
            if not coordinator:
                self.send_event("collab-groups", {"success": False, "error": "協作協調器未初始化"})
                return
            
            result = await coordinator.get_all_collab_groups(
                status=payload.get('status'),
                purpose=payload.get('purpose'),
                limit=payload.get('limit', 50)
            )
            
            self.send_event("collab-groups", result)
            
        except Exception as e:
            self.send_event("collab-groups", {"success": False, "error": str(e)})
    
    async def handle_update_collab_status(self, payload: Dict[str, Any]):
        """更新協作狀態"""
        try:
            coordinator = get_collaboration_coordinator()
            if not coordinator:
                self.send_event("collab-group-updated", {"success": False, "error": "協作協調器未初始化"})
                return
            
            result = await coordinator.update_group_status(
                collab_id=payload.get('collabId'),
                status=payload.get('status'),
                outcome=payload.get('outcome')
            )
            
        except Exception as e:
            self.send_event("collab-group-updated", {"success": False, "error": str(e)})
    
    async def handle_get_collab_stats(self, payload: Dict[str, Any]):
        """獲取協作統計"""
        try:
            coordinator = get_collaboration_coordinator()
            if not coordinator:
                self.send_event("collab-stats", {"success": False, "error": "協作協調器未初始化"})
                return
            
            result = await coordinator.get_collab_stats()
            self.send_event("collab-stats", result)
            
        except Exception as e:
            self.send_event("collab-stats", {"success": False, "error": str(e)})
    
    # ==================== Resource Discovery Handlers ====================
    
    async def handle_init_resource_discovery(self):
        """初始化資源發現系統"""
        try:
            self.send_log("🚀 開始初始化資源發現系統...", "info")
            
            # 設置搜索服務（不需要額外初始化數據庫，共用主數據庫）
            group_search_service.set_clients(self.telegram_manager.clients)
            group_search_service.set_event_callback(self.send_event)
            resource_discovery.set_event_callback(self.send_event)
            resource_discovery._initialized = True  # 標記為已初始化
            
            # 統計已連接的帳號
            connected_clients = {phone: client for phone, client in self.telegram_manager.clients.items() if client.is_connected}
            self.send_log(f"📱 找到 {len(connected_clients)} 個已連接的帳號", "info")
            
            self.send_log("✅ 資源發現系統初始化完成", "success")
            self.send_event("resource-discovery-initialized", {"success": True})
            
        except Exception as e:
            import traceback
            self.send_log(f"❌ 資源發現系統初始化失敗: {e}", "error")
            traceback.print_exc()
            self.send_event("resource-discovery-initialized", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_search_resources(self, payload: Dict[str, Any]):
        """搜索群組/頻道資源"""
        try:
            query = payload.get('query', '')
            phone = payload.get('phone')
            limit = payload.get('limit', 50)
            keywords = payload.get('keywords', [])
            search_type = payload.get('searchType', 'all')  # all, group, channel, supergroup
            min_members = payload.get('minMembers', 0)
            language = payload.get('language')
            
            if not query:
                self.send_event("search-resources-complete", {
                    "success": False,
                    "error": "搜索關鍵詞不能為空"
                })
                return
            
            self.send_log(f"🔍 開始搜索: '{query}'", "info")
            
            # 檢查是否有已連接的客戶端
            connected_clients = {p: c for p, c in self.telegram_manager.clients.items() if c.is_connected}
            if not connected_clients:
                self.send_log("⚠️ 沒有已連接的帳號，無法搜索", "warning")
                self.send_event("search-resources-complete", {
                    "success": False,
                    "error": "沒有已連接的帳號，請先登錄帳號"
                })
                return
            
            self.send_log(f"📱 使用 {len(connected_clients)} 個帳號進行搜索", "info")
            
            # 確保客戶端已設置
            group_search_service.set_clients(self.telegram_manager.clients)
            group_search_service.set_event_callback(self.send_event)
            
            # 發送搜索開始事件
            self.send_event("search-progress", {
                "query": query,
                "status": "searching",
                "message": f"正在搜索 '{query}'..."
            })
            
            # 使用超時保護
            try:
                stats = await asyncio.wait_for(
                    group_search_service.search_and_save(
                        query=query,
                        phone=phone,
                        limit=limit,
                        keywords=keywords if keywords else [query],
                        search_type=search_type,
                        min_members=min_members,
                        language=language
                    ),
                    timeout=60.0  # 60秒超時
                )
            except asyncio.TimeoutError:
                self.send_log("⏱️ 搜索超時（60秒）", "warning")
                self.send_event("search-resources-complete", {
                    "success": False,
                    "error": "搜索超時，請稍後再試"
                })
                return
            
            self.send_log(f"🔍 搜索完成: 找到 {stats['found']} 個, 新增 {stats['new']} 個", "success")
            self.send_event("search-resources-complete", {
                "success": True,
                "query": query,
                **stats
            })
            
        except Exception as e:
            self.send_log(f"❌ 搜索資源失敗: {e}", "error")
            import traceback
            traceback.print_exc()
            self.send_event("search-resources-complete", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_get_resources(self, payload: Dict[str, Any]):
        """獲取資源列表"""
        try:
            status = payload.get('status')
            resource_type = payload.get('type')
            limit = payload.get('limit', 50)
            offset = payload.get('offset', 0)
            order_by = payload.get('orderBy', 'overall_score DESC')
            
            resources = await resource_discovery.list_resources(
                status=status,
                resource_type=resource_type,
                limit=limit,
                offset=offset,
                order_by=order_by
            )
            
            total = await resource_discovery.count_resources(status=status, resource_type=resource_type)
            
            self.send_event("resources-list", {
                "success": True,
                "resources": resources,
                "total": total,
                "limit": limit,
                "offset": offset
            })
            
        except Exception as e:
            self.send_log(f"❌ 獲取資源列表失敗: {e}", "error")
            self.send_event("resources-list", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_get_resource_stats(self):
        """獲取資源統計"""
        try:
            stats = await resource_discovery.get_statistics()
            
            self.send_event("resource-stats", {
                "success": True,
                **stats
            })
            
        except Exception as e:
            self.send_log(f"❌ 獲取資源統計失敗: {e}", "error")
            self.send_event("resource-stats", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_add_resource_manually(self, payload: Dict[str, Any]):
        """手動添加資源"""
        try:
            resource = DiscoveredResource(
                resource_type=payload.get('type', 'group'),
                telegram_id=payload.get('telegramId', ''),
                username=payload.get('username', ''),
                title=payload.get('title', ''),
                description=payload.get('description', ''),
                member_count=payload.get('memberCount', 0),
                invite_link=payload.get('inviteLink', ''),
                discovery_source='manual',
                tags=payload.get('tags', []),
                notes=payload.get('notes', '')
            )
            
            # 如果有 username 或邀請鏈接，嘗試獲取詳細信息
            if resource.username or resource.invite_link:
                group_search_service.set_clients(self.telegram_manager.clients)
                chat_id = resource.username or resource.invite_link
                info = await group_search_service.get_chat_info(chat_id)
                if info and 'error' not in info:
                    resource.telegram_id = info.get('telegram_id', resource.telegram_id)
                    resource.title = info.get('title', resource.title)
                    resource.description = info.get('description', resource.description)
                    resource.member_count = info.get('member_count', resource.member_count)
                    resource.is_public = info.get('is_public', True)
                    resource.has_discussion = info.get('has_discussion', False)
            
            resource_id = await resource_discovery.add_resource(resource)
            
            self.send_log(f"➕ 手動添加資源: {resource.title or resource.username}", "success")
            self.send_event("resource-added", {
                "success": True,
                "resourceId": resource_id
            })
            
        except Exception as e:
            self.send_log(f"❌ 添加資源失敗: {e}", "error")
            self.send_event("resource-added", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_delete_resource(self, payload: Dict[str, Any]):
        """刪除資源"""
        try:
            resource_id = payload.get('resourceId')
            if not resource_id:
                raise ValueError("資源 ID 不能為空")
            
            await resource_discovery.delete_resource(resource_id)
            
            self.send_log(f"🗑️ 已刪除資源 ID: {resource_id}", "success")
            self.send_event("resource-deleted", {
                "success": True,
                "resourceId": resource_id
            })
            
        except Exception as e:
            self.send_log(f"❌ 刪除資源失敗: {e}", "error")
            self.send_event("resource-deleted", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_add_to_join_queue(self, payload: Dict[str, Any]):
        """添加資源到加入隊列"""
        try:
            resource_ids = payload.get('resourceIds', [])
            priority = payload.get('priority', 5)
            
            if not resource_ids:
                raise ValueError("資源 ID 列表不能為空")
            
            added = 0
            for rid in resource_ids:
                await resource_discovery.add_to_join_queue(rid, priority)
                added += 1
            
            self.send_log(f"📋 已添加 {added} 個資源到加入隊列", "success")
            self.send_event("join-queue-updated", {
                "success": True,
                "added": added
            })
            
        except Exception as e:
            self.send_log(f"❌ 添加到隊列失敗: {e}", "error")
            self.send_event("join-queue-updated", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_process_join_queue(self, payload: Dict[str, Any]):
        """處理加入隊列"""
        try:
            limit = payload.get('limit', 5)
            
            group_search_service.set_clients(self.telegram_manager.clients)
            stats = await group_search_service.process_join_queue(limit=limit)
            
            self.send_log(f"🚀 處理隊列: 成功 {stats['successCount']}, 失敗 {stats['failed']}", "success")
            self.send_event("join-queue-processed", {
                "success": True,
                **stats
            })
            
        except Exception as e:
            self.send_log(f"❌ 處理隊列失敗: {e}", "error")
            self.send_event("join-queue-processed", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_batch_join_resources(self, payload: Dict[str, Any]):
        """批量加入資源"""
        try:
            resource_ids = payload.get('resourceIds', [])
            delay_min = payload.get('delayMin', 30)
            delay_max = payload.get('delayMax', 60)
            
            if not resource_ids:
                raise ValueError("資源 ID 列表不能為空")
            
            group_search_service.set_clients(self.telegram_manager.clients)
            
            # 在後台運行批量加入
            async def batch_join_task():
                stats = await group_search_service.batch_join(
                    resource_ids=resource_ids,
                    delay_range=(delay_min, delay_max)
                )
                self.send_event("batch-join-complete", {
                    "success": True,
                    **stats
                })
            
            task = asyncio.create_task(batch_join_task())
            self.background_tasks.append(task)
            
            self.send_log(f"🚀 開始批量加入 {len(resource_ids)} 個資源", "info")
            self.send_event("batch-join-started", {
                "success": True,
                "count": len(resource_ids)
            })
            
        except Exception as e:
            self.send_log(f"❌ 批量加入失敗: {e}", "error")
            self.send_event("batch-join-started", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_join_and_monitor_resource(self, payload: Dict[str, Any]):
        """加入群組並添加到監控"""
        try:
            resource_id = payload.get('resourceId')
            username = payload.get('username')
            telegram_id = payload.get('telegramId')
            title = payload.get('title', '')
            
            if not username and not telegram_id:
                raise ValueError("需要 username 或 telegramId")
            
            # 設置客戶端
            group_search_service.set_clients(self.telegram_manager.clients)
            
            # 加入群組
            self.send_log(f"🚀 正在加入: {title}", "info")
            join_result = await group_search_service.join_group(
                username=username,
                resource_id=resource_id
            )
            
            if join_result.get('success'):
                self.send_log(f"✅ 已加入群組: {title}", "success")
                
                # 添加到監控
                self.send_log(f"📊 正在添加監控: {title}", "info")
                
                # 獲取群組信息添加到監控
                from database import db
                await db.connect()
                
                # 檢查是否已在監控列表
                existing = await db.fetch_one(
                    "SELECT id FROM monitored_groups WHERE url LIKE ?",
                    (f"%{username}%" if username else f"%{telegram_id}%",)
                )
                
                if not existing:
                    # 添加到監控群組
                    await db._connection.execute("""
                        INSERT INTO monitored_groups (url, name, keyword_set_ids, is_active, created_at)
                        VALUES (?, ?, '[]', 1, datetime('now'))
                    """, (f"https://t.me/{username}" if username else telegram_id, title))
                    await db._connection.commit()
                    self.send_log(f"✅ 已添加到監控: {title}", "success")
                else:
                    self.send_log(f"ℹ️ 群組已在監控列表中", "info")
                
                self.send_event("join-and-monitor-complete", {
                    "success": True,
                    "resourceId": resource_id,
                    "joined": True,
                    "monitored": True
                })
            else:
                raise Exception(join_result.get('error', '加入失敗'))
                
        except Exception as e:
            self.send_log(f"❌ 加入並監控失敗: {e}", "error")
            self.send_event("join-and-monitor-complete", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_batch_join_and_monitor(self, payload: Dict[str, Any]):
        """批量加入並監控"""
        try:
            resource_ids = payload.get('resourceIds', [])
            
            if not resource_ids:
                raise ValueError("資源 ID 列表不能為空")
            
            self.send_log(f"🚀 開始批量加入並監控 {len(resource_ids)} 個資源", "info")
            
            # 獲取資源詳情
            from database import db
            await db.connect()
            
            success_count = 0
            fail_count = 0
            
            for resource_id in resource_ids:
                try:
                    resource = await db.fetch_one(
                        "SELECT * FROM discovered_resources WHERE id = ?",
                        (resource_id,)
                    )
                    
                    if resource:
                        await self.handle_join_and_monitor_resource({
                            'resourceId': resource_id,
                            'username': resource.get('username'),
                            'telegramId': resource.get('telegram_id'),
                            'title': resource.get('title', '')
                        })
                        success_count += 1
                        
                        # 延遲避免頻率限制
                        import random
                        await asyncio.sleep(random.uniform(30, 60))
                        
                except Exception as e:
                    self.send_log(f"❌ 處理資源 {resource_id} 失敗: {e}", "error")
                    fail_count += 1
            
            self.send_log(f"✅ 批量加入監控完成: 成功 {success_count}, 失敗 {fail_count}", "success")
            self.send_event("batch-join-and-monitor-complete", {
                "success": True,
                "total": len(resource_ids),
                "successCount": success_count,
                "failed": fail_count
            })
            
        except Exception as e:
            self.send_log(f"❌ 批量加入並監控失敗: {e}", "error")
            self.send_event("batch-join-and-monitor-complete", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_analyze_group_link(self, payload: Dict[str, Any]):
        """分析群組鏈接"""
        try:
            link = payload.get('link', '').strip()
            
            if not link:
                raise ValueError("鏈接不能為空")
            
            self.send_log(f"🔍 正在分析鏈接: {link}", "info")
            
            # 解析鏈接
            import re
            username = None
            
            # 匹配 t.me/username 或 @username
            patterns = [
                r't\.me/([a-zA-Z0-9_]+)',
                r'@([a-zA-Z0-9_]+)',
                r'^([a-zA-Z0-9_]+)$'
            ]
            
            for pattern in patterns:
                match = re.search(pattern, link)
                if match:
                    username = match.group(1)
                    break
            
            if not username:
                # 可能是私有鏈接
                if 't.me/+' in link or 't.me/joinchat' in link:
                    self.send_event("link-analysis-complete", {
                        "success": True,
                        "isPrivate": True,
                        "link": link,
                        "message": "這是私有邀請鏈接，需要加入後才能分析"
                    })
                    return
                else:
                    raise ValueError("無法解析鏈接格式")
            
            # 獲取群組信息
            group_search_service.set_clients(self.telegram_manager.clients)
            phone, client = group_search_service._get_available_client()
            
            if not client:
                raise ValueError("沒有可用的帳號")
            
            from pyrogram.enums import ChatType
            
            chat = await client.get_chat(username)
            
            if not chat:
                raise ValueError("找不到該群組")
            
            # 基礎信息
            analysis = {
                "success": True,
                "isPrivate": False,
                "basic": {
                    "id": str(chat.id),
                    "title": chat.title or chat.first_name or username,
                    "username": chat.username or "",
                    "type": str(chat.type.name) if chat.type else "unknown",
                    "memberCount": chat.members_count or 0,
                    "description": chat.description or "",
                    "isPublic": bool(chat.username),
                    "createdAt": chat.date.isoformat() if chat.date else None
                }
            }
            
            self.send_log(f"✅ 分析完成: {chat.title or username}", "success")
            self.send_event("link-analysis-complete", analysis)
            
        except Exception as e:
            self.send_log(f"❌ 分析鏈接失敗: {e}", "error")
            self.send_event("link-analysis-complete", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_get_ollama_models(self, payload: Dict[str, Any]):
        """獲取 Ollama 可用模型列表"""
        try:
            import aiohttp
            
            endpoint = payload.get('endpoint', 'http://localhost:11434')
            
            async with aiohttp.ClientSession() as session:
                async with session.get(f"{endpoint}/api/tags", timeout=aiohttp.ClientTimeout(total=10)) as resp:
                    if resp.status == 200:
                        data = await resp.json()
                        models = [model['name'] for model in data.get('models', [])]
                        
                        self.send_log(f"🦙 Ollama 模型列表: {models}", "info")
                        self.send_event("ollama-models", {
                            "success": True,
                            "models": models
                        })
                    else:
                        raise Exception(f"Ollama API 返回 {resp.status}")
                        
        except Exception as e:
            self.send_log(f"❌ 獲取 Ollama 模型失敗: {e}", "error")
            self.send_event("ollama-models", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_test_ollama_connection(self, payload: Dict[str, Any]):
        """測試 Ollama 連接"""
        try:
            import aiohttp
            
            endpoint = payload.get('endpoint', 'http://localhost:11434')
            
            async with aiohttp.ClientSession() as session:
                async with session.get(f"{endpoint}/api/version", timeout=aiohttp.ClientTimeout(total=5)) as resp:
                    if resp.status == 200:
                        data = await resp.json()
                        version = data.get('version', 'unknown')
                        
                        self.send_log(f"🦙 Ollama 連接成功 (版本: {version})", "success")
                        self.send_event("local-ai-test-result", {
                            "success": True,
                            "message": f"Ollama 連接成功 (v{version})"
                        })
                    else:
                        raise Exception(f"連接失敗: HTTP {resp.status}")
                        
        except Exception as e:
            self.send_log(f"❌ Ollama 連接失敗: {e}", "error")
            self.send_event("local-ai-test-result", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_ollama_generate(self, payload: Dict[str, Any]):
        """使用 Ollama 生成文本"""
        try:
            import aiohttp
            
            endpoint = payload.get('endpoint', 'http://localhost:11434')
            model = payload.get('model', 'qwen2:7b')
            prompt = payload.get('prompt', '')
            system = payload.get('system', '')
            
            if not prompt:
                raise ValueError("prompt 不能為空")
            
            request_data = {
                "model": model,
                "prompt": prompt,
                "stream": False
            }
            
            if system:
                request_data["system"] = system
            
            async with aiohttp.ClientSession() as session:
                async with session.post(
                    f"{endpoint}/api/generate",
                    json=request_data,
                    timeout=aiohttp.ClientTimeout(total=60)
                ) as resp:
                    if resp.status == 200:
                        data = await resp.json()
                        response = data.get('response', '')
                        
                        self.send_event("ollama-response", {
                            "success": True,
                            "response": response,
                            "model": model
                        })
                    else:
                        raise Exception(f"生成失敗: HTTP {resp.status}")
                        
        except Exception as e:
            self.send_log(f"❌ Ollama 生成失敗: {e}", "error")
            self.send_event("ollama-response", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_get_discovery_keywords(self):
        """獲取搜索關鍵詞列表"""
        try:
            keywords = await resource_discovery.get_search_keywords()
            
            self.send_event("discovery-keywords", {
                "success": True,
                "keywords": keywords
            })
            
        except Exception as e:
            self.send_log(f"❌ 獲取關鍵詞失敗: {e}", "error")
            self.send_event("discovery-keywords", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_add_discovery_keyword(self, payload: Dict[str, Any]):
        """添加搜索關鍵詞"""
        try:
            keyword = payload.get('keyword', '').strip()
            category = payload.get('category', 'general')
            priority = payload.get('priority', 5)
            
            if not keyword:
                raise ValueError("關鍵詞不能為空")
            
            keyword_id = await resource_discovery.add_search_keyword(keyword, category, priority)
            
            self.send_log(f"➕ 添加搜索關鍵詞: {keyword}", "success")
            self.send_event("discovery-keyword-added", {
                "success": True,
                "keywordId": keyword_id,
                "keyword": keyword
            })
            
        except Exception as e:
            self.send_log(f"❌ 添加關鍵詞失敗: {e}", "error")
            self.send_event("discovery-keyword-added", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_get_discovery_logs(self, payload: Dict[str, Any]):
        """獲取搜索日誌"""
        try:
            limit = payload.get('limit', 50)
            logs = await resource_discovery.get_discovery_logs(limit=limit)
            
            self.send_event("discovery-logs", {
                "success": True,
                "logs": logs
            })
            
        except Exception as e:
            self.send_log(f"❌ 獲取搜索日誌失敗: {e}", "error")
            self.send_event("discovery-logs", {
                "success": False,
                "error": str(e)
            })
    
    # ==================== Discussion Watcher Handlers ====================
    
    async def handle_init_discussion_watcher(self):
        """初始化討論組監控服務"""
        try:
            await discussion_watcher.initialize()
            discussion_watcher.set_clients(self.telegram_manager.clients)
            discussion_watcher.set_event_callback(self.send_event)
            
            # 設置關鍵詞匹配器
            from keyword_matcher import keyword_matcher
            discussion_watcher.set_keyword_matcher(keyword_matcher)
            
            self.send_log("✅ 討論組監控服務初始化完成", "success")
            self.send_event("discussion-watcher-initialized", {"success": True})
            
        except Exception as e:
            self.send_log(f"❌ 討論組監控服務初始化失敗: {e}", "error")
            self.send_event("discussion-watcher-initialized", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_discover_discussion(self, payload: Dict[str, Any]):
        """發現頻道的討論組"""
        try:
            channel_id = payload.get('channelId', '')
            phone = payload.get('phone')
            
            if not channel_id:
                raise ValueError("頻道 ID 不能為空")
            
            discussion_watcher.set_clients(self.telegram_manager.clients)
            discussion = await discussion_watcher.discover_discussion(channel_id, phone)
            
            if discussion:
                self.send_log(f"✅ 發現討論組: {discussion.discussion_title}", "success")
                self.send_event("discussion-discovered", {
                    "success": True,
                    "discussion": {
                        "id": discussion.id,
                        "channel_id": discussion.channel_id,
                        "channel_title": discussion.channel_title,
                        "discussion_id": discussion.discussion_id,
                        "discussion_title": discussion.discussion_title
                    }
                })
            else:
                self.send_event("discussion-discovered", {
                    "success": False,
                    "error": "未找到討論組或頻道無討論區"
                })
            
        except Exception as e:
            self.send_log(f"❌ 發現討論組失敗: {e}", "error")
            self.send_event("discussion-discovered", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_discover_discussions_from_resources(self):
        """從已發現的資源中發現討論組"""
        try:
            discussion_watcher.set_clients(self.telegram_manager.clients)
            discussions = await discussion_watcher.discover_from_resources()
            
            self.send_log(f"✅ 從資源發現了 {len(discussions)} 個討論組", "success")
            self.send_event("discussions-batch-discovered", {
                "success": True,
                "count": len(discussions),
                "discussions": [
                    {
                        "id": d.id,
                        "channel_title": d.channel_title,
                        "discussion_title": d.discussion_title
                    } for d in discussions
                ]
            })
            
        except Exception as e:
            self.send_log(f"❌ 批量發現討論組失敗: {e}", "error")
            self.send_event("discussions-batch-discovered", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_get_channel_discussions(self, payload: Dict[str, Any]):
        """獲取頻道-討論組列表"""
        try:
            active_only = payload.get('activeOnly', True)
            discussions = await discussion_watcher.list_channel_discussions(active_only)
            
            self.send_event("channel-discussions-list", {
                "success": True,
                "discussions": discussions
            })
            
        except Exception as e:
            self.send_log(f"❌ 獲取討論組列表失敗: {e}", "error")
            self.send_event("channel-discussions-list", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_start_discussion_monitoring(self, payload: Dict[str, Any]):
        """開始監控討論組"""
        try:
            discussion_id = payload.get('discussionId', '')
            phone = payload.get('phone')
            
            if not discussion_id:
                raise ValueError("討論組 ID 不能為空")
            
            discussion_watcher.set_clients(self.telegram_manager.clients)
            success = await discussion_watcher.start_monitoring(discussion_id, phone)
            
            if success:
                self.send_log(f"🟢 開始監控討論組: {discussion_id}", "success")
            
            self.send_event("discussion-monitoring-status", {
                "success": success,
                "discussion_id": discussion_id,
                "status": "monitoring" if success else "error"
            })
            
        except Exception as e:
            self.send_log(f"❌ 啟動監控失敗: {e}", "error")
            self.send_event("discussion-monitoring-status", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_stop_discussion_monitoring(self, payload: Dict[str, Any]):
        """停止監控討論組"""
        try:
            discussion_id = payload.get('discussionId', '')
            
            if not discussion_id:
                raise ValueError("討論組 ID 不能為空")
            
            success = await discussion_watcher.stop_monitoring(discussion_id)
            
            if success:
                self.send_log(f"🔴 停止監控討論組: {discussion_id}", "success")
            
            self.send_event("discussion-monitoring-status", {
                "success": success,
                "discussion_id": discussion_id,
                "status": "stopped"
            })
            
        except Exception as e:
            self.send_log(f"❌ 停止監控失敗: {e}", "error")
            self.send_event("discussion-monitoring-status", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_get_discussion_messages(self, payload: Dict[str, Any]):
        """獲取討論組消息"""
        try:
            discussion_id = payload.get('discussionId', '')
            limit = payload.get('limit', 50)
            matched_only = payload.get('matchedOnly', False)
            
            if not discussion_id:
                raise ValueError("討論組 ID 不能為空")
            
            messages = await discussion_watcher.get_discussion_messages(
                discussion_id, limit, matched_only
            )
            
            self.send_event("discussion-messages", {
                "success": True,
                "discussion_id": discussion_id,
                "messages": messages
            })
            
        except Exception as e:
            self.send_log(f"❌ 獲取討論組消息失敗: {e}", "error")
            self.send_event("discussion-messages", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_reply_to_discussion(self, payload: Dict[str, Any]):
        """回復討論組消息"""
        try:
            discussion_id = payload.get('discussionId', '')
            message_id = payload.get('messageId')
            reply_text = payload.get('replyText', '')
            phone = payload.get('phone')
            
            if not discussion_id or not message_id or not reply_text:
                raise ValueError("討論組 ID、消息 ID 和回復內容不能為空")
            
            discussion_watcher.set_clients(self.telegram_manager.clients)
            result = await discussion_watcher.reply_to_message(
                discussion_id, message_id, reply_text, phone
            )
            
            if result['success']:
                self.send_log(f"✅ 已回復消息 {message_id}", "success")
            
            self.send_event("discussion-reply-result", result)
            
        except Exception as e:
            self.send_log(f"❌ 回復失敗: {e}", "error")
            self.send_event("discussion-reply-result", {
                "success": False,
                "error": str(e)
            })
    
    async def handle_get_discussion_stats(self):
        """獲取討論組監控統計"""
        try:
            stats = await discussion_watcher.get_statistics()
            
            self.send_event("discussion-stats", {
                "success": True,
                **stats
            })
            
        except Exception as e:
            self.send_log(f"❌ 獲取統計失敗: {e}", "error")
            self.send_event("discussion-stats", {
                "success": False,
                "error": str(e)
            })

    async def handle_graceful_shutdown(self):
        """Handle graceful shutdown - disconnect all clients and close database"""
        import sys
        print("[Backend] Graceful shutdown initiated...", file=sys.stderr)
        
        try:
            # 1. Stop monitoring if running
            if self.is_monitoring:
                print("[Backend] Stopping monitoring...", file=sys.stderr)
                try:
                    await self.telegram_manager.stop_monitoring()
                    self.is_monitoring = False
                except Exception as e:
                    print(f"[Backend] Error stopping monitoring: {e}", file=sys.stderr)
            
            # 2. Disconnect all Telegram clients
            print("[Backend] Disconnecting all Telegram clients...", file=sys.stderr)
            try:
                await self.telegram_manager.disconnect_all()
            except Exception as e:
                print(f"[Backend] Error disconnecting clients: {e}", file=sys.stderr)
            
            # 3. Stop scheduler
            print("[Backend] Stopping scheduler...", file=sys.stderr)
            try:
                await scheduler.stop()
            except Exception as e:
                print(f"[Backend] Error stopping scheduler: {e}", file=sys.stderr)
            
            # 4. Close database connection
            print("[Backend] Closing database connection...", file=sys.stderr)
            try:
                await db.close()
            except Exception as e:
                print(f"[Backend] Error closing database: {e}", file=sys.stderr)
            
            # 5. Cancel all background tasks
            print("[Backend] Cancelling background tasks...", file=sys.stderr)
            for task in self.background_tasks:
                try:
                    task.cancel()
                except Exception as e:
                    pass
            
            print("[Backend] Graceful shutdown completed", file=sys.stderr)
            
            # Send confirmation and exit
            self.send_event("shutdown-complete", {"success": True})
            
            # Stop the running flag to exit the main loop
            self.running = False
            
        except Exception as e:
            print(f"[Backend] Error during graceful shutdown: {e}", file=sys.stderr)
            import traceback
            traceback.print_exc(file=sys.stderr)
        
        # Exit the process
        print("[Backend] Exiting...", file=sys.stderr)
        sys.exit(0)


async def main():
    """Main entry point"""
    service = BackendService()
    await service.run()


if __name__ == "__main__":
    # Run the async main function
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        pass

